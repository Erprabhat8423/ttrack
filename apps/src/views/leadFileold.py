import sys
import os
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings


def leadReport(request):
    today = date.today()

    with connection.cursor() as cursor:
        query = f"""
        SELECT slb.id,slb.created_by_id, slb.basic_date , slb.company_name ,slb.desk_no , slb.contact_person_name ,  
        slb.mobile_no , slb.email, slb.address ,slb.total_no_of_employee, slb.turnover 
        , slb.core_business_area , slb.phase ,slb.remark, slis.iso_created_id , slis.iso_amount , 
        slis.iso_created_status ,slo.other_production_pitch,slo.software_or_erp , slo.sales_person , slo.visit_date , 
        slo.other_resource ,slo.reminder , slo.remark FROM sp_lead_basic as slb
        Left Join sp_lead_iso_save as slis on slis.created_by_id = slb.created_by_id and slis.status = 1 and slis.last_lead_id = slb.id
        Left Join sp_lead_other as slo on slo.created_by_id = slb.created_by_id and slo.status = 1 and slo.last_lead_id = slis.id
        where slb.status= 1 AND Date(slb.created_at) between {today} and {today}
        """
        cursor.execute(query)
            
        data = cursor.fetchall()   
    context = {}
    context['lead_data'] = data
  
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['page_title'] = "Lead Report"


    template = 'lead-report/lead-report.html'
    return render(request, template, context)


# ajax order list

def ajaxleadReportLists(request):
    context = {}
    start_date              = request.GET['start_date']
    end_date                = request.GET['end_date']
    
    start_date              = datetime.strptime(str(start_date),'%Y-%m-%d').strftime('%Y-%m-%d')
    end_date                = datetime.strptime(end_date, '%Y-%m-%d')
    end_dates               = end_date
    end_dates               = str(end_dates).replace(' 00:00:00', '')
    end_dates                = end_date + timedelta(days=1)
    end_dates               = end_date.strftime('%Y-%m-%d')
    if  str(start_date) != str(end_date):
        end_date  = end_dates
    with connection.cursor() as cursor:
        query = f"""
        SELECT slb.id,slb.created_by_id, slb.basic_date , slb.company_name ,slb.desk_no , slb.contact_person_name ,  
        slb.mobile_no , slb.email, slb.address ,slb.total_no_of_employee, slb.turnover 
        , slb.core_business_area , slb.phase ,slb.remark, slis.iso_created_id , slis.iso_amount , 
        slis.iso_created_status ,slo.other_production_pitch,slo.software_or_erp , slo.sales_person , slo.visit_date , 
        slo.other_resource ,slo.reminder , slo.remark FROM sp_lead_basic as slb
        Left Join sp_lead_iso_save as slis on slis.created_by_id = slb.created_by_id and slis.status = 1 and slis.last_lead_id = slb.id
        Left Join sp_lead_other as slo on slo.created_by_id = slb.created_by_id and slo.status = 1 and slo.last_lead_id = slis.id
        where slb.status= 1 AND Date(slb.created_at) between '{start_date}' and '{end_date}'
        """
        cursor.execute(query)
        data = cursor.fetchall()   
    context['lead_data'] = data
    template = 'lead-report/ajax-lead-report.html'
    return render(request, template, context)


    

def exportLeadReportListToExcel(request, start_date, end_date):
    start_date  = start_dates = start_date
    end_date    = end_dates    = end_date

    start_date              = datetime.strptime(str(start_date),'%Y-%m-%d').strftime('%Y-%m-%d')
    end_date                = datetime.strptime(end_date, '%Y-%m-%d')
    end_dates               = end_date
    end_dates               = str(end_dates).replace(' 00:00:00', '')
    end_date                = end_date + timedelta(days=1)
    end_date                = end_date.strftime('%Y-%m-%d')
    if  str(start_date) != str(end_date):
        end_date  = end_dates
    with connection.cursor() as cursor:
        query = f"""
        SELECT slb.id,slb.created_by_id, slb.basic_date , slb.company_name ,slb.desk_no , slb.contact_person_name ,  
        slb.mobile_no , slb.email, slb.address ,slb.total_no_of_employee, slb.turnover 
        , slb.core_business_area , slb.phase ,slb.remark, slis.iso_created_id , slis.iso_amount , 
        slis.iso_created_status ,slo.other_production_pitch,slo.software_or_erp , slo.sales_person , slo.visit_date , 
        slo.other_resource ,slo.reminder , slo.remark FROM sp_lead_basic as slb
        Left Join sp_lead_iso_save as slis on slis.created_by_id = slb.created_by_id and slis.status = 1 and slis.last_lead_id = slb.id
        Left Join sp_lead_other as slo on slo.created_by_id = slb.created_by_id and slo.status = 1 and slo.last_lead_id = slis.id
        where slb.status= 1 AND Date(slb.created_at) between '{start_date}' and '{end_date}'
        """
        cursor.execute(query)
        data = cursor.fetchall()   
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=lead-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Lead-reports'

    # Define the titles for columns
    columns = []

    columns += ['Basic Date']  
    columns += ['Company Name']  
    columns += ['Desk No.']
    columns += ['Contact Person Name']
    columns += ['Mobile No']
    columns += ['Email']
    columns += ['Address']
    columns += ['Total No. Of Employee']
    columns += ['Turnover']
    columns += ['Phase']
    columns += ['Remark']
    columns += ['ISO Amount']
    columns += ['Other Production Pitch']
    columns += ['Software OR Erp']
    columns += ['Sales Person']
    columns += ['Visit Date']
    columns += ['Other Resource']
    columns += ['Reminder']
    columns += ['Remark']
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    for results in data:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        if results[2]:
            row += [results[2]]
        else:
            row += ['']
        if results[3]:
            row += [results[3]]
        else:
            row += ['']
        if results[4]:
            row += [results[4]]
        else:
            row += ['']
        if results[5]:
            row += [results[5]]
        else:
            row += ['']
        if results[6]:
            row += [results[6]]
        else:
            row += ['']
        if results[7]:
            row += [results[7]]
        else:
            row += ['']
        if results[8]:
            row += [results[8]]
        else:
            row += ['']
        if results[9]:
            row += [results[9]]
        else:
            row += ['']
        if results[10]:
            row += [results[10]]
        else:
            row += ['']
        if results[12] == 1:
            phase = 'Progress'
            row += [phase]
        else:
            phase = 'Success'
            row += [phase]
        if results[13]:
            row += [results[13]]
        else:
            row += ['']
        if results[15]:
            row += [results[15]]
        else:
            row += ['']
        if results[17]:
            row += [results[17]]
        else:
            row += ['']
        if results[18]:
            row += [results[18]]
        else:
            row += ['']
        if results[19]:
            row += [results[19]]
        else:
            row += ['']
        if results[20]:
            row += [results[20]]
        else:
            row += ['']
        if results[21]:
            row += [results[21]]
        else:
            row += ['']
        if results[22]:
            row += [results[22]]
        else:
            row += ['']
        if results[23]:
            row += [results[23]]
        else:
            row += ['']
 
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
    workbook.save(response)

    return response
import json
def getCoreBusLists(request):
    basic_deatail_id = request.GET['basic_deatail_id']
    getCoreBusniessIds = SpLeadBasic.objects.filter(id = basic_deatail_id,status =1).values_list('core_business_area',flat= True)
    converted_list = [int(x.strip(" '")) for x in getCoreBusniessIds[0].strip("[]").split(',')]
    getCoreBusLists = SpCoreBusinessArea.objects.filter(id__in = converted_list)
    context ={}
    context['getCoreBusLists'] = getCoreBusLists
    context['basic_lead'] = "fdtf"
    template = 'lead-report/core-bus-list.html'
    return render(request, template, context)

def getIsoLists(request):
    basic_deatail_id = request.GET['basic_deatail_id']
    created_by_id =  request.GET['created_by_id']
    getIsoIds = SpLeadIsoSave.objects.get(last_lead_id = basic_deatail_id,created_by_id = created_by_id,status= 1)
    if getIsoIds.iso_created_status == 1:

        master_iso_id = SpLeadIso.objects.filter(created_by_id = getIsoIds.created_by_id,last_lead_id = getIsoIds.last_lead_id,status = 1).values_list('master_iso_id',flat=True)
        if master_iso_id:
            IsoList = SpIsoMaster.objects.filter(iso_id__in = master_iso_id)
        else:
            IsoList =[]
    else:
        converted_list = [int(x.strip(" '")) for x in getIsoIds.iso_created_id.strip("[]").split(',')]
        IsoList = SpIsoMaster.objects.filter(id__in = converted_list)
      
    context ={}
    context['IsoList'] = IsoList
    context['basic_lead'] = "fdtf"
    template = 'lead-report/iso-list.html'
    return render(request, template, context)
    
# ----------------------------------------------------------------------------------------------------------------------------------------------
def index(request):
    today = date.today()

    # with connection.cursor() as cursor:
    #     query = f"""
    #     SELECT slb.id,slb.created_by_id, slb.basic_date , slb.company_name ,slb.desk_no , slb.contact_person_name ,  
    #     slb.mobile_no , slb.email, slb.address ,slb.total_no_of_employee, slb.turnover 
    #     , slb.core_business_area , slb.phase ,slb.remark, slis.iso_created_id , slis.iso_amount , 
    #     slis.iso_created_status ,slo.other_production_pitch,slo.software_or_erp , slo.sales_person , slo.visit_date , 
    #     slo.other_resource ,slo.reminder , slo.remark FROM sp_lead_basic as slb
    #     Left Join sp_lead_iso_save as slis on slis.created_by_id = slb.created_by_id and slis.status = 1 and slis.last_lead_id = slb.id
    #     Left Join sp_lead_other as slo on slo.created_by_id = slb.created_by_id and slo.status = 1 and slo.last_lead_id = slis.id
    #     where slb.status= 1 AND Date(slb.created_at) between {today} and {today}
    #     """
    #     cursor.execute(query)
            
    #     data = cursor.fetchall()   
    context                 = {}
    context['today_date']   = today.strftime("%d/%m/%Y")
    context['page_title']   = "Lead Report"
    template = 'lead-report/index.html'
    return render(request, template, context)