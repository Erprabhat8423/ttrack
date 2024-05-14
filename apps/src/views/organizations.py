import sys
import os
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from ..decorators import *


# Create your views here.

# Organizations View
@login_required
@has_par(sub_module_id=1,permission='list')
def index(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    if organization_details:
        departments = SpDepartments.objects.all().filter(organization_id=organization_details.id).order_by('-id')
    else:
        departments = {}
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['page_limit']             = getConfigurationResult('page_limit')
    context['organization_details']   = organization_details
    context['departments']            = departments
    context['page_title']             = "Manage Organizations"

    template = 'role-permission/organizations.html'
    return render(request, template, context)

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
@has_par(sub_module_id=1,permission='export')
def exportToPDF(request, columns):
    column_list = columns.split (",")
    context = {}
    organizations = SpOrganizations.objects.all().values().order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('role-permission/organization_pdf_template.html', {'organizations': organizations, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)+1})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'organizations.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
@has_par(sub_module_id=1,permission='export')
def exportToXlsx(request, columns):
    column_list = columns.split (",")
    organizations = SpOrganizations.objects.all().order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=organizations.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Organizations'
    
    # Define the titles for columns
    columns = []

    if 'org_name' in column_list:
        columns += [ 'Organization Name' ]

    if 'landline_no' in column_list:
        columns += [ 'Landline No.' ]
    
    if 'mobile_no' in column_list:
        columns += [ 'Mobile No.' ] 

    if 'email_id' in column_list:
        columns += [ 'Email Id' ]

        columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for organization in organizations:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'org_name' in column_list:
            row += [ organization.organization_name ]

        if 'landline_no' in column_list:
            row += [ organization.landline_country_code + ' ' + organization.landline_state_code + ' ' + organization.landline_number ]
        
        if 'mobile_no' in column_list:
            row += [ organization.mobile_country_code + ' ' + organization.mobile_number ] 

        if 'email_id' in column_list:
            row += [ organization.email ]       
       
        row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response


@login_required
@has_par(sub_module_id=1,permission='list')
def ajaxOrganizationList(request):
    page = request.GET.get('page')
    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))   
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    organization_details = SpOrganizations.objects.order_by('-id').first()
    
    context = {}
    context['organizations']          = organizations
    context['total_pages']            = total_pages
    context['organization_details']   = organization_details

    template = 'role-permission/ajax-organizations.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=1,permission='list')
def ajaxOrganizationLists(request):
    page = request.GET.get('page')

    organizations = SpOrganizations.objects.all().order_by('-id')
    paginator = Paginator(organizations, getConfigurationResult('page_limit'))

    try:
        organizations = paginator.page(page)
    except PageNotAnInteger:
        organizations = paginator.page(1)
    except EmptyPage:
        organizations = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    
    context = {}
    context['organizations']     = organizations
    context['total_pages']       = total_pages

    template = 'role-permission/ajax-organization-lists.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=1,permission='add')
def addOrganization(request):
    template = 'role-permission/add-organization.html'
    return render(request, template)


@login_required
@has_par(sub_module_id=1,permission='add')
def saveOrganization(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            organization_name = request.POST.get('organization_name')
            
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = ''
            else:
                land_line_code = request.POST.get('land_line_code')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')
            address = request.POST.get('address')
            pincode = request.POST.get('pincode')

            error_count = 0
            if organization_name == '':
                error_count = 1
                error_response['organization_name_error'] = "Please enter organization name"
            if SpOrganizations.objects.filter(organization_name=organization_name).exists():
                error_count = 1
                error_response['organization_name_error'] = "Organization name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpOrganizations.objects.filter(email=email_id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"    
            if address == '':
                error_count = 1
                error_response['address_error'] = "Please enter address"
            if pincode == '':
                error_count = 1
                error_response['pincode_error'] = "Please enter pincode"    
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                organization                        =   SpOrganizations()
                organization.organization_name      =   organization_name
                organization.landline_country_code  =   land_line_code
                organization.landline_state_code    =   phone_code
                organization.landline_number        =   landline_no
                organization.mobile_country_code    =   mobile_code
                organization.mobile_number          =   mobile_no
                organization.email                  =   email_id
                organization.address                =   address
                organization.pincode                =   pincode
                organization.status                 =   1
                organization.save() 

                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'New '+organization_name+' added'
                activity    = organization_name+' added by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                
                saveActivity('Roles & Permission', 'Oragnizations', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
                response['error'] = False
                response['message'] = "Record has been saved successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
@has_par(sub_module_id=1,permission='edit')
def editOrganization(request):
    id = request.GET.get('id')

    context = {}
    context['organization']     = SpOrganizations.objects.get(id=id)

    template = 'role-permission/edit-organization.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=1,permission='list')
def updateOrganization(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            organization_name = request.POST.get('organization_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
            else:
                land_line_code = request.POST.get('land_line_code')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')
            address = request.POST.get('address')
            pincode = request.POST.get('pincode')

            error_count = 0
            if organization_name == '':
                error_count = 1
                error_response['organization_name_error'] = "Please enter organization name"
            if SpOrganizations.objects.filter(organization_name=organization_name).exclude(id=id).exists():
                error_count = 1
                error_response['organization_name_error'] = "Organization name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpOrganizations.objects.filter(email=email_id).exclude(id=id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"    
            if address == '':
                error_count = 1
                error_response['address_error'] = "Please enter address"
            if pincode == '':
                error_count = 1
                error_response['pincode_error'] = "Please enter pincode"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                data = SpOrganizations.objects.get(id=id)
                data.organization_name = organization_name
                data.landline_country_code = land_line_code
                data.landline_state_code = phone_code
                data.landline_number = landline_no
                data.mobile_country_code = mobile_code
                data.mobile_number = mobile_no
                data.email = email_id
                data.address = address
                data.pincode = pincode
                data.save()

                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = organization_name+' updated'
                activity    = organization_name+' updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                
                saveActivity('Roles & Permission', 'Oragnizations', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
                response['error'] = False
                response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
@has_par(sub_module_id=1,permission='view')
def getOrganizationRecord(request):
    id = request.GET.get('id')
    
    context = {}
    context['organization_details']     = SpOrganizations.objects.get(id=id)
    context['departments']              = SpDepartments.objects.all().filter(organization_id=id).order_by('-id')

    template = 'role-permission/get-organization-record.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=1,permission='delete')
def updateOrganizationStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpOrganizations.objects.get(id=id)
            data.status = is_active
            data.save()
            
            try:
                department_data = SpDepartments.objects.filter(organization_id=id)
            except SpDepartments.DoesNotExist:
                department_data = None
                
            if department_data:
                for department in department_data:
                    department_data = SpDepartments.objects.get(id=department.id)
                    department_data.status = is_active
                    department_data.save()

            try:
                user_data = SpUsers.objects.filter(organization_id=id)
            except SpUsers.DoesNotExist:
                user_data = None

            if user_data:
                for user in user_data:
                    AuthtokenToken.objects.filter(user_id=user.id).delete()
                    user_data = SpUsers.objects.get(id=user.id)
                    user_data.status = is_active
                    user_data.save()

            if is_active == '1':
                status = 'Unblock'
            else:
                status = 'Block'

            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = getModelColumnById(SpOrganizations, id, 'organization_name')+' '+status
            activity    = getModelColumnById(SpOrganizations, id, 'organization_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Roles & Permission', 'Oragnizations', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
@has_par(sub_module_id=1,permission='add')
def addDepartment(request):
    context = {}
    context['organization_id']          = request.GET.get('organization_id')
    context['organization_name']        = request.GET.get('organization_name')
    context['landline']                 = request.GET.get('landline')
    context['country_code']             = request.GET.get('country_code')
    context['code']                     = request.GET.get('code')
    context['country_codes']            = SpCountryCodes.objects.filter(status=1)

    template = 'role-permission/add-department.html'
    return render(request, template, context)


@login_required
@has_par(sub_module_id=1,permission='add')
def saveDepartment(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            organization_id = request.POST.get('organization_id')
            department_name = request.POST.get('department_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
                dept_ext = '';
            else:
                land_line_code = request.POST.get('land_line_code')
                dept_ext = request.POST.get('dept_ext')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')

            error_count = 0
            if department_name == '':
                error_count = 1
                error_response['department_name_error'] = "Please enter department name"
            if SpDepartments.objects.filter(department_name=department_name).filter(Q(organization_id = organization_id)).exists():
                error_count = 1
                error_response['department_name_error'] = "Department name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpDepartments.objects.filter(email=email_id).filter(Q(organization_id = organization_id)).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists" 
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                department = SpDepartments()
                department.organization_id=organization_id
                department.organization_name=getModelColumnById(SpOrganizations,organization_id,'organization_name')
                department.department_name=department_name
                department.landline_country_code=land_line_code
                department.landline_state_code=phone_code
                department.landline_number=landline_no
                department.extension_number=dept_ext
                department.mobile_country_code=mobile_code
                department.mobile_number=mobile_no
                department.email=email_id
                department.status=1
                department.save()

                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = department_name+' added'
                activity    = department_name+' added by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                
                saveActivity('Roles & Permission', 'Departments', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
                response['error'] = False
                response['message'] = "Record has been saved successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
@has_par(sub_module_id=1,permission='edit')
def editDepartment(request):
    organization_id = request.GET.get('organization_id')
    organization_name = request.GET.get('organization_name')
    department_name = request.GET.get('department_name')
    id = request.GET.get('id')
    department = SpDepartments.objects.get(id=id)
    return render(request, 'role-permission/edit-department.html', {'organization_id': organization_id, 'department_name': department_name, 'department': department, 'organization_name':organization_name})


@login_required
@has_par(sub_module_id=1,permission='edit')
def updateDepartment(request):

    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            organization_id = request.POST.get('organization_id')
            department_name = request.POST.get('department_name')
            phone_code = request.POST.get('phone_code')
            landline_no = request.POST.get('landline_no')
            if landline_no == '':
                land_line_code = '';
                dept_ext = '';
            else:
                land_line_code = request.POST.get('land_line_code')
                dept_ext = request.POST.get('dept_ext')
            mobile_code = request.POST.get('mobile_code')
            mobile_no = request.POST.get('mobile_no')
            email_id = request.POST.get('email_id')

            error_count = 0
            if department_name == '':
                error_count = 1
                error_response['department_name_error'] = "Please enter department name"
            if SpDepartments.objects.filter(department_name=department_name).filter(Q(organization_id = organization_id)).exclude(id=id).exists():
                error_count = 1
                error_response['department_name_error'] = "Department name already exists"
            if mobile_no == '':
                error_count = 1
                error_response['mobile_no_error'] = "Please enter mobile no."
            if email_id == '':
                error_count = 1
                error_response['email_id_error'] = "Please enter email"
            if SpDepartments.objects.filter(email=email_id).filter(Q(organization_id = organization_id)).exclude(id=id).exists():
                error_count = 1
                error_response['email_id_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
            else:
                data = SpDepartments.objects.get(id=id)
                data.organization_id = organization_id
                data.organization_name = getModelColumnById(SpOrganizations,organization_id,'organization_name')
                data.department_name = department_name
                data.landline_country_code = land_line_code
                data.extension_number = dept_ext
                data.landline_state_code = phone_code
                data.landline_number = landline_no
                data.mobile_country_code = mobile_code
                data.mobile_number = mobile_no
                data.email = email_id
                data.save()

                user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = department_name+' updated'
                activity    = department_name+' updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                
                saveActivity('Roles & Permission', 'Departments', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
                response['error'] = False
                response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')


@login_required
@has_par(sub_module_id=1,permission='delete')
def updateDepartmentStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpDepartments.objects.get(id=id)
            data.status = is_active
            data.save()

            try:
                user_data = SpUsers.objects.filter(department_id=id)
            except SpUsers.DoesNotExist:
                user_data = None

            if user_data:
                for user in user_data:
                    AuthtokenToken.objects.filter(user_id=user.id).delete()
                    user_data = SpUsers.objects.get(id=user.id)
                    user_data.status = is_active
                    user_data.save()

            if is_active == '1':
                status = 'Unblock'
            else:
                status = 'Block'
                
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = getModelColumnById(SpDepartments, id, 'department_name')+' '+status
            activity    = getModelColumnById(SpDepartments, id, 'department_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Roles & Permission', 'Departments', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/organizations')
