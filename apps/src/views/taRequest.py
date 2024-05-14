import sys
import os
import openpyxl
import collections, functools, operator 
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from ..decorators import has_par
import operator
from functools import reduce


@login_required
def index(request):
    page                                = request.GET.get('employee_page')
    today                               = date.today()
    users                                = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')
    tarequest_lists                        = SpTaRequest.objects.filter(created_at__icontains=today.strftime("%Y-%m-%d")).order_by('-id')
    #tarequest_lists                        = SpTaRequest.objects.filter().order_by('-id')
    tarequest_list                        = SpTaRequest.objects.all().count()
    paginator = Paginator(tarequest_lists, getConfigurationResult('page_limit'))
    for request_list in tarequest_lists:
      
        request_list.user_name     = getUserName(request_list.user_id)
    first_order_id                      = tarequest_lists[0].id if len(tarequest_lists)>0 else 0  
    try:
        tarequest_lists = paginator.page(page)
    except PageNotAnInteger:
        tarequest_lists = paginator.page(1)
    except EmptyPage:
        tarequest_lists = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    context = {
        'today_date': today.strftime("%d/%m/%Y"),
        'page_title': "TA Request",
        'tarequest_lists':tarequest_lists,
        'tarequest_list':tarequest_list,
        'first_order_id':first_order_id,
        'users':users,
        'employee_total_pages' : total_pages,

    }
    
    template = 'ta-request/index.html'
    return render(request, template, context)

@login_required
def getTaRequestDetails(request): 
    id                                 = request.GET.get('id')
   
    assets_details                     = SpTaRequest.objects.get(id=id)
    
    assets_details.user_name          = getUserName(assets_details.user_id)
   
    assets_details.visit_from_date        = datetime.strptime(str(assets_details.visit_from_date), '%Y-%m-%d').strftime('%d/%m/%Y') 
    assets_details.visit_to_date        = datetime.strptime(str(assets_details.visit_to_date), '%Y-%m-%d').strftime('%d/%m/%Y')
    assets_details.stay_items        = SpTaRequestDetails.objects.filter(ta_request_id = id, ta_details_type=0)
    assets_details.travelling_history      = SpTaRequestDetails.objects.filter(ta_request_id = id, ta_details_type=1)
    assets_details.food_history      = SpTaRequestDetails.objects.filter(ta_request_id = id, ta_details_type=2)
    assets_details.misc_history      = SpTaRequestDetails.objects.filter(ta_request_id = id, ta_details_type=3)
    emp_sap_id        =  getModelColumnById(SpUsers, assets_details.user_id, 'emp_sap_id')
   
    context = {
        'assets_details':assets_details,
        'emp_sap_id': emp_sap_id,
    }
    template = 'ta-request/get-tarequest-details.html'
    return render(request, template, context)

@login_required
@has_par(sub_module_id=8,permission='list')
def ajaxTARequestLists(request):
    page                = request.GET.get('employee_page')
    customer_id         = request.GET['customer_id']
    
    order_date          = request.GET['order_date']
    tarequest_lists        = SpTaRequest.objects.all().order_by('-id')
    if customer_id:
        tarequest_lists    = tarequest_lists.filter(user_id = customer_id)
    
    if order_date:
        order_date      = datetime.strptime(str(order_date), '%d/%m/%Y').strftime('%Y-%m-%d')
        tarequest_lists    = tarequest_lists.filter(created_at__icontains=order_date)
    for request_list in tarequest_lists:
      
        request_list.user_name     = getUserName(request_list.user_id)
   
    first_order_id                      = tarequest_lists[0].id if len(tarequest_lists)>0 else 0
    paginator = Paginator(tarequest_lists, getConfigurationResult('page_limit'))

    try:
        tarequest_lists = paginator.page(page)
    except PageNotAnInteger:
        tarequest_lists = paginator.page(1)
    except EmptyPage:
        tarequest_lists = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    context = {
        'page_title': "TA Request Management",
        'tarequest_lists':tarequest_lists,
        'first_order_id':first_order_id,
        'page_loading_type': request.GET.get('page_loading_type'),
       
    }
    template = 'ta-request/ajex-taRequest-lists.html'
    return render(request, template, context)

       

@login_required
def exportToXlsx(request, customer_id,  order_date):
   
    assets_lists        = SpTaRequest.objects.all().order_by('-id')
    if customer_id!='0':
        assets_lists    = assets_lists.filter(user_id = customer_id)
    if order_date!='0':
        
        assets_lists    = assets_lists.filter(created_at__icontains=order_date)
    for request_list in assets_lists:
      
        request_list.user_name     = getUserName(request_list.user_id)
    context = {}
    context['assets_lists']             = assets_lists

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=TaRequest.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'TA Request'
    
    # Define the titles for columns
    columns = []

    columns += [ 'Employee Name' ]

    columns += [ 'Visit Place' ]
    
    columns += [ 'Visit From Date' ] 
    columns += [ 'Visit To Date' ] 

    columns += [ 'Total Expenses' ]

    columns += [ 'Company Paid' ] 

    columns += [ 'Balance' ]
    columns += [ 'Status' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(name='Arial Nova Cond Light',size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    
    for order in assets_lists:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
   
        row += [ order.user_name ]
        
        row += [ order.visit_place ]
   
        row += [ order.visit_from_date ] 
        row += [ order.visit_to_date ] 

        row += [ order.total_expenses ]

        row += [ order.company_paid ]
        
        row += [ order.balance ]
        #row += [ order.status ]
        if order.status == 0:
            row.append('pending')
        elif order.status == 1:
            row.append('approved')

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    workbook.save(response)

    return response

@login_required
def changeTARequestStatusRequest(request):
    today     = date.today()
    if request.method == "POST":
        response = {}
        try:
            user        = SpTaRequest.objects.get(id = request.POST['request_assets_id'])
            user.status =  request.POST['assets_status']
            user.save()
            
            user_name                           = getUserName(request.user.id)
            heading                             = f'Request({user_name}) request status is updated'
            activity                            = f'Request({user_name}) request status is updated by {user_name} on ' + datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            print(activity)
            saveActivity('TA Request Management', 'Request status updated', heading, activity, request.user.id, user_name, 'noti.png', '2', 'web.png')
            response['flag'] = True
            response['message'] = "Record has been updated successfully."
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)
    else:
        context = {'request_assets_id': request.GET['id'],'today_date': today.strftime("%d/%m/%Y"),'user':SpTaRequest.objects.get(id = request.GET['id'])}
        template = 'ta-request/change-taRequest-status.html'
        return render(request, template, context)
    
#update user status
@login_required
def updateUserStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            data = SpTaRequest.objects.get(id=id)
            data.status = 1
            data.save()
 
            user_name                           = getUserName(request.user.id)
            heading                             = f'Request({user_name}) request status is updated'
            activity                            = f'Request({user_name}) request status is updated by {user_name} on ' + datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Status has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
        
    return redirect('/ta-request/index')





 