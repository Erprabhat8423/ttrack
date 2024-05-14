import ast
import json
from django.contrib.auth.decorators import login_required

# from apps.src.views.api.teachers import individualVisit
from ..models import *
from utils import *
from datetime import datetime
from django.shortcuts import render
from django.db.models import F
from django.conf import settings
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

@login_required
def schoolVisitMapview(request):

    school_visits = TblSchoolVisit.objects.all()
    total_school_visits = school_visits.count()
    total_students = school_visits.aggregate(Sum('high_school_students'))['high_school_students__sum']

    all_faculty = SpUsers.objects.filter(role_id=1, status = 1).values('id') #.exclude(role_id = 0)

    for _faculty in all_faculty:
        _faculty['username'] = getUserName(_faculty['id'])
        _faculty['total_visits'] = TblSchoolVisit.objects.filter(visited_by = _faculty['id']).count()

    all_coord = TblSchoolVisit.objects.values('id', 'school_name','high_school_students','visited_datetime', 'visited_by', 'latitude', 'longitude')
    for each_coord in all_coord:
        each_coord['username'] = getUserName(each_coord['visited_by'])

    context = {}
    context['users']                = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['page_title']           = "Visits"
    context['total_school_visits']  = total_school_visits
    context['total_students']       = total_students
    context['all_faculty']          = all_faculty
    context['all_coordinates']      = all_coord

    template = 'widgets/user-tracking-report.html'
    return render(request, template, context)

@login_required
def ajaxschoolVisit(request):
    user_id     = request.POST['user_id']
    date_from   = request.POST['date_from']
    date_to     = request.POST['date_to']
    visit_type  = request.POST['visit_type']

    all_faculty = SpUsers.objects.filter(role_id = 1, status = 1).values('id')
    if visit_type == "1":
        all_coord = TblSchoolVisit.objects
    
    elif visit_type == "2":
        all_coord = TblIndividualVisit.objects

    if user_id:
        all_coord = all_coord.filter(visited_by = user_id)
        all_faculty = all_faculty.filter(id = user_id)

    if date_from:
        date_from = datetime.strptime(str(date_from), "%d/%m/%Y").strftime('%Y-%m-%d')
        all_coord = all_coord.filter(visited_datetime__gte = date_from)

    if date_to:
        date_to = datetime.strptime(str(date_to), "%d/%m/%Y") + timedelta(days=1)
        all_coord = all_coord.filter(visited_datetime__lte = date_to)

    if visit_type == "1":
        all_coord = all_coord.values('id','visited_datetime', 'visited_by', 'latitude', 'longitude', name=F('school_name'))
    elif visit_type == "2":
        all_coord = all_coord.values('id', 'visited_datetime', 'visited_by', 'latitude', 'longitude', name=F('student_name'))

    for each_coord in all_coord:
        each_coord['username'] = getUserName(each_coord['visited_by'])

    for _faculty in all_faculty:
        _faculty['username'] = getUserName(_faculty['id'])

        if visit_type == "1":
            _faculty['total_visits'] = TblSchoolVisit.objects.filter(visited_by = _faculty['id'])
        elif visit_type == "2":
            _faculty['total_visits'] = TblIndividualVisit.objects.filter(visited_by = _faculty['id'])

        if date_from:
            _faculty['total_visits'] = _faculty['total_visits'].filter(visited_datetime__gte = date_from)
        if date_to:
            _faculty['total_visits'] = _faculty['total_visits'].filter(visited_datetime__lte = date_to)

        _faculty['total_visits'] = (_faculty['total_visits']).count()

    context = {}
    context['user']                 = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['all_coordinates']      = all_coord
    context['total_school_visits']  = all_coord.count()
    if visit_type == "1":
        context['total_students']   = all_coord.aggregate(Sum('high_school_students'))['high_school_students__sum']
    context['all_faculty']          = all_faculty
    context['visit_type']           = visit_type

    template = 'widgets/ajax-user-tracking.html'

    return render(request, template, context)



@login_required
def schoolVisitDetails(request):
    user_id = request.POST['user_id']
    visit_type = request.POST['visit_type']
    id = request.POST['id']

    context = {}

    BASE_URL = settings.BASE_URL
    if visit_type == "1":
        school_data = TblSchoolVisit.objects.filter(id = id).first()
        # support_staff = ''
        # staff = json.loads(school_data.support_staff)
        # for each_staff in staff:
        #     support_staff += getUserName(each_staff) 
        #     if each_staff == staff[-2]:
        #         support_staff += " & "
        #     elif each_staff != staff[-1]:
        #         support_staff += ", "

        selfie = []
        if school_data.selfie is None or school_data.selfie == "":
            pass
        else:
            all_selfie = ast.literal_eval(school_data.selfie)
            for each_selfie in all_selfie:
                selfie.append(BASE_URL + each_selfie)

        school_contact = TblSchoolContact.objects.filter(school_id = id).values('contact_name', 'contact_number', 'contact_type', 'referred_by')

        if school_data.school_image is None or school_data.school_image == '':
            school_image = (BASE_URL + '/static/img/svg/org.svg')
        else:
            school_image = BASE_URL + school_data.school_image
        
        school_data.name = school_data.school_name
        
        context['school_cont']      = school_contact
        context['selfie']           = selfie
        
    elif visit_type == "2":
        school_data = TblIndividualVisit.objects.filter(id = id).first()

        # selfie = []
        # if school_data.selfie is None or school_data.selfie == "":
        #     pass
        # else:
        #     all_selfie = ast.literal_eval(school_data.selfie)
        #     for each_selfie in all_selfie:
        #         selfie.append(BASE_URL + each_selfie)

        if school_data.student_image is None or school_data.student_image == '':
            school_image = (BASE_URL + '/static/img/svg/profile.svg')
        else:
            school_image = BASE_URL + school_data.student_image

        school_data.name = school_data.student_name

    # context['support_staff']    = support_staff
    context['school_data']      = school_data
    context['visited_by']       = getUserName(school_data.visited_by)
    context['school_image']     = school_image
    context['visit_type']       = visit_type
    

    template = 'widgets/school-visit-details.html'
    return render(request, template, context)


@login_required
def exportSchoolVisits(request, visit_type, user_id, date_from, date_to):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=visit.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font         = Font(name='Calibri', bold=True)
    centered_alignment  = Alignment(horizontal='center')
    border_bottom       = Border(bottom=Side(border_style='medium', color='FF000000'),)
    wrapped_alignment   = Alignment(vertical='top', wrap_text=True)

    # Get active worksheet/tab
    worksheet       = workbook.active
    if visit_type == "1":
        worksheet.title = 'School Visits'
    elif visit_type == "2":
        worksheet.title = 'Individual Visits'

    columns = []
    columns += [ 'S No' ]
    columns += [ 'Name' ]
    columns += [ 'Contact' ]
    if visit_type == "1":
        # columns += [ 'School Name' ]
        # columns += [ 'School Contact' ]
        columns += [ 'High School Strength' ]
    # elif visit_type == "2":
        # columns += [ 'Student Contact' ]
        # columns += [ 'Student Contact' ]

    columns += [ 'Visit Datetime' ]
    columns += [ 'Visited By' ]
    columns += [ 'Address: H.No.' ]
    columns += [ 'Address: Locality' ]
    columns += [ 'Address: Village Name' ]
    columns += [ 'Address: Tehsil Name' ]
    columns += [ 'Address: District Name' ]
    columns += [ 'Address: State Name' ]
    if visit_type == "1":
        columns += [ 'Contact' ]
    elif visit_type == "2":
        columns += [ 'Referred By' ]

    row_num = 1
    counter = 1
    if visit_type == "1":
        school_visit_data = TblSchoolVisit.objects
    elif visit_type == "2":
        school_visit_data = TblIndividualVisit.objects

    if user_id != '0':
        school_visit_data = school_visit_data.filter(visited_by = user_id)
    if date_from != "0":
        date_from = datetime.strptime(str(date_from), "%d-%m-%Y").strftime('%Y-%m-%d')
        school_visit_data = school_visit_data.filter(visited_datetime__gte = date_from)
    if date_to != "0":
        date_to = datetime.strptime(str(date_to), "%d-%m-%Y") + timedelta(days=1)
        school_visit_data = school_visit_data.filter(visited_datetime__lte = date_to)

    if user_id == '0' and date_from == '0' and date_to == '0':
        school_visit_data = school_visit_data.all()

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        if col_num == len(columns):
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 50
            cell.alignment = Alignment(horizontal='center')
        elif col_num == 1:
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 10
            cell.alignment = Alignment(horizontal='center')
        else:
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 25
            cell.alignment = Alignment(horizontal='center')

    # School Visit
    if visit_type == "1":
        for each_visit_data in school_visit_data:
            each_visit_data.name = each_visit_data.school_name
            each_visit_data.contact = each_visit_data.school_contact
            row_num += 1
            all_contacts = TblSchoolContact.objects.filter(school_id = each_visit_data.id).values('contact_name', 'contact_number', 'contact_type', 'is_referred', 'referred_by')
            contact_list = []
            contact_data = ''
            for each_contact in all_contacts:
                contact_data = ''
                contact_data += 'Contact Person: '
                contact_data += each_contact['contact_name']+"("+each_contact['contact_number']+"),"
                contact_data += each_contact['contact_type']
                if each_contact['is_referred'] == 1:
                    contact_data += " (Reference by - "+each_contact['referred_by']+") "

            contact_list.append(contact_data)

            contact_data = ""
            for each_contact in contact_list:
                contact_data += each_contact +"\n"
        
            # Define the data for each cell in the row
            row = [ counter ]
            row += [ each_visit_data.name ]
            row += [ each_visit_data.contact ]
            if visit_type == "1":
                row += [ each_visit_data.high_school_students ]
            row += [ each_visit_data.visited_datetime ]
            row += [ getUserName(each_visit_data.visited_by) ]
            row += [ each_visit_data.address_hno ]
            row += [ each_visit_data.address_locality ]
            row += [ each_visit_data.village_name ]
            row += [ each_visit_data.tehsil_name ]
            row += [ each_visit_data.district_name ]
            row += [ each_visit_data.state_name ]
            row += [ contact_data ]
            counter = counter+1

            for col_num, cell_value in enumerate(row, 1):
                cell            = worksheet.cell(row=row_num, column=col_num)
                cell.value      = cell_value
                cell.alignment  = Alignment(horizontal='center', wrap_text=True)


    # Individual Visit
    elif visit_type == "2":
        for each_visit_data in school_visit_data:
            row_num += 1
            each_visit_data.name            = each_visit_data.student_name
            each_visit_data.contact         = each_visit_data.student_contact
            each_visit_data.village_name    = each_visit_data.address_village
            each_visit_data.tehsil_name     = each_visit_data.address_tehsil
            each_visit_data.district_name   = each_visit_data.address_district
            each_visit_data.state_name      = each_visit_data.address_state
            contact_data = " Reference by - "+each_visit_data.referred_by_teacher +" (Teacher)\n"+each_visit_data.referred_by_student+" (Student)"

            # Define the data for each cell in the row
            row = [ counter ]
            row += [ each_visit_data.name ]
            row += [ each_visit_data.contact ]
            row += [ each_visit_data.visited_datetime ]
            row += [ getUserName(each_visit_data.visited_by) ]
            row += [ each_visit_data.address_hno ]
            row += [ each_visit_data.address_locality ]
            row += [ each_visit_data.village_name ]
            row += [ each_visit_data.tehsil_name ]
            row += [ each_visit_data.district_name ]
            row += [ each_visit_data.state_name ]
            row += [ contact_data ]

            counter = counter+1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell            = worksheet.cell(row=row_num, column=col_num)
                cell.value      = cell_value
                cell.alignment  = Alignment(horizontal='center', wrap_text=True)

    workbook.save(response)
    
    return response

