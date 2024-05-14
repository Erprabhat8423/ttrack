
def leaveReport(request):
    today = date.today()
    
    if request.user.role_id == 0:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')

    else:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')
    user_type = SpPermissionWorkflowRoles.objects.filter(sub_module_id=49,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context = {}

    user_ids = SpUserLeaves.objects.all().distinct().values_list('user_id',flat=True)
    users = SpUsers.objects.filter(id__in=user_ids).values('id','first_name','middle_name','last_name','emp_sap_id')

    context['leaveReport'] = leaveReport
    context['role_id'] = request.user.role_id
    context['level_id'] = level_id
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['page_title'] = "Leaves Report"
    context['users'] = users

    template = 'holidays/attendance-report/leave-report.html'
    return render(request, template, context)


# ajax order list

def ajaxLeaveReportLists(request):
    context = {}
    today = date.today()
    user_id = request.GET['user_id']
    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            context['leave_status'] = leave_status
            leaveReport = leaveReport.filter(leave_status=leave_status).order_by('-id')

        # if 'leave_from_date' in request.GET and request.GET['leave_from_date'] and request.GET['leave_to_date']:
        #     leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
        #     leave_to_date = datetime.strptime(request.GET['leave_to_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
            
        #     leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        # if request.GET['leave_from_date']:
        #     if request.GET['leave_to_date'] in today.strftime("%d/%m/%Y"):
        #         leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
                
        #         leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        if user_id:
                leaveReport = leaveReport.filter(user_id=user_id).order_by('-id')
    else:
        condition = ''
        
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status
        if user_id:
            condition += ' and sp_user_leaves.user_id = "%s"' % user_id
            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    

    user_type = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add',
                                                             workflow_level_role_id=request.user.role_id).exclude(
            role_id=request.user.role_id).values('level_id').order_by('-id').first()
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context['leaveReport'] = leaveReport
    context['level_id'] = level_id

    context['role_id'] = request.user.role_id
    template = 'holidays/attendance-report/ajax-leave-report-lists.html'
    return render(request, template, context)


def leaveReport(request):
    today = date.today()
    
    if request.user.role_id == 0:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')
        for report in leaveReport:
            report.handover_name = getUserName(report.handover_user_id)
            documents = SpUserLeaveDocument.objects.filter(user_leave_id = report.id)
            for document in documents:
                document.doc_name = getModelColumnById(SpLeaveTypeDocuments,document.leave_type_document_id,'document')
            report.documents = documents
    
    else:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')
    #     leaveReport = SpUserLeaves.objects.raw('''SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    # FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
    # left join sp_users on sp_users.id = sp_user_leaves.user_id 
    # where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' order by id desc ''',[request.user.id])
        for report in leaveReport:
            report.handover_name = getUserName(report.handover_user_id)
            documents = SpUserLeaveDocument.objects.filter(user_leave_id = report.id)
            for document in documents:
                document.doc_name = getModelColumnById(SpLeaveTypeDocuments,document.leave_type_document_id,'document')
            report.documents = documents

    user_type = SpPermissionWorkflowRoles.objects.filter(sub_module_id=49,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context = {}

    user_ids = SpUserLeaves.objects.all().distinct().values_list('user_id',flat=True)
    users = SpUsers.objects.filter(id__in=user_ids).values('id','first_name','middle_name','last_name','emp_sap_id')

    context['leaveReport'] = leaveReport
    context['role_id'] = request.user.role_id
    context['level_id'] = level_id
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['page_title'] = "Leaves Report"
    context['users'] = users

    template = 'holidays/attendance-report/leave-report.html'
    return render(request, template, context)


# ajax order list

def ajaxLeaveReportLists(request):
    context = {}
    today = date.today()
    user_id = request.GET['user_id']
    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            context['leave_status'] = leave_status
            leaveReport = leaveReport.filter(leave_status=leave_status).order_by('-id')

        # if 'leave_from_date' in request.GET and request.GET['leave_from_date'] and request.GET['leave_to_date']:
        #     leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
        #     leave_to_date = datetime.strptime(request.GET['leave_to_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
            
        #     leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        # if request.GET['leave_from_date']:
        #     if request.GET['leave_to_date'] in today.strftime("%d/%m/%Y"):
        #         leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
                
        #         leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        if user_id:
                leaveReport = leaveReport.filter(user_id=user_id).order_by('-id')
        for report in leaveReport:
            report.handover_name = getUserName(report.handover_user_id)
            documents = SpUserLeaveDocument.objects.filter(user_leave_id = report.id)
            for document in documents:
                document.doc_name = getModelColumnById(SpLeaveTypeDocuments,document.leave_type_document_id,'document')
            report.documents = documents
    else:
        condition = ''
        
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status
        if user_id:
            condition += ' and sp_user_leaves.user_id = "%s"' % user_id
            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)
        for report in leaveReport:
            report.handover_name = getUserName(report.handover_user_id)
            documents = SpUserLeaveDocument.objects.filter(user_leave_id = report.id)
            for document in documents:
                document.doc_name = getModelColumnById(SpLeaveTypeDocuments,document.leave_type_document_id,'document')
            report.documents = documents
        
    

    user_type = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add',
                                                             workflow_level_role_id=request.user.role_id).exclude(
            role_id=request.user.role_id).values('level_id').order_by('-id').first()
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context['leaveReport'] = leaveReport
    context['level_id'] = level_id

    context['role_id'] = request.user.role_id
    template = 'holidays/attendance-report/ajax-leave-report-lists.html'
    return render(request, template, context)



def leaveStatusDetails(request):
    leave_id = request.GET.get('leave_id')
    initiate_leave_details = SpUserLeaves.objects.get(id=leave_id)
    leave_details = SpApprovalStatus.objects.filter(row_id=leave_id, model_name='SpUserLeaves', status=1).values(
        'final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')

    context = {}
    context['initiate_leave_details'] = initiate_leave_details
    context['leave_details'] = leave_details
    template = 'holidays/attendance-report/leave-status-details.html'

    return render(request, template, context)



def leaveExportToXlsx(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=leave-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Leave-reports'

    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += ['Name']
        
    if 'leave_from_date' in column_list:
        columns += ['Leave Apply From Date']

    

    if 'leave_to_date' in column_list:
        columns += ['Leave Apply To Date']

    if 'status' in column_list:
        columns += ['Status']

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    for results in leaveReport:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        if 'user_name' in column_list:
            row += [results.user_name]
            
        

        if 'leave_from_date' in column_list:
            if None in [results.leave_from_date]:
                leave_from_date = ['']
            else:
                leave_from_date = [results.leave_from_date]
            row += leave_from_date

        if 'leave_to_date' in column_list:
            if None in [results.leave_to_date]:
                leave_to_date = ['']
            else:
                leave_to_date = [results.leave_to_date]
            row += leave_to_date

        if 'status' in column_list:
            if results.leave_status == 1:
                status = 'Pending'
                row += [status]
            elif results.leave_status == 2:
                status = 'Forwarded'
                row += [status]
            elif results.leave_status == 3:
                status = 'Approved'
                row += [status]
            elif results.leave_status == 4:
                status = 'Declined'
                row += [status]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
    workbook.save(response)

    return response

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


# Automaticly downloads to PDF file

def leaveExportToPDF(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)
    
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('holidays/attendance-report/leave_pdf_template.html',
                        {'leaveReport': leaveReport, 'url': baseurl, 'columns': column_list,
                         'columns_length': len(column_list)})
    
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'leave-report.pdf'
    content = "attachment; filename=%s" % filename
    response['Content-Disposition'] = content
    return response



def editLeaveStatus(request):
    id = request.GET.get('id')
    context = {}
    context['leaveData'] = SpUserLeaves.objects.get(id=id)
    template = 'holidays/attendance-report/edit-leave-status.html'
    return render(request, template, context)


def updateLeaveRemark(request): 
    context = {}
    context['level_id']     = request.GET.get('level_id')
    context['leave_status'] = request.GET.get('leave_status')
    template = 'holidays/attendance-report/update-leave-remark.html'
    return render(request, template, context)

#update order status

def updateLeaveStatus(request):
    response = {}
    leave_id        = request.POST.getlist('leave_id[]')
    level_id        = request.POST['level_id']
    leave_status    = request.POST['leave_status']
    
    
    if request.user.role_id == 0:
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.level_id                  = leave_status
                    if leave_status == '2':
                        approval_data.level                    = 'Forward'
                    elif leave_status == '3':
                        approval_data.level                    = 'Approve'
                        
                    elif leave_status == '4':
                        approval_data.level                    = 'Declined'         
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    leave                   = SpUserLeaves.objects.get(id=leave)   
                    
                    leave.leave_status      = leave_status
                    if request.POST['remark']:
                        leave.remark      = request.POST['remark']
                    leave.save()
                    if leave_status == '3':
                        last_leave_type_value = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave.leave_type_id,user_id = leave.user_id).last()
                        last_leave_balance = SpUserLeavePolicyLedger.objects.filter(user_id = leave.user_id).last()
                        
                        leave_from_date             = leave.leave_from_date
                        leave_to_date               = leave.leave_to_date
                        total_no_of_leave_days      = leave_to_date - leave_from_date
                        total_no_of_leave_days      = total_no_of_leave_days.days 
                        total_no_of_leave_days      = total_no_of_leave_days+1
                        if leave.is_first_half_day == 1:
                            total_no_of_leave_days = total_no_of_leave_days - 0.5
                        if leave.is_last_half_day == 1:
                            total_no_of_leave_days = total_no_of_leave_days - 0.5
                        
                        last_leave_type_values  = float(last_leave_type_value.year_leave_count) - total_no_of_leave_days
                        last_leave_balances     = float(last_leave_balance.balance) - total_no_of_leave_days
                        
                        leave_ledger                            = SpUserLeavePolicyLedger()
                        leave_ledger.user_id                    = leave.user_id
                        leave_ledger.leave_policy_id            = last_leave_type_value.leave_policy_id
                        leave_ledger.leave_type_id              = last_leave_type_value.leave_type_id
                        leave_ledger.year_leave_count           = last_leave_type_values
                        leave_ledger.month_leave_count          = last_leave_type_value.month_leave_count
                        leave_ledger.consecutive_leave          = last_leave_type_value.consecutive_leave
                        leave_ledger.debit                      = total_no_of_leave_days
                        leave_ledger.balance                    = last_leave_balances
                        leave_ledger.save()
                        
                        
                        
            else:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                leave.save()
                today   = date.today()
                
                
                if leave_status == '3':
                    last_leave_type_value = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave.leave_type_id,user_id = leave.user_id).last()
                    last_leave_balance = SpUserLeavePolicyLedger.objects.filter(user_id = leave.user_id).last()
                    
                    leave_from_date             = leave.leave_from_date
                    leave_to_date               = leave.leave_to_date
                    total_no_of_leave_days      = leave_to_date - leave_from_date
                    total_no_of_leave_days      = total_no_of_leave_days.days 
                    total_no_of_leave_days      = total_no_of_leave_days+1
                    if leave.is_first_half_day == 1:
                        total_no_of_leave_days = total_no_of_leave_days - 0.5
                    if leave.is_last_half_day == 1:
                        total_no_of_leave_days = total_no_of_leave_days - 0.5
                    
                    last_leave_type_values  = float(last_leave_type_value.year_leave_count) - total_no_of_leave_days
                    last_leave_balances     = float(last_leave_balance.balance) - total_no_of_leave_days
                    
                    leave_ledger                            = SpUserLeavePolicyLedger()
                    leave_ledger.user_id                    = leave.user_id
                    leave_ledger.leave_policy_id            = last_leave_type_value.leave_policy_id
                    leave_ledger.leave_type_id              = last_leave_type_value.leave_type_id
                    leave_ledger.year_leave_count           = last_leave_type_values
                    leave_ledger.month_leave_count          = last_leave_type_value.month_leave_count
                    leave_ledger.consecutive_leave          = last_leave_type_value.consecutive_leave
                    leave_ledger.debit                      = total_no_of_leave_days
                    leave_ledger.balance                    = last_leave_balances
                    leave_ledger.save()
                        
                 
    
    else:    
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', role_id=request.user.role_id, level_id=level_id)
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.save()   
                
                if leave_status == '3':
                    last_leave_type_value = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave.leave_type_id,user_id = leave.user_id).last()
                    last_leave_balance = SpUserLeavePolicyLedger.objects.filter(user_id = leave.user_id).last()
                    
                    leave_from_date             = leave.leave_from_date
                    leave_to_date               = leave.leave_to_date
                    total_no_of_leave_days      = leave_to_date - leave_from_date
                    total_no_of_leave_days      = total_no_of_leave_days.days 
                    total_no_of_leave_days      = total_no_of_leave_days+1
                    if leave.is_first_half_day == 1:
                        total_no_of_leave_days = total_no_of_leave_days - 0.5
                    if leave.is_last_half_day == 1:
                        total_no_of_leave_days = total_no_of_leave_days - 0.5
                    
                    last_leave_type_values  = float(last_leave_type_value.year_leave_count) - total_no_of_leave_days
                    last_leave_balances     = float(last_leave_balance.balance) - total_no_of_leave_days
                    
                    leave_ledger                            = SpUserLeavePolicyLedger()
                    leave_ledger.user_id                    = leave.user_id
                    leave_ledger.leave_policy_id            = last_leave_type_value.leave_policy_id
                    leave_ledger.leave_type_id              = last_leave_type_value.leave_type_id
                    leave_ledger.year_leave_count           = last_leave_type_values
                    leave_ledger.month_leave_count          = last_leave_type_value.month_leave_count
                    leave_ledger.consecutive_leave          = last_leave_type_value.consecutive_leave
                    leave_ledger.debit                      = total_no_of_leave_days
                    leave_ledger.balance                    = last_leave_balances
                    leave_ledger.save()
                        

    
    if leave_status == '2':
        for leave in leave_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpUserLeaves'
                    notification.notification           = 'leave '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()

    if leave_status == '2':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been forwarded'
            activity    = 'Leave Request has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)
            
            message_title = "Leave request forwarded"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been forwarded  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------# 
    elif leave_status == '3':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been approved'
            activity    = 'Leave Request has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            #update user leave count
            
            user_basic_details              = SpBasicDetails.objects.get(user_id=user_id)
            
           
            # leave_count    = get_user_leave_count(leave)
           
            # leave_count    = float(user_basic_details.leave_count)-float(leave_count)
             
            # user_basic_details.leave_count  = leave_count
            # user_basic_details.save()

            message_title = "Leave request approved"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been approved  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#    

    elif leave_status == '4':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been declined'
            activity    = 'Leave Request has been declined by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'declined.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Leave request declined"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been declined  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request declined',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Leave status has been updated successfully."
    return JsonResponse(response)















































































import sys
import os
import json
from django.core import serializers
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import getConfigurationResult,getModelColumnById,clean_data
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from django.conf import settings
from ..decorators import *
from datetime import datetime
from calendar import monthrange
from datetime import date
import calendar
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.utils.html import strip_tags
from django.forms.models import model_to_dict
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm

@login_required
def index(request):
    page = request.GET.get('page')
    context = {}
    months = [12,11,10,9,8,7,6,5,4,3,2,1]
    month_holidays = []
    for month in months :
        temp = {}
        temp['month'] = month
        temp['month_name'] = calendar.month_name[month]
        holidays = SpHolidays.objects.filter(start_date__month=month).order_by('-id')
        if len(holidays):
            temp['month_holidays'] = holidays
            month_holidays.append(temp)
        else:
            holidays = SpHolidays.objects.filter(end_date__month=month).order_by('-id')
            if len(holidays):
                temp['month_holidays'] = holidays
                month_holidays.append(temp)

    context['total_holidays'] = SpHolidays.objects.all().count()
    current_date = date.today() 
    context['current_year'] = current_year = current_date.year
    context['current_month'] = current_month = current_date.month
    
    if int(current_month) == 12:
        context['next_month'] = 1
    else:
        context['next_month'] = int(current_month) + 1
    if int(current_month) == 1:
        context['previous_month'] = 12
    else:
        context['previous_month'] = int(current_month) - 1
    
    context['current_month_name'] = calendar.month_name[current_month]

    month_holiday_dates = SpHolidays.objects.filter().order_by('id')#.first()
   
    calendarObj = calendar.Calendar()
    calendar_dates = []
    for week in calendarObj.monthdatescalendar(current_year, current_month):
        dates = []
        for week_date in week:
            calendar_datass = []
            tmp = {}
            tmp['full_date'] = week_date
            tmp['day'] = week_date.strftime('%A')
            tmp['short_date'] = week_date.strftime('%d')
            tmp['month'] = week_date.strftime('%m')
            holiday_datas = SpHolidays.objects.filter(start_date__lte=week_date,end_date__gte=week_date,status=1)
            if holiday_datas:
                tmp['is_holiday'] = 1
                store_multi_holidays = []
                for holiday_data in holiday_datas:
                    holiday_temp = {}
                    holiday_temp['holiday'] = holiday_data.holiday
                    holiday_temp['organization_name'] = holiday_data.organization_name
                    if len(store_multi_holidays)<5:
                        store_multi_holidays.append(holiday_temp)
                    tmp['holiday'] = holiday_data.holiday
                    tmp['organization_name'] = holiday_data.organization_name
                    if len(store_multi_holidays)>1:
                        tmp['multi'] = store_multi_holidays
            else:
                tmp['is_holiday'] = 0
            
            dates.append(tmp)               
        calendar_dates.append(dates)
    level = SpRoleWorkflowPermissions.objects.filter(role_id=request.user.role_id,permission_slug='add',sub_module_id=33)
    if level:
        context['level'] = level[0].level_id
        print(level[0].level_id,"level[0].level_id")
    context['calendar_dates'] = calendar_dates
    context['holiday'] = month_holiday_dates
    context['holidays'] = month_holidays
    context['holiday_types'] = SpHolidayTypes.objects.all()
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_title'] = "Manage Holidays"
    template = 'holiday/index.html'
    return render(request, template, context)

@login_required
def filterHoliday(request,filter_status):
    if request.method == 'POST':
        context = {}
        months = [1,2,3,4,5,6,7,8,9,10,11,12]
        month_holidays = []
        for month in months :
            temp = {}
            temp['month'] = month
            temp['month_name'] = calendar.month_name[month]
            if int(filter_status) < 0:
                holidays = SpHolidays.objects.filter(start_date__month=month)
                if len(holidays):
                    temp['month_holidays'] = holidays
                    month_holidays.append(temp)
                else:
                    holidays = SpHolidays.objects.filter(end_date__month=month)
                    if len(holidays):
                        temp['month_holidays'] = holidays
                        month_holidays.append(temp)
            else:
                holidays = SpHolidays.objects.filter(holiday_status=filter_status,start_date__month=month)
                if len(holidays):
                    temp['month_holidays'] = holidays
                    month_holidays.append(temp)
                else:
                    holidays = SpHolidays.objects.filter(holiday_status=filter_status,end_date__month=month)
                    if len(holidays):
                        temp['month_holidays'] = holidays
                        month_holidays.append(temp)


        if int(filter_status) < 0:
            context['total_holidays'] = SpHolidays.objects.all().count()
        else:
            context['total_holidays'] = SpHolidays.objects.filter(holiday_status=filter_status).count()
        current_date = date.today() 
        context['current_year'] = current_year = current_date.year
        context['current_month'] = current_month = current_date.month
        context['holidays'] = month_holidays       
        template = 'holiday/ajax-holiday-filter.html'
        return render(request, template, context)



@login_required
def holidayCalendar(request):
    import datetime
    context = {}
    calendarObj = calendar.Calendar()
    current_date = date.today() 
    action = request.GET.get('action')

    if 'year' in request.GET and int(request.GET.get('year')) != "":
        context['current_year'] = current_year = int(request.GET.get('year'))
    else:
        context['current_year'] = current_year = current_date.year
    
    if 'month' in request.GET and int(request.GET.get('month')) != "":
        context['current_month'] = current_month = int(request.GET.get('month'))
    else:
        context['current_month'] = current_month = current_date.month

    if action == 'next' and int(current_month) == 1:
            context['current_year'] = current_year = int(current_year) + 1

    if action == 'previous' and int(current_month) == 12:
        context['current_year'] = current_year = int(current_year) - 1

    if int(current_month) == 12:
        context['next_month'] = 1

    else:
        context['next_month'] = int(current_month) + 1
        
        
    if int(current_month) == 1:
        context['previous_month'] = 12
    else:
        context['previous_month'] = int(current_month) - 1
    

    context['current_month_name'] = calendar.month_name[current_month]
    holiday_dates = []
    time = datetime.datetime.now()
    currentMonth = time.strftime("%m")
    
    last_holidays = SpHolidays.objects.filter(start_date__icontains=str(request.GET['month'])).order_by('-id')
    
    if last_holidays:
        for last_holiday in last_holidays:
            start_date = last_holiday.start_date
            delta = last_holiday.end_date - last_holiday.start_date
            for i in range(delta.days + 1):
                holiday_date = start_date + timedelta(days=i)
                holiday_date = holiday_date.strftime('%Y-%m-%d')
                holiday_dates.append(holiday_date)
    # print("338 Holidays: ", holiday_dates)
    
    calendarObj = calendar.Calendar()
    calendar_dates = []
    for week in calendarObj.monthdatescalendar(current_year, current_month):
        dates = []
        for week_date in week:
            calendar_datass = []
            tmp = {}
            tmp['full_date'] = week_date
            tmp['day'] = week_date.strftime('%A')
            tmp['short_date'] = week_date.strftime('%d')
            tmp['month'] = week_date.strftime('%m')
            if current_month <9:
                current_mon = "0"+str(current_month)
            else:
                current_mon = str(current_month)
            holiday_datas = SpHolidays.objects.filter(start_date__lte=week_date,end_date__gte=week_date,status=1).order_by('id')   
            if str(currentMonth) == current_mon:                            
                if holiday_datas:
                    tmp['is_holiday'] = 1
                    store_multi_holidays = []
                    for holiday_data in holiday_datas:
                        holiday_temp = {}
                        holiday_temp['holiday'] = holiday_data.holiday
                        holiday_temp['organization_name'] = holiday_data.organization_name
                        if len(store_multi_holidays)<5:
                            store_multi_holidays.append(holiday_temp)
                        tmp['holiday'] = holiday_data.holiday
                        tmp['organization_name'] = holiday_data.organization_name
                        if len(store_multi_holidays)>1:
                            tmp['multi'] = store_multi_holidays
                else:
                    tmp['is_holiday'] = 0
            else:
                if str(week_date) in holiday_dates:
                    tmp['is_holiday'] = 1
                    store_multi_holidays = []
                    for holiday_data in holiday_datas:
                        holiday_temp = {}
                        holiday_temp['holiday'] = holiday_data.holiday
                        holiday_temp['organization_name'] = holiday_data.organization_name
                        if len(store_multi_holidays)<5:
                            store_multi_holidays.append(holiday_temp)
                        tmp['holiday'] = holiday_data.holiday
                        tmp['organization_name'] = holiday_data.organization_name
                        if len(store_multi_holidays)>1:
                            tmp['multi'] = store_multi_holidays
                else:
                    tmp['is_holiday'] = 0
            
            dates.append(tmp)               
        calendar_dates.append(dates)
    context['calendar_dates'] = calendar_dates
    context['page_limit'] = getConfigurationResult('page_limit')
    context['page_title'] = "Manage Holidays"
    template = 'holiday/holiday-calendar.html'
    return render(request, template, context)

@login_required
def addHoliday(request):
    if request.method == "POST":
        context = {}
        holiday_dates = []
        
        if request.POST['filter_org_name[]']:

            for org_id in request.POST.getlist('filter_org_name[]'):
                org_name = getModelColumnById(SpOrganizations,org_id,'organization_name')
                applicable_to = {}
                
                role = request.POST.getlist('filter_role[]')
                role = ','.join(str(e) for e in role)
                holiday = SpHolidays()
                holiday.holiday = clean_data(request.POST['holidayName'])
                holiday.holiday_type_id = request.POST['holiday_type_id'] 
                holiday.holiday_type = getModelColumnById(SpHolidayTypes,request.POST['holiday_type_id'],'holiday_type')                
                holiday.organization_name = org_name
                holiday.organization_id = org_id
                holiday.applicable_to=role
                holiday.start_time = "00:00:00" 
                holiday.start_date = datetime.strptime(clean_data(request.POST['from_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
                holiday.end_date = datetime.strptime(clean_data(request.POST['to_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
                holiday.end_time = "00:00:00"
                holiday.description = clean_data(request.POST['holiday_description'])
                holiday.status = 1
                holiday.holiday_status = 3
                holiday.created_at = datetime.now()
                holiday.updated_at = datetime.now()                
                holiday.save()

                user_name   = request.user.first_name+" "+request.user.middle_name+" "+request.user.last_name#getUserName(request.user.id)

                if request.user.role_id == 0:
                    SpHolidays.objects.filter(id=holiday.id).update(holiday_status=3)
                else:
                    sendNotificationToUsers(holiday.id,clean_data(request.POST['holidayName']), 'add', 0, request.user.id, user_name, 'SpHolidays',request.user.role_id)

                if request.POST['filter_role[]']:
                    role_id_to_mapped = request.POST.getlist('filter_role[]')   
                    if 'all' in role_id_to_mapped:
                        for roleId in SpRoles.objects.filter(organization_id=org_id):
                            mapping =  SpRoleEntityMapping()
                            mapping.entity_type = "Holiday"
                            mapping.role_id = roleId.id
                            mapping.entity_id = holiday.id
                            mapping.save()
                    else:
                        for roleId in SpRoles.objects.filter(organization_id=org_id,id__in=role_id_to_mapped):
                            role_mapping = SpRoleEntityMapping()
                            role_mapping.entity_type = "Holiday" 
                            role_mapping.role_id = roleId.id
                            role_mapping.entity_id = holiday.id
                            role_mapping.save()

                heading     = 'New holiday has been created.'
                activity    = 'New holiday has been created by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('Leave Management', 'Leave Request', heading, activity, request.user.id, user_name, 'add.png', '2', 'web.png')
            context['flag'] = True
            context['message'] = "Record has been save successfully."
            return JsonResponse(context)                
    else:
        context = {}
        if request.user.role_id == 0:
            context['institutions'] = SpOrganizations.objects.all()
        else:            
            context['institutions'] = SpOrganizations.objects.filter(id=request.user.organization_id)

        context['holiday_types'] = SpHolidayTypes.objects.filter(status=1)
        context['semester']=TblSemester.objects.all().filter(type__startswith="semester")
       

        # organisation_id=SpOrganizations.objects.get(id=id)
        # branch=TblBranch.object.get(id=organisation_id)
        # context['branch']=branch

        template = 'holiday/add-holiday.html'
        return render(request, template, context)
    


@login_required
def roleByOrganization(request):
    response = {}
    filter_role_option = ''
    branch_option = ''

    if request.method == "POST":
        if request.POST.getlist('filter_org_id[]'):
            filter_org_id = request.POST.getlist('filter_org_id[]')
            filter_org_id = ','.join(str(e) for e in filter_org_id)
            # filter_role_option += '<option value="all">All Role</option>'  # Add "All Role" option
            
            for role in SpRoles.objects.filter(organization_id__in=filter_org_id):
                filter_role_option += '<option value="'+str(role.id)+'">'+str(role.role_name)+'</option>'
                
            response['filter_role_option'] = filter_role_option
        else:
            # filter_role_option += '<option value="all">All Role</option>'  # Add "All Role" option
            
            for role in SpRoles.objects.filter(organization_id=request.POST['filter_org_id']).filter(role_name__regex=r'[s|S]tudent'):
                filter_role_option += '<option value="'+str(role.id)+'">'+str(role.role_name)+'</option>'
                
            # for branch in TblBranch.objects.filter(college_id=request.POST['filter_org_id']):
            #     branch_option += '<option value="'+str(branch.id)+'">'+str(branch.abbr)+'</option>'
                
            # response['branch_option'] = branch_option
            response['filter_role_option'] = filter_role_option
    return JsonResponse(response)

@login_required
def roleByEditOrganization(request):
    response = {}
    # filter_role_option = '<option value="all">All Role</option>'
    filter_role_option = ''
    branch_option = '<option value="all">All Branch</option>'
    # branch_option = ''

    if request.method == "POST":
        if request.POST.getlist('filter_org_id[]'):
            for org_id in request.POST.getlist('filter_org_id[]'):
                for role in SpRoles.objects.filter(organization_id=org_id).filter(role_name__regex = r'%s'%'[s|S]tudent'):
                    filter_role_option += '<option id="role_'+str(role.id)+' value="'+str(role.id)+'">'+str(role.role_name)+'</option>'
                for branch in TblBranch.objects.filter(college_id=org_id):
                    branch_option += '<option id="branch_"'+str(branch.id)+' value="'+str(branch.id)+'">'+str(branch.abbr)+'</option>'
            response['branch_option'] = branch_option
            response['filter_role_option'] = filter_role_option
        else:
            for role in SpRoles.objects.filter(organization_id=request.POST['filter_org_id']).filter(role_name__regex = r'%s'%'[s|S]tudent'):
                    # filter_role_option += '<option value="'+str(role.id)+'">'+str(role.role_name)+'</option>'
                    filter_role_option += '<option id="role_'+str(role.id)+' value="'+str(role.id)+'">'+str(role.role_name)+'</option>'
            for branch in TblBranch.objects.filter(college_id=request.POST['filter_org_id']):
                    # branch_option += '<option value="'+str(branch.id)+'">'+str(branch.abbr)+'</option>'
                    branch_option += '<option id="branch_"'+str(branch.id)+' value="'+str(branch.id)+'">'+str(branch.abbr)+'</option>'
            response['branch_option'] = branch_option
            response['filter_role_option'] = filter_role_option
    return JsonResponse(response)



@login_required
def editHoliday(request,holiday_id):
    if request.method == "POST":
        context = {}
        holiday_name    = clean_data(request.POST['holidayName'])
        holiday_id      = request.POST['holiday_id']
        filter_org_name = request.POST['filter_org_name']
        if SpHolidays.objects.filter(holiday=holiday_name,organization_id=filter_org_name).exclude(id=holiday_id).exists():
            context['flag'] = False
            context['message'] = "Leave Policy already exists."
        else:
            applicable_to = {}
            branch_1 = []
            semester_1 = []
            role = request.POST.getlist('filter_role[]')
            role = ','.join(str(e) for e in role)
            
            holiday = SpHolidays.objects.get(id=holiday_id)
            holiday.holiday = clean_data(request.POST['holidayName'])
            holiday.organization_name = getModelColumnById(SpOrganizations,filter_org_name,'organization_name')
            holiday.organization_id = filter_org_name
            holiday.applicable_to=role
            holiday.holiday_type_id = request.POST['holiday_type_id']
            holiday.holiday_type = getModelColumnById(SpHolidayTypes,request.POST['holiday_type_id'],'holiday_type')
            holiday.start_time = "00:00:00" 
            holiday.start_date = datetime.strptime(clean_data(request.POST['from_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
            holiday.end_date = datetime.strptime(clean_data(request.POST['to_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
            holiday.end_time = "00:00:00" 
            holiday.description = clean_data(request.POST['holiday_description'])
            holiday.status = 1 
            holiday.holiday_status = 1 
            holiday.save()
            if holiday.id:
                if request.POST['filter_role[]']:

                    SpRoleEntityMapping.objects.filter(entity_id=holiday.id,entity_type='holiday').delete()

                    role_id_to_mapped = request.POST.getlist('filter_role[]')   
                    if 'all' in role_id_to_mapped:
                        for org_id in request.POST.getlist('filter_org_name[]'):
                            for roleId in SpRoles.objects.filter(organization_id=org_id):
                                mapping =  SpRoleEntityMapping()
                                mapping.entity_type = "holiday"
                                mapping.role_id = roleId.id
                                mapping.entity_id = holiday.id
                                mapping.save()
                    else:
                        for roleId in role_id_to_mapped:
                            role_mapping = SpRoleEntityMapping()
                            role_mapping.entity_type = "holiday"
                            role_mapping.role_id = roleId
                            role_mapping.entity_id = holiday.id
                            role_mapping.save()
                            
            context['flag'] = True
            context['message'] = "Record has been updated successfully."

        return JsonResponse(context)

    else:
        context = {}
        if request.user.role_id == 0:
            context['institutions'] = SpOrganizations.objects.all()            
        else:
            context['institutions'] = SpOrganizations.objects.filter(id=request.user.organization_id)
                
        context['role_mappings'] = SpRoleEntityMapping.objects.filter(entity_type='holiday',entity_id=holiday_id).values_list('role_id',flat=True)
        context['holiday'] = holiday = SpHolidays.objects.get(id=holiday_id)
        context['applicable_to'] = holiday.applicable_to
        context['holiday_org'] = holiday.organization_id
        context['roles'] = SpRoles.objects.filter(status = 1)
        context['holiday_types'] = SpHolidayTypes.objects.filter(status=1)
        roles = holiday.applicable_to
        roles_list = list(roles.split(','))
        roles= SpRoles.objects.filter(organization_id=holiday.organization_id).values('id', 'role_name')
        for role in roles:
            if str(role['id']) in roles_list:
                role['selected'] = 'selected'
            else:
                role['selected'] = ''        
        context['roles'] = roles
       
        template = 'holiday/edit-holiday.html'
       
        return render(request, template, context)


@login_required
def updateHolidayStatus(request):
    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            if bool(request.POST.get('is_active',False)) == True:
                is_active = request.POST.get('is_active')
                data = SpHolidays.objects.get(id=id)
                data.status = is_active
                data.save()
                response['error'] = False

                response['message'] = "Record has been updated successfully."
            else:
                holidayID = request.POST.getlist('holiday_id[]')
                if bool(request.POST.get('description', False)) == True:
                    for holiday_id in holidayID:
                        update_holiday_status = SpHolidays.objects.get(id=holiday_id)
                        update_holiday_status.holiday_status = request.POST['statusId']
                        update_holiday_status.holiday_description = request.POST['description']
                        update_holiday_status.save()
                    response['error'] = False
                    response['message'] = "Record has been updated successfully."
                else:
                    for holiday_id in holidayID:
                        updateholidayStatus = SpHolidays.objects.get(id=holiday_id)
                        updateholidayStatus.holiday_status = request.POST['statusId']
                        updateholidayStatus.save()
                    response['error'] = False
                    response['message'] = "Record has been updated successfully."
            return JsonResponse(response)
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return JsonResponse(response)


def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None

@login_required
def exportToPDF(request, filter_status):
    if int(filter_status) < 0:
        holidays = SpHolidays.objects.all().values().order_by('-id')
    else:
        holidays = SpHolidays.objects.filter(holiday_status=filter_status).values().order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('holiday/holiday_pdf_template.html', {'holidays': holidays, 'url': baseurl})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'holidays.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

@login_required
def exportToXlsx(request, filter_status):

    if int(filter_status) < 0:
        holidays = SpHolidays.objects.all().order_by('-id')
    else:
        holidays = SpHolidays.objects.filter(holiday_status=filter_status).order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=holidays.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Holidays'
    
    # Define the titles for columns
    columns = []
    columns += [ 'Holiday' ]
    columns += [ 'Type' ]
    columns += [ 'Duration' ] 
    columns += [ 'Description' ]
    
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for holiday in holidays:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        row += [ holiday.holiday ]
        row += [ holiday.holiday_type ]
        duration = holiday.start_date.strftime('%d %b') +' '+holiday.start_date.strftime('%a')+' - '+holiday.end_date.strftime('%d %b') +' '+holiday.end_date.strftime('%a')+', '+holiday.end_date.strftime('%Y')
        row += [ duration ]
        row += [ strip_tags(holiday.description) ] 
       
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response
    
@login_required
def addpdf(request, holiday_id):
    if int(holiday_id) < 0:
        holidays = SpHolidays.objects.all().values().order_by('-id')
    else:
        holidays = SpHolidays.objects.filter(id=holiday_id).first()
        roles = SpRoleEntityMapping.objects.filter(entity_type="Holiday",entity_id=holiday_id)
        role_names =[]
        for role in roles:
            if SpRoles.objects.filter(id=role.role_id).exists():
                role_name = getModelColumnById(SpRoles,role.role_id,'role_name') 
                role_names.append(role_name)
        
        org_address = getModelColumnById(TblColleges,holidays.organization_id,'college_address')
        profile_image = getModelColumnById(TblColleges,holidays.organization_id,'college_logo')
        org_name = getModelColumnById(TblColleges,holidays.organization_id,'college_name')
    holidays.role_names=role_names
    holidays.org_address=org_address
    holidays.org_name=org_name
    baseurl = settings.BASE_URL
    if profile_image != None:
        baseurl= baseurl + profile_image
    else:
        baseurl=baseurl + '/static/img/png/dashboardLogo.png'
    context={}
    today = date.today()
    holidays.today=today
    description = holidays.description
    description = re.sub(r"<?.[a-z]?.>", "", description, flags = re.I)
    holidays.description = description
    org = SpHolidays.objects.get(id=holiday_id)
    try:
        count = len(org.applicable_to['semester'])
    except :
        count = 0

    try:
        counts = len(org.applicable_to['branch'])
    except :
        counts = 0
    semester = TblSemester.objects.filter(type__startswith="semester")
    branch = TblBranch.objects.filter(college_id = org.organization_id)
    branch_count =TblBranch.objects.filter(college_id = org.organization_id).count()
    semester_1 = ''
    branch_1 = ''
    if count == len(semester):
        semester_1 = "All Semester"
    else:
        for i, holiday in enumerate(holidays.applicable_to['semester']):
            semester_1 += TblSemester.objects.filter(semester_id = holiday).values('sem_name')[0]['sem_name']
            if i == len(holidays.applicable_to['semester'])-1:
                semester_1 += " & "
            elif i < len(holidays.applicable_to['semester'])-1:
                semester_1 += ", "
            elif i == len(holidays.applicable_to['semester']):
                pass
            else:
                pass

    if counts == len(branch):
        branch_1 = "All Branches"
    else:
        for holiday in holidays.applicable_to['branch']:
            branch_1 += " " + TblBranch.objects.filter(id = holiday).values('branch')[0]['branch']
    semester = semester_1
    branch = branch_1
    website=TblColleges.objects.filter(id=org.organization_id).values('college_website')[0]['college_website']
    email=SpOrganizations.objects.filter(id=org.organization_id).values('email')[0]['email']
    today = date.today()
    doc = DocxTemplate("./static/holiday_notice.docx")
    context = {
        'holiday_type': holidays.holiday_type,
        'from_date': holidays.start_date,
        'to_date': holidays.end_date,
        'subject': holidays.holiday,
        'description' : holidays.description,
        'semester': semester, 
        'branch': branch, 
        'today':today, 
        'website':website, 
        'email':email
    }
    context['url'] = InlineImage(doc, os.path.abspath(profile_image), width=Mm(30), height=Mm(30))
    doc_io = BytesIO()
    doc.render(context)
    doc.save(doc_io)
    doc_io.seek(0)
    response = HttpResponse(doc_io.read())
    response["Content-Disposition"] = "attachment; filename=holiday.docx"
    response["Content-Type"] = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    return response




        
    
    