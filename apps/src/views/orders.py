import sys
import os
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings

# Create your views here.

# Order View
@login_required
def index(request):
    page    = request.GET.get('page')
    today   = date.today()
    if request.user.role_id == 0:
        orders  = SpOrders.objects.all().order_by('-id').filter(order_date__icontains=today.strftime("%Y-%m-%d"))
    else:
        orders = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' order by id desc ''',[today.strftime("%Y-%m-%d"), request.user.id])
    
    towns       = SpTowns.objects.all()
    routes      = SpRoutes.objects.all()
    user_list   = SpUsers.objects.filter(user_type=2)

    distributor_order_count     = SpOrders.objects.filter(user_type='Distributor').count()
    super_stockist_order_count  = SpOrders.objects.filter(user_type='SuperStockist').count()
    total_orders                = distributor_order_count+super_stockist_order_count
    user_type                   = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add', workflow_level_role_id=request.user.role_id).exclude(role_id=request.user.role_id).values('level_id').order_by('-id').first()    
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    
    week_dates   = get_week_date(today.year, today.month)
    week_days    = get_week_day(today.year, today.month)

    order_counts = []
    for id, dates in enumerate(week_dates):    
        order_count_details = {}
        order_count         = SpOrders.objects.filter(order_date__icontains=dates).count()
        order_count_details['count'] = order_count
        order_count_details['day']   = week_days[id]
        order_counts.append(order_count_details)

    products_list = []
    total_order_quantity = []
    product_classes = SpProductClass.objects.filter(status=1).order_by('-id')
    for product_class in product_classes:
        product_list = {}
        product_list['id']          = product_class.id
        product_list['class_name']  = product_class.product_class
        products = SpProducts.objects.filter(product_class_id=product_class.id).values_list('id', flat=True)
        if products:
            quantity  = SpOrderDetails.objects.filter(product_id__in=products).aggregate(Sum('quantity'))['quantity__sum']
            if quantity:
                total_quantity = quantity
            else:
                total_quantity = 0
        total_order_quantity.append(total_quantity)         
        product_list['quantity']  = total_quantity       
        products_list.append(product_list)    

    if total_order_quantity:
        total_order_quantity = sum(total_order_quantity)
    else:
        total_order_quantity = 0
            
    context = {}
    context['orders']                           = orders
    context['user_list']                        = user_list
    context['products_list']                    = products_list
    context['total_order_quantity']             = total_order_quantity
    context['level_id']                         = level_id
    context['role_id']                          = request.user.role_id
    context['order_counts']                     = order_counts
    context['total_orders']                     = total_orders
    context['distributor_order_count']          = distributor_order_count
    context['super_stockist_order_count']       = super_stockist_order_count
    context['towns']                            = towns
    context['routes']                           = routes
    context['today_date']                       = today.strftime("%d/%m/%Y")
    context['week_start_date']                  = datetime.strptime(str(week_dates[0]), "%Y-%m-%d").strftime('%d/%m/%Y')
    context['week_end_date']                    = datetime.strptime(str(week_dates[6]), "%Y-%m-%d").strftime('%d/%m/%Y')
    context['total_pages']                      = 'all'
    context['page_limit']                       = getConfigurationResult('page_limit')
    context['page_title']                       = "Order Management"

    template = 'order-management/index.html'
    return render(request, template, context)

#ajax order list
@login_required
def ajaxOrdersLists(request):
    order_status    = request.GET['order_status']
    town_id         = request.GET['town_id']
    route_id        = request.GET['route_id']
    user_sap_id     = request.GET['user_sap_id']
    order_date      = request.GET['order_date']

    if request.user.role_id == 0:
        orders = SpOrders.objects.all().order_by('-id')
        if order_status:
            orders = orders.filter(order_status=order_status)
        if town_id:
            orders = orders.filter(town_id=town_id)
        if route_id:
            orders = orders.filter(route_id=route_id)
        if user_sap_id:
            orders = orders.filter(user_sap_id=user_sap_id) 
        if order_date:
            order_date  = datetime.strptime(str(order_date), '%d/%m/%Y').strftime('%Y-%m-%d')
            orders      = orders.filter(order_date__icontains=order_date)
    else:
        
        condition = ''
        if order_status:
            condition += ' and sp_orders.order_status = %s' % order_status
        if town_id:
            condition += ' and sp_orders.town_id = %s' % town_id
        if route_id:
            condition += ' and sp_orders.route_id = %s' % route_id    
        if user_sap_id:
            user_sap_id = '"'+user_sap_id+'"'
            condition += ' and sp_orders.user_sap_id = %s' % user_sap_id    
        
        order_date  = '"'+datetime.strptime(str(order_date), '%d/%m/%Y').strftime('%Y-%m-%d')+'"'
        query = """ SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id 
    where DATE(sp_orders.order_date) = %s and sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpOrders' %s order by id desc """ % (order_date, request.user.id, condition)
        
        orders = SpOrders.objects.raw(query)

    context = {}
    context['orders']           = orders
    context['role_id']          = request.user.role_id
    context['order_status']     = order_status
    template = 'order-management/ajax-order-lists.html'
    return render(request, template, context)

#get order details view
@login_required
def getOrderDetails(request):
    id                          = request.GET.get('id')

    if request.user.role_id == 0:
        order_details  = SpOrders.objects.get(id=id)
    else:
        order_details = SpOrders.objects.raw('''SELECT sp_orders.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name
    FROM sp_orders left join sp_approval_status on sp_approval_status.row_id = sp_orders.id 
    where sp_orders.id = %s order by id desc LIMIT 1 ''',[request.GET.get('id')])[0] 

    order_details.profile_image = getModelColumnById(SpUsers, order_details.user_id, 'profile_image')
    order_details.user_details  = SpBasicDetails.objects.get(user_id=order_details.user_id)
    order_details.user_address  = SpAddresses.objects.get(user_id=order_details.user_id, type='permanent')

    order_item_list = SpOrderDetails.objects.filter(order_id=id)
    
    crate_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Crate').aggregate(Sum('quantity'))['quantity__sum']
    matki_sum  = SpOrderDetails.objects.filter(order_id=id,product_container_type='Matki').aggregate(Sum('quantity'))['quantity__sum']
    
    context = {}
    context['order_details']    = order_details
    context['order_item_list']  = order_item_list
    context['crate_sum']        = crate_sum
    context['matki_sum']        = matki_sum
    context['role_id']          = request.user.role_id

    template = 'order-management/get-order-details.html'
    return render(request, template, context)

#update order status
@login_required
def updateOrderStatus(request):
    response = {}
    order_id = request.POST.getlist('order_id[]')
    level_id = request.POST['level_id']
    if request.user.role_id == 0:
        for order in order_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    order                   = SpOrders.objects.get(id=order)   
                    order.order_status      = level_id
                    order.save()
            else:
                order                   = SpOrders.objects.get(id=order)   
                order.order_status      = level_id
                order.save()
    else:    
        for order in order_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', role_id=request.user.role_id, level_id=level_id)
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                order                   = SpOrders.objects.get(id=order)   
                order.order_status      = level_id
                order.save()   

    
    if level_id == '2':
        for order in order_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=order, model_name='SpOrders', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpOrders'
                    notification.notification           = 'Order '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()

    if level_id == '2':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been forwarded'
            activity    = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'add.png', '1', 'web') 
    elif level_id == '3':
        for order in order_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been approved'
            activity    = 'Order('+getModelColumnById(SpOrders, order, 'order_code')+') has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            
            saveActivity('Order Management', 'Order Summary', heading, activity, request.user.id, user_name, 'add.png', '1', 'web')

    response['error'] = False
    response['message'] = "Order status has been updated successfully."
    return JsonResponse(response)

@login_required
def exportToXlsx(request, columns, town_id, route_id, user_sap_id, order_date):
    column_list = columns.split (",")
    orders = SpOrders.objects.all().order_by('-id')
    if town_id != 0:
        orders = orders.filter(town_id=town_id)
    if route_id != 0:
        orders = orders.filter(route_id=route_id)
    if user_sap_id != '0':
        orders = orders.filter(user_sap_id=user_sap_id) 
    if order_date:
        orders = orders.filter(order_date__icontains=order_date)
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Orders'
    
    # Define the titles for columns
    columns = []

    if 'order_id' in column_list:
        columns += [ 'Order ID' ]

    if 'distributor_ss' in column_list:
        columns += [ 'Distributor/Super-Stockist Name' ]
    
    if 'shift' in column_list:
        columns += [ 'Shift' ] 

    if 'route' in column_list:
        columns += [ 'Route' ]

    if 'amount' in column_list:
        columns += [ 'Amount' ] 

    if 'status' in column_list:
        columns += [ 'Status' ]       

    if 'order_date' in column_list:
        columns += [ 'Order Date' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    
    for order in orders:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'order_id' in column_list:
            row += [ order.order_code ]

        if 'distributor_ss' in column_list:
            row += [ order.user_name ]
        
        if 'shift' in column_list:
            row += [ order.order_shift_name ] 

        if 'route' in column_list:
            row += [ order.route_name ]

        if 'amount' in column_list:
            row += [ order.order_total_amount ]

        if 'status' in column_list:
            if order.order_status == 0:
                order_status = 'Intiate'
            elif order.order_status == 1:
                order_status ='Forward'
            else:
                order_status='Delivered'

            row += [ order_status ]

        if 'order_date' in column_list:
            order_date = str(order.order_date).replace('+00:00', '')
            row += [ datetime.strptime(str(order_date), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %I:%M:%p') ]                   
       
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    workbook.save(response)

    return response

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#Automaticly downloads to PDF file
@login_required
def exportToPDF(request, columns, town_id, route_id, user_sap_id, order_date):
    column_list = columns.split (",")
    context = {}
    orders = SpOrders.objects.all().order_by('-id')
    if town_id != 0:
        orders = orders.filter(town_id=town_id)
    if route_id != 0:
        orders = orders.filter(route_id=route_id)
    if user_sap_id != '0':
        orders = orders.filter(user_sap_id=user_sap_id) 
    if order_date:
        orders = orders.filter(order_date__icontains=order_date)

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('order-management/order_pdf_template.html', {'orders': orders, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'orders.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#get order details view
@login_required
def editOrder(request):
    id                          = request.GET.get('id')
    order_details               = SpOrders.objects.get(id=id)
    order_details.profile_image = getModelColumnById(SpUsers, order_details.user_id, 'profile_image')
    order_details.user_details  = SpBasicDetails.objects.get(user_id=order_details.user_id)
    order_details.user_address  = SpAddresses.objects.get(user_id=order_details.user_id, type='permanent')

    order_item_list = SpOrderDetails.objects.filter(order_id=id)

    context = {}
    context['order_details']        = order_details
    context['order_item_list']      = order_item_list

    template = 'order-management/edit-order.html'

    return render(request, template, context)

#get order status view
@login_required
def orderStatusDetails(request):
    order_id                    = request.GET.get('order_id')
    initiate_order_details      = SpOrders.objects.get(id=order_id)  
    order_details               = SpApprovalStatus.objects.filter(row_id=order_id, model_name='SpOrders', status=1).values('final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')
    
    context = {}
    context['initiate_order_details']   = initiate_order_details
    context['order_details']            = order_details
    template = 'order-management/order-status-details.html'

    return render(request, template, context)

#get order indent Report
@login_required
def indentReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()

    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_basic_details.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s ''',[today.strftime("%Y-%m-%d"), 3])

    product_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='3' order by sp_product_variants.product_class_id asc ''')
    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='3' order by sp_product_variants.product_class_id asc ''')
    
    indent_lists = []
    total_crate_sum = []
    total_without_milk_crate_sum = []
    total_milk_crates_quantity = []
    for user in user_list:
        indent = {}
        product_milk_variants_list = []
        total_milk_crates = []
        
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                total = total + Item_in_liters
            else:
                Item_in_liters = 0
                total = total + Item_in_liters 
                        
            product_variant_milk_lists['id'] = product_variant_milk.id    
            product_variant_milk_lists['milk_items'] = Item_in_liters
            product_variant_milk_lists['product_color_code'] = product_variant_milk.product_color_code
            total_milk_crates.append(total)
            total_crates = total_milk_crates  
            product_milk_variants_list.append(product_variant_milk_lists)

        total_crate_sum.append(total_crates)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                totals = totals + Item_in_liters
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            product_variant_without_milk_lists['id'] = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            product_without_milk_variants_list.append(product_variant_without_milk_lists)    
        total_without_milk_crate_sum.append(total_without_milk_crates)

        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)

    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()

    context = {}
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Indent Report"
    template                                        = 'order-management/indent-report.html'
    
    return render(request, template, context)    


#get order indent Report
@login_required
def ajaxIndentReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_basic_details.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s ''',[today, 3])

    product_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='3' order by sp_product_variants.product_class_id asc ''')
    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='3' order by sp_product_variants.product_class_id asc ''')
    
    indent_lists = []
    total_crate_sum = []
    total_without_milk_crate_sum = []
    total_milk_crates_quantity = []
    for user in user_list:
        indent = {}
        product_milk_variants_list = []
        total_milk_crates = []
        
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                total = total + Item_in_liters
            else:
                Item_in_liters = 0
                total = total + Item_in_liters 
                        
            product_variant_milk_lists['id'] = product_variant_milk.id    
            product_variant_milk_lists['milk_items'] = Item_in_liters
            product_variant_milk_lists['product_color_code'] = product_variant_milk.product_color_code
            total_milk_crates.append(total)
            total_crates = total_milk_crates  
            product_milk_variants_list.append(product_variant_milk_lists)

        total_crate_sum.append(total_crates)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                totals = totals + Item_in_liters
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            product_variant_without_milk_lists['id'] = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            product_without_milk_variants_list.append(product_variant_without_milk_lists)    
        total_without_milk_crate_sum.append(total_without_milk_crates)

        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)

    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()
    
    today_date   = date.today()
    context = {}
    context['user_list']                            = user_list
    context['indent_lists']                         = indent_lists
    context['total_milk_crates']                    = total_milk_crates_quantity
    context['total_milk_crates_qty']                = total_milk_crates_qty
    context['total_without_milk_crates_qty']        = total_without_milk_crates_qty
    context['product_milk_variant_list']            = product_milk_variant_list
    context['product_without_milk_variant_list']    = product_without_milk_variant_list
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Indent Report"
    template                                        = 'order-management/ajax-indent-report.html'
    
    return render(request, template, context)

#get order indent Report
@login_required
def generateIndentReport(request):
    response = {}
    today   = date.today()
    if request.GET['id'] == '0':
        msg = 'generating'
    else:
        msg = 're-generating'      
    try:
        approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today.strftime("%Y-%m-%d")).count()
        today_order = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()
        if today_order == 0:
            response['error']       = True
            response['error_type']  = 'danger'
            response['message']     = "No Order Found."
        elif approved_order == today_order:
            orders   = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d"))
            for order in orders:
                data = SpOrders.objects.get(id=order.id)
                data.indent_status = 1
                data.save()

            response['error']       = False
            response['error_type']  = 'success'
            response['message']     = "Indent Report is "+msg+". Please wait..."
        else:
            response['error']       = True
            response['error_type']  = 'danger'
            response['message']     = "All the orders for the date("+today.strftime("%d/%m/%Y")+") are not approved. Kindly approved the orders first."   
    except ObjectDoesNotExist:
        response['error']       = True
        response['error_type']  = 'danger'
        response['message']     = "Method not allowed"
    except Exception as e:
        response['error']       = True
        response['error_type']  = 'danger'
        response['message']     = e
    return JsonResponse(response)

#get order indent Report
@login_required
def exportIndentReport(request, order_date):
    
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()

    user_list = SpUsers.objects.raw('''SELECT sp_users.id, sp_users.first_name, sp_users.middle_name, sp_users.last_name, sp_users.emp_sap_id, sp_basic_details.outstanding_amount, sp_basic_details.security_amount, sp_orders.id as order_id, sp_orders.order_status, sp_orders.town_name, sp_orders.order_total_amount, sp_orders.mode_of_payment, sp_orders.amount_to_be_paid
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_orders on sp_orders.user_id = sp_users.id
    where DATE(sp_orders.order_date) = %s and sp_orders.order_status >= %s ''',[today, 3])

    product_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id='3' order by sp_product_variants.product_class_id asc ''')
    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_products.product_color_code FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id!='3' order by sp_product_variants.product_class_id asc ''')
    
    indent_lists = []
    total_crate_sum = []
    total_without_milk_crate_sum = []
    total_milk_crates_quantity = []
    for user in user_list:
        indent = {}
        product_milk_variants_list = []
        total_milk_crates = []
        
        for id, product_variant_milk in enumerate(product_milk_variant_list):
            total = 0
            product_variant_milk_lists = {}
            total_crates = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                total = total + Item_in_liters
            else:
                Item_in_liters = 0
                total = total + Item_in_liters 
                        
            product_variant_milk_lists['id'] = product_variant_milk.id    
            product_variant_milk_lists['milk_items'] = Item_in_liters
            product_variant_milk_lists['product_color_code'] = product_variant_milk.product_color_code
            total_milk_crates.append(total)
            total_crates = total_milk_crates  
            product_milk_variants_list.append(product_variant_milk_lists)

        total_crate_sum.append(total_crates)
        sum_total_crates = sum(total_milk_crates)
        total_milk_crates_quantity.append(sum_total_crates)

        product_without_milk_variants_list = []
        total_without_milk_crates = []
        for product_variant_without_milk in product_without_milk_variant_list:
            totals = 0
            product_variant_without_milk_lists = {}
            try:
                order_details = SpOrderDetails.objects.get(product_variant_id=product_variant_without_milk.id, order_id=user.order_id)
            except SpOrderDetails.DoesNotExist:
                order_details = None
            if order_details:
                Item_in_liters = int(order_details.quantity)
                totals = totals + Item_in_liters
            else:
                Item_in_liters = 0
                totals = totals + Item_in_liters
            
            product_variant_without_milk_lists['id'] = product_variant_without_milk.id    
            product_variant_without_milk_lists['without_milk_items'] = Item_in_liters
            product_variant_without_milk_lists['product_color_code'] = product_variant_without_milk.product_color_code
            total_without_milk_crates.append(totals)
            total_without_milk_crates = total_without_milk_crates

            product_without_milk_variants_list.append(product_variant_without_milk_lists)    
        total_without_milk_crate_sum.append(total_without_milk_crates)

        balance_after_deposit = (float(user.order_total_amount)+float(user.outstanding_amount))-float(user.amount_to_be_paid)
        balance_security      = float(user.security_amount)-balance_after_deposit
        indent['id']                    = user.id
        indent['first_name']            = user.first_name
        indent['middle_name']           = user.middle_name
        indent['last_name']             = user.last_name
        indent['emp_sap_id']            = user.emp_sap_id
        indent['town_name']             = user.town_name
        indent['order_total_amount']    = user.order_total_amount
        indent['outstanding_amount']    = user.outstanding_amount
        indent['security_amount']       = user.security_amount
        indent['mode_of_payment']       = user.mode_of_payment
        indent['amount_to_be_paid']     = user.amount_to_be_paid
        indent['balance_after_deposit'] = balance_after_deposit
        indent['balance_security']      = balance_security
        indent['milk_items']            = product_milk_variants_list
        indent['without_milk_items']    = product_without_milk_variants_list
        indent_lists.append(indent)

    #sum of milk product qty
    total_milk_crates_qty = []
    if total_crate_sum:
        for column in enumerate(total_crate_sum[0]):
            count = sum([x[column[0]] for x in total_crate_sum])
            total_milk_crates_qty.append(count)

    #sum of without milk product qty
    total_without_milk_crates_qty = []
    if total_without_milk_crate_sum:
        for column in enumerate(total_without_milk_crate_sum[0]):
            count = sum([x[column[0]] for x in total_without_milk_crate_sum])
            total_without_milk_crates_qty.append(count)
    
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=indent-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'INDENT REPORT'
    worksheet.merge_cells('A1:B1') 
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=3)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+13
    
    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=4).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH'
    worksheet.cell(row=1, column=4).font = header_font
    worksheet.cell(row=1, column=4).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=4).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=4).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    columns = []
    columns += [ 'Customer Code' ]
    columns += [ 'Name of Distributor/SS' ]
    columns += [ 'Town' ]

    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ product_variant.variant_name ]

    columns += [ 'TOTAL MILK CRATES' ]
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ product_variant.variant_name ]
    columns += [ 'FLAT INCENTIVE' ]
    columns += [ 'BULK PACK INCENTIVE' ]
    columns += [ 'TOTAL INCENTIVE AMOUNT' ]
    columns += [ 'INVOICE AMOUNT' ]
    columns += [ 'PAYMENT DETAILS' ]
    columns += [ 'OUTSTANDING DETAILS' ]
    columns += [ 'SECURITY' ]
    columns += [ 'BALANCE AFTER DEPOSIT' ]
    columns += [ 'BALANCE SECURITY' ]

    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num <=3:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        elif col_num <= len(product_milk_variant_list)+3:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type = "solid")
        elif col_num == len(product_milk_variant_list)+4:
            for product_variant in product_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=14, color='000000', bold=True)
                cell.fill = PatternFill()        
        elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+4:
            for product_variant in product_without_milk_variant_list:
                color_code = str(product_variant.product_color_code).replace('#', '')
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type = "solid")        
        else:
            cell.fill = PatternFill()
            cell.font = Font(size=14, color='000000', bold=True)              
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 40

    
    for id, indent in enumerate(indent_lists):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ indent['emp_sap_id'] ]
        row += [ indent['first_name'] +' '+ indent['middle_name'] +' '+ indent['last_name'] ]
        row += [ indent['town_name'] ]
        if indent['milk_items']:
            for item in indent['milk_items']:
                row += [ item['milk_items'] ]
        row += [ total_milk_crates_quantity[id] ]
        if indent['without_milk_items']:
            for item in indent['without_milk_items']:
                row += [ item['without_milk_items'] ]
        row += [ '-' ]
        row += [ '-' ]
        row += [ '-' ]
        row += [ indent['order_total_amount'] ]
        row += [ indent['amount_to_be_paid'] ]
        row += [ indent['outstanding_amount'] ]
        row += [ indent['security_amount'] ]
        row += [ indent['balance_after_deposit'] ]
        row += [ indent['balance_security'] ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num <=3:
                cell.font = Font(size=12, color='000000')
            elif col_num <= len(product_milk_variant_list)+3:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type = "solid")
            elif col_num == len(product_milk_variant_list)+4:
                for product_variant in product_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=12, color='000000')
                    cell.fill = PatternFill()        
            elif col_num <= len(product_milk_variant_list)+len(product_without_milk_variant_list)+4:
                for product_variant in product_without_milk_variant_list:
                    color_code = str(product_variant.product_color_code).replace('#', '')
                    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type = "solid")        
            else:
                cell.fill = PatternFill()
                cell.font = Font(size=12, color='000000')
    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ 'GRAND TOTAL(CRATES)' ]

    if total_milk_crates_qty:
        for total_milk_crates in total_milk_crates_qty:
            bottom_columns += [ total_milk_crates ]

    bottom_columns += [ ' ' ]
    if total_without_milk_crates_qty:
        for total_without_milk_crates in total_without_milk_crates_qty:
            bottom_columns += [ total_without_milk_crates ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=11, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Define the titles for bottom_columns
    row_num += 1
    bottom_columns = []
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' TRADE SCHEME POUCHES ' ]

    if total_milk_crates_qty:
        for total_milk_crates in total_milk_crates_qty:
            bottom_columns += [ ' ' ]

    bottom_columns += [ ' ' ]
    if total_without_milk_crates_qty:
        for total_without_milk_crates in total_without_milk_crates_qty:
            bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]
    bottom_columns += [ ' ' ]

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(bottom_columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrapped_alignment
        cell.font = Font(size=11, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    workbook.save(response)

    return response

#get summary Report
@login_required
def summaryReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    
    product_classes = SpProductClass.objects.all().order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today.strftime("%Y-%m-%d")).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0

                    pouch_wo_scheme                                     = int(quantity_wo_scheme/float(product_variant.variant_size))
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                
                product_variant.total_product_quantity_with_scheme        = product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)

            product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
            product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
            product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
            product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

            product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
            product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
            product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
            product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

    products = SpProducts.objects.filter(status=1).order_by('-id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        
        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes = 0
            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today.strftime("%Y-%m-%d")).aggregate(quantity = Sum('quantity'))
                product_total_quantity = product_total_quantity['quantity']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0
                  
            if product_total_quantity:
                quantity_wo_schemes  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes) 
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = sum(total_quantity_wo_scheme_crate_sums)
        products_list.append(product_list)
           
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today.strftime("%Y-%m-%d")).count()

    context = {}
    context['product_classes']                      = product_classes
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['products_list']                        = products_list
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Summary Report"
    template                                        = 'order-management/summary-report.html'

    return render(request, template, context) 
             
#get summary Report
@login_required
def ajaxSummaryReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    
    product_classes = SpProductClass.objects.all().order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0

                    pouch_wo_scheme                                     = int(quantity_wo_scheme/float(product_variant.variant_size))
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                
                product_variant.total_product_quantity_with_scheme        = product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)

            product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
            product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
            product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
            product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

            product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
            product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
            product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
            product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
            product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
            product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

    products = SpProducts.objects.filter(status=1).order_by('-id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        
        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes = 0
            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                product_total_quantity = product_total_quantity['quantity']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0
                  
            if product_total_quantity:
                quantity_wo_schemes  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes) 
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = sum(total_quantity_wo_scheme_crate_sums)
        products_list.append(product_list)
           
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['products_list']                        = products_list
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Summary Report"
    template                                        = 'order-management/ajax-summary-report.html'

    return render(request, template, context) 

#get order summary Report
@login_required
def exportSummaryReport(request, order_date):
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    
    product_classes = SpProductClass.objects.all().order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_product_quantity_with_scheme = []
        total_quantity_wo_scheme_crate_sum = []
        total_scheme_quantity_crate_sum = []
        total_bonus_scheme_quantity_crate_sum = []
        total_employee_sale_crate_sum = []
        total_foc_quantity_crate_sum = []

        total_pouch_product_quantity_with_scheme = []
        total_pouch_wo_scheme_crate_sum = []
        total_pouch_scheme_crate_sum = []
        total_pouch_bonus_scheme_crate_sum = []
        total_pouch_employee_sale_crate_sum = []
        total_pouch_foc_crate_sum = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_quantity_wo_scheme = 0
                total_scheme = 0
                total_bonus_scheme = 0
                total_employee_sale = 0
                total_foc_quantity = 0

                total_pouch_wo_scheme = 0
                total_pouch_scheme = 0
                total_pouch_bonus_scheme = 0
                total_pouch_employee_sale = 0
                total_pouch_foc = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                                  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_total_quantity_wo_scheme    = quantity_wo_scheme
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + quantity_wo_scheme

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0

                    pouch_wo_scheme                                     = int(quantity_wo_scheme/float(product_variant.variant_size))
                    product_variant.product_total_pouch_wo_scheme       = pouch_wo_scheme
                    total_pouch_wo_scheme                               = total_pouch_wo_scheme + pouch_wo_scheme
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                else:
                    product_variant.product_total_quantity_wo_scheme    = 0
                    total_quantity_wo_scheme                            = total_quantity_wo_scheme + 0

                    product_variant.product_total_scheme_quantity       = 0
                    total_scheme                                        = total_scheme + 0

                    product_variant.product_total_bonus_scheme_quantity = 0
                    total_bonus_scheme                                  = total_bonus_scheme + 0

                    product_variant.product_total_employee_sale_quantity = 0
                    total_employee_sale                                  = total_employee_sale + 0

                    product_variant.product_total_foc_quantity           = 0
                    total_foc_quantity                                   = total_foc_quantity + 0
                    
                    product_variant.product_total_pouch_wo_scheme        = 0
                    total_pouch_wo_scheme                                = total_pouch_wo_scheme + 0
                    
                    product_variant.product_total_pouch_scheme           = 0
                    total_pouch_scheme                                   = total_pouch_scheme + 0

                    product_variant.product_total_pouch_bonus_scheme      = 0
                    total_pouch_bonus_scheme                              = total_pouch_bonus_scheme + 0

                    product_variant.product_total_pouch_employee_sale     = 0
                    total_pouch_employee_sale                             = total_pouch_employee_sale + 0

                    product_variant.product_total_pouch_foc               = 0
                    total_pouch_foc                                       = total_pouch_foc + 0
                
                product_variant.total_product_quantity_with_scheme        = product_variant.product_total_quantity_wo_scheme+product_variant.product_total_scheme_quantity+product_variant.product_total_bonus_scheme_quantity+product_variant.product_total_employee_sale_quantity+product_variant.product_total_foc_quantity

                total_product_quantity_with_scheme.append(product_variant.total_product_quantity_with_scheme)
                total_quantity_wo_scheme_crate_sum.append(total_quantity_wo_scheme) 
                total_scheme_quantity_crate_sum.append(total_scheme)
                total_bonus_scheme_quantity_crate_sum.append(total_bonus_scheme)
                total_employee_sale_crate_sum.append(total_employee_sale)
                total_foc_quantity_crate_sum.append(total_foc_quantity)

                product_variant.total_pouch_product_quantity_with_scheme   =  product_variant.product_total_pouch_wo_scheme+product_variant.product_total_pouch_scheme+product_variant.product_total_pouch_bonus_scheme+product_variant.product_total_pouch_employee_sale+product_variant.product_total_pouch_foc

                total_pouch_product_quantity_with_scheme.append(product_variant.total_pouch_product_quantity_with_scheme)
                total_pouch_wo_scheme_crate_sum.append(total_pouch_wo_scheme)
                total_pouch_scheme_crate_sum.append(total_pouch_scheme)
                total_pouch_bonus_scheme_crate_sum.append(total_pouch_bonus_scheme)
                total_pouch_employee_sale_crate_sum.append(total_pouch_employee_sale)
                total_pouch_foc_crate_sum.append(total_pouch_foc)

        product_class.total_product_quantity_with_scheme             = sum(total_product_quantity_with_scheme)
        product_class.total_quantity_wo_scheme_crate_sum             = sum(total_quantity_wo_scheme_crate_sum) 
        product_class.total_scheme_quantity_crate_sum                = sum(total_scheme_quantity_crate_sum)
        product_class.total_bonus_scheme_quantity_crate_sum          = sum(total_bonus_scheme_quantity_crate_sum)
        product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
        product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)

        product_class.total_pouch_product_quantity_with_scheme       = sum(total_pouch_product_quantity_with_scheme)
        product_class.total_pouch_wo_scheme_crate_sum                = sum(total_pouch_wo_scheme_crate_sum)
        product_class.total_pouch_scheme_crate_sum                   = sum(total_pouch_scheme_crate_sum)
        product_class.total_pouch_bonus_scheme_crate_sum             = sum(total_pouch_bonus_scheme_crate_sum)
        product_class.total_employee_sale_crate_sum                  = sum(total_employee_sale_crate_sum)
        product_class.total_foc_quantity_crate_sum                   = sum(total_foc_quantity_crate_sum)
        
    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    products = SpProducts.objects.filter(status=1).order_by('-id')
    products_list = []
    for product in products:
        total_quantity_wo_scheme_crate_sums = []
        
        product_variant_list    =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_id=%s order by sp_product_variants.product_id desc ''', [product.id])
        
        product_list = {}
        for product_variant in product_variant_list:
            total_quantity_wo_schemes = 0
            try:
                product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                product_total_quantity = product_total_quantity['quantity']
            except SpOrderDetails.DoesNotExist:
                product_total_quantity = 0
                  
            if product_total_quantity:
                quantity_wo_schemes  = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                product_variant.product_total_quantity_wo_scheme    = quantity_wo_schemes
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + quantity_wo_schemes
            else:
                product_variant.product_total_quantity_wo_scheme    = 0
                total_quantity_wo_schemes                            = total_quantity_wo_schemes + 0

            total_quantity_wo_scheme_crate_sums.append(total_quantity_wo_schemes) 
            
        product_list['product_name'] = product.product_name
        product_list['total_quantity'] = sum(total_quantity_wo_scheme_crate_sums)
        products_list.append(product_list)

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=summary-report.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'SUMMARY REPORT'
    worksheet.merge_cells('A1:A1') 
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    length = []
    for product_variant in product_classes:
                length.append(len(product_variant.product_variant_list))            

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    column_length = len(length)+sum(length)+1
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=24, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    columns = []
    columns += [ 'PARTICULARS' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.variant_name ]
            columns += [ product_variant.product_variant_total ]
    
    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 50
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                    cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            cell.font = Font(size=14, color='000000', bold=True)
            cell.fill = PatternFill()                
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 20

    columns = []
    columns += [ 'TOTAL QUANTITY W/O SCHEME (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_quantity_wo_scheme ]
            columns += [ product_variant.total_quantity_wo_scheme_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
        
    
    columns = []
    columns += [ 'TOTAL SCHEME QUANTITY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_scheme_quantity ]
            columns += [ product_variant.total_scheme_quantity_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')

    columns = []
    columns += [ 'TOTAL BONUS SCHEME QUANTITY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_bonus_scheme_quantity ]
            columns += [ product_variant.total_bonus_scheme_quantity_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
             

    columns = []
    columns += [ 'TOTAL EMPLOYEE SALE (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_employee_sale_quantity ]
            columns += [ product_variant.total_employee_sale_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
         

    columns = []
    columns += [ 'FOC QTY (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_foc_quantity ]
            columns += [ product_variant.total_foc_quantity_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'TOTAL QUANTITY WITH SCHEME (LTR/KG)' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.total_product_quantity_with_scheme ]
            columns += [ product_variant.total_product_quantity_with_scheme ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000', bold=True)
        

    columns = []
    columns += [ ' ' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ ' ' ]
            columns += [ ' ' ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    color_code = str(product.product_color_code).replace('#', '')
                    
            

    columns = []
    columns += [ 'TOTAL POUCH W/O SCHEME' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_pouch_wo_scheme ]
            columns += [ product_variant.total_pouch_wo_scheme_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'TOTAL SCHEME POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_pouch_scheme ]
            columns += [ product_variant.total_pouch_scheme_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'BONUS SCHEME POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_pouch_bonus_scheme ]
            columns += [ product_variant.total_pouch_bonus_scheme_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
            
    
    columns = []
    columns += [ 'EMPLOYEE SALE POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_pouch_employee_sale ]
            columns += [ product_variant.total_employee_sale_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
        

    columns = []
    columns += [ 'FOC POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.product_total_pouch_foc ]
            columns += [ product_variant.total_foc_quantity_crate_sum ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000')
      

    columns = []
    columns += [ 'TOTAL VARIANT WISE POUCH' ]
    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.total_pouch_product_quantity_with_scheme ]
            columns += [ product_variant.total_pouch_product_quantity_with_scheme ]
    
    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:
            for cols, product_variant in enumerate(product_classes):
                for product in product_variant.product_variant_list:
                    cell.font = Font(size=14, color='000000', bold=True)
        

    columns = []
    columns += [ 'PARTICULARS' ]
    columns += [ 'QUANTITY' ]
    row_num += 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        cell.font = Font(size=14, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

    for products in products_list:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        row += [ products['product_name'] ]
        row += [ products['total_quantity'] ]                 
       
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=14, color='000000')      

    workbook.save(response)

    return response

#get summary Report
@login_required
def packingStationReport(request):
    today                   = date.today()
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today.strftime("%Y-%m-%d")).count()
    
    product_classes = SpProductClass.objects.filter(id=3).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=3).order_by('-id')
    for product_classs in product_classess:
        product_classs.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_classs.id])
        product_classs.product_variant_total             = 'Total '+product_classs.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []

        if product_classs.product_variant_list:
            for id, product_variant in enumerate(product_classs.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_classs.total_quantity                    = sum(total_quantity)
            product_classs.total_pouches_packed              = sum(total_pouches_packed) 
            product_classs.total_crates_packed               = sum(total_crates_packed) 
            product_classs.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)

    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['today_date']                           = today.strftime("%d/%m/%Y")
    context['page_title']                           = "Packing Station Report"
    template                                        = 'order-management/packing-station-report.html'

    return render(request, template, context)   
    
#get packing station Report
@login_required
def ajaxPackingStationReport(request):
    today                   = request.GET['order_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    
    product_classes = SpProductClass.objects.filter(id=3).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=3).order_by('-id')
    for product_classs in product_classess:
        product_classs.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_classs.id])
        product_classs.product_variant_total             = 'Total '+product_classs.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_classs.product_variant_list:
            for id, product_variant in enumerate(product_classs.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_classs.total_quantity                    = sum(total_quantity)
            product_classs.total_pouches_packed              = sum(total_pouches_packed) 
            product_classs.total_crates_packed               = sum(total_crates_packed) 
            product_classs.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)

    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")
    context['page_title']                           = "Packing Station Report"
    template                                        = 'order-management/ajax-packing-station-report.html'

    return render(request, template, context)      

#export packing station Report
@login_required
def exportPackingStationReport(request, order_date):
    today                   = order_date
    today_order_status      = SpOrders.objects.filter(indent_status=1, order_date__icontains=today).count()
    order_regenerate_status = SpOrders.objects.filter(indent_status=0, order_date__icontains=today).count()
    
    product_classes = SpProductClass.objects.filter(id=3).order_by('-id')
    for product_class in product_classes:
        product_class.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_class.id])
        product_class.product_variant_total             = 'Total '+product_class.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_class.product_variant_list:
            for id, product_variant in enumerate(product_class.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_class.total_quantity                    = sum(total_quantity)
            product_class.total_pouches_packed              = sum(total_pouches_packed) 
            product_class.total_crates_packed               = sum(total_crates_packed) 
            product_class.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)
    
    product_classess = SpProductClass.objects.filter().exclude(id=3).order_by('-id')
    for product_classs in product_classess:
        product_classs.product_variant_list              =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, sp_product_variants.variant_name, sp_product_variants.variant_size, sp_product_variants.no_of_pouch, sp_products.product_color_code  FROM sp_product_variants left join sp_products on sp_products.id = sp_product_variants.product_id WHERE sp_product_variants.product_class_id=%s order by sp_product_variants.product_class_id desc ''', [product_classs.id])
        product_classs.product_variant_total             = 'Total '+product_classs.product_class

        total_quantity              = []
        total_pouches_packed        = []
        total_crates_packed         = []
        total_quantity_in_ltr_kg    = []
        if product_classs.product_variant_list:
            for id, product_variant in enumerate(product_classs.product_variant_list):
                total_pouche_packed = 0
                total_quantitys_in_ltr_kg = 0

                try:
                    product_total_quantity = SpOrderDetails.objects.filter(product_variant_id=product_variant.id, order_date__icontains=today).aggregate(quantity = Sum('quantity'))
                    product_total_quantity = product_total_quantity['quantity']
                except SpOrderDetails.DoesNotExist:
                    product_total_quantity = 0

                if product_total_quantity:
                    quantity_wo_scheme                              = int(product_variant.no_of_pouch)*int(product_total_quantity)
                    product_variant.product_total_pouches_packed    = quantity_wo_scheme
                    total_pouche_packed                             = total_pouche_packed + quantity_wo_scheme

                    quantity_in_ltr_kg                              = (float(product_variant.variant_size)*float(product_variant.no_of_pouch))*float(product_total_quantity)
                    product_variant.product_quantity_in_ltr_kg      = quantity_in_ltr_kg
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + quantity_in_ltr_kg
                else:
                    product_variant.product_total_pouches_packed    = 0
                    total_pouche_packed                             = total_pouche_packed + 0

                    product_variant.product_quantity_in_ltr_kg      = 0
                    total_quantitys_in_ltr_kg                        = total_quantitys_in_ltr_kg + 0

                product_variant.total_quantity             = product_variant.product_total_pouches_packed    
                product_variant.total_crates_packed        = int(int(product_variant.product_total_pouches_packed)/int(product_variant.no_of_pouch))
                product_variant.total_quantity_in_ltr_kg   = product_variant.product_quantity_in_ltr_kg
                
                total_quantity.append(product_variant.total_quantity)
                total_pouches_packed.append(total_pouche_packed)
                total_crates_packed.append(int(product_variant.total_crates_packed))
                total_quantity_in_ltr_kg.append(product_variant.total_quantity_in_ltr_kg)

            product_classs.total_quantity                    = sum(total_quantity)
            product_classs.total_pouches_packed              = sum(total_pouches_packed) 
            product_classs.total_crates_packed               = sum(total_crates_packed) 
            product_classs.total_quantity_in_ltr_kg          = sum(total_quantity_in_ltr_kg)

    approved_order = SpOrders.objects.filter(order_status = 3, order_date__icontains=today).count()
    today_order = SpOrders.objects.filter(order_date__icontains=today).count()

    today_date   = date.today()
    context = {}
    context['product_classes']                      = product_classes
    context['product_classess']                     = product_classess
    context['today_order_status']                   = today_order_status
    context['order_regenerate_status']              = order_regenerate_status
    context['approved_order']                       = approved_order
    context['today_order']                          = today_order
    context['order_date']                           = today
    context['today_date']                           = today_date.strftime("%Y-%m-%d")

#export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=packing-station-report.xlsx'.format(
        date=today,
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Packing Station Report'
    worksheet.merge_cells('A1:A1') 
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    product_classes_length = []
    for product_variant in product_classes:
                for product in product_variant.product_variant_list:
                    product_classes_length.append(len(product_variant.product_variant_list))
                product_classes_length.append(len(product_variant.product_variant_list))    
    product_classess_length = []
    for product_variant in product_classess:
                for product in product_variant.product_variant_list:
                    product_classess_length.append(len(product_variant.product_variant_list))
                product_classess_length.append(len(product_variant.product_variant_list)) 

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")
    
    column_length = len(product_classes_length)+1
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH PACKING STATION : VDLLP, KHURJA, BULANDSHAHR UTTAR PRADESH'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=16, color='FFFFFFFF', bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type = "solid")

    columns = []
    columns += [ 'PARTICULARS' ]

    if product_classes:
        for product_variant in product_classes:
            for product in product_variant.product_variant_list:
                columns += [ product.variant_name ]
            columns += [ product_variant.product_variant_total ]    

    row_num = 2
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = wrapped_alignment
        cell.border = black_border
        if col_num == 1:
            cell.font = Font(size=14, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
        else:    
            cell.font = Font(size=14, color='000000', bold=True)
            cell.fill = PatternFill()             
        
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 40

    
    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL POUCH TO BE PACKED' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_quantity ]
            row += [ product_variant.total_quantity ]

            # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=12, color='000000')     
    
    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'POUCH PER CRATE' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.no_of_pouch ]
            row += [ ' ' ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=12, color='000000')   

    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL CRATES TO BE PACKED' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_crates_packed ]
            row += [ product_variant.total_crates_packed ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=12, color='000000')

    for id, product_variant in enumerate(product_classes):
        row_num += 1
        # Define the data for each cell in the row 
           
        row = []
        row += [ 'TOTAL QTY (LTR / KG)' ]
        if product_classes:
            for product in product_variant.product_variant_list:
                row += [ product.total_quantity_in_ltr_kg ]
            row += [ product_variant.total_quantity_in_ltr_kg ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:
                cell.font = Font(size=12, color='000000')    

    row_num += 2
    bottom_columns = []
    bottom_columns += [ 'PARTICULARS' ]

    if product_classess:
        for product_variant in product_classess:
            for product in product_variant.product_variant_list:
                bottom_columns += [ product.variant_name ]
            bottom_columns += [ product_variant.product_variant_total ]    

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(bottom_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:    
                cell.font = Font(size=14, color='000000', bold=True)
                cell.fill = PatternFill()             
            
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 40

    bottom_columns = []
    bottom_columns += [ 'TOTAL POUCH TO BE PACKED' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            for product in product_variant.product_variant_list:
                bottom_columns += [ product.total_quantity ]
            bottom_columns += [ product_variant.total_quantity ]    

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(bottom_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:    
                cell.font = Font(size=12, color='000000')
                cell.fill = PatternFill() 

    bottom_columns = []
    bottom_columns += [ 'POUCH PER CRATE' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            for product in product_variant.product_variant_list:
                bottom_columns += [ product.no_of_pouch ]
            bottom_columns += [ ' ' ]    

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(bottom_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:    
                cell.font = Font(size=12, color='000000')
                cell.fill = PatternFill()    

    bottom_columns = []
    bottom_columns += [ 'TOTAL CRATES TO BE PACKED' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            for product in product_variant.product_variant_list:
                bottom_columns += [ product.total_crates_packed ]
            bottom_columns += [ product_variant.total_crates_packed ]    

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(bottom_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:    
                cell.font = Font(size=12, color='000000')
                cell.fill = PatternFill()

    bottom_columns = []
    bottom_columns += [ 'TOTAL QTY (LTR / KG)' ]

    if product_classess:
        row_num += 1
        for product_variant in product_classess:
            for product in product_variant.product_variant_list:
                bottom_columns += [ product.total_quantity_in_ltr_kg ]
            bottom_columns += [ product_variant.total_quantity_in_ltr_kg ]    

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(bottom_columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = wrapped_alignment
            cell.border = black_border
            if col_num == 1:
                cell.font = Font(size=14, color='FFFFFFFF', bold=True)
                cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")
            else:    
                cell.font = Font(size=12, color='000000')
                cell.fill = PatternFill()            
    workbook.save(response)
    return response




#get order HO Report
@login_required
def hoReport(request):
    context = {}
    today                   = date.today()
    user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
    WHERE user_type = 1 AND role_id != 0 ''')

    product_milk_variant_list = SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code  
    FROM sp_product_variants LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id
    WHERE sp_product_variants.product_class_id = '3' 
    order by sp_product_variants.product_class_id asc 
    ''')
    context['product_milk_variant_list'] =  product_milk_variant_list

    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code FROM sp_product_variants 
    LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id 
    WHERE sp_product_variants.product_class_id != '3' 
    order by sp_product_variants.product_class_id asc 
    ''')

    context['product_without_milk_variant_list'] =  product_without_milk_variant_list
    context['page_title']                       = "Order Management > HO Report"
    context['product_variants'] = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
    context['users'] = user_list
    context['last_row'] = last_row = SpHoReport.objects.first()
    template                                    = 'order-management/ho-report.html'
    return render(request, template, context)

#update HO Report
@login_required
def updateHoReport(request):
    response = {}
    if request.method == "POST":

        # update history and delete data
        ho_old_records =  SpHoReport.objects.all()
        if(len(ho_old_records)):
            for ho_old_record in ho_old_records:
                ho_report_history = SpHoReportHistory()
                if ho_old_record.user_id is not None :
                    ho_report_history.user_id = ho_old_record.user_id
                    
                ho_report_history.product_variant_id = ho_old_record.product_variant_id

                if ho_old_record.user_id is not None :
                    ho_report_history.quantity = ho_old_record.quantity
                if ho_old_record.user_id is not None :
                    ho_report_history.foc_pouch = ho_old_record.foc_pouch

                ho_report_history.save()

            SpHoReport.objects.all().delete()


        #insert new data
        user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
            WHERE user_type = 1 AND role_id != 0 ''')
        for user in user_list:
            product_variant_list = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
            for product_variant in product_variant_list :
                quantity_var = "quantity_"+str(user.id)+"_"+str(product_variant.id)
                if quantity_var in request.POST :
                    ho_report = SpHoReport()
                    ho_report.user_id = user.id
                    ho_report.product_variant_id = product_variant.id
                    ho_report.quantity = request.POST[quantity_var]
                    ho_report.save() 

        product_variant_list = SpProductUnits.objects.raw(''' SELECT id FROM sp_product_variants ''')
        for product_variant in product_variant_list :
            quantity_var = "foc_quantity_"+str(product_variant.id)
            if quantity_var in request.POST :
                ho_report = SpHoReport()
                ho_report.user_id = None
                ho_report.product_variant_id = product_variant.id
                ho_report.foc_pouch = request.POST[quantity_var]
                ho_report.save()     

        response['flag'] = True
        response['message'] = "Report has been updated successfully."
    else:
        response['flag'] = False
        response['message'] = "Method not allowed"
    return JsonResponse(response)

@login_required
def exportHoReport(request):

    user_list = SpUsers.objects.raw('''SELECT id, first_name, middle_name, last_name FROM sp_users
    WHERE user_type = 1 AND role_id != 0 ''')

    product_milk_variant_list = SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code  
    FROM sp_product_variants LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id
    WHERE sp_product_variants.product_class_id = '3' 
    order by sp_product_variants.product_class_id asc 
    ''')

    product_without_milk_variant_list =  SpProductUnits.objects.raw(''' SELECT sp_product_variants.id, 
    sp_product_variants.variant_name,sp_product_variants.sp_employee, sp_products.product_color_code FROM sp_product_variants 
    LEFT JOIN sp_products on sp_products.id = sp_product_variants.product_id 
    WHERE sp_product_variants.product_class_id != '3' 
    order by sp_product_variants.product_class_id asc 
    ''')

    #export code
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Ho-report.xlsx'.format(
        date=datetime.now().strftime('%d-%m-%Y'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'HO REPORT'
    worksheet.merge_cells('A1:A1') 
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.png')
    img.height = 50
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = 'DATE('+datetime.now().strftime('%d-%m-%Y')+')'
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    column_length = len(product_milk_variant_list)+len(product_without_milk_variant_list)+2
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = 'SAAHAJ MILK PRODUCER COMPANY LIMITED, AGRA, UTTAR PRADESH'
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    columns = []
    columns += [ 'EMPLOYEE NAME' ]

    if product_milk_variant_list:
        for product_variant in product_milk_variant_list:
            columns += [ product_variant.variant_name ]

    
    if product_without_milk_variant_list:
        for product_variant in product_without_milk_variant_list:
            columns += [ product_variant.variant_name ]
    
    columns += [ 'AMOUNT' ]

    row_num = 2


    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=10, bold=True)
        cell.border = black_border

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

# price/unit row
    row_num += 1
    row = []
    row +=['PRICE/UNIT']
    total_unit_price = 0
    for product_variant in product_milk_variant_list:
        row += [ product_variant.sp_employee ]
        total_unit_price = total_unit_price + product_variant.sp_employee
    
    for product_variant in product_without_milk_variant_list:
        row += [ product_variant.sp_employee ]
        total_unit_price = total_unit_price + product_variant.sp_employee

    row += [ total_unit_price ]
    for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 

# employees row
    total_qty_amount = 0
    for user in user_list:
        row_num += 1
        row = []
        row += [ user.first_name+" "+user.middle_name+" "+user.last_name ]
        variant_total_amount = 0
        if product_milk_variant_list:
            for product_variant in product_milk_variant_list:
                row += [ get_ho_report_quantity(user.id,product_variant.id) ]
                print(float(get_ho_report_quantity(user.id,product_variant.id) * product_variant.sp_employee))
                variant_total_amount = float(variant_total_amount + float(get_ho_report_quantity(user.id,product_variant.id) * product_variant.sp_employee))

        if product_without_milk_variant_list:
            for product_variant_wm in product_without_milk_variant_list:
                row += [ get_ho_report_quantity(user.id,product_variant_wm.id) ]
                variant_total_amount = float(variant_total_amount + float(get_ho_report_quantity(user.id,product_variant_wm.id) * product_variant_wm.sp_employee))

        row += [ variant_total_amount ]
        total_qty_amount = float(total_qty_amount + variant_total_amount)
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border 
    
    # total row
    row_num += 1
    row = []
    row +=['TOTAL NO. OF POUCHES / UNITS']
    total_qty = 0
    for product_variant in product_milk_variant_list:
        row += [ get_ho_report_total_variant_qty(product_variant.id) ]
    for product_variant in product_without_milk_variant_list:
        row += [ get_ho_report_total_variant_qty(product_variant.id) ]
        
    row += [ total_qty_amount ]

    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = wrapped_alignment
        cell.border = black_border

# FOC row
    row_num += 1
    row = []
    row +=['FOC POUCHES']
    total_foc_qty = 0
    for product_variant in product_milk_variant_list:
        row += [ get_ho_report_foc(product_variant.id) ]
        total_foc_qty = total_foc_qty + get_ho_report_foc(product_variant.id)
    
    for product_variant in product_without_milk_variant_list:
        row += [ get_ho_report_foc(product_variant.id) ]
        total_foc_qty = total_foc_qty + get_ho_report_foc(product_variant.id)
        
    row += [ total_foc_qty ]

    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = cell_value
        cell.alignment = wrapped_alignment
        cell.border = black_border 

    workbook.save(response)

    return response
