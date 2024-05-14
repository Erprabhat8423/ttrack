import sys
import os
import time
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from utils import *
from datetime import datetime, date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.core.files.storage import FileSystemStorage
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings


@login_required
def locateLeadOnMap(request):
    baseurl = settings.BASE_URL
    context = {}
    users      = SpLeadBasic.objects.filter(latitude__isnull=False,longitude__isnull=False)
    for user in users:
        country_code = getModelColumnById(SpCountryCodes, user.contry_code_id, "country_code")
        user.phone_code = country_code
        
        last_follow_up_date = SpFollowUp.objects.filter(lead_id=user.id).order_by('-id').first()
        if last_follow_up_date:
            user.last_follow_up_date = last_follow_up_date.created_at.strftime("%d/%m/%Y %I:%M %p")
        else:
            user.last_follow_up_date = None
            
        created = user.created_at
        user.created_at_date = created.strftime("%d/%m/%Y %I:%M %p")
        first_name = getModelColumnById(SpUsers, user.created_by_id, "first_name")
        last_name = getModelColumnById(SpUsers, user.created_by_id, "last_name")
        created_by_name = first_name + " " + last_name
        user.created_by_name = created_by_name
    context['users'] = users
    context['baseurl'] = baseurl
    template = 'lead-report/locate-lead-on-map.html'

    return render(request, template, context)

@login_required
def locateSingleLeadOnMap(request):
    baseurl = settings.BASE_URL
    context = {}
    users      = SpLeadBasic.objects.filter(latitude__isnull=False,longitude__isnull=False,id =request.GET['lead_id'])
    for user in users:
        country_code = getModelColumnById(SpCountryCodes, user.contry_code_id, "country_code")
        user.phone_code = country_code
        
        last_follow_up_date = SpFollowUp.objects.filter(lead_id=user.id).order_by('-id').first()
        if last_follow_up_date:
            user.last_follow_up_date = last_follow_up_date.created_at.strftime("%d/%m/%Y %I:%M %p")
        else:
            user.last_follow_up_date = None
        created = user.created_at
        user.created_at_date = created.strftime("%d/%m/%Y %I:%M %p")
        first_name = getModelColumnById(SpUsers, user.created_by_id, "first_name")
        last_name = getModelColumnById(SpUsers, user.created_by_id, "last_name")
        created_by_name = first_name + " " + last_name
        user.created_by_name = created_by_name
    context['users'] = users
    context['baseurl'] = baseurl
    template = 'lead-report/locate-lead-on-map.html'

    return render(request, template, context)
    

@login_required
def filterLead(request,status):
    today = date.today() 
    page = request.GET.get('page')
    if status <= 0:  
        lead_lists              = SpLeadBasic.objects.filter().order_by('-id')
    else:
        lead_lists              = SpLeadBasic.objects.filter(status = status).order_by('-id')
    lead_count                  = lead_lists.count()
    for lead_list in lead_lists:
        first_name              = getModelColumnById(SpUsers, lead_list.created_by_id, "first_name")
        last_name               = getModelColumnById(SpUsers, lead_list.created_by_id, "last_name")
        created_by_name         = first_name + " " + last_name
        lead_list.created_by_name = created_by_name
        last_follow_up_date     = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
        if last_follow_up_date:
            lead_list.formatted_last_follow_up_date = last_follow_up_date.created_at
        else:
            lead_list.formatted_last_follow_up_date = None
    
    paginator = Paginator(lead_lists, 20)
    try:
        lead_lists = paginator.page(page)
    except PageNotAnInteger:
        lead_lists = paginator.page(1)
    except EmptyPage:
        lead_lists = paginator.page(paginator.num_pages)  
    if page is not None:
        page = page
    else:
        page = 1
    total_pages = math.ceil(paginator.count/20) 
    
    if(paginator.count == 0):
        paginator.count = 1
  
    temp = total_pages%paginator.count
    if(temp > 0 and 20!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages
    context                 = {}
    context['lead_lists']   = lead_lists
    context['lead_status']   = status
    context['lead_id']      = lead_lists[0].id if len(lead_lists)>0 else 0
    context['user_lists']    = SpUsers.objects.filter(status=1,user_type=1)
    context['today_date']   = today.strftime("%d/%m/%Y")
    context['page_title']   = "Lead Management"
    context['lead_count']   = lead_count
    context['total_pages'] = total_pages
    context['page_limit'] = 20
    template = 'lead-report/index.html'
    return render(request, template, context)
    
@login_required
def filterLeads(request,ids):
    page = request.GET.get('page')
    today = date.today() 
    try:
        lead_lists              = SpLeadBasic.objects.filter(id = ids).order_by('-id')
        lead_ids                = "TTRACK"+str(ids)
    except:
        lead_lists = []
        lead_ids = ""
    lead_count                  = lead_lists.count()
    for lead_list in lead_lists:
        first_name = getModelColumnById(SpUsers, lead_list.created_by_id, "first_name")
        last_name = getModelColumnById(SpUsers, lead_list.created_by_id, "last_name")
        
        created_by_name = first_name + " " + last_name
        lead_list.created_by_name = created_by_name
        last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
        if last_follow_up_date:
            lead_list.formatted_last_follow_up_date = last_follow_up_date.created_at
        else:
            lead_list.formatted_last_follow_up_date = None
    paginator = Paginator(lead_lists, 20)
    try:
        lead_lists = paginator.page(page)
    except PageNotAnInteger:
        lead_lists = paginator.page(1)
    except EmptyPage:
        lead_lists = paginator.page(paginator.num_pages)  
    if page is not None:
        page = page
    else:
        page = 1
    total_pages = math.ceil(paginator.count/20) 
    
    if(paginator.count == 0):
        paginator.count = 1
  
    temp = total_pages%paginator.count
    if(temp > 0 and 20!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context                 = {}
    context['lead_lists']   = lead_lists
    context['lead_ids']   = lead_ids
    context['lead_id']      = lead_lists[0].id if len(lead_lists)>0 else 0
    context['user_lists']    = SpUsers.objects.filter(status=1,user_type=1)
    context['today_date']   = today.strftime("%d/%m/%Y")
    context['lead_count']   = lead_count
    context['page_title']   = "Lead Management"
    context['total_pages'] = total_pages
    context['page_limit'] = 20
    template = 'lead-report/index.html'
    return render(request, template, context)


@login_required
def index(request):
    page = request.GET.get('page')
    today = date.today()  
    lead_lists              = SpLeadBasic.objects.filter().order_by('-id')
    lead_count                  = lead_lists.count()
    for lead_list in lead_lists:
        first_name = getModelColumnById(SpUsers, lead_list.created_by_id, "first_name")
        last_name = getModelColumnById(SpUsers, lead_list.created_by_id, "last_name")
        
        created_by_name = first_name + " " + last_name
        lead_list.created_by_name = created_by_name
        last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
        if last_follow_up_date:
            lead_list.formatted_last_follow_up_date = last_follow_up_date.created_at
        else:
            lead_list.formatted_last_follow_up_date = None
            
    paginator = Paginator(lead_lists, 20)
    try:
        lead_lists = paginator.page(page)
    except PageNotAnInteger:
        lead_lists = paginator.page(1)
    except EmptyPage:
        lead_lists = paginator.page(paginator.num_pages)  
    if page is not None:
        page = page
    else:
        page = 1
    total_pages = math.ceil(paginator.count/20) 
    
    if(paginator.count == 0):
        paginator.count = 1
  
    temp = total_pages%paginator.count
    if(temp > 0 and 20!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context                 = {}
    context['lead_lists']   = lead_lists
    context['lead_id']      = lead_lists[0].id if len(lead_lists)>0 else 0
    context['user_lists']    = SpUsers.objects.filter(status=1,user_type=1)
    context['today_date']   = today.strftime("%d/%m/%Y")
    context['page_title']   = "Lead Management"
    context['lead_count']   = lead_count
    context['total_pages'] = total_pages
    context['page_limit'] = 20
    template = 'lead-report/index.html'
    return render(request, template, context)

@login_required
def ajaxLeadList(request):
    page                    = request.GET.get('page')
    order_date              = request.GET['order_date']
    customer_id             = request.GET['customer_id']
    order_status            = request.GET['order_status']
    if order_date:
        start_date          = datetime.strptime(order_date.split(" - ")[0],'%d/%m/%Y').strftime('%Y-%m-%d')
        start_end           = datetime.strptime(order_date.split(" - ")[1],'%d/%m/%Y').strftime('%Y-%m-%d')
        to_date             = datetime.strptime(str(start_end), "%Y-%m-%d")+ timedelta(days=1)
        to_date             = str(to_date).replace(' 00:00:00', '')
        lead_lists          = SpLeadBasic.objects.filter(created_at__range = [start_date,to_date] ).order_by('-id')
    else:
        lead_lists          = SpLeadBasic.objects.filter().order_by('-id')
    if customer_id:
        lead_lists          = lead_lists.filter(created_by_id = customer_id)
    if order_status:
        lead_lists          = lead_lists.filter(status = order_status)
    lead_count              = lead_lists.count()
    for lead_list in lead_lists:
        first_name = getModelColumnById(SpUsers, lead_list.created_by_id, "first_name")
        last_name = getModelColumnById(SpUsers, lead_list.created_by_id, "last_name")
        
        created_by_name = first_name + " " + last_name
        lead_list.created_by_name = created_by_name
        last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
        if last_follow_up_date:
            lead_list.formatted_last_follow_up_date = last_follow_up_date.created_at
        else:
            lead_list.formatted_last_follow_up_date = None
            
        country_code = getModelColumnById(SpCountryCodes, lead_list.contry_code_id, "country_code")
        lead_list.phone_code = country_code
    
    paginator = Paginator(lead_lists, 20)
    try:
        lead_lists = paginator.page(page)
    except PageNotAnInteger:
        lead_lists = paginator.page(1)
    except EmptyPage:
        lead_lists = paginator.page(paginator.num_pages)  
    if page is not None:
        page = page
    else:
        page = 1
    total_pages = math.ceil(paginator.count/20) 
    
    if(paginator.count == 0):
        paginator.count = 1
  
    temp = total_pages%paginator.count
    if(temp > 0 and 20!= paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context                 = {}
    context['lead_lists']   = lead_lists
    context['lead_count']   = lead_count
    context['total_pages'] = total_pages
    context['page_loading_type'] = 1
    context['lead_id']      = lead_lists[0].id if len(lead_lists)>0 else 0
    template = 'lead-report/ajax-lead-list.html'
    return render(request, template, context)

@login_required
def getLeadDetails(request):  # sourcery skip: avoid-builtin-shadow
    
    id                                     = request.GET.get('id')
    status_dict                            = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    lead_list                              = SpLeadBasic.objects.filter(id = id).first()
    if lead_list:
        lead_list.status_name              = status_dict[lead_list.status]
        list_of_strings                    = lead_list.core_business_area.split(',')
        core_business_area                 = [int(item) for item in list_of_strings ]
        lead_list.core_business_area_list  = SpCoreBusinessArea.objects.filter(id__in = core_business_area).values_list('core_business_area_name',flat=True)
    else:
        lead_list                          = None
    last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
    if last_follow_up_date:
        lead_list.formatted_last_follow_up_date = last_follow_up_date.created_at
    else:
        lead_list.formatted_last_follow_up_date = None
    country_code = getModelColumnById(SpCountryCodes, lead_list.contry_code_id, "country_code")
    lead_list.phone_code = country_code
    first_name = getModelColumnById(SpUsers, lead_list.created_by_id, "first_name")
    last_name = getModelColumnById(SpUsers, lead_list.created_by_id, "last_name")
        
    created_by_name = first_name + " " + last_name
    lead_list.created_by_name = created_by_name
    lead_iso                               = SpLeadIso.objects.filter(last_lead_id = id)
    
    sp_lead_iso_saves                      = SpLeadIsoSave.objects.filter(last_lead_id = id)
    
    for sp_lead_iso_save in sp_lead_iso_saves:
        list_of_strings                    = sp_lead_iso_save.iso_created_id.split(',')
        iso_ids                            = [int(item) for item in list_of_strings ]
        iso_lists                          = SpIsoMaster.objects.filter(id__in = iso_ids).values()
        result_string                      = ""
        count                              = 0
        for asset_data in iso_lists:
            count += 1
            asset_name                     = asset_data['iso_name']
            result_string                  += f"{count}.{asset_name}\n"
        sp_lead_iso_save.item_list         = result_string

    leade_others                                = SpLeadOther.objects.filter(last_lead_id = id)
    for leade_other in leade_others:
        if leade_other.software_or_erp:
            leade_other.product_pitch           = "NO" if leade_other.other_production_pitch == 0 else "Yes"
            list_of_strings                     = leade_other.software_or_erp.split(',')
            software_or_erp                     = [int(item) for item in list_of_strings ]
            software_or_erp                     = TtrackService.objects.filter(id__in = software_or_erp).values('service_name') 
            result_string                       = ""
            count                               = 0
            for asset_data in software_or_erp:
                count += 1
                asset_name                      = asset_data['service_name']
                result_string                   += f"{count}.{asset_name}\n"
            leade_other.item_list               = result_string
        else:
            leade_other.item_list               = None
        leade_other.sales_person                = getUserName(leade_other.sales_person) 
        leade_other.other_resource              = getUserName(leade_other.other_resource) if leade_other.other_resource else None
    context = {
        'lead_lists':lead_list,
        'lead_iso':lead_iso,
        'sp_lead_iso_saves':sp_lead_iso_saves,
        'leade_others':leade_others,
        
    }
    template = 'lead-report/get-lead-details.html'
    return render(request, template, context)



# @login_required
# def exportToXlsx(request, customer_id,order_status,start_date,end_date):
#     today = date.today() 
#     lead_lists                          = SpLeadBasic.objects.filter()
#     if start_date!='0' and end_date !='0':
#         to_date                         = datetime.strptime(str(end_date), "%Y-%m-%d")+ timedelta(days=1)
#         to_date                         = str(to_date).replace(' 00:00:00', '')
#         lead_lists                      = SpLeadBasic.objects.filter(created_at__range = [start_date,to_date]) 
#     if customer_id!='0':
#         lead_lists                      = lead_lists.filter(created_by_id = customer_id)
#     if order_status!='0':
#         lead_lists                      = lead_lists.filter(status = order_status)
#     status_dict                         = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
#     for lead_list in lead_lists:
#         lead_list.created_by_name       = getUserName(lead_list.created_by_id)
#         lead_list.basic_date            = datetime.strptime(str(lead_list.basic_date), '%Y-%m-%d').strftime('%d/%m/%Y')
#         lead_list.created_at            = datetime.strptime(str(lead_list.created_at), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')
#         lead_list.status_name           = status_dict[lead_list.status]

#         list_of_strings                 = lead_list.core_business_area.split(',')
#         core_business_area              = [int(item) for item in list_of_strings ]
#         core_business_area_list         = SpCoreBusinessArea.objects.filter(id__in = core_business_area).values('core_business_area_name')
#         result_string                   = ""
#         count                           = 0
#         country_code                    = getModelColumnById(SpCountryCodes, lead_list.contry_code_id, "country_code")
#         lead_list.phone_code           = country_code
#         for asset_data in core_business_area_list:
#             count += 1
#             asset_name                  = asset_data['core_business_area_name']
#             result_string += f"{count}.{asset_name}\n"
#         lead_list.item_list             = result_string

#         last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
#         if last_follow_up_date:
#             lead_list.formatted_last_follow_up_date = datetime.strptime(str(last_follow_up_date.created_at), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')
#         else:
#             lead_list.formatted_last_follow_up_date = None
#     context = {}
#     context['lead_lists']               = lead_lists

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     # response['Content-Disposition'] = 'attachment; filename=Lead Management.xlsx'.format(
#     #     date=datetime.now().strftime('%Y-%m-%d'),
#     # )
#     response['Content-Disposition'] = 'attachment; filename=Lead Report {}.xlsx'.format(
#         datetime.now().strftime('%d-%m-%Y')
#     )
#     workbook = Workbook()

#     # Define some styles and formatting that will be later used for cells
#     header_font = Font(name='Calibri', bold=True)
#     centered_alignment = Alignment(horizontal='left')
#     thin = Side(border_style="thin", color="303030") 
#     black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#     wrapped_alignment = Alignment(
#         vertical='top',
#         horizontal='left',
#         wrap_text=True
#     )

#     header_alignment = Alignment(
#         vertical='top',
#         horizontal='center',
#         wrap_text=True
#     )
    
#     # Get active worksheet/tab
#     worksheet = workbook.active
#     worksheet.title = 'Lead Report'
#     worksheet.merge_cells('A1:A1') 
    
#     worksheet.page_setup.orientation = 'landscape'
#     worksheet.page_setup.paperSize = 9
#     worksheet.page_setup.fitToPage = True
    
#     worksheet = workbook.worksheets[0]
#     img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
#     img.height = 48
#     img.width = 95
#     img.alignment = 'center'
#     img.anchor = 'A1'
#     worksheet.add_image(img)
    
#     column_length = 13
    
#     worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#     worksheet.cell(row=1, column=2).value = 'Lead Report as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
#     worksheet.cell(row=1, column=2).font = header_font
#     worksheet.cell(row=1, column=2).alignment = header_alignment
#     worksheet.cell(row=1, column=column_length).border = black_border
#     worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
#     worksheet.cell(row=1, column=2).fill = PatternFill()

#     # Define the titles for columns
#     # columns = []
#     row_num = 1
#     worksheet.row_dimensions[1].height = 40
    
#     # Define the titles for columns
#     columns = []

#     columns += [ 'Lead ID' ]
#     columns += [ 'Company Name' ]
#     columns += [ 'Turnover' ] 
#     columns += [ 'Total Employees' ]
#     columns += [ 'Contact Person' ]
#     columns += [ 'Contact No.' ]
#     columns += [ 'Desk No.' ]
#     columns += [ 'Email' ]
#     columns += [ 'Address' ]
#     columns += [ 'Current Status' ]
#     columns += [ 'Created By' ]
#     columns += [ 'Last FollowUp' ]
#     columns += [ 'Created At' ]

#     row_num = 2

#     # Assign the titles for each cell of the header
#     for col_num, column_title in enumerate(columns, 1):
#         cell = worksheet.cell(row=row_num, column=col_num)
#         cell.value = column_title
#         cell.font = header_font
#         cell.alignment = centered_alignment
#         cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#         cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#         column_letter = get_column_letter(col_num)
#         column_dimensions = worksheet.column_dimensions[column_letter]
#         column_dimensions.width = 20

     
#     for lead in lead_lists:
        
#         row_num += 1
#         row = []
#         row += [ "TTRACK"+str(lead.id) ]
#         row += [ lead.company_name ]
#         if lead.currency_code:
#             row += [ "{} {}".format(lead.currency_code, lead.turnover) ]
#         else:
#             row += [ lead.turnover ]
#         row += [ lead.total_no_of_employee ]
#         row += [ lead.contact_person_name ]
#         row += [ "'{}-{}".format(lead.phone_code, lead.mobile_no) ]
#         row += [ lead.desk_no ]
#         row += [ lead.email ]
#         row += [ lead.address ]
#         row += [ lead.status_name ]
#         row += [ lead.created_by_name ]
#         row += [ lead.formatted_last_follow_up_date ]
#         row += [ lead.created_at ]

#         # Assign the data for each cell of the row 
#         for col_num, cell_value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = cell_value
#             cell.alignment = wrapped_alignment
#             cell.border = black_border    
#     wrapped_alignment = Alignment(
#         horizontal='center',
#         wrap_text=True
#         )
#     row_num += 1
#     last_row = row_num
#     worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=13)
#     worksheet.row_dimensions[last_row].height = 20
#     worksheet.cell(row=last_row, column=1).value = 'Generated By Emobic'
#     worksheet.cell(row=last_row, column=1).font = header_font
#     worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#     worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
#     worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

#     workbook.save(response)
#     return response

@login_required
def exportToXlsx(request, customer_id,order_status,start_date,end_date):
    today = date.today() 
    lead_lists                          = SpLeadBasic.objects.filter()
    if start_date!='0' and end_date !='0':
        to_date                         = datetime.strptime(str(end_date), "%Y-%m-%d")+ timedelta(days=1)
        to_date                         = str(to_date).replace(' 00:00:00', '')
        lead_lists                      = SpLeadBasic.objects.filter(created_at__range = [start_date,to_date]) 
    if customer_id!='0':
        lead_lists                      = lead_lists.filter(created_by_id = customer_id)
    if order_status!='0':
        lead_lists                      = lead_lists.filter(status = order_status)
    status_dict                         = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    for lead_list in lead_lists:
        lead_list.created_by_name       = getUserName(lead_list.created_by_id)
        lead_list.basic_date            = datetime.strptime(str(lead_list.basic_date), '%Y-%m-%d').strftime('%d/%m/%Y')
        lead_list.created_at            = datetime.strptime(str(lead_list.created_at), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')
        lead_list.status_name           = status_dict[lead_list.status]

        list_of_strings                 = lead_list.core_business_area.split(',')
        core_business_area              = [int(item) for item in list_of_strings ]
        core_business_area_list         = SpCoreBusinessArea.objects.filter(id__in = core_business_area).values('core_business_area_name')
        result_string                   = ""
        count                           = 0
        country_code                    = getModelColumnById(SpCountryCodes, lead_list.contry_code_id, "country_code")
        lead_list.phone_code           = country_code
        for asset_data in core_business_area_list:
            count += 1
            asset_name                  = asset_data['core_business_area_name']
            result_string += f"{count}.{asset_name}\n"
        lead_list.item_list             = result_string

        last_follow_up_date = SpFollowUp.objects.filter(lead_id=lead_list.id).order_by('-id').first()
        if last_follow_up_date:
            lead_list.formatted_last_follow_up_date = datetime.strptime(str(last_follow_up_date.created_at), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M:%S')
        else:
            lead_list.formatted_last_follow_up_date = None
        lead_iso = SpLeadIso.objects.filter(last_lead_id=lead_list.id).values_list('master_iso_id', flat=True)
        iso_with_string = ", ".join([f"ISO {iso}" for iso in lead_iso])
        lead_list.existing_iso = iso_with_string
        sp_lead_iso_saves                      = SpLeadIsoSave.objects.filter(last_lead_id = lead_list.id)
        
        for sp_lead_iso_save in sp_lead_iso_saves:
            
            list_of_strings                    = sp_lead_iso_save.iso_created_id.split(',')
            iso_ids                            = [int(item) for item in list_of_strings ]
            iso_lists                          = SpIsoMaster.objects.filter(id__in = iso_ids).values()
            result_string                      = ""
            count                              = 0
            for asset_data in iso_lists:
                count += 1
               
               
                iso_code = asset_data['iso_id'] 
                result_string += f"ISO {iso_code},"
            item_list         = result_string
            item_cost = sp_lead_iso_save.iso_amount
            item_currency_code = sp_lead_iso_save.currency_code
        lead_list.suggested_iso = item_list
        lead_list.suggested_iso_cost = item_cost
        lead_list.currency_code_val = item_currency_code
        print(item_currency_code)
    context = {}
    context['lead_lists']               = lead_lists

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # response['Content-Disposition'] = 'attachment; filename=Lead Management.xlsx'.format(
    #     date=datetime.now().strftime('%Y-%m-%d'),
    # )
    response['Content-Disposition'] = 'attachment; filename=Lead Report {}.xlsx'.format(
        datetime.now().strftime('%d-%m-%Y')
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        horizontal='left',
        wrap_text=True
    )

    header_alignment = Alignment(
        vertical='top',
        horizontal='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Lead Report'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
    img.height = 48
    img.width = 95
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)
    
    column_length = 16
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'Lead Report as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = header_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill()

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    
    # Define the titles for columns
    columns = []

    columns += [ 'Lead ID' ]
    columns += [ 'Company Name' ]
    columns += [ 'Turnover' ] 
    columns += [ 'Total Employees' ]
    columns += [ 'Contact Person' ]
    columns += [ 'Contact No.' ]
    columns += [ 'Desk No.' ]
    columns += [ 'Email' ]
    columns += [ 'Address' ]
    columns += [ 'Current Status' ]
    columns += [ 'Created By' ]
    columns += [ 'Last FollowUp' ]
    columns += ['Suggested ISO']
    columns += ['Suggested Cost']
    columns += ['Existing ISO']
    columns += [ 'Created At' ]



    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

     
    for lead in lead_lists:
        
        row_num += 1
        row = []
        row += [ "TTRACK"+str(lead.id) ]
        row += [ lead.company_name ]
        if lead.currency_code:
            row += [ "{} {}".format(lead.currency_code, lead.turnover) ]
        else:
            row += [ lead.turnover ]
        row += [ lead.total_no_of_employee ]
        row += [ lead.contact_person_name ]
        row += [ "'{}-{}".format(lead.phone_code, lead.mobile_no) ]
        row += [ lead.desk_no ]
        row += [ lead.email ]
        row += [ lead.address ]
        row += [ lead.status_name ]
        row += [ lead.created_by_name ]
        row += [ lead.formatted_last_follow_up_date ]
        row += [ lead.suggested_iso]
        row += ["{} {}".format(lead.currency_code_val, lead.suggested_iso_cost)]
        row += [ lead.existing_iso ]
        row += [ lead.created_at ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
    wrapped_alignment = Alignment(
        horizontal='center',
        wrap_text=True
        )
    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=13)
    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Emobic'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

    workbook.save(response)
    return response




@login_required
def showActivityLog(request):
    lead_id = request.GET.get('lead_id')
    lead_folloups = SpFollowUp.objects.filter(lead_id = lead_id).order_by('-created_at')
    for lead_folloup in lead_folloups:
        lead_folloup.reson_name  = getModelColumnById(SpReasons,lead_folloup.reason_id,'reason') if lead_folloup.reason_id else ""
        lead_folloup.company_name  = getModelColumnById(SpLeadBasic,lead_folloup.lead_id,'company_name') if lead_folloup.lead_id else ""
        lead_folloup.lead_code      = "TTRACK"+str(lead_folloup.lead_id)
        lead_folloup.employee_name  = getUserName(lead_folloup.created_by)
    template = 'lead-report/show-activity-log.html'
    context = {'lead_folloups':lead_folloups,'lead_id':lead_id}
    return render(request, template,context)


def leadFullDetail(request,lead_id):
    status_dict                         = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    lead_list                              = SpLeadBasic.objects.filter(id = lead_id).first()
    if lead_list:
        lead_list.status_name              = status_dict[lead_list.status]
        list_of_strings                    = lead_list.core_business_area.split(',')
        core_business_area                 = [int(item) for item in list_of_strings ]
        lead_list.core_business_area_list  = SpCoreBusinessArea.objects.filter(id__in = core_business_area).values_list('core_business_area_name',flat=True)
    else:
        lead_list                          = None
    country_code = getModelColumnById(SpCountryCodes, lead_list.contry_code_id, "country_code")
    lead_list.phone_code = country_code
    lead_iso                               = SpLeadIso.objects.filter(last_lead_id = lead_id)
    
    sp_lead_iso_saves                      = SpLeadIsoSave.objects.filter(last_lead_id = lead_id)
    result_list =[]
    for sp_lead_iso_save in sp_lead_iso_saves:
        list_of_strings                    = sp_lead_iso_save.iso_created_id.split(',')
        iso_ids                            = [int(item) for item in list_of_strings ]
        iso_lists                          = SpIsoMaster.objects.filter(id__in = iso_ids).values()
        result_string                      = ""
        count                              = 0
        for asset_data in iso_lists:
            count += 1
            # asset_name                     = asset_data['iso_name']
            # result_string                  += f"{count}.{asset_name}\n"
            asset_name                     = asset_data['iso_name']
            iso_code = asset_data['iso_id'] 
            result_string = f"ISO {iso_code} - {asset_name}"
            result_list.append(result_string)
        sp_lead_iso_save.item_list         = result_string
    
    leade_others                                = SpLeadOther.objects.filter(last_lead_id = lead_id)
    intrested_result_list = []
    for leade_other in leade_others:
        if leade_other.software_or_erp:
           
            leade_other.product_pitch           = "NO" if leade_other.other_production_pitch == 0 else "Yes"
            list_of_strings                     = leade_other.software_or_erp.split(',')
            software_or_erp                     = [int(item) for item in list_of_strings ]
            software_or_erp                     = TtrackService.objects.filter(id__in = software_or_erp).values('service_name') 
            result_string                       = ""
            count                               = 0
            for asset_data in software_or_erp:
                count += 1
                asset_name                      = asset_data['service_name']
                result_string                   = f"{asset_name}\n"
                intrested_result_list.append(result_string)
        #     leade_other.item_list               = result_string
        # else:
        #     leade_other.item_list               = None
        
        leade_other.sales_person                = getUserName(leade_other.sales_person) 
        leade_other.other_resource              = getUserName(leade_other.other_resource) if leade_other.other_resource else None
    if intrested_result_list != []:
        intrested_result_list               = intrested_result_list
    else:
        intrested_result_list              = None
    lead_folloups = SpFollowUp.objects.filter(lead_id = lead_id).order_by('-created_at')
    for lead_folloup in lead_folloups:
        lead_folloup.reson_name  = getModelColumnById(SpReasons,lead_folloup.reason_id,'reason') if lead_folloup.reason_id else ""
        lead_folloup.company_name  = getModelColumnById(SpLeadBasic,lead_folloup.lead_id,'company_name') if lead_folloup.lead_id else ""
        lead_folloup.lead_code      = "TTRACK"+str(lead_folloup.lead_id)
        lead_folloup.employee_name  = getUserName(lead_folloup.created_by)
    
    context = {
        'lead_lists':lead_list,
        'lead_iso':lead_iso,
        'sp_lead_iso_saves':sp_lead_iso_saves,
        'leade_others':leade_others,
        'lead_folloups':lead_folloups,
        'result_list':result_list,
        'intrested_result_list':intrested_result_list
        
    }
    
    template = 'lead-report/get-full-lead-details.html'
    return render(request, template, context)
    

@login_required
def leadReport(request):
    sp_lead_list = []
    status_dict  = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    spleadiso = SpLeadIso.objects.filter().values_list('last_lead_id', flat=True).distinct()
    if spleadiso:
        spleadbasic = SpLeadBasic.objects.filter(id__in=spleadiso)
        for spleadbasic in leads:
            sp_lead_dict = {
                "id": spleadbasic.id,
                "lead_code": "TTRACK"+str(spleadbasic.id),
                "country_code":getModelColumnById(SpCountryCodes, spleadbasic.contry_code_id,'country_code') if leads_basic.contry_code_id else "",
                "created_by_id": spleadbasic.created_by_id,
                "basic_date": spleadbasic.basic_date,
                "company_name": spleadbasic.company_name,
                "turnover": spleadbasic.turnover,
                "contact_person_name": spleadbasic.contact_person_name,
                "desk_no": spleadbasic.desk_no,
                "mobile_no": spleadbasic.mobile_no,
                "contry_code_id": spleadbasic.contry_code_id,
                "email": spleadbasic.email,
                "address": spleadbasic.address,
                "total_no_of_employee": spleadbasic.total_no_of_employee,
                "currency_code": spleadbasic.currency_code,
                "deal_amount": spleadbasic.deal_amount,
                "deal_date_time": spleadbasic.deal_date_time,
                "deal_currency_code": spleadbasic.deal_currency_code,
                "created_at": spleadbasic.created_at,
                "updated_at": spleadbasic.updated_at,
                "status": spleadbasic.status,
                "longitude": spleadbasic.longitude,
                "latitude": spleadbasic.latitude,
                "lead_status":status_dict[lead_list.status],
                "lead_iso_list": list(SpLeadIso.objects.filter(last_lead_id=spleadbasic.id).values())
            }
            sp_lead_list.append(sp_lead_dict)
    
    template = 'lead-report/lead-report.html'
    context = {}
    context['user_lists']    = SpUsers.objects.filter(status=1,user_type=1)
    context['sp_lead_list']  = sp_lead_list
    context['page_title']    = "Lead Report"
    return render(request, template,context)

@login_required
def editLeadBasicDetails(request,lead_id):
    response = {}
    if request.method == "POST":
        try:
          
            lead_basic = SpLeadBasic.objects.get(id =lead_id)

            lead_basic.company_name     = request.POST['company_name']
            lead_basic.turnover         = request.POST['turnover']
            lead_basic.contact_person_name  = request.POST['contact_person_name']
            lead_basic.desk_no          = request.POST['desk_no']
            lead_basic.contry_code_id   = request.POST['country_code']
            lead_basic.mobile_no        = request.POST['contact_no']
            lead_basic.email            = request.POST['official_email']
            lead_basic.address          = request.POST['address']
            lead_basic.total_no_of_employee = request.POST['total_no_of_employee']
            core_business_area_list         = request.POST.getlist('core_business_area[]') 
            core_business_area_str          = ','.join(core_business_area_list)
            lead_basic.core_business_area   = core_business_area_str
            lead_basic.currency_code        = request.POST['currancy_code']
            lead_basic.save() 
          
            message  = "Lead Basic Deatils Saved Successfully"
            response = {}
            response['error'] = False
            response['message'] = message
            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    else:
        core_business_area  = SpCoreBusinessArea.objects.filter(status = 1).order_by('core_business_area_name')
        country_code        = SpCountryCodes.objects.filter(status = 1).order_by('country_code')
        lead_basic_details  = SpLeadBasic.objects.filter(id = lead_id).first()
        currancy_code       = SpCurrencyCode.objects.filter(status = 1)
        core_business_ids   = lead_basic_details.core_business_area
    
        context                 = {}
        
        context['core_business_area']   = core_business_area
        context['user_core_area']       = core_business_ids.split(',')
        context['country_codes']        = country_code
        context['currancy_code']        = currancy_code
        context['lead_id']              = lead_id
        context['lead_basic_details']   = lead_basic_details

        template = 'lead-report/edit-lead-basic-details-new.html'

        return render(request, template, context)

@login_required


def editLeadIsoDetails(request,lead_id,iso_status):
    response = {}
    if request.method == "POST":
     
        try:
            
            if  iso_status != 0 and iso_status != 2:
               
                iso_applicable_id           = request.POST.getlist('suggested_iso[]')
                existing_iso_id             = request.POST.getlist('existing_iso_id[]')
                date_of_issue               = request.POST.getlist('date_of_issue[]')
                date_of_survilance1       = request.POST.getlist('date_of_survilance1[]')
                date_of_survilance2       = request.POST.getlist('date_of_survilance2[]')
                date_of_expiry          = request.POST.getlist('date_of_expiry[]')
                copy_of_iso_files       = request.FILES.getlist('iso_doc[]')   
                file_ids                = request.POST.getlist('file_ids[]')
                file_ids                = [item for item in file_ids if item] 
                iso_issued_agency       = request.POST.getlist('iso_issued_agency[]')
                currency_code           = request.POST.getlist('currency_code[]')
                iso_issued_consultant     = request.POST.getlist('iso_issued_consultant[]')
                price_of_existing_iso     = request.POST.getlist('price_of_existing_iso[]')
              
                i = 0

                for id in range(0, len(existing_iso_id)):
                    if existing_iso_id[id] != '0':
                        lead = SpLeadIso.objects.get(id = existing_iso_id[id] )
                        lead.created_by_id                   = int(request.POST['created_by_id'])
                        lead.last_lead_id                    = lead_id
                        lead.iso_applicable_id               = iso_applicable_id[id]
                        lead.date_of_issue                   = datetime.strptime(date_of_issue[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_survilance1             = datetime.strptime(date_of_survilance1[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_survilance2             = datetime.strptime(date_of_survilance2[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_expiry                  = datetime.strptime(date_of_expiry[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.master_iso_id                   = getModelColumnById(SpIsoMaster, iso_applicable_id[id], 'iso_id')
                        lead.master_iso_name                 = getModelColumnById(SpIsoMaster, iso_applicable_id[id], 'iso_name')
                        lead.currency_code                   = currency_code[id]

                        if str(id) in file_ids:
                            if copy_of_iso_files[i]:
                                uploaded_iso = copy_of_iso_files[i]
                                storage = FileSystemStorage()
                                timestamp = int(time.time())
                                iso_name = uploaded_iso.name
                                temp = iso_name.split('.')
                                iso_name = 'media/iso_image/'+str(timestamp)+"."+temp[(len(temp) - 1)] # Use negative indexing to get the file extension
                                iso = storage.save(iso_name, uploaded_iso)
                                lead.copy_of_iso = storage.url(iso) 
                                i = i+1 
                            else:
                                lead.copy_of_iso = lead.copy_of_iso
                        else:
                            lead.copy_of_iso = lead.copy_of_iso
                        
                        lead.iso_issued_agency               = iso_issued_agency[id]
                        lead.iso_issued_consultant           = iso_issued_consultant[id]
                        lead.price_of_existing_iso           = price_of_existing_iso[id]
                        lead.iso_status                      = 1
                        lead.status  = 1
                        
                        try:
                            sp_lead_iso_save_instance = SpLeadIsoSave.objects.get(last_lead_id=lead_id)
                            sp_lead_iso_save_instance.delete()
                        except SpLeadIsoSave.DoesNotExist:
                            # Handle the case where no instance is found
                            pass  
                        
                        lead.save()
                    else:
                       
                        lead = SpLeadIso()
                        lead.created_by_id                   = int(request.POST['created_by_id'])
                        lead.last_lead_id                    = lead_id
                        lead.iso_applicable_id               = iso_applicable_id[id]
                        lead.date_of_issue                   = datetime.strptime(date_of_issue[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_survilance1             = datetime.strptime(date_of_survilance1[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_survilance2             = datetime.strptime(date_of_survilance2[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.date_of_expiry                  = datetime.strptime(date_of_expiry[id], '%d/%m/%Y').strftime('%Y-%m-%d')
                        lead.master_iso_id                   = getModelColumnById(SpIsoMaster, int(iso_applicable_id[id]), 'iso_id')
                        lead.master_iso_name                 = getModelColumnById(SpIsoMaster, int(iso_applicable_id[id]), 'iso_name')
                        lead.currency_code                   = currency_code[id]

                        if str(id) in file_ids:
                            if copy_of_iso_files[i]:
                        
                                uploaded_iso = copy_of_iso_files[i]
                                storage = FileSystemStorage()
                                timestamp = int(time.time())
                                iso_name = uploaded_iso.name
                                temp = iso_name.split('.')
                                iso_name = 'media/iso_image/'+str(timestamp)+"."+temp[(len(temp) - 1)]
                                iso = storage.save(iso_name, uploaded_iso)
                                lead.copy_of_iso = storage.url(iso) 
                                i = i+1
                            else:
                                lead.copy_of_iso = None 
                        else:
                            lead.copy_of_iso = None
                        
                        lead.iso_issued_agency               = iso_issued_agency[id]
                        lead.iso_issued_consultant           = iso_issued_consultant[id]
                        lead.price_of_existing_iso           = price_of_existing_iso[id]
                        lead.iso_status                      = 1
                        lead.status  = 1
                        
                        try:
                            sp_lead_iso_save_instance = SpLeadIsoSave.objects.get(last_lead_id=lead_id)
                            sp_lead_iso_save_instance.delete()
                        except SpLeadIsoSave.DoesNotExist:
                            pass  
                        lead.save()

                message  = "Lead ISO Details Saved Successfully"
                response = {}
                response['error'] = False
                response['message'] = message
                return JsonResponse(response)
                
            else:
                
                try:
                    lead = SpLeadIsoSave.objects.get(last_lead_id=lead_id)
                    suggested_iso_created_id        = request.POST.getlist('suggested_iso_zero') 
                    core_business_area_str          = ','.join(suggested_iso_created_id)
                    lead.iso_created_id             = core_business_area_str
                    lead.iso_amount                 = request.POST['enter_suggested_cost']
                    lead.currency_code              = request.POST['suggested_currancy_code']
                    lead.iso_created_status         = 0
                    lead.created_by_id              = request.POST['created_by_id']
                    lead.last_lead_id               = lead_id
                    lead.status =1

                    try:
                        sp_lead_iso_save_instance = SpLeadIso.objects.get(last_lead_id=lead_id)
                        sp_lead_iso_save_instance.delete()
                    except SpLeadIso.DoesNotExist:
                       
                        pass  
                
                    lead.save()
                except SpLeadIsoSave.DoesNotExist:
                    lead = SpLeadIsoSave()
                    suggested_iso_created_id        =  request.POST.getlist('suggested_iso_zero') 
                    core_business_area_str          = ','.join(suggested_iso_created_id)
                    lead.iso_created_id             = core_business_area_str
                    lead.iso_amount                 = request.POST['enter_suggested_cost']
                    lead.currency_code              = request.POST['suggested_currancy_code']
                    lead.iso_created_status         = 0
                    lead.created_by_id              =   request.POST['created_by_id']
                    lead.last_lead_id               = lead_id
                    lead.status =1

                    try:
                        sp_lead_iso_save_instance = SpLeadIso.objects.filter(last_lead_id=lead_id)
                        sp_lead_iso_save_instance.delete()
                    except SpLeadIso.DoesNotExist:
                        pass

                    lead.save()
                message  = "Lead ISO Details Saved Successfully"
                response = {}
                response['error'] = False
                response['message'] = message
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    else:
        created_by = SpLeadBasic.objects.get(id = lead_id)
        suggested_iso_details = SpLeadIso.objects.filter(last_lead_id = lead_id)
        creatrd_lead_iso_details = SpLeadIsoSave.objects.filter(last_lead_id = lead_id).first()
        if suggested_iso_details:
            iso_status = 1
        else:
            iso_status = 0

        isos_master = SpIsoMaster.objects.filter()
        currancy_code = SpCurrencyCode.objects.filter(status = 1)

        context = {}
        context['iso_list_deatils']         = creatrd_lead_iso_details
        context['created_iso_details']      = suggested_iso_details
        context['iso_status']       = iso_status
        context['isos_master']      = isos_master
        context['currancy_code']    = currancy_code
        context['lead_id']          = lead_id
        context['created_by_id']    = created_by.created_by_id

        template = 'lead-report/edit-lead-iso-details-new.html'

        return render(request, template, context)
        
@login_required
def editLeadOtherDetails(request,lead_id):
    response = {}
    if request.method == "POST":
        try:
            if SpLeadOther.objects.filter(last_lead_id = lead_id).exists():
                lead_other = SpLeadOther.objects.get(last_lead_id = lead_id)
            else:
                lead_other = SpLeadOther()
            if str(lead_other.created_by_id) != request.POST['created_by_id']:
                lead = SpLeadBasic.objects.get(id=int(lead_id))
                lead.created_by_id = int(request.POST['created_by_id'])
                lead.save()
               
                if SpLeadIso.objects.filter(last_lead_id= int(lead_id)).exists():
                    SpLeadIso.objects.filter(last_lead_id= int(lead_id)).update(created_by_id=int(request.POST['created_by_id']))
                else:
                    SpLeadIsoSave.objects.filter(last_lead_id= int(lead_id)).update(created_by_id=int(request.POST['created_by_id']))
                lead_other.created_by_id = request.POST['created_by_id']
            if (request.POST['other_product_pitch']) == 'on':
                lead_other.other_production_pitch = 1
                software_or_erp_list    = request.POST.getlist('software_or_erp[]') 
                software_or_erp         = ','.join(software_or_erp_list)
                lead_other.software_or_erp = software_or_erp
            else:
                lead_other.other_production_pitch = 0
                lead_other.software_or_erp = None

            lead_other.sales_person     = request.POST['created_by_id']
            lead_other.visit_date       = datetime.strptime(request.POST['any_follow_up'], '%d/%m/%Y').strftime('%Y-%m-%d')
            lead_other.other_resource   = request.POST['other_resource_required']
            lead_other.reminder         = datetime.strptime(request.POST['reminder'], '%d/%m/%Y').strftime('%Y-%m-%d')
            lead_other.remark           = request.POST['remark']
            lead_other.status           = 1
            lead_other.last_lead_id     = lead_id
            lead_other.save() 
           
            message  = "Lead Other Details Saved Successfully"
            response = {}
            response['error'] = False
           
            response['message'] = message
            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    else:

        context                 = {}

        created_by          = SpLeadBasic.objects.get(id = lead_id)
        other_lead_deatils  = SpLeadOther.objects.filter(status =1,last_lead_id = lead_id).first()
        service         = TtrackService.objects.all()
        sales_person    = SpUsers.objects.filter(status=1)
        context['other_lead_deatils'] = other_lead_deatils
        try:
            if other_lead_deatils :
                context['other_resource']   = int(other_lead_deatils.other_resource)
        except (ValueError, TypeError) as e:
            context['other_resource'] = None
        context['service']          = service
        context['sales_persons']    = sales_person
        context['created_by_id']    = created_by.created_by_id
        context['lead_id']          = lead_id

        template = 'lead-report/edit-lead-other-details.html'
        return render(request, template, context)
        

@login_required
def leadReport(request):
    today       = date.today()
    sp_lead_list = []
    status_dict  = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    spleadiso = SpLeadIso.objects.filter(date_of_survilance1__icontains = today.strftime("%Y-%m"),date_of_survilance1__gte=today.strftime("%Y-%m-%d")).values_list('last_lead_id', flat=True).distinct()
    count_lead = SpLeadIso.objects.filter(date_of_survilance1__icontains = today.strftime("%Y-%m"),date_of_survilance1__gte=today.strftime("%Y-%m-%d")).values_list('last_lead_id', flat=True).count()
    if spleadiso:
        spleadbasics = SpLeadBasic.objects.filter(id__in=spleadiso)
        for spleadbasic in spleadbasics:
            sp_lead_dict = {
                "id": spleadbasic.id,
                "created_at":spleadbasic.created_at,
                "lead_code": "TTRACK"+str(spleadbasic.id),
                "country_code":getModelColumnById(SpCountryCodes, spleadbasic.contry_code_id,'country_code') if spleadbasic.contry_code_id else "",
                "created_by_name": getUserName(spleadbasic.created_by_id),
                "basic_date": spleadbasic.basic_date,
                "company_name": spleadbasic.company_name,
                "turnover": spleadbasic.turnover,
                "contact_person_name": spleadbasic.contact_person_name,
                "desk_no": spleadbasic.desk_no,
                "mobile_no": spleadbasic.mobile_no,
                "contry_code_id": spleadbasic.contry_code_id,
                "email": spleadbasic.email,
                "address": spleadbasic.address,
                "total_no_of_employee": spleadbasic.total_no_of_employee,
                "currency_code": spleadbasic.currency_code,
                "deal_amount": spleadbasic.deal_amount,
                "deal_date_time": spleadbasic.deal_date_time,
                "deal_currency_code": spleadbasic.deal_currency_code,
                "created_at": spleadbasic.created_at,
                "updated_at": spleadbasic.updated_at,
                "status": spleadbasic.status,
                "longitude": spleadbasic.longitude,
                "latitude": spleadbasic.latitude,
                "lead_status":status_dict[spleadbasic.status],
                "lead_iso_list": list(SpLeadIso.objects.filter(last_lead_id=spleadbasic.id,date_of_survilance1__icontains = today.strftime("%Y-%m"),date_of_survilance1__gte=today.strftime("%Y-%m-%d")).values())
            }
            sp_lead_list.append(sp_lead_dict)
    
    template = 'lead-report/lead-report.html'
    context = {}
    context['user_lists']    = SpUsers.objects.filter(status=1,user_type=1)
    context['sp_lead_list']  = sp_lead_list
    context['record_count']  = count_lead
    context['page_title']    = "Upcoming Renewals"
    return render(request, template,context)

@login_required
def ajaxLeadReport(request):
    today               = date.today()
    customer_id         = request.GET.get('customer_id')
    filter_by           = request.GET.get('filter_by')
    order_date          = request.GET.get('order_date')
    sp_lead_list = []
    status_dict  = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    if order_date:
        start_date          = datetime.strptime(order_date.split(" - ")[0],'%d/%m/%Y').strftime('%Y-%m-%d')
        to_date             = datetime.strptime(order_date.split(" - ")[1],'%d/%m/%Y').strftime('%Y-%m-%d')
        # start_end           = datetime.strptime(order_date.split(" - ")[1],'%d/%m/%Y').strftime('%Y-%m-%d')
        # to_date             = datetime.strptime(str(start_end), "%Y-%m-%d")+ timedelta(days=1)
        # to_date             = str(to_date).replace(' 00:00:00', '')
        date_field = "date_of_survilance1" if filter_by == "1" else "date_of_survilance2" if filter_by == "2" else "date_of_expiry"
        if customer_id:
            spleadiso = SpLeadIso.objects.filter(created_by_id = customer_id, **{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).distinct() 
            count_lead = SpLeadIso.objects.filter(created_by_id = customer_id, **{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).count() 
        else:
            spleadiso = SpLeadIso.objects.filter(**{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).distinct() 
            count_lead = SpLeadIso.objects.filter(**{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).count() 
    else:
        if customer_id:
            spleadiso = SpLeadIso.objects.filter(created_by_id = customer_id,date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).distinct()
            count_lead = SpLeadIso.objects.filter(created_by_id = customer_id,date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).count()
        else:
            spleadiso = SpLeadIso.objects.filter(date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).distinct()
            count_lead = SpLeadIso.objects.filter(date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).count()
    if spleadiso:
        spleadbasics = SpLeadBasic.objects.filter(id__in=spleadiso)
        for spleadbasic in spleadbasics:
            sp_lead_dict = {
                "id": spleadbasic.id,
                "lead_code": "TTRACK"+str(spleadbasic.id),
                "created_at":spleadbasic.created_at,
                "country_code":getModelColumnById(SpCountryCodes, spleadbasic.contry_code_id,'country_code') if spleadbasic.contry_code_id else "",
                "created_by_name": getUserName(spleadbasic.created_by_id),
                "basic_date": spleadbasic.basic_date,
                "company_name": spleadbasic.company_name,
                "turnover": spleadbasic.turnover,
                "contact_person_name": spleadbasic.contact_person_name,
                "desk_no": spleadbasic.desk_no,
                "mobile_no": spleadbasic.mobile_no,
                "contry_code_id": spleadbasic.contry_code_id,
                "email": spleadbasic.email,
                "address": spleadbasic.address,
                "total_no_of_employee": spleadbasic.total_no_of_employee,
                "currency_code": spleadbasic.currency_code,
                "deal_amount": spleadbasic.deal_amount,
                "deal_date_time": spleadbasic.deal_date_time,
                "deal_currency_code": spleadbasic.deal_currency_code,
                "created_at": spleadbasic.created_at,
                "updated_at": spleadbasic.updated_at,
                "status": spleadbasic.status,
                "longitude": spleadbasic.longitude,
                "latitude": spleadbasic.latitude,
                "lead_status":status_dict[spleadbasic.status],
                "lead_iso_list": list(SpLeadIso.objects.filter(last_lead_id=spleadbasic.id,**{f"{date_field}__range": [start_date, to_date]}).values())
            }
            sp_lead_list.append(sp_lead_dict)
    
    template = 'lead-report/ajax-lead-report.html'
    context = {}
    context['sp_lead_list']  = sp_lead_list
    context['record_count']  = count_lead
    return render(request, template,context)

@login_required
def exportLeadReportToXlsx(request, customer_id,filter_by,start_date,end_date):
    today               = date.today()
    customer_id         = customer_id
    filter_by           = filter_by
    start_date          = start_date
    end_date            = end_date
    sp_lead_list = []
    status_dict  = {1: "Initiated", 2: "In Progress", 3: "Closed(Win)", 4: "Closed(Lost)"}
    if start_date!='0' and end_date !='0':
        to_date         = end_date
        # to_date                         = datetime.strptime(str(end_date), "%Y-%m-%d")+ timedelta(days=1)
        # to_date                         = str(to_date).replace(' 00:00:00', '')
        date_field = "date_of_survilance1" if filter_by == "1" else "date_of_survilance2" if filter_by == "2" else "date_of_expiry"
        if customer_id!='0':
            spleadiso = SpLeadIso.objects.filter(created_by_id = customer_id, **{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).distinct() 
        else:
            spleadiso = SpLeadIso.objects.filter(**{f"{date_field}__range": [start_date, to_date]}).values_list('last_lead_id', flat=True).distinct() 
    else:
        if customer_id!='0':
            spleadiso = SpLeadIso.objects.filter(created_by_id = customer_id,date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).distinct()
        else:
            spleadiso = SpLeadIso.objects.filter(date_of_survilance1__icontains = today.strftime("%Y-%m")).values_list('last_lead_id', flat=True).distinct()
    if spleadiso:
        spleadbasics = SpLeadBasic.objects.filter(id__in=spleadiso)
        for spleadbasic in spleadbasics:
            sp_lead_dict = {
                "id": spleadbasic.id,
                "lead_code": "TTRACK"+str(spleadbasic.id),
                "created_at":spleadbasic.created_at,
                "country_code":getModelColumnById(SpCountryCodes, spleadbasic.contry_code_id,'country_code') if spleadbasic.contry_code_id else "",
                "created_by_name": getUserName(spleadbasic.created_by_id),
                "basic_date": spleadbasic.basic_date,
                "company_name": spleadbasic.company_name,
                "turnover": spleadbasic.turnover,
                "contact_person_name": spleadbasic.contact_person_name,
                "desk_no": spleadbasic.desk_no,
                "mobile_no": spleadbasic.mobile_no,
                "contry_code_id": spleadbasic.contry_code_id,
                "email": spleadbasic.email,
                "address": spleadbasic.address,
                "total_no_of_employee": spleadbasic.total_no_of_employee,
                "currency_code": spleadbasic.currency_code,
                "deal_amount": spleadbasic.deal_amount,
                "deal_date_time": spleadbasic.deal_date_time,
                "deal_currency_code": spleadbasic.deal_currency_code,
                "created_at": spleadbasic.created_at,
                "updated_at": spleadbasic.updated_at,
                "status": spleadbasic.status,
                "longitude": spleadbasic.longitude,
                "latitude": spleadbasic.latitude,
                "lead_status":status_dict[spleadbasic.status],
                "lead_iso_list": list(SpLeadIso.objects.filter(last_lead_id=spleadbasic.id, **{f"{date_field}__range": [start_date, to_date]}).values())
            }
            sp_lead_list.append(sp_lead_dict)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    # response['Content-Disposition'] = 'attachment; filename=Lead Management.xlsx'.format(
    #     date=datetime.now().strftime('%Y-%m-%d'),
    # )
    response['Content-Disposition'] = 'attachment; filename=Upcoming Renewals.xlsx'.format(
        datetime.now().strftime('%d-%m-%Y')
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        horizontal='left',
        wrap_text=True
    )

    header_alignment = Alignment(
        vertical='top',
        horizontal='center',
        wrap_text=True
    )
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Upcoming Renewals'
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True
    
    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
    img.height = 48
    img.width = 95
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)
    
    column_length = 10
    
    worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'Upcoming Renewals'
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = header_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill()

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40
    
    # Define the titles for columns
    columns = []

    columns += [ 'Lead ID' ]
    columns += [ 'Company Name' ]
    columns += [ 'Turnover' ] 
    columns += [ 'Contact Person' ]
    columns += [ 'Contact No.' ]
    columns += [ 'Desk No.' ]
    columns += [ 'Email' ]
    columns += [ 'Created Date' ]
    columns += [ 'Created By' ]
    columns += [ 'Lead Status' ]

    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="02529c", end_color="02529c", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

     
    for lead in sp_lead_list:
        row_num += 1
        row = []
        row += [ "TTRACK"+str(lead['id']) ]
        row += [ lead["company_name"] ]
        if lead["currency_code"]:
            row += [ "{} {}".format(lead["currency_code"], lead["turnover"]) ]
        else:
            row += [ lead["turnover"] ]
        row += [ lead["contact_person_name"] ]
        new_contact = str(lead["country_code"])+" "+str(lead["mobile_no"])
        row += [ new_contact]
        row += [ lead["desk_no"] ]
        row += [ lead["email"]]
        created_at_str = lead["created_at"].strftime('%Y-%m-%d %H:%M:%S')
        created_at_str = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
        row += [ created_at_str ]
        row += [ lead["created_by_name"] ]
        row += [ lead["lead_status"] ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border    
            wrapped_alignment = Alignment(wrap_text=True)
        row_num += 1

        columns = []
        columns += [ '' ]
        columns += [ '' ]
        columns += [ 'ISO Name' ]
        columns += [ 'Issue Date' ]
        columns += [ 'Expiry Date' ] 
        columns += [ 'Survilance1' ]
        columns += [ 'Survilance2' ]
        columns += [ 'Agency' ]
        columns += [ 'Consultant' ]
        columns += [ 'Amount' ]

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="55b3c0", end_color="55b3c0", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 20
        
        for distributor in lead["lead_iso_list"]:
            row_num += 1
            row = []
            row += ['']
            row += ['']
            names= "ISO "+str(distributor["master_iso_id"])+" - "+str(distributor["master_iso_name"])
            row += [ names ]
            date_of_issue_str = distributor["date_of_issue"].strftime('%Y-%m-%d')
            date_of_issue_str = datetime.strptime(date_of_issue_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            row += [ date_of_issue_str ]
            date_of_expiry_str = distributor["date_of_expiry"].strftime('%Y-%m-%d')
            date_of_expiry_str = datetime.strptime(date_of_expiry_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            row += [ date_of_expiry_str ]
            date_of_survilance1_str = distributor["date_of_survilance1"].strftime('%Y-%m-%d')
            date_of_survilance1_str = datetime.strptime(date_of_survilance1_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            row += [ date_of_survilance1_str ]
            date_of_survilance2_str = distributor["date_of_survilance2"].strftime('%Y-%m-%d')
            date_of_survilance2_str = datetime.strptime(date_of_survilance2_str, '%Y-%m-%d').strftime('%d/%m/%Y')
            row += [ date_of_survilance2_str ]
            row += [ distributor["iso_issued_agency"] ]
            row += [ distributor["iso_issued_consultant"] ]
            new_name = str(distributor["currency_code"])+" "+str(distributor["price_of_existing_iso"])
            row += [ new_name ]
           
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border    
                wrapped_alignment = Alignment(wrap_text=True)


    # last_row = row_num
    # worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=10)
    # worksheet.row_dimensions[last_row].height = 20
    # worksheet.cell(row=last_row, column=1).value = 'Generated By Emobic'
    # worksheet.cell(row=last_row, column=1).font = header_font
    # worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    # worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
    # worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

    workbook.save(response)
    return response

    
def assigendBulkLead(request):
    lead_ids = request.POST.getlist('lead_id[]')  # Corrected parameter name
    
    created_by_ids = request.POST.getlist('created_by_id[]')  # Corrected parameter name
   
    context = {}
    employee_list = SpUsers.objects.filter(status=1).exclude(id__in=created_by_ids)
    
    context['employee_list'] = employee_list
    context['lead_ids'] = lead_ids
    template = 'lead-report/bulk-assign-lead.html'
    return render(request, template, context)
import ast
def updateAssignEmployee(request):
    lead_ids = request.POST['lead_id_list']
    python_list = ast.literal_eval(lead_ids)
    assign_employee_id = request.POST['employee_lead_id_list']
    basic_leads =SpLeadBasic.objects.filter(id__in= python_list)
    for basic_lead in basic_leads:
        
        lead = SpLeadBasic.objects.get(id=basic_lead.id)
        
        lead.created_by_id = int(assign_employee_id)
        lead.save()
        
        if SpLeadIso.objects.filter(last_lead_id= basic_lead.id).exists():
            SpLeadIso.objects.filter(last_lead_id= basic_lead.id).update(created_by_id=int(assign_employee_id))
        else:
            SpLeadIsoSave.objects.filter(last_lead_id= basic_lead.id).update(created_by_id=int(assign_employee_id))
        if SpLeadOther.objects.filter(last_lead_id=basic_lead.id).exists():
            other_lead = SpLeadOther.objects.get(last_lead_id=basic_lead.id)
            other_lead.created_by_id = int(assign_employee_id)
            other_lead.save()

    
    message  = "Lead has been successfully Assigned"
    response = {}
    response['error'] = False
    
    response['message'] = message
    return JsonResponse(response)
