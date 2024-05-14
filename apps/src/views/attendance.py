import sys
import os
from numpy.lib.function_base import append
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound, FileResponse, response
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import *
from django.forms.models import model_to_dict
from utils import *
from datetime import datetime, date, timedelta
from io import BytesIO
from django.views import View
from django.conf import settings
from gtts import gTTS
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import calendar
from rest_framework.response import Response
import ast
import dateutil.relativedelta


@csrf_exempt
@login_required
def markStudentAttendanceNew(request):
    if request.method == "GET":
        context = {}
        template = 'attendance/mark-student-attendance-new.html'
        return render(request, template, context)
    else:
        context = {}
        if 'student_id' in request.POST and request.POST['student_id'] != "":
            current_date = datetime.now().strftime('%Y-%m-%d')
            if TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).exists():
                last_record = TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).last()
                if last_record.end_datetime is None:
                    TblAttendance.objects.filter(id=last_record.id).update(end_datetime=datetime.now())
                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')

                    # send notification to web
                    sendWebNotification(request.POST['student_id'])

                else:
                    attendance = TblAttendance()
                    attendance.student_id = request.POST['student_id']
                    attendance.start_datetime = datetime.now()
                    attendance.end_datetime = None
                    if 'latitude' in request.POST and request.POST['latitude'] != "":
                        attendance.latitude = request.POST['latitude']
                    if 'longitude' in request.POST and request.POST['longitude'] != "":
                        attendance.longitude = request.POST['longitude']
                    attendance.save()

                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')

                    # send notification to web
                    sendWebNotification(request.POST['student_id'])
            else:
                attendance = TblAttendance()
                attendance.student_id = request.POST['student_id']
                attendance.start_datetime = datetime.now()
                attendance.end_datetime = None
                if 'latitude' in request.POST and request.POST['latitude'] != "":
                    attendance.latitude = request.POST['latitude']
                if 'longitude' in request.POST and request.POST['longitude'] != "":
                    attendance.longitude = request.POST['longitude']
                attendance.save()

                context['flag'] = True
                context['message'] = "Attendance marked successfully."
                context['current_time'] = datetime.now().strftime('%H:%M')

                # send notification to web
                sendWebNotification(request.POST['student_id'])
        else:
            context['flag'] = False
            context['message'] = "Parameter missing. please try again."

        return JsonResponse(context)

@csrf_exempt
@login_required
def newmarkStudentAttendance(request):
    if request.method == "GET":
        context = {}
        template = 'attendance/new-mark-student-attendance.html'
        return render(request, template, context)
    else:
        context = {}
        if 'student_id' in request.POST and request.POST['student_id'] != "":
            current_date = datetime.now().strftime('%Y-%m-%d')
            if TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).exists():
                last_record = TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).last()
                if last_record.end_datetime is None:
                    TblAttendance.objects.filter(id=last_record.id).update(end_datetime=datetime.now())
                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')

                    # send notification to web
                    sendWebNotification(request.POST['student_id'])

                else:
                    attendance = TblAttendance()
                    attendance.student_id = request.POST['student_id']
                    attendance.start_datetime = datetime.now()
                    attendance.end_datetime = None
                    if 'latitude' in request.POST and request.POST['latitude'] != "":
                        attendance.latitude = request.POST['latitude']
                    if 'longitude' in request.POST and request.POST['longitude'] != "":
                        attendance.longitude = request.POST['longitude']
                    attendance.save()

                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')

                    # send notification to web
                    sendWebNotification(request.POST['student_id'])
            else:
                attendance = TblAttendance()
                attendance.student_id = request.POST['student_id']
                attendance.start_datetime = datetime.now()
                attendance.end_datetime = None
                if 'latitude' in request.POST and request.POST['latitude'] != "":
                    attendance.latitude = request.POST['latitude']
                if 'longitude' in request.POST and request.POST['longitude'] != "":
                    attendance.longitude = request.POST['longitude']
                attendance.save()

                context['flag'] = True
                context['message'] = "Attendance marked successfully."
                context['current_time'] = datetime.now().strftime('%H:%M')

                # send notification to web
                sendWebNotification(request.POST['student_id'])
        else:
            context['flag'] = False
            context['message'] = "Parameter missing. please try again."

        return JsonResponse(context)


# @csrf_exempt
# @login_required
# def markStudentAttendance(request):
#     if request.method == "GET":
#         context = {}
#         template = 'attendance/new-mark-student-attendance.html'
#         return render(request, template, context)
#     else:
#         context = {}
#         if 'student_id' in request.POST and request.POST['student_id'] != "" :
#             current_date = datetime.now().strftime('%Y-%m-%d')
#             if TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).exists():
#                 last_record = TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).last()
#                 if last_record.end_datetime is None:
#                     TblAttendance.objects.filter(id=last_record.id).update(end_datetime=datetime.now())
#                     context['flag'] = True
#                     context['message'] = "Attendance marked successfully."
#                     context['current_time'] = datetime.now().strftime('%H:%M')
                    
#                     # send notification to web
#                     sendWebNotification(request.POST['student_id'])

#                 else:
#                     attendance = TblAttendance()
#                     attendance.student_id = request.POST['student_id']
#                     attendance.start_datetime =  datetime.now()
#                     attendance.end_datetime =  None
#                     if 'latitude' in request.POST and request.POST['latitude'] != "" :
#                         attendance.latitude = request.POST['latitude']
#                     if 'longitude' in request.POST and request.POST['longitude'] != "" :
#                         attendance.longitude = request.POST['longitude']

#                     if student_info.semester_id is not None:
#                         attendance.semester_id = student_info.semester_id
#                     elif student_info.year_id is not None:
#                         attendance.semester_id = student_info.year_id
#                     attendance.save()

#                     context['flag'] = True
#                     context['message'] = "Attendance marked successfully."
#                     context['current_time'] = datetime.now().strftime('%H:%M')

#                     # send notification to web
#                     sendWebNotification(request.POST['student_id'])
#             else:
#                 attendance = TblAttendance()
#                 attendance.student_id = request.POST['student_id']
#                 attendance.start_datetime =  datetime.now()
#                 attendance.end_datetime =  None
#                 if 'latitude' in request.POST and request.POST['latitude'] != "" :
#                     attendance.latitude = request.POST['latitude']
#                 if 'longitude' in request.POST and request.POST['longitude'] != "" :
#                     attendance.longitude = request.POST['longitude']
                
#                 if student_info.semester_id is not None:
#                     attendance.semester_id = student_info.semester_id
#                 elif student_info.year_id is not None:
#                     attendance.semester_id = student_info.year_id

#                 attendance.save()

#                 context['flag'] = True
#                 context['message'] = "Attendance marked successfully."
#                 context['current_time'] = datetime.now().strftime('%H:%M')

#                 # send notification to web
#                 sendWebNotification(request.POST['student_id'])
#         else:
#             context['flag'] = False
#             context['message'] = "Parameter missing. please try again."

#         return JsonResponse(context)

@csrf_exempt
@login_required
def markStudentAttendance(request):
    if request.method == "GET":
        context = {}
        template = 'attendance/new-mark-student-attendance.html'
        return render(request, template, context)
    else:
        context = {}
        if 'student_id' in request.POST and request.POST['student_id'] != "" :
            current_date = datetime.now().strftime('%Y-%m-%d')
            if TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).exists():
                last_record = TblAttendance.objects.filter(student_id=request.POST['student_id'],created_at__contains=current_date).last()

                if last_record.end_datetime is None:
                    TblAttendance.objects.filter(id=last_record.id).update(end_datetime=datetime.now())
                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')
                    
                    # send notification to web
                    sendWebNotification(request.POST['student_id'])

                else:
                    attendance = TblAttendance()
                    attendance.student_id = request.POST['student_id']
                    attendance.start_datetime =  datetime.now()
                    attendance.end_datetime =  None
                    if 'latitude' in request.POST and request.POST['latitude'] != "" :
                        attendance.latitude = request.POST['latitude']
                    if 'longitude' in request.POST and request.POST['longitude'] != "" :
                        attendance.longitude = request.POST['longitude']
                    
                    student_info = TblStudents.objects.filter(reg_no = request.POST['student_id']).first()
                    if student_info.semester_id is not None:
                        attendance.semester_id = student_info.semester_id
                    elif student_info.year_id is not None:
                        attendance.semester_id = student_info.year_id
                    attendance.save()

                    context['flag'] = True
                    context['message'] = "Attendance marked successfully."
                    context['current_time'] = datetime.now().strftime('%H:%M')

                    # send notification to web
                    sendWebNotification(request.POST['student_id'])
            else:
                attendance = TblAttendance()
                attendance.student_id = request.POST['student_id']
                attendance.start_datetime =  datetime.now()
                attendance.end_datetime =  None
                if 'latitude' in request.POST and request.POST['latitude'] != "" :
                    attendance.latitude = request.POST['latitude']
                if 'longitude' in request.POST and request.POST['longitude'] != "" :
                    attendance.longitude = request.POST['longitude']
                student_info = TblStudents.objects.filter(id = request.POST['student_id']).first()
                
                if student_info.semester_id:
                    attendance.semester_id = student_info.semester_id
                elif student_info.year_id:
                    attendance.semester_id = student_info.year_id
                attendance.save()

                context['flag'] = True
                context['message'] = "Attendance marked successfully."
                context['current_time'] = datetime.now().strftime('%H:%M')

                # send notification to web
                sendWebNotification(request.POST['student_id'])
        else:
            context['flag'] = False
            context['message'] = "Parameter missing. please try again."

        return JsonResponse(context)



@login_required
def getStudentData(request,student_id):
    context = {}
    if request.method == "POST":
        context['flag'] = False
        context['message'] = "Method not allowed"
    else:
        student_details = TblStudents.objects.raw(''' SELECT tbl_students.*,tbl_colleges.college_name,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
        WHERE tbl_students.id=%s ''',[student_id])
        
        if student_details:
            context['student_details'] = student_details[0]
            if student_details[0].college_id == 1:
                context['college_base_url'] = "http://bipe.sortstring.co.in/"
            elif student_details[0].college_id == 2:
                context['college_base_url'] = "http://bite.sortstring.co.in/"

            current_date = datetime.now().strftime('%Y-%m-%d')
            if TblAttendance.objects.filter(student_id=student_id,created_at__contains=current_date).exists():
                last_record = TblAttendance.objects.filter(student_id=student_id,created_at__contains=current_date).last()
                if last_record.end_datetime is None:
                    context['punch_in_status'] = 1
                    context['last_punch_in'] = last_record.start_datetime
                else:
                    context['punch_in_status'] = 0
                    context['last_punch_in'] = last_record.end_datetime
            else:
                context['punch_in_status'] = 0
                context['last_punch_in'] = None
            
            student_html = render_to_string('attendance/student-details.html', context)

            student_name = student_details[0].first_name+' '
            if student_details[0].middle_name is not None:
                student_name += student_details[0].middle_name+' '
            if student_details[0].last_name is not None:
                student_name += student_details[0].last_name+' '

            # generate audio
            student_file_name = "student_id_"+student_id+".mp3"
            if os.path.isdir(str(settings.MEDIA_ROOT)+'/attendance_audio/qr_scan/'+student_file_name):
                audio = '/media/attendance_audio/qr_scan/'+student_file_name
            else:
                Text = "Welcome "+student_name+". Please scan your thumb."
                speech = gTTS(text = Text)
                file_name = str(settings.MEDIA_ROOT) + '/attendance_audio/qr_scan/'+student_file_name
                speech.save(file_name)
                audio = '/media/attendance_audio/qr_scan/'+student_file_name
            
            response = {}
            response['flag'] = True
            response['student_id'] = student_id
            response['student_html'] = student_html
            response['audio'] = audio

            return JsonResponse(response)
        else:
            context['flag'] = False
            context['message'] = "Record not found"

    return JsonResponse(context)





@login_required
def studentAttendanceReport(request):

    page = request.GET.get('page')

    students = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
    tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
    tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.alias,tbl_students.college_id,
    tbl_students.branch_id,tbl_students.semester_id,tbl_students.year_id,tbl_branch.branch FROM tbl_students
    LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
    LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
    LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id
    WHERE tbl_students.is_registered = 1 AND  date(tbl_attendance.created_at)=CURDATE() GROUP BY tbl_attendance.student_id ORDER BY  tbl_students.id DESC
    ''')

    totat_record = len(students)
    
    paginator = Paginator(list(students), getConfigurationResult('page_limit'))

    try:
        students = paginator.page(page)
    except PageNotAnInteger:
        students = paginator.page(1)
    except EmptyPage:
        students = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}

    for student in students:
        college = TblColleges.objects.get(id=student.college_id)
        course = TblBranch.objects.get(id=student.branch_id)
        student.college_name = college.college_name
        student.college_alias = college.alias
        student.branch = course.branch
        if student.semester_id:
            student.semester_year = str(student.semester_id)+ " Sem"
       
        elif student.year_id:
            student.semester_year = str(student.year_id)+ " Year"
       
        
        current_date = datetime.now().strftime('%Y-%m-%d')
        if TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).exists():
            last_record = TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).last()
            if last_record.end_datetime is None:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = None
            else:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = last_record.end_datetime
        else:
            student.punch_in_time = None
            student.punch_out_time = None

    student_details = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
    tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
    tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.college_name,tbl_students.college_id,
    tbl_students.branch_id,tbl_students.semester_id,tbl_branch.branch
    FROM tbl_students
    LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
    LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
    LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id
    WHERE date(tbl_attendance.created_at)=CURDATE()
    ORDER BY  tbl_students.id DESC LIMIT 1
    ''')
    if student_details:
        context['student_details']        = student_details[0]
        if student_details[0].college_id == 1:
            context['college_base_url'] = "http://bipe.sortstring.co.in/"
        elif student_details[0].college_id == 2:
            context['college_base_url'] = "http://bite.sortstring.co.in/"
        elif student_details[0].college_id == 3:
            context['college_base_url'] = "http://bip.sortstring.co.in/"
        
        current_date = datetime.now().strftime('%Y-%m-%d')
        if TblAttendance.objects.filter(student_id=student_details[0].id,created_at__contains=current_date).exists():
            last_record = TblAttendance.objects.filter(student_id=student_details[0].id,created_at__contains=current_date).last()
            if last_record.end_datetime is None:
                context['punch_in_status'] = 1
                context['last_punch_in'] = last_record.start_datetime
            else:
                context['punch_in_status'] = 0
                context['last_punch_in'] = last_record.end_datetime
        else:
            context['punch_in_status'] = 0
            context['last_punch_in'] = None

    
    context['colleges'] = TblColleges.objects.all().values('id','college_name')
    context['mentors'] = TblStudents.objects.all().values_list('teacher_gaurdian_name',flat=True).distinct()

    context['students']          = students
    context['totat_record']         = totat_record
    context['total_pages']            = total_pages
    context['page_limit']             = getConfigurationResult('page_limit')
    context['page_title']             = "Student Attendance Report"
    context['current_date']            = datetime.now().strftime('%Y-%m-%d')
    template = 'attendance/student-attendance-report.html'
    return render(request, template, context)


@login_required
def filterStudentAttendanceReport(request):
    context = {}

    page = request.GET.get('page')
    condition = ''

    if 'search' in request.POST and request.POST['search'] != "":
        condition += " and (tbl_students.first_name LIKE '%%"+request.POST['search']+"%%' OR tbl_students.primary_contact_no LIKE '%%"+request.POST['search']+"%%') " 
    
    if 'college_id' in request.POST and request.POST['college_id'] != "":
        condition += " and tbl_students.college_id="+request.POST['college_id']

    if 'mentor' in request.POST and request.POST['mentor'] != "":
        condition += " and tbl_students.teacher_gaurdian_name='"+request.POST['mentor']+"'"

    if 'course' in request.POST and request.POST['course'] != "":
        condition += " and tbl_students.branch_id="+request.POST['course']

    if 'sem_year' in request.POST and request.POST['sem_year'] != "":
        condition += " and tbl_students.semester_id='"+request.POST['sem_year']+"'"

    if 'date' in request.POST and request.POST['date'] != "":
        date = datetime.strptime(request.POST['date'], "%d/%m/%Y").strftime('%Y-%m-%d')
        condition += " and date(tbl_attendance.created_at)='"+date+"'"

    if condition == "":
        students = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
        tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
        tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.alias,tbl_students.college_id,
        tbl_students.branch_id,tbl_students.semester_id,tbl_students.year_id,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id 
        LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id
        WHERE tbl_students.is_registered = 1 AND  date(tbl_attendance.created_at)=CURDATE()  ORDER BY tbl_students.id DESC ''')
    else:
        students = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
        tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
        tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.alias,tbl_students.college_id,
        tbl_students.branch_id,tbl_students.semester_id,tbl_students.year_id,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
        LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id
        WHERE tbl_students.is_registered = 1  AND 1 {condition} ORDER BY tbl_students.id DESC '''.format(condition=condition))


    total_record = len(students)
    for student in students:
        college = TblColleges.objects.get(id=student.college_id)
        course = TblBranch.objects.get(id=student.branch_id)
        student.college_name = college.college_name
        student.college_alias = college.alias
        student.branch = course.branch
        if student.semester_id:
            student.semester_year = student.semester_id+ " Sem"
        elif student.year_id:
            student.semester_year = student.year_id+ " Year"
        # tmp = student.semester_id.split('_')
        # suffix = ""
        # if tmp[1] == "1":
        #     suffix = "st"
        # elif tmp[1] == "2":
        #     suffix = "nd"
        # elif tmp[1] == "3":
        #     suffix = "rd"
        # elif tmp[1] == "4" or tmp[1] == "5" or tmp[1] == "6":
        #     suffix = "th"

        # if tmp[0] == "sem":
        #     student.semester_year = str(tmp[1])+suffix+" Sem"
        # else:
        #     student.semester_year = str(tmp[1])+suffix+" Year"
        
        current_date = datetime.now().strftime('%Y-%m-%d')
        if 'date' in request.POST and request.POST['date'] != "":
            current_date = datetime.strptime(request.POST['date'], "%d/%m/%Y").strftime('%Y-%m-%d')
        
        context['current_date']            = current_date

        if TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).exists():
            last_record = TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).last()
            if last_record.end_datetime is None:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = None
            else:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = last_record.end_datetime
        else:
            student.punch_in_time = None
            student.punch_out_time = None


    paginator = Paginator(list(students), getConfigurationResult('page_limit'))

    try:
        students = paginator.page(page)
    except PageNotAnInteger:
        students = paginator.page(1)
    except EmptyPage:
        students = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    

    context['students']          = students
    context['total_record']          = total_record
    context['total_pages']            = total_pages
    context['page_limit']             = getConfigurationResult('page_limit')
    template = 'attendance/filter-student-attendance-report.html'
    return render(request, template, context)

@login_required
def ajaxStudentAttendanceLists(request):
    page = request.POST['page']

    condition = ''
    if 'search' in request.POST and request.POST['search'] != "":
        condition += " and (tbl_students.first_name LIKE '%%"+request.POST['search']+"%%' OR tbl_students.primary_contact_no LIKE '%%"+request.POST['search']+"%%') " 
        
    if 'college_id' in request.POST and request.POST['college_id'] != "":
        condition += " and tbl_students.college_id="+request.POST['college_id']

    if 'mentor' in request.POST and request.POST['mentor'] != "":
        condition += " and tbl_students.teacher_gaurdian_name='"+request.POST['mentor']+"'"

    if 'course' in request.POST and request.POST['course'] != "":
        condition += " and tbl_students.branch_id="+request.POST['course']

    if 'sem_year' in request.POST and request.POST['sem_year'] != "":
        condition += " and tbl_students.semester_id='"+request.POST['sem_year']+"'"

    if condition == "":
        students = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
        tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
        tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.alias,tbl_students.college_id,
        tbl_students.branch_id,tbl_students.semester_id,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id
        LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id WHERE tbl_students.is_registered = 1 AND  date(tbl_attendance.created_at)=CURDATE() ORDER BY tbl_students.id DESC ''')
    else:
        students = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.reg_no,tbl_students.first_name,
        tbl_students.middle_name,tbl_students.last_name,tbl_students.primary_contact_no,tbl_students.teacher_gaurdian_name,
        tbl_students.is_registered,tbl_students.is_mobile_verified,tbl_colleges.alias,tbl_students.college_id,
        tbl_students.branch_id,tbl_students.semester_id,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
        LEFT JOIN tbl_attendance on tbl_attendance.student_id = tbl_students.id
        WHERE tbl_students.is_registered = 1  AND  date(tbl_attendance.created_at)=CURDATE()  AND 1 {condition} ORDER BY tbl_students.id DESC '''.format(condition=condition))

    for student in students:
        college = TblColleges.objects.get(id=student.college_id)
        course = TblBranch.objects.get(id=student.branch_id)
        student.college_name = college.college_name
        student.college_alias = college.alias
        student.branch = course.branch

        tmp = student.semester_id.split('_')
        suffix = ""
        if tmp[1] == "1":
            suffix = "st"
        elif tmp[1] == "2":
            suffix = "nd"
        elif tmp[1] == "3":
            suffix = "rd"
        elif tmp[1] == "4" or tmp[1] == "5" or tmp[1] == "6":
            suffix = "th"

        if tmp[0] == "sem":
            student.semester_year = str(tmp[1])+suffix+" Sem"
        else:
            student.semester_year = str(tmp[1])+suffix+" Year"
        
        current_date = datetime.now().strftime('%Y-%m-%d')
        if 'date' in request.POST and request.POST['date'] != "":
            current_date = datetime.strptime(request.POST['date'], "%d/%m/%Y").strftime('%Y-%m-%d')

        if TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).exists():
            last_record = TblAttendance.objects.filter(student_id=student.id,created_at__contains=current_date).last()
            if last_record.end_datetime is None:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = None
            else:
                student.punch_in_time = last_record.start_datetime
                student.punch_out_time = last_record.end_datetime
        else:
            student.punch_in_time = None
            student.punch_out_time = None


    paginator = Paginator(list(students), getConfigurationResult('page_limit'))

    try:
        students = paginator.page(page)
    except PageNotAnInteger:
        students = paginator.page(1)
    except EmptyPage:
        students = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    
    context = {}
    context['students']     = students
    context['total_pages']       = total_pages

    template = 'attendance/ajax-student-attendance-list.html'
    return render(request, template, context)



@login_required
def studentAttendanceDetail(request,student_id):
    context = {}
    if request.method == "GET":
        context['flag'] = False
        context['message'] = "Method not allowed"
    else:
        student_details = TblStudents.objects.raw(''' SELECT tbl_students.*,tbl_colleges.college_name,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_colleges on tbl_colleges.id = tbl_students.college_id 
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
        WHERE tbl_students.id=%s ''',[student_id])
        
        if student_details:
            context['student_details'] = student_details[0]
            if student_details[0].college_id == 1:
                context['college_base_url'] = "http://bipe.sortstring.co.in/"
            elif student_details[0].college_id == 2:
                context['college_base_url'] = "http://bite.sortstring.co.in/"
            elif student_details[0].college_id == 3:
                context['college_base_url'] = "http://bip.sortstring.co.in/"

            current_filter_date = datetime.now().strftime('%Y-%m-%d')
            if 'date' in request.POST and request.POST['date'] != "":
                current_filter_date = datetime.strptime(request.POST['date'], "%d/%m/%Y").strftime('%Y-%m-%d')

            if TblAttendance.objects.filter(student_id=student_id,created_at__contains=current_filter_date).exists():
                last_record = TblAttendance.objects.filter(student_id=student_id,created_at__contains=current_filter_date).last()
                if last_record.end_datetime is None:
                    context['punch_in_status'] = 1
                    context['last_punch_in'] = last_record.start_datetime
                else:
                    context['punch_in_status'] = 0
                    context['last_punch_in'] = last_record.end_datetime
            else:
                context['punch_in_status'] = 0
                context['last_punch_in'] = None

            context['current_date'] = datetime.now().strftime('%Y-%m-%d')
            context['current_filter_date'] = current_filter_date
            template = "attendance/student-attendance-details.html"
           
            return render(request,template,context)
        else:
            context['flag'] = False
            context['message'] = "Record not found"

    return JsonResponse(context)

def sendWebNotification(student_id):
    registration_ids = []
    tokens = TblUserWebTokens.objects.all().values_list('token',flat=True).distinct()
    for token in tokens:
        registration_ids.append(token)
    if len(tokens):

        student_details = TblStudents.objects.raw(''' SELECT tbl_students.id,tbl_students.first_name,tbl_students.middle_name,
        tbl_students.last_name,tbl_students.semester_id,tbl_branch.branch FROM tbl_students
        LEFT JOIN tbl_branch on tbl_branch.id = tbl_students.branch_id
        WHERE tbl_students.id=%s ''',[student_id])


        
        student_name = student_details[0].first_name+' '
        if student_details[0].middle_name is not None:
            student_name += student_details[0].middle_name+' '
        if student_details[0].last_name is not None:
            student_name += student_details[0].last_name

        tmp = student_details[0].semester_id.split('_')
        suffix = ""
        if tmp[1] == "1":
            suffix = "st"
        elif tmp[1] == "2":
            suffix = "nd"
        elif tmp[1] == "3":
            suffix = "rd"
        elif tmp[1] == "4" or tmp[1] == "5" or tmp[1] == "6":
            suffix = "th"

        if tmp[0] == "sem":
            semester_year = str(tmp[1])+suffix+" Sem"
        else:
            semester_year = str(tmp[1])+suffix+" Year"

        title = "Attendance marked Successfully."
        message = student_name+" "
        message += "("+ student_details[0].branch +" - "+semester_year+") marked his/her attendance successfully at " +str(datetime.now().strftime('%H:%M'))+"." 
        
        result = sendWebPushNotification(title,message,registration_ids)
        
#Summary part

# List View
@login_required
def indexSummary(request):
    from datetime import date
    today   = date.today()  
    year  = today.year
    month = today.month
    date = today.day
    
    month_list = days_in_months(year,month)
    
    context = {}
    context['today_date']                   = today.strftime("%m/%Y")
    nameMonth = []
    for month_date in month_list:
        month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
        if int(date) >= int(month_date.strftime('%d')):
            nameMonth.append(month_date.strftime('%d %b'))

    context['month_list']                   = nameMonth
    context['colleges'] = colleges = TblColleges.objects.all().values('id','college_name')
    context['courses'] = courses = TblBranch.objects.filter(college_id=colleges[0]['id']).values('id','alias','abbr')
    
    context['page_title']                   = "Attendance Summary"
    template = 'attendance/attendance-summary.html'
    return render(request, template, context)

def checkAttendance(date,student_id):
    if TblAttendance.objects.filter(student_id=student_id,start_datetime__contains=date).exists():
        return True
    else:
        return False

def getPunchTime(date, student_id):
    context = {}
    first_record = TblAttendance.objects.filter(student_id=student_id,created_at__contains=date).first()
    context['start_time'] = first_record.start_datetime
    last_record = TblAttendance.objects.filter(student_id=student_id,created_at__contains=date).last()
    if last_record.end_datetime is None:
        context['end_time'] = ""
    else:
        context['end_time'] = last_record.end_datetime
    return context

def checkHoliday(date,organization_id,):
    if SpHolidays.objects.filter(organization_id=organization_id,start_date__gte=date,start_date__lte=date).exists():
        first_holiday = SpHolidays.objects.filter(organization_id=organization_id,start_date__gte=date,start_date__lte=date).first()
        return first_holiday.holiday
    else:
        return False 

# ajax List View
@login_required
def filterSummary(request):
    if request.method == "GET":
        from datetime import date
        today   = date.today()  
        year  = today.year
        month = today.month
        current_month = today.month

        date = today.day
        if request.GET['date'] !='':
            filter_date = str(request.GET['date']).split('/')
            if len(filter_date) >2:
                year = int(filter_date[2])
                month = int(filter_date[1])
            else:
                year = int(filter_date[1])
                month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date.strftime('%d %b'))
            else:
                nameMonth.append(month_date.strftime('%d %b'))
        
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = nameMonth
        
        condition = ''
        if request.GET['college_id'] !='':
            condition += " and college_id= "+request.GET['college_id']
        if request.GET['course'] != '':
            condition += " and branch_id= "+request.GET['course']
        if request.GET['sem_year'] != '':
            condition += " and semester_id= '"+request.GET['sem_year']+"'"
        students = TblStudents.objects.raw(''' SELECT id,first_name,middle_name,last_name,reg_no FROM tbl_students WHERE is_registered = 1 and 1 {condition} '''.format(condition=condition))
        
        total_class = 0
        total_present_students = 0

        for student in students:
            student_attendance = []
            present_count = 0
            absent_count = 0
            student_total_class = 0
            for months in month_list[:len(nameMonth)]:
                
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                in_time = ''
                out_time = ''

                if datetime.strptime(str(months), '%d/%m/%Y').strftime('%a') == 'Sun':
                    attendance_status = 'SUN'
                elif checkHoliday(month_date,student.college_id):
                    attendance_status = checkHoliday(month_date,student.college_id)
                else:
                    total_class = total_class +1
                    student_total_class = student_total_class + 1

                    if checkAttendance(month_date, student.id):
                        attendance_status         = "P"
                        present_count = present_count +1
                        total_present_students = total_present_students +1
                        
                        punch_time = getPunchTime(month_date, student.id)
                        in_time = punch_time['start_time']
                        out_time = punch_time['end_time']
                    else:
                        attendance_status         = "A"
                        absent_count = absent_count +1

                tmp               = {}
                tmp['attendance_status']   = attendance_status
                tmp['in_time']   = in_time
                tmp['out_time']   = out_time

                student_attendance.append(tmp) 

            student.attendances             = student_attendance
            student.present_count           = present_count
            student.absent_count            = absent_count
            student.attendance_percentage   = float((present_count / student_total_class ) * 100)
        
        context['total_class'] = total_class
        context['total_present_students'] = total_present_students
        if total_present_students == 0:
            context['attendance_percentage'] = 0
        else:
            context['attendance_percentage'] = float((total_present_students / total_class ) * 100)

        context['students'] = students
        context['page_title']                   = "Attendance Summary"
        template = 'attendance/ajax-attendance-summary.html'
        return render(request, template, context)  
    

@login_required
def exportToXlsx(request, college_id, course, sem_year, filter_date):

    if request.method == "GET":
        from datetime import date
        today   = date.today()
        year  = today.year
        month = today.month
        current_month = today.month
        date = today.day
        if filter_date !='':
            filter_date = filter_date.split('-')
            
            year = int(filter_date[1])
            month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date)
            else:
                nameMonth.append(month_date)
        
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = days_in_months(year,month)
        
        condition = ''
        filter_string = ''
        if int(college_id) > 0:
            condition += " and college_id= "+college_id
            filter_string += getModelColumnById(SpOrganizations,college_id,'organization_name')
        if course != "course":
            condition += " and branch_id= "+course
            filter_string += ' - '+getModelColumnById(TblBranch,course,'abbr')
        if sem_year != '0' :
            condition += " and semester_id= '"+sem_year+"'"
            if 'year' in sem_year:
                filter_string += ' - '+sem_year.split('_')[1] + ' Year'
            if 'sem' in sem_year:
                filter_string += ' - '+sem_year.split('_')[1] + ' Semester'

        students = TblStudents.objects.raw(''' SELECT id,first_name,middle_name,last_name,reg_no FROM tbl_students WHERE  is_registered = 1 and 1 {condition} '''.format(condition=condition))
        
        total_class = 0
        total_present_students = 0

        for student in students:
            student_attendance = []
            present_count = 0
            absent_count = 0
            student_total_class = 0
            for months in month_list[:len(nameMonth)]:
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                if datetime.strptime(str(months), '%d/%m/%Y').strftime('%a') == 'Sun':
                    attendance_status = 'SUN'
                elif checkHoliday(month_date,student.college_id):
                    attendance_status = checkHoliday(month_date,student.college_id)
                else:
                    student_total_class = student_total_class + 1
                    total_class = total_class + 1

                    if checkAttendance(month_date, student.id):
                        attendance_status         = "P"
                        present_count = present_count +1
                        total_present_students = total_present_students +1
                    else:
                        attendance_status         = "A"
                        absent_count = absent_count +1

                tmp               = {}
                tmp['attendance_status']   = attendance_status

                student_attendance.append(tmp) 

            student.attendances =  student_attendance
            student.present_count =  present_count
            student.absent_count =  absent_count
            student.attendance_percentage   = float((present_count / student_total_class ) * 100)

        if total_present_students == 0:
            attendance_percentage = 0
        else:
            attendance_percentage = float((total_present_students / total_class ) * 100)

        filter_string += ' - ('+str(round(attendance_percentage,2))+')%'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=student-attendance-list.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Summary Report'
    
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True

    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sansthaa.png')
    img.height = 63
    img.width = 80
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)
    worksheet['A1'].alignment = wrapped_alignment

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = calendar.month_name[int(filter_date[0])] + ' '+filter_date[1]
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    column_length = len(nameMonth) + 4
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = filter_string
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    row_num = 2
    # Define the titles for columns
    columns = ['S.No']
    columns += [ 'Student name' ]
    for month_date in nameMonth:
        # month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')#.strftime('%Y-%m-%d')
        columns += [ month_date.strftime('%d %b') ]

    columns += [ 'Total Present' ]
    columns += [ 'Total Absent' ]
    

    # Assign the titles for each cell of the header
    header_column_counter = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = thin_border
        
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        if header_column_counter == 0:
            cell.alignment = centered_alignment

        elif header_column_counter == 1:
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = centered_alignment


        header_column_counter = header_column_counter + 1

    # Iterate through all movies
    for student in students:
        row_num += 1
        # Define the data for each cell in the row 
        column = [str(int(row_num)-2)]
        name = ''
        if student.first_name:
            name += student.first_name
        if student.middle_name:
            name += " "+student.middle_name
        if student.last_name:
            name += " "+student.last_name
        column += [ name +"\n("+ student.reg_no+")" ]
        for attendance in student.attendances:
            column += [ attendance['attendance_status'] ]
        column += [ str(student.present_count)+'('+ str(round(student.attendance_percentage,2)) +'%)' ]
        column += [ student.absent_count ]
        
        content_column_counter = 0
        for col_num, cell_value in enumerate(column, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = header_font
            cell.border = thin_border
            
            if content_column_counter == 0:
                column_dimensions.width = 5

            elif content_column_counter == 1:
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 30
                cell.alignment = Alignment(horizontal='left',wrap_text=True)
            else:
                cell.alignment = centered_alignment

            if content_column_counter > 1 and content_column_counter < len(nameMonth)+2:
                if cell_value == "P":
                    cell.fill = PatternFill(start_color="7EC857", end_color="7EC857", fill_type = "solid")
                elif cell_value == "SUN":
                    cell.fill = PatternFill(start_color="fd8823", end_color="fd8823", fill_type = "solid")
                elif cell_value == "A":
                    cell.fill = PatternFill(start_color="FF4859", end_color="FF4859", fill_type = "solid")
                else:
                    cell.fill = PatternFill(start_color="c7fd2370", end_color="c7fd2370", fill_type = "solid")
            
            content_column_counter = content_column_counter + 1

    workbook.save(response)

    return response


@login_required
def attendanceStat(request):
    context = {}
    context['page_title'] = "Attendance Stats"
    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        to_date   = datetime.now().date()
        from_date = to_date  - dateutil.relativedelta.relativedelta(months=1)
    

    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    moderate            = 0
    moderate_percent    = 0.0
    good                = 0
    good_percent        = 0.0
    excellent           = 0
    excellent_percent   = 0.0

    if from_date != "" or to_date != "" :
        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.now().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves +=  (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date, to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date, to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves 
        days_between_dates = days_between_dates + timedelta(1)
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date +timedelta(1) ]).values('student_id').distinct()
        list_1 = []
        if len(all_students) != 0:
            for each_student in all_students:
                list_1.append(each_student['student_id'])
                attendance_date = TblAttendance.objects.filter(student_id = each_student['student_id']).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student['student_id']).count()
                if attendance_count == 1:
                    if attendance_date.exists():
                        for each_attend in attendance_date:
                            if datetime.today().date() == each_attend['start_datetime'].date():
                                first_time += 1
                            # break
                    else:
                        very_poor += 1
                else:
                    if days_between_dates.days == 0:
                        percent = (attendance_count / 1) * 100
                    else:
                        percent = (attendance_count / days_between_dates.days) * 100

                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                    elif (percent >30 and percent <= 50):
                        poor += 1
                    elif (percent >50 and percent <= 80):
                        moderate += 1
                    elif (percent >80 and percent <= 90):
                        good += 1
                    elif (percent>90):
                        excellent += 1
                    else:
                        pass

            first_time_percent = round((first_time / registered_student) * 100,2)
            very_poor_percent = round((very_poor / registered_student) * 100,2)
            poor_percent = round((poor / registered_student) * 100,2)
            moderate_percent = round((moderate / registered_student) * 100,2)
            good_percent = round((good / registered_student) * 100,2)
            excellent_percent = round((excellent / registered_student) * 100,2)
        zero_attendance = TblStudents.objects.filter(is_registered=1).exclude(id__in = list_1).count()
        zero_percent = round((zero_attendance / (registered_student)) * 100,2)

        context['zero_attendance']      = zero_attendance
        context['zero_percent']         = zero_percent
        context['semester1']            = TblSemester.objects.all().filter(type__startswith="semester")
        context['branch']               =TblBranch.objects.all().filter(college_id = college_id)
        context['from_date']            = from_date.strftime(("%d-%m-%Y"))
        context['to_date']              = to_date.strftime(("%d-%m-%Y"))
        context['total_registered']     = registered_student
        context['total_unregistered']   = unregistered_student
        context['today_registered']     = registered_today
        context['first_time']           = first_time
        context['first_time_percent']   = first_time_percent
        context['very_poor']            = very_poor
        context['very_poor_percent']    = very_poor_percent
        context['poor']                 = poor
        context['poor_percent']         = poor_percent
        context['moderate']             = moderate
        context['moderate_percent']     = moderate_percent
        context['good']                 = good
        context['good_percent']         = good_percent
        context['excellent']            = excellent
        context['excellent_percent']    = excellent_percent

        template = 'attendance/attendance-stat.html'
        return render(request, template, context)

@login_required
def ajaxAttendanceStat(request):
    context = {}
    context['page_title'] = "Attendance Stats"

    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        to_date   = datetime.now()
        from_date = to_date  - dateutil.relativedelta.relativedelta(months=1)

    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    good                = 0
    good_percent        = 0
    excellent           = 0
    excellent_percent   = 0.0

    if from_date != "" and to_date != "" :
        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
        content = []
        raw_data = {}
        if len(all_students) != 0:
            for each_student in all_students:
                attendance_date = TblAttendance.objects.filter(student_id = each_student['student_id']).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student['student_id']).count()
                if attendance_count == 1:
                    for each_attend in attendance_date:
                        print("DEF: ", each_attend['start_datetime'].date())
                        if datetime.today().date() == each_attend['start_datetime'].date():
                            first_time += 1
                            name = getStudentName(each_student['student_id'])
                            each_data = TblStudents.objects.filter(each_student['student_id']).values('reg_no', 'branch_id', 'semester_id')
                            reg_no = each_data[0]['reg_no']
                            raw_data['student_name'] = name
                            raw_data['reg_no'] = reg_no
                            raw_data['branch'] = TblBranch.objects.filter(id = each_data[0]['branch_id']).values('abbr')[0]['abbr']
                            raw_data['semester'] = TblSemester.objects.filter(semester_id = each_data[0]['semester_id']).values('sem_name')[0]['sem_name']
                            content.append(raw_data)
                            break
                else:
                    percent = (attendance_count / days_between_dates.days) * 100
                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                        name = getStudentName(each_student['student_id'])
                        raw_data['student_name'] = each_student['student_id']
                        content.append(raw_data)
                    elif (percent > 30 and percent <= 50):
                        poor += 1
                    elif (percent >= 80 and percent <= 90):
                        good += 1
                    elif (percent>90):
                        excellent += 1
                    else:
                        pass

            first_time_percent = round((first_time / len(all_students)) * 100, 2)
            very_poor_percent = round((very_poor / len(all_students)) * 100, 2)
            poor_percent = round((poor / len(all_students)) * 100, 2)
            good_percent = round((good / len(all_students)) * 100, 2)
            excellent_percent = round((excellent / len(all_students)) * 100, 2)
    

        context['total_registered']     = registered_student
        context['total_unregistered']   = unregistered_student
        context['today_registered']     = registered_today
        context['first_time']           = first_time
        context['first_time_percent']   = first_time_percent
        context['very_poor']            = very_poor
        context['very_poor_percent']    = very_poor_percent
        context['poor']                 = poor
        context['poor_percent']         = poor_percent
        context['good']                 = good
        context['good_percent']         = good_percent
        context['excellent']            = excellent
        context['excellent_percent']    = excellent_percent
        template = 'attendance/ajax-attendance-stat.html'
        return render(request, template, context)

@login_required
def attendanceStatsExportToXlsx(request,from_date,to_date,branch,semester):
    

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Attendance-Stats.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    branch1 = branch
    if branch1 != 'null':
        branch  = ast.literal_eval(branch1)

    semester1 = semester
    if semester1 != 'null':
        semester  = ast.literal_eval(semester1)
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Stats'
    columns = []
    columns += [ 'From Date : ' +from_date ]
    columns += [ 'To Date : ' +to_date ]

    row_num =1


    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20



    branch_names=""
    if branch == 'null':
        branch_names = "All Branches"
    else:
        for i, br in enumerate(branch, 1):
            if br == "all":
                branch_names = "All Branches"
            else:
                if i == len(branch):
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" "
                elif i == len(branch)-1:
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" & "
                else:
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" , "

    semester_names=""
    if semester  == "null":
        semester_names = "All semester"
    else:
        for j, sem in enumerate(semester, 1):
            if sem  == "all":
                semester_names = "All semester"
            else:
                if j == len(semester):
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" "
                elif j == len(semester)-1:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" & "
                else:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" , "


    columns_1 = []
    columns_1 += [ 'Branch:'  +branch_names  ]
    columns_1 += [ 'Semester:' +semester_names  ]

    row_num =2


    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20



    columns = []
    columns += [ 'Annotation' ]
    columns += [ 'Number of Students' ]
    
    row_num = 3
    from_date = datetime.strptime(from_date, "%d-%m-%Y") 
    to_date = datetime.strptime(to_date, "%d-%m-%Y") + timedelta(1)
    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    moderate            = 0
    moderate_percent    = 0.0
    good                = 0
    good_percent        = 0.0
    excellent           = 0
    excellent_percent   = 0.0

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20


    college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
    registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)
    if  "all" in branch or branch == 'null':
        registered_student = registered_student
    elif branch:
        registered_student = registered_student.filter(branch_id__in = [id for id in branch])
    else:
        pass
    if "all" in semester or semester == 'null':
        registered_student = registered_student
    elif semester:
        registered_student = registered_student.filter(semester_id__in = [id for id in semester])
    else:
        pass
    unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0)
    if "all" in  branch or branch == 'null':
        unregistered_student = unregistered_student
    elif branch:
        unregistered_student = unregistered_student.filter(branch_id__in = [id for id in branch])
    else:
        pass
    if "all" in  semester or semester == 'null':
        unregistered_student = unregistered_student
    elif semester:
        unregistered_student = unregistered_student.filter(semester_id__in = [id for id in semester])
    else:
        pass
    for i in range(4):
        row_num += 1
        # Define the data for each cell in the row 
        if i == 0:
            row = [ "Registered Students" ]
            row += [ registered_student.count() ]
        elif i == 1:
            row = [ "Unregistered Students" ]
            row += [ unregistered_student.count() ]
        elif i == 2:
            row = [ "Today's Registered Students" ]
            row += [ TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count() ]
       
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    columns_1 = []
    columns_1 += [ 'Attendance Zone' ]
    columns_1 += [ 'Number of Students' ]
    columns_1 += [ 'Percentage ' ]
    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
        
    # branch = request.GET.getlist('branch[]')
    # semester = request.GET.getlist('semester[]')
    holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
    leaves=timedelta(0)
    for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)
    if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
    else:
        sundays_count = 0
    days_between_dates = (to_date - from_date) + timedelta(1)
    days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
    all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
    stu_id = []
    for abc in all_students:
        stu = TblStudents.objects.filter(id =  abc['student_id'])
        if "all" in  branch or branch == 'null':
            stu = stu
        elif branch:
            stu = stu.filter(branch_id__in = [id for id in branch])
        if "all" in semester or semester == 'null':
            stu = stu
        elif semester:
            stu = stu.filter(semester_id__in = [id for id in semester])
        for xyz in stu:
            stu_id.append(xyz.id)
    if len(stu_id) != 0:
        for each_student in stu_id:
            attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
            attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
            if attendance_count == 1:
                if attendance_date.exists():
                    for each_attend in attendance_date:
                        if datetime.today().date() == each_attend['start_datetime'].date():
                            first_time += 1
                            # break
                else:
                    very_poor += 1
            else:
                if days_between_dates.days == 0:
                    percent = (attendance_count / 1) * 100
                else:
                    percent = (attendance_count / days_between_dates.days) * 100

                if (percent > 0 and percent <= 30):
                    very_poor += 1
                elif (percent > 30 and percent <= 50):
                    poor += 1
                elif (percent >50 and percent <= 80):
                    moderate += 1
                elif (percent > 80 and percent <= 90):
                    good += 1
                elif (percent>90):
                    excellent += 1
                else:
                    pass

            first_time_percent = round((first_time / registered_student.count()) * 100,2)
            very_poor_percent = round((very_poor / registered_student.count()) * 100,2)
            poor_percent = round((poor / registered_student.count()) * 100,2)
            moderate_percent = round((moderate / registered_student.count()) * 100,2)
            good_percent = round((good / registered_student.count()) * 100,2)
            excellent_percent = round((excellent / registered_student.count()) * 100,2)


    zero_attendance = TblStudents.objects.filter(is_registered=1)
    if "all" in  branch:
        zero_attendance = zero_attendance
    elif branch:
        zero_attendance = zero_attendance.filter(branch_id__in = [id for id in branch])
    if "all" in semester:
        zero_attendance = zero_attendance
    elif semester:
        zero_attendance = zero_attendance.filter(semester_id__in = [id for id in semester])

    zero_attendance = zero_attendance.exclude(id__in = stu_id)
    zero_percent = round((zero_attendance.count() / (registered_student.count())) * 100,2)
    for i in range(7):
        row_num += 1
        # Define the data for each cell in the row 
        if i == 0:
            row = [ "First Time (0%)" ]
            row += [ first_time ]
            row += [ first_time_percent ]
        elif i == 1:
            row = [ "Very Poor (0-30%)" ]
            row += [ very_poor ]
            row += [ very_poor_percent ]
        elif i == 2:
            row = [ "Poor (31-50%)" ]
            row += [ poor ]
            row += [ poor_percent ]
        elif i == 3:
            row = [ "Moderate (51-80%)" ]
            row += [ moderate ]
            row += [ moderate_percent ]
        elif i == 4:
            row = [ "Good (81-90%)" ]
            row += [ good ]
            row += [ good_percent ]
        elif i == 5:
            row = [ "Excellent (90%+)" ]
            row += [ excellent ]
            row += [ excellent_percent ]
        elif i == 6:
            row = [ "Never Attended" ]
            row += [ zero_attendance.count() ]
            row += [ zero_percent ]
       
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)
    return response

@login_required
def getRegisterZone(request):
    context = {}
    branch = request.GET.getlist('branch[]')
    semester = request.GET.getlist('semester[]')

    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        to_date   = datetime.now()
        from_date = to_date  - dateutil.relativedelta.relativedelta(months=1)

    if from_date != "" or to_date != "" :  
        #=====================================================register-count 
     
        if request.GET['check'] == "0":

                registered_student1 = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)  
                if  "all" in branch:
                    registered_student1 = registered_student1
                elif branch:
                    registered_student1 = registered_student1.filter(branch_id__in = [id for id in branch])
                else:
                    pass

                if "all" in semester:
                    registered_student1 = registered_student1
                elif semester:
                    registered_student1 = registered_student1.filter(semester_id__in = [id for id in semester])
                else:
                    pass
                branch              = TblBranch.objects.all().filter(college_id = college_id)
                semester            = TblSemester.objects.all().filter(type__startswith="semester")
                for name in registered_student1:
                    name.full_name = getStudentName(name.id)

        #==================================================unregister-count

        elif request.GET['check'] == "1":


                registered_student1 = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0)
                if "all" in  branch:
                    registered_student1 = registered_student1
                elif branch:
                    registered_student1 = registered_student1.filter(branch_id__in = [id for id in branch])
                else:
                    pass

                if "all" in  semester:
                    registered_student1 = registered_student1
                elif semester:
                    registered_student1 = registered_student1.filter(semester_id__in = [id for id in semester])
                else:
                    pass

                branch              = TblBranch.objects.all().filter(college_id = college_id)
                semester            = TblSemester.objects.all().filter(type__startswith="semester")
                for name in registered_student1:
                    name.full_name = getStudentName(name.id)
        #==========================today-count
        elif request.GET['check'] == "2":
                registered_student1 = TblStudents.objects.filter(college_id = college_id).filter(is_registered=1 , registered_date__icontains = datetime.today().date() )
                branch              = TblBranch.objects.all().filter(college_id = college_id)
                semester            = TblSemester.objects.all().filter(type__startswith="semester") 
                for name in registered_student1:
                    name.full_name = getStudentName(name.id)
    
    context['counter']              = len(registered_student1)
    context['registered_student1']  = registered_student1
    context['branch']               = branch
    context['semester']             = semester
    context['check']                = request.GET['check']
    context['table_head']                = request.GET['table_head']
    template = 'attendance/get-attendance-zone.html'
    return render(request, template, context)

@login_required
def getAttendaceZone(request):
    context = {}
    branch = request.GET.getlist('branch[]')
    semester = request.GET.getlist('semester[]')

    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.GET['from_date'] != '':
            from_date = datetime.strptime(request.GET['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.GET['to_date'] != '':
            to_date = datetime.strptime(request.GET['to_date'], "%d-%m-%Y")
            to_date = to_date + timedelta(1)
        else:
            to_date = datetime.now() + timedelta(1)
        

    first_time = 0
    first_time_list = []
    first_time_percent = 0.0
    very_poor = 0
    very_poor_list = []
    very_poor_percent = 0.0
    poor = 0
    poor_list = []
    poor_percent = 0.0
    moderate            = 0
    moderate_list       = []
    moderate_percent    = 0.0
    good = 0
    good_list = []
    good_percent = 0.0
    excellent = 0
    excellent_list = []
    excellent_percent = 0.0

    if from_date != "" or to_date != "" :
        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
        stu_id = []
        for abc in all_students:
            stu = TblStudents.objects.filter(id = abc['student_id'])
            if "all" in  branch or branch == 'null':
                stu = stu
            elif branch:
                stu = stu.filter(branch_id__in = [id for id in branch])
            if "all" in  semester or semester == 'null' :
                stu = stu
            elif semester:
                stu = stu.filter(semester_id__in = [id for id in semester])
            for xyz in stu: 
                stu_id.append(xyz.id)
        if len(stu_id) != 0:
            for each_student in stu_id:
                attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                if attendance_count == 1:
                    if attendance_date.exists():
                        for each_attend in attendance_date:
                            if datetime.today().date() == each_attend['start_datetime'].date():
                                first_time += 1
                                first_time_list.append(each_student)
                                # break 
                    else:
                        very_poor += 1
                        very_poor_list.append(each_student)
                else:
                    # percent = (attendance_count / days_between_dates.days) * 100
                    if days_between_dates.days == 0:
                        percent = (attendance_count / 1) * 100
                    else:
                        percent = (attendance_count / days_between_dates.days) * 100
                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                        very_poor_list.append(each_student)
                    elif (percent > 30 and percent <= 50):
                        poor += 1
                        poor_list.append(each_student)
                    elif (percent >50 and percent <= 80):
                        moderate += 1
                        moderate_list.append(each_student)
                    elif (percent >= 80 and percent <= 90):
                        good += 1
                        good_list.append(each_student)
                    elif (percent>90):
                        excellent += 1
                        excellent_list.append(each_student)
                    else:
                        pass

            first_time_percent = round((first_time / registered_student) * 100, 2)
            very_poor_percent = round((very_poor / registered_student) * 100, 2)
            poor_percent = round((poor / registered_student) * 100, 2)
            moderate_percent = round((moderate / registered_student) * 100,2)
            good_percent = round((good / registered_student) * 100, 2)
            excellent_percent = round((excellent / registered_student) * 100, 2)
        zero_attendance = TblStudents.objects.filter(is_registered=1)
        if "all" in  branch:
            zero_attendance = zero_attendance
        elif branch:
            zero_attendance = zero_attendance.filter(branch_id__in = [id for id in branch])
        if "all" in semester:
            zero_attendance = zero_attendance
        elif semester:
            zero_attendance = zero_attendance.filter(semester_id__in = [id for id in semester])

        zero_attendance = zero_attendance.exclude(id__in = stu_id)
        zero_percent = round((len(zero_attendance) / (registered_student)) * 100,2)
        reg_data = {}
        content = []
        semester            = TblSemester.objects.all().filter(type__startswith="semester")
        branch              = TblBranch.objects.all().filter(college_id = college_id)
        #====================attendance-count========================================
        if request.GET['check'] == "3":
            for each_student in first_time_list:
                reg_data = {}
                name = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)#.values('reg_no')[0]['reg_no']
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "4":
            for each_student in very_poor_list:
                reg_data = {}
                name                = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "5":
            for each_student in poor_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)



        elif request.GET['check'] == "6":
            for each_student in moderate_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)



        elif request.GET['check'] == "7":
            for each_student in good_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)
                
        elif request.GET['check'] == "8":
            for each_student in excellent_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)
        elif request.GET['check'] == "9":
            for each_student in zero_attendance:
                reg_data     = {}
                name         = getStudentName(each_student.id)
                regs         = TblStudents.objects.filter(id = each_student.id)
                for newone in regs:
                    reg_data['reg_no']      = newone.reg_no
                    reg_data['branch_id']   = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student.id
                content.append(reg_data)
        else:
            pass

        
        context['counter']                  = len(content)
        context['registered_student1']      = content
        context['branch']                   = branch
        context['semester']                 = semester
        context['total_registered']         = registered_student
        context['total_unregistered']       = unregistered_student
        context['today_registered']         = registered_today
        context['first_time']               = first_time
        context['first_time_percent']       = first_time_percent
        context['very_poor']                = very_poor
        context['very_poor_percent']        = very_poor_percent
        context['poor']                     = poor
        context['poor_percent']             = poor_percent
        context['moderate']                 = moderate
        context['moderate_percent']         = moderate_percent
        context['good']                     = good
        context['good_percent']             = good_percent
        context['excellent']                = excellent
        context['excellent_percent']        = excellent_percent
        context['check']                    = request.GET['check']
        context['table_head']               = request.GET['table_head']
        template = 'attendance/get-attendance-zone.html'
        return render(request, template, context)

@login_required
def attendanceReportExportToXlsx(request,list,branch,semester,from_date,to_date):
    response = BytesIO()

    array = list
    student_array  = ast.literal_eval(array)
    branch1 = branch
    if branch1 != 'null':
        branch  = ast.literal_eval(branch1)
    semester1 = semester
    if semester1 != 'null':
        semester  = ast.literal_eval(semester1)
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Report'
    columns = []
    columns += [ 'From Date : ' +from_date ]
    columns += [ 'To Date : ' +to_date ]

    row_num =1


    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20



    branch_names=""
    if branch == 'null':
        branch_names = "All Branches"
    else:
        for i, br in enumerate(branch, 1):
            if br == "all":
                branch_names = "All Branches"
            else:
                if i == len(branch):
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" "
                elif i == len(branch)-1:
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" & "
                else:
                    branch_names += TblBranch.objects.filter(id = br).values("abbr")[0]["abbr"]+" , "

    semester_names=""
    if semester  == "null":
        semester_names = "All semester"
    else:
        for j, sem in enumerate(semester, 1):
            if sem  == "all":
                semester_names = "All semester"
            else:
                if j == len(semester):
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" "
                elif j == len(semester)-1:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" & "
                else:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" , "


    


    columns_1 = []
    columns_1 += [ 'Branch : ' +branch_names ]
    columns_1 += [ 'Semester : ' +semester_names ]
    row_num =2


    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
    
    # Define the titles for columns
    columns = []
    columns += [ 'S.No.' ]
    columns += [ 'Name' ]
    columns += [ 'Registration number' ]
    columns += [ 'Branch' ]
    columns += [ 'Semester' ]
    row_num  =  3
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for i,each_id in enumerate(student_array, 1):
        row_num += 1
        each_data = TblStudents.objects.filter(id = each_id)
        row = [ i ]
        row += [ getStudentName(each_id) ]
        row += [ each_data[0].reg_no ]
        row += [ TblBranch.objects.filter(id = each_data[0].branch_id).values('abbr')[0]['abbr'] ]
        row += [ TblSemester.objects.filter(semester_id = each_data[0].semester_id).values('sem_name')[0]['sem_name'] ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)
    response.seek(0)
    return FileResponse(response, as_attachment=True, filename="Student-List.xlsx")


@login_required
def filterRegisterZone(request):
    context = {}
    branch = request.GET.getlist('branch[]')
    semester = request.GET.getlist('semester[]')



    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
        
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.GET['from_date'] != '':
            from_date = datetime.strptime(request.GET['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.GET['to_date'] != '':
            to_date = datetime.strptime(request.GET['to_date'], "%d-%m-%Y")
            to_date = to_date + timedelta(1)
        else:
            to_date = datetime.now() + timedelta(1)

    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    moderate            = 0
    moderate_percent    = 0.0
    good                = 0
    good_percent        = 0.0
    excellent           = 0
    excellent_percent   = 0.0

    if from_date != "" or to_date != "" :
        #=======================register student 

        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)
        if  "all" in branch:
            registered_student = registered_student
        elif branch:
            registered_student = registered_student.filter(branch_id__in = [id for id in branch])
        else:
            pass

        if "all" in semester:
            registered_student = registered_student
        elif semester:
            registered_student = registered_student.filter(semester_id__in = [id for id in semester])
        else:
            pass
        registered_student = registered_student.count()

        #==============================unregister student

        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0)
        if "all" in  branch:
            unregistered_student = unregistered_student
        elif branch:
            unregistered_student = unregistered_student.filter(branch_id__in = [id for id in branch])
        else:
            pass
        if "all" in  semester:
            unregistered_student = unregistered_student
        elif semester:
            unregistered_student = unregistered_student.filter(semester_id__in = [id for id in semester])
        else:
            pass
        unregistered_student = unregistered_student.count()

        #==========================================today
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count()

        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)
        #=====================Attendance zone
        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()

        stu_id = []
        for abc in all_students:
            stu = TblStudents.objects.filter(id =  abc['student_id'])
            if "all" in  branch:
                stu = stu
            elif branch:
                stu = stu.filter(branch_id__in = [id for id in branch])
            if "all" in semester:
                stu = stu
            elif semester:
                stu = stu.filter(semester_id__in = [id for id in semester])
            for xyz in stu:
                stu_id.append(xyz.id)
        if len(stu_id) != 0:
            for each_student in stu_id:
                attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                if attendance_count == 1:
                    if attendance_date.exists():
                        for each_attend in attendance_date:
                            if datetime.today().date() == each_attend['start_datetime'].date():
                                first_time += 1
                            # break
                    else:
                        very_poor += 1        
                else:
                    # percent = (attendance_count / days_between_dates.days) * 100
                    if days_between_dates.days == 0:
                        percent = (attendance_count / 1) * 100
                    else:
                        percent = (attendance_count / days_between_dates.days) * 100

                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                    elif (percent > 30 and percent <= 50):
                        poor += 1
                    elif (percent >50 and percent <= 80):
                        moderate += 1
                    elif (percent > 80 and percent <= 90):
                        good += 1
                    elif (percent>90):
                        excellent += 1
                    else:
                        pass
            if registered_student != 0 :
                first_time_percent = round((first_time / registered_student) * 100, 2)
                very_poor_percent = round((very_poor / registered_student) * 100, 2)
                poor_percent = round((poor / registered_student) * 100, 2)
                moderate_percent = round((moderate / registered_student) * 100,2)
                good_percent = round((good / registered_student) * 100, 2)
                excellent_percent = round((excellent / registered_student) * 100, 2)

        zero_attendance = TblStudents.objects.filter(is_registered=1)
        if "all" in  branch:
            zero_attendance = zero_attendance
        elif branch:
            zero_attendance = zero_attendance.filter(branch_id__in = [id for id in branch])
        if "all" in semester:
            zero_attendance = zero_attendance
        elif semester:
            zero_attendance = zero_attendance.filter(semester_id__in = [id for id in semester])

        zero_attendance = zero_attendance.exclude(id__in = stu_id)
        if registered_student != 0:
            zero_percent = round((len(zero_attendance) / (registered_student)) * 100,2)
        else:
            zero_percent = 0.0

        context['total_registered']     = registered_student
        context['total_unregistered']   = unregistered_student
        context['today_registered']     = registered_today
        context['zero_attendance']      = len(zero_attendance)
        context['zero_percent']         = zero_percent
        context['first_time']           = first_time
        context['first_time_percent']   = first_time_percent
        context['very_poor']            = very_poor
        context['very_poor_percent']    = very_poor_percent
        context['poor']                 = poor
        context['poor_percent']         = poor_percent
        context['moderate']             = moderate
        context['moderate_percent']     = moderate_percent
        context['good']                 = good
        context['good_percent']         = good_percent
        context['excellent']            = excellent
        context['excellent_percent']    = excellent_percent
        template = 'attendance/ajax-attendance-stat.html'
        return render(request, template, context)


@login_required
def attendanceReport(request):
    context = {}
    context['page_title'] = "Attendance Report"
    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        to_date   = datetime.now().date()
        from_date = to_date  - dateutil.relativedelta.relativedelta(months=1)

    if from_date != "" or to_date != "" :
        # registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        # registered_student_details = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)
        # unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        # registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.now().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves +=  (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date, to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date, to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves 
        days_between_dates = days_between_dates + timedelta(1)
        Attendance = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date +timedelta(1) ]).values('student_id').distinct()

        civil_engineering = []
        electrical_engineering = []
        mechanical_engineering_Automobile = []
        mechanical_engineering_production = []
        branches= {}
        response = []
        for each in Attendance:
            if str(18) == TblStudents.objects.filter(id = each['student_id']).values('branch_id')[0]['branch_id']:  
                civil_engineering.append(each['student_id'])
            elif str(19) == TblStudents.objects.filter(id = each['student_id']).values('branch_id')[0]['branch_id']: 
                electrical_engineering.append(each['student_id'])
            elif str(20) == TblStudents.objects.filter(id = each['student_id']).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_Automobile.append(each['student_id'])
            elif str(21) == TblStudents.objects.filter(id = each['student_id']).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_production.append(each['student_id'])

        branches['Civil Engineering'] = civil_engineering
        branches['Electrical Engineering'] = electrical_engineering
        branches['ME(Automobile)'] = mechanical_engineering_Automobile
        branches['ME(Production)'] = mechanical_engineering_production

        if len(Attendance) != 0:
            for key,values in branches.items():
                first_time          = 0
                first_time_percent  = 0.0
                very_poor           = 0
                very_poor_percent   = 0.0
                poor                = 0
                poor_percent        = 0.0
                moderate            = 0
                moderate_percent    = 0.0
                good                = 0
                good_percent        = 0.0
                excellent           = 0
                excellent_percent   = 0.0

                for each_student in values:
                    attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                    attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                    if attendance_count == 1:
                        if attendance_date.exists():
                            for each_attend in attendance_date:
                                if datetime.today().date() == each_attend['start_datetime'].date():
                                    first_time += 1
                        else:
                            very_poor += 1
                    else:
                        if days_between_dates.days == 0:
                            percent = (attendance_count / 1) * 100
                        else:
                            percent = (attendance_count / days_between_dates.days) * 100

                        if (percent > 0 and percent <= 30):
                            very_poor += 1
                        elif (percent >30 and percent <= 50):
                            poor += 1
                        elif (percent >50 and percent <= 80):
                            moderate += 1
                        elif (percent >80 and percent <= 90):
                            good += 1
                        elif (percent>90):
                            excellent += 1
                        else:
                            pass
                leng = TblStudents.objects.filter(is_registered=1, branch_id = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id']).count()
                first_time_percent = round((first_time / (leng)) * 100,2)
                very_poor_percent = round((very_poor / (leng)) * 100,2)
                poor_percent = round((poor / (leng)) * 100,2)
                moderate_percent = round((moderate / (leng)) * 100,2)
                good_percent = round((good / leng) * 100,2)
                excellent_percent = round((excellent / leng) * 100,2)

                zero_attendance = TblStudents.objects.filter(is_registered=1, branch_id = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id'])
                zero_attendance = zero_attendance.exclude(id__in = values)
                zero_percent = round((len(zero_attendance) / (leng)) * 100,2)
                res = {}
                res['branch'] = key
                res['branch_id'] = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id']
                res['registered'] = leng
                res['first'] = first_time
                res['first_percent'] = first_time_percent
                res['very_poor'] = very_poor
                res['very_poor_percent'] = very_poor_percent
                res['poor'] = poor
                res['poor_percent'] = poor_percent
                res['moderate'] = moderate
                res['moderate_percent'] = moderate_percent
                res['good'] = good
                res['good_percent'] = good_percent
                res['excellent'] = excellent
                res['excellent_percent'] = excellent_percent
                res['zero'] = zero_attendance.count()
                res['zero_percent'] = zero_percent
                response.append(res)

        context['response']             = response
        context['semester1']            = TblSemester.objects.all().filter(type__startswith="semester")
        context['branch']               = TblBranch.objects.all().filter(college_id = college_id)
        context['from_date']            = from_date.strftime(("%d-%m-%Y"))
        context['to_date']              = to_date.strftime(("%d-%m-%Y"))

        template = 'attendance/attendance-report.html'
        return render(request, template, context)

@login_required
def ajaxattendanceReport(request):
    context = {}
    semester = request.GET.getlist('semester[]')
    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
        
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.GET['from_date'] != '':
            from_date = datetime.strptime(request.GET['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.GET['to_date'] != '':
            to_date = datetime.strptime(request.GET['to_date'], "%d-%m-%Y")
            to_date = to_date + timedelta(1)
        else:
            to_date = datetime.now() + timedelta(1)

    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    moderate            = 0
    moderate_percent    = 0.0
    good                = 0
    good_percent        = 0.0
    excellent           = 0
    excellent_percent   = 0.0

    if from_date != "" or to_date != "" :
        #=======================register student 

        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)

        if "all" in semester:
            registered_student = registered_student
        elif semester:
            registered_student = registered_student.filter(semester_id__in = [id for id in semester])
        else:
            pass
        # registered_student = registered_student.count()

        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)
        #=====================Attendance zone
        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        Attendance = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
        stu_id = []
        civil_engineering = []
        electrical_engineering = []
        mechanical_engineering_Automobile = []
        mechanical_engineering_production = []
        branches= {}
        response = []
        for abc in Attendance:
            stu = TblStudents.objects.filter(id =  abc['student_id'])
            if "all" in semester:
                stu = stu
            elif semester:
                stu = stu.filter(semester_id__in = [id for id in semester])
            for xyz in stu:
                stu_id.append(xyz.id)

        for each in stu_id:
            if str(18) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']:  
                civil_engineering.append(each)
            elif str(19) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                electrical_engineering.append(each)
            elif str(20) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_Automobile.append(each)
            elif str(21) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_production.append(each)

        branches['Civil Engineering'] = civil_engineering
        branches['Electrical Engineering'] = electrical_engineering
        branches['ME(Automobile)'] = mechanical_engineering_Automobile
        branches['ME(Production)'] = mechanical_engineering_production

        if len(stu_id) != 0:
            for key,values in branches.items():
                first_time          = 0
                first_time_percent  = 0.0
                very_poor           = 0
                very_poor_percent   = 0.0
                poor                = 0
                poor_percent        = 0.0
                moderate            = 0
                moderate_percent    = 0.0
                good                = 0
                good_percent        = 0.0
                excellent           = 0
                excellent_percent   = 0.0

                for each_student in values:
                    attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                    attendance_count = TblAttendance.objects.filter(student_id = each_student).count()

                    if attendance_count == 1:
                        if attendance_date.exists():
                            for each_attend in attendance_date:
                                if datetime.today().date() == each_attend['start_datetime'].date():
                                    first_time += 1
                        else:
                            very_poor += 1        
                    else:
                        if days_between_dates.days == 0:
                            percent = (attendance_count / 1) * 100
                        else:
                            percent = (attendance_count / days_between_dates.days) * 100

                        if (percent > 0 and percent <= 30):
                            very_poor += 1
                        elif (percent > 30 and percent <= 50):
                            poor += 1
                        elif (percent >50 and percent <= 80):
                            moderate += 1
                        elif (percent > 80 and percent <= 90):
                            good += 1
                        elif (percent>90):
                            excellent += 1
                        else:
                            pass
                # if reg_leng != 0:
                reg_leng = TblStudents.objects.filter(is_registered=1, branch_id = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id'])
                if "all" in semester:
                    reg_leng = reg_leng
                elif semester:
                    reg_leng = reg_leng.filter(semester_id__in = [id for id in semester])
                leng = reg_leng.count()
                first_time_percent = round((first_time / leng) * 100,2)
                very_poor_percent = round((very_poor / leng) * 100,2)
                poor_percent = round((poor / leng) * 100,2)
                moderate_percent = round((moderate / leng) * 100,2)
                good_percent = round((good / leng) * 100,2)
                excellent_percent = round((excellent / leng) * 100,2)

                zero_attendance = reg_leng
                zero_attendance = zero_attendance.exclude(id__in = values)
                zero_percent = round((len(zero_attendance) / leng) * 100,2)
                res = {}
                res['branch'] = key
                res['branch_id'] = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id']
                res['registered'] = leng
                res['first'] = first_time
                res['first_percent'] = first_time_percent
                res['very_poor'] = very_poor
                res['very_poor_percent'] = very_poor_percent
                res['poor'] = poor
                res['poor_percent'] = poor_percent
                res['moderate'] = moderate
                res['moderate_percent'] = moderate_percent
                res['good'] = good
                res['good_percent'] = good_percent
                res['excellent'] = excellent
                res['excellent_percent'] = excellent_percent
                res['zero'] = zero_attendance.count()
                res['zero_percent'] = zero_percent
                response.append(res)
                # else:
                #     res['branch'] = 0.0
                #     res['branch_id'] = 0.0
                #     res['registered'] = 0.0
                #     res['first'] = 0.0
                #     res['first_percent'] = 0.0
                #     res['very_poor'] = 0.0
                #     res['very_poor_percent'] = 0.0
                #     res['poor'] = 0.0
                #     res['poor_percent'] = 0.0
                #     res['moderate'] = 0.0
                #     res['moderate_percent'] = 0.0
                #     res['good'] = 0.0
                #     res['good_percent'] = 0.0
                #     res['excellent'] = 0.0
                #     res['excellent_percent'] = 0.0
                #     res['zero'] = 0.0
                #     res['zero_percent'] = 0.0
                #     response.append(res)

        context['response']  = response
        template = 'attendance/ajax-attendance-report.html'
        return render(request, template, context)

@login_required
def attendanceReportExportToXlsx(request,from_date,to_date,semester):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Attendance-Report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )

    semester1 = semester
    if semester1 != 'null':
        semester  = ast.literal_eval(semester1)

    workbook = Workbook()
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    worksheet = workbook.active
    worksheet.title = 'Attendance Report'
    columns = []
    columns += [ 'From Date : ' +from_date ]
    columns += [ 'To Date : ' +to_date ]

    row_num =1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    semester_names = ""
    if semester  == "null":
        semester_names = "All semester"
    else:
        for j, sem in enumerate(semester, 1):
            if sem  == "all":
                semester_names = "All semester"
            else:
                if j == len(semester):
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" "
                elif j == len(semester)-1:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" & "
                else:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" , "

    columns_1 = []
    columns_1 += [ 'Semester:' +semester_names  ]
    row_num =2

    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    
    from_date           = datetime.strptime(from_date, "%d-%m-%Y") 
    to_date             = datetime.strptime(to_date, "%d-%m-%Y") + timedelta(1)
    first_time          = 0
    first_time_percent  = 0.0
    very_poor           = 0
    very_poor_percent   = 0.0
    poor                = 0
    poor_percent        = 0.0
    moderate            = 0
    moderate_percent    = 0.0
    good                = 0
    good_percent        = 0.0
    excellent           = 0
    excellent_percent   = 0.0

    # Assign the titles for each cell of the header
    # for col_num, column_title in enumerate(columns, 1):
    #     cell = worksheet.cell(row=row_num, column=col_num)
    #     cell.value = column_title
    #     cell.font = header_font
    #     cell.alignment = centered_alignment
    #     cell.font = Font(size=12, color='FFFFFFFF', bold=True)
    #     cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

    #     column_letter = get_column_letter(col_num)
    #     column_dimensions = worksheet.column_dimensions[column_letter]
    #     column_dimensions.width = 20

    college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
    registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1)
    if "all" in semester or semester == 'null':
        registered_student = registered_student
    elif semester:
        registered_student = registered_student.filter(semester_id__in = [id for id in semester])
    else:
        pass
    unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0)
    if "all" in  semester or semester == 'null':
        unregistered_student = unregistered_student
    elif semester:
        unregistered_student = unregistered_student.filter(semester_id__in = [id for id in semester])
    else:
        pass

    columns_1 = []
    columns_1 += [ 'Branch' ]
    columns_1 += [ 'Registered-students' ]
    columns_1 += [ '0%' ]
    columns_1 += [ '0-30%' ]
    columns_1 += [ '31-50%' ]
    columns_1 += [ '51-80%' ]
    columns_1 += [ '81-90%' ]
    columns_1 += [ '90%' ]
    columns_1 += [ 'Never Attended' ]
    row_num = 3
    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
        
    holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
    leaves=timedelta(0)
    for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)
    if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
    else:
        sundays_count = 0
    days_between_dates = (to_date - from_date) + timedelta(1)
    days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
    Attendance = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
    stu_id = []
    civil_engineering = []
    electrical_engineering = []
    mechanical_engineering_Automobile = []
    mechanical_engineering_production = []
    branches= {}
    response1 = []
    for abc in Attendance:
        stu = TblStudents.objects.filter(id =  abc['student_id'])
        if "all" in semester:
            stu = stu
        elif semester:
            stu = stu.filter(semester_id__in = [id for id in semester])
        for xyz in stu:
            stu_id.append(xyz.id)

    for each in stu_id:
            if str(18) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']:  
                civil_engineering.append(each)
            elif str(19) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                electrical_engineering.append(each)
            elif str(20) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_Automobile.append(each)
            elif str(21) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_production.append(each)

    branches['Civil Engineering'] = civil_engineering
    branches['Electrical Engineering'] = electrical_engineering
    branches['ME(Automobile)'] = mechanical_engineering_Automobile
    branches['ME(Production)'] = mechanical_engineering_production

    if len(stu_id) != 0:
        for key,values in branches.items():
                first_time          = 0
                first_time_percent  = 0.0
                very_poor           = 0
                very_poor_percent   = 0.0
                poor                = 0
                poor_percent        = 0.0
                moderate            = 0
                moderate_percent    = 0.0
                good                = 0
                good_percent        = 0.0
                excellent           = 0
                excellent_percent   = 0.0
                never_attended      = 0
                never_attended_percent = 0.0
                for each_student in values:
                    attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                    attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                    if attendance_count == 0:
                            never_attended += 1
                    if attendance_count == 1:
                        if attendance_date.exists():
                            for each_attend in attendance_date:
                                if datetime.today().date() == each_attend['start_datetime'].date():
                                    first_time += 1
                                # break
                        else:
                            very_poor += 1    
                    else:
                        if days_between_dates.days == 0:
                            percent = (attendance_count / 1) * 100
                        else:
                            percent = (attendance_count / days_between_dates.days) * 100

                        if (percent > 0 and percent <= 30):
                            very_poor += 1
                        elif (percent > 30 and percent <= 50):
                            poor += 1
                        elif (percent >50 and percent <= 80):
                            moderate += 1
                        elif (percent > 80 and percent <= 90):
                            good += 1
                        elif (percent>90):
                            excellent += 1
                        else:
                            pass

                reg_leng = TblStudents.objects.filter(is_registered=1, branch_id = TblBranch.objects.filter(abbr__icontains = key).values('id')[0]['id'])
                if "all" in semester:
                    reg_leng = reg_leng
                elif semester:
                    reg_leng = reg_leng.filter(semester_id__in = [id for id in semester])
                leng = reg_leng.count()
                first_time_percent = round((first_time / (leng)) * 100,2)
                very_poor_percent = round((very_poor / (leng)) * 100,2)
                poor_percent = round((poor / (leng)) * 100,2)
                moderate_percent = round((moderate / (leng)) * 100,2)
                good_percent = round((good / leng) * 100,2)
                excellent_percent = round((excellent / leng) * 100,2)

                zero_attendance = reg_leng
                zero_attendance = zero_attendance.exclude(id__in = values)
                zero_percent = round((len(zero_attendance) / (leng)) * 100,2)
                res = {}
                res['branch'] = key
                res['registered'] = leng
                res['first'] = first_time
                res['first_percent'] = first_time_percent
                res['very_poor'] = very_poor
                res['very_poor_percent'] = very_poor_percent
                res['poor'] = poor
                res['poor_percent'] = poor_percent
                res['moderate'] = moderate
                res['moderate_percent'] = moderate_percent
                res['good'] = good
                res['good_percent'] = good_percent
                res['excellent'] = excellent
                res['excellent_percent'] = excellent_percent
                res['zero'] = zero_attendance.count()
                res['zero_percent'] = zero_percent
                response1.append(res)

    for i in response1:
        row_num += 1 
        row = ""
        row = [ i['branch']]
        row += [ i['registered' ]]
        row += [ str(i['first']) + "("+str(i['first_percent'])+"%)" ]
        row += [ str(i['very_poor']) + "("+str(i['very_poor_percent'])+"%)" ]
        row += [ str(i['poor']) + "("+str(i['poor_percent'])+"%)" ]
        row += [ str(i['moderate']) + "("+str(i['moderate_percent'])+"%)" ]
        row += [ str(i['good']) + "("+str(i['good_percent'])+"%)" ]
        row += [ str(i['excellent']) + "("+str(i['excellent_percent'])+"%)" ]
        row += [ str(i['zero']) + "("+str(i['zero_percent'])+"%)" ]

        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)
    return response

@login_required
def getAttendaceReport(request):
    context = {}
    semester = request.GET.getlist('semester[]')
    branch_id = request.GET['branch_id']
    # print(branch_id)

    if request.method == "POST":
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.POST['from_date'] != '':
            from_date = datetime.strptime(request.POST['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.POST['to_date'] != '':
            to_date = datetime.strptime(request.POST['to_date'], "%d-%m-%Y")
        else:
            to_date = datetime.now() 
    else:
        college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
        if request.GET['from_date'] != '':
            from_date = datetime.strptime(request.GET['from_date'], "%d-%m-%Y")
        else:
            from_date = datetime.now() - dateutil.relativedelta.relativedelta(months=1)
        if request.GET['to_date'] != '':
            to_date = datetime.strptime(request.GET['to_date'], "%d-%m-%Y")
            to_date = to_date + timedelta(1)
        else:
            to_date = datetime.now() + timedelta(1)

    first_time = 0
    first_time_list = []
    first_time_percent = 0.0
    very_poor = 0
    very_poor_list = []
    very_poor_percent = 0.0
    poor = 0
    poor_list = []
    poor_percent = 0.0
    moderate            = 0
    moderate_list       = []
    moderate_percent    = 0.0
    good = 0
    good_list = []
    good_percent = 0.0
    excellent = 0
    excellent_list = []
    excellent_percent = 0.0

    if from_date != "" or to_date != "" :
        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
        stu_id = []
        civil_engineering = []
        electrical_engineering = []
        mechanical_engineering_Automobile = []
        mechanical_engineering_production = []
        branches= {}
        response = []
        for abc in all_students:
            stu = TblStudents.objects.filter(id = abc['student_id'], branch_id = branch_id)
            if "all" in  semester or semester == 'null' :
                stu = stu
            elif semester:
                stu = stu.filter(semester_id__in = [id for id in semester])
            for xyz in stu: 
                stu_id.append(xyz.id)
        for each in stu_id:
            if str(18) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']:  
                civil_engineering.append(each)
            elif str(19) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                electrical_engineering.append(each)
            elif str(20) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_Automobile.append(each)
            elif str(21) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_production.append(each)

        branches['Civil Engineering'] = civil_engineering
        branches['Electrical Engineering'] = electrical_engineering
        branches['ME(Automobile)'] = mechanical_engineering_Automobile
        branches['ME(Production)'] = mechanical_engineering_production

        if len(stu_id) != 0:
            for each_student in stu_id:
                attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                
                if attendance_count == 1:
                    if attendance_date.exists():
                        for each_attend in attendance_date:
                            if datetime.today().date() == each_attend['start_datetime'].date():
                                first_time += 1
                                first_time_list.append(each_student)
                    else:
                        very_poor += 1
                        very_poor_list.append(each_student)
                else:
                    if days_between_dates.days == 0:
                        percent = (attendance_count / 1) * 100
                    else:
                        percent = (attendance_count / days_between_dates.days) * 100
                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                        very_poor_list.append(each_student)
                    elif (percent > 30 and percent <= 50):
                        poor += 1
                        poor_list.append(each_student)
                    elif (percent >50 and percent <= 80):
                        moderate += 1
                        moderate_list.append(each_student)
                    elif (percent >= 80 and percent <= 90):
                        good += 1
                        good_list.append(each_student)
                    elif (percent>90):
                        excellent += 1
                        excellent_list.append(each_student)
                    else:
                        pass

            first_time_percent = round((first_time / registered_student) * 100, 2)
            very_poor_percent = round((very_poor / registered_student) * 100, 2)
            poor_percent = round((poor / registered_student) * 100, 2)
            moderate_percent = round((moderate / registered_student) * 100,2)
            good_percent = round((good / registered_student) * 100, 2)
            excellent_percent = round((excellent / registered_student) * 100, 2)
        zero_attendance = TblStudents.objects.filter(is_registered=1, branch_id = branch_id)
        if "all" in semester:
            zero_attendance = zero_attendance
        elif semester:
            zero_attendance = zero_attendance.filter(semester_id__in = [id for id in semester])

        zero_attendance = zero_attendance.exclude(id__in = stu_id)
        zero_percent = round((len(zero_attendance) / (registered_student)) * 100,2)
        reg_data = {}
        content = []
        semester_fetch      = TblSemester.objects.all().filter(type__startswith="semester")
        branch              = TblBranch.objects.all().filter(college_id = college_id)

        #====================attendance-count========================================
        if request.GET['check'] == "1":
            reg_list = TblStudents.objects.filter(is_registered=1, branch_id = branch_id)

            if "all" in semester:
                reg_list = reg_list
            elif semester:
                reg_list = reg_list.filter(semester_id__in = [id for id in semester])

            for each_student in reg_list:
                reg_data = {}
                name = getStudentName(each_student.id)
                reg_data['reg_no'] = each_student.reg_no
                reg_data['branch_id'] = each_student.branch_id
                reg_data['semester_id'] = each_student.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "3":
            for each_student in first_time_list:
                reg_data = {}
                name = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)#.values('reg_no')[0]['reg_no']
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "4":
            for each_student in very_poor_list:
                reg_data = {}
                name                = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "5":
            for each_student in poor_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "6":
            for each_student in moderate_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "7":
            for each_student in good_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)
                
        elif request.GET['check'] == "8":
            for each_student in excellent_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif request.GET['check'] == "9":
            for each_student in zero_attendance:
                reg_data     = {}
                name         = getStudentName(each_student.id)
                regs         = TblStudents.objects.filter(id = each_student.id)
                for newone in regs:
                    reg_data['reg_no']      = newone.reg_no
                    reg_data['branch_id']   = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student.id
                content.append(reg_data)
        else:
            pass

        # for content in content:
        #     content.branch_id


        context['counter']                  = len(content)
        context['registered_student1']      = content
        context['branch_id']                = branch_id
        context['branch']                   = branch
        context['semester']                 = semester_fetch
        context['total_registered']         = registered_student
        context['total_unregistered']       = unregistered_student
        context['today_registered']         = registered_today
        context['first_time']               = first_time
        context['first_time_percent']       = first_time_percent
        context['very_poor']                = very_poor
        context['very_poor_percent']        = very_poor_percent
        context['poor']                     = poor
        context['poor_percent']             = poor_percent
        context['moderate']                 = moderate
        context['moderate_percent']         = moderate_percent
        context['good']                     = good
        context['good_percent']             = good_percent
        context['excellent']                = excellent
        context['excellent_percent']        = excellent_percent
        context['check']                    = request.GET['check']
        context['table_head']               = request.GET['table_head']
        template = 'attendance/get-attendance-students-report.html'
        return render(request, template, context)

@login_required
def attendanceStudentsReportExportToXlsx(request,check,semester,branch_id,from_date,to_date):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Register-Students-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    semester1 = semester
    if semester1 != 'null':
        semester  = ast.literal_eval(semester1)
    workbook = Workbook()
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Student Report'
    columns = []
    columns += [ 'From Date : ' +from_date ]
    columns += [ 'To Date : ' +to_date ]

    row_num =1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    semester_names=""
    if semester  == "null":
        semester_names = "All semester"
    else:
        for j, sem in enumerate(semester, 1):
            if sem  == "all":
                semester_names = "All semester"
            else:
                if j == len(semester):
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" "
                elif j == len(semester)-1:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" & "
                else:
                    semester_names += TblSemester.objects.filter(semester_id=sem).values("sem_name")[0]["sem_name"]+" , "

    columns_1 = []
    # columns_1 += [ 'Branch : ' +branch_names ]
    columns_1 += [ 'Semester : ' +semester_names ]
    row_num =2

    for col_num, column_title in enumerate(columns_1, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20
    
    # Define the titles for columns
    columns = []
    columns += [ 'S.No.' ]
    columns += [ 'Name' ]
    columns += [ 'Registration number' ]
    columns += [ 'Branch' ]
    columns += [ 'Semester' ]
    row_num  =  3
    
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    from_date           = datetime.strptime(from_date, "%d-%m-%Y") 
    to_date             = datetime.strptime(to_date, "%d-%m-%Y") + timedelta(1)    
    first_time = 0
    first_time_list = []
    first_time_percent = 0.0
    very_poor = 0
    very_poor_list = []
    very_poor_percent = 0.0
    poor = 0
    poor_list = []
    poor_percent = 0.0
    moderate            = 0
    moderate_list       = []
    moderate_percent    = 0.0
    good = 0
    good_list = []
    good_percent = 0.0
    excellent = 0
    excellent_list = []
    excellent_percent = 0.0


    college_id = SpUsers.objects.filter(id = request.user.id).values('organization_id')[0]['organization_id']
    if from_date != "" or to_date != "" :
        registered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).count()
        unregistered_student = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 0).count()
        registered_today = TblStudents.objects.filter(college_id = college_id).filter(is_registered = 1).filter(registered_date__icontains = datetime.today().date()).count()
        holidays = SpHolidays.objects.filter(organization_id = college_id).filter(start_date__gte = from_date).filter(start_date__lte = to_date )
        leaves=timedelta(0)
        for holiday in holidays:
            leaves += (holiday.end_date - holiday.start_date) + timedelta(1)

        if weekday_count(from_date,to_date) != {} and "Sunday" in (weekday_count(from_date,to_date)):
            sundays_count = (weekday_count(from_date,to_date))['Sunday']
        else:
            sundays_count = 0
        days_between_dates = (to_date - from_date) + timedelta(1)
        days_between_dates = (days_between_dates) - timedelta(sundays_count) - leaves
        all_students = TblAttendance.objects.filter(start_datetime__range=[from_date, to_date]).values('student_id').distinct()
        stu_id = []
        civil_engineering = []
        electrical_engineering = []
        mechanical_engineering_Automobile = []
        mechanical_engineering_production = []
        branches= {}
        # response = []
        for abc in all_students:
            stu = TblStudents.objects.filter(id = abc['student_id'], branch_id = branch_id)
            if "all" in  semester or semester == 'null' :
                stu = stu
            elif semester:
                stu = stu.filter(semester_id__in = [id for id in semester])
            for xyz in stu: 
                stu_id.append(xyz.id)
        for each in stu_id:
            if str(18) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']:  
                civil_engineering.append(each)
            elif str(19) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                electrical_engineering.append(each)
            elif str(20) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_Automobile.append(each)
            elif str(21) == TblStudents.objects.filter(id = each).values('branch_id')[0]['branch_id']: 
                mechanical_engineering_production.append(each)

        branches['Civil Engineering'] = civil_engineering
        branches['Electrical Engineering'] = electrical_engineering
        branches['ME(Automobile)'] = mechanical_engineering_Automobile
        branches['ME(Production)'] = mechanical_engineering_production

        if len(stu_id) != 0:
            for each_student in stu_id:
                attendance_date = TblAttendance.objects.filter(student_id = each_student).filter(start_datetime__icontains = datetime.today().date()).values('start_datetime')
                attendance_count = TblAttendance.objects.filter(student_id = each_student).count()
                
                if attendance_count == 1:
                    if attendance_date.exists():
                        for each_attend in attendance_date:
                            if datetime.today().date() == each_attend['start_datetime'].date():
                                first_time += 1
                                first_time_list.append(each_student)
                    else:
                        very_poor += 1
                        very_poor_list.append(each_student)
                else:
                    if days_between_dates.days == 0:
                        percent = (attendance_count / 1) * 100
                    else:
                        percent = (attendance_count / days_between_dates.days) * 100
                    if (percent > 0 and percent <= 30):
                        very_poor += 1
                        very_poor_list.append(each_student)
                    elif (percent > 30 and percent <= 50):
                        poor += 1
                        poor_list.append(each_student)
                    elif (percent >50 and percent <= 80):
                        moderate += 1
                        moderate_list.append(each_student)
                    elif (percent >= 80 and percent <= 90):
                        good += 1
                        good_list.append(each_student)
                    elif (percent>90):
                        excellent += 1
                        excellent_list.append(each_student)
                    else:
                        pass

            first_time_percent = round((first_time / registered_student) * 100, 2)
            very_poor_percent = round((very_poor / registered_student) * 100, 2)
            poor_percent = round((poor / registered_student) * 100, 2)
            moderate_percent = round((moderate / registered_student) * 100,2)
            good_percent = round((good / registered_student) * 100, 2)
            excellent_percent = round((excellent / registered_student) * 100, 2)


        zero_attendance = TblStudents.objects.filter(is_registered=1, branch_id = branch_id)
        if "all" in semester:
            zero_attendance = zero_attendance
        elif semester:
            zero_attendance = zero_attendance.filter(semester_id__in = [id for id in semester])

        zero_attendance = zero_attendance.exclude(id__in = stu_id)
        zero_percent = round((len(zero_attendance) / (registered_student)) * 100,2)
        reg_data = {}
        content = []
        semester_fetch      = TblSemester.objects.all().filter(type__startswith="semester")
        branch              = TblBranch.objects.all().filter(college_id = college_id)

        if check == "1":
            reg_list = TblStudents.objects.filter(is_registered=1, branch_id = branch_id)

            if "all" in semester:
                reg_list = reg_list
            elif semester:
                reg_list = reg_list.filter(semester_id__in = [id for id in semester])

            for each_student in reg_list:
                reg_data = {}
                name = getStudentName(each_student.id)
                reg_data['reg_no'] = each_student.reg_no
                reg_data['branch_id'] = each_student.branch_id
                reg_data['semester_id'] = each_student.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student.id
                content.append(reg_data)

        elif check == "3":
            for each_student in first_time_list:
                reg_data = {}
                name = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)#.values('reg_no')[0]['reg_no']
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif check == "4":
            for each_student in very_poor_list:
                reg_data = {}
                name                = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif check == "5":
            for each_student in poor_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif check == "6":
            for each_student in moderate_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif check == "7":
            for each_student in good_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)
                
        elif check == "8":
            for each_student in excellent_list:
                reg_data                  = {}
                name                      = getStudentName(each_student)
                regs                = TblStudents.objects.filter(id = each_student)
                for newone in regs:
                    reg_data['reg_no'] = newone.reg_no
                    reg_data['branch_id'] = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student
                content.append(reg_data)

        elif check == "9":
            for each_student in zero_attendance:
                reg_data     = {}
                name         = getStudentName(each_student.id)
                regs         = TblStudents.objects.filter(id = each_student.id)
                for newone in regs:
                    reg_data['reg_no']      = newone.reg_no
                    reg_data['branch_id']   = newone.branch_id
                    reg_data['semester_id'] = newone.semester_id
                reg_data['full_name'] = name
                reg_data['id'] = each_student.id
                content.append(reg_data)
        else:
            pass 

    for i,each_id in enumerate(content,1):
        row_num += 1
        each_data = TblStudents.objects.filter(id = each_id['id'])
        row = [ i ]
        row += [ each_id['full_name'] ]
        row += [ each_id['reg_no'] ]
        row += [ TblBranch.objects.filter(id = each_data[0].branch_id).values('abbr')[0]['abbr'] ]
        row += [ TblSemester.objects.filter(semester_id = each_data[0].semester_id).values('sem_name')[0]['sem_name']]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)
    return response





# List View
@login_required
def EmployeeSummary(request):
    from datetime import date
    today   = date.today()  
    year  = today.year
    month = today.month
    date = today.day
    
    month_list = days_in_months(year,month)
    
    context = {}
    context['today_date']                   = today.strftime("%m/%Y")
    nameMonth = []
    for month_date in month_list:
        month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
        if int(date) >= int(month_date.strftime('%d')):
            nameMonth.append(month_date.strftime('%d %b'))

    context['month_list']                   = nameMonth
    context['colleges'] = colleges = TblColleges.objects.all().values('id','college_name')
    context['courses'] = courses = ''
    
    context['page_title']                   = "Attendance Summary"
    template = 'attendance/employee-attendance-summary.html'
    return render(request, template, context)


def getEmployeePunchTime(date, employee_id):
    context = {}
    first_record = SpUserAttendance.objects.filter(user_id=employee_id, attendance_date_time__contains=date).first()
    context['start_time'] = first_record.start_time
    last_record = SpUserAttendance.objects.filter(user_id=employee_id, attendance_date_time__contains=date).last()
    if last_record.end_time is None:
        context['end_time'] = ""
    else:
        context['end_time'] = last_record.end_time
    return context

def checkEmployeeAttendance(date,student_id):
    if SpUserAttendance.objects.filter(user_id=student_id,attendance_date_time__contains=date).exists():
        return True
    else:
        return False


# ajax List View
@login_required
def filterEmployeeAttendanceSummary(request):
    if request.method == "GET":
        from datetime import date
        today   = date.today()  
        year  = today.year
        month = today.month
        current_month = today.month

        date = today.day
        if request.GET['date'] !='':
            filter_date = str(request.GET['date']).split('/')
            if len(filter_date) >2:
                year = int(filter_date[2])
                month = int(filter_date[1])
            else:
                year = int(filter_date[1])
                month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date.strftime('%d %b'))
            else:
                nameMonth.append(month_date.strftime('%d %b'))
        
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = nameMonth
        
        condition = ''
        if request.GET['college_id'] !='':
            condition += " and organization_id= "+request.GET['college_id']
        if request.GET['department'] != '':
            condition += " and department_id= "+request.GET['department']
        # if request.GET['sem_year'] != '':
        #     condition += " and semester_id= '"+request.GET['sem_year']+"'"
        employees = SpUsers.objects.raw(''' SELECT id,first_name,middle_name,last_name,emp_sap_id,organization_id FROM sp_users WHERE status = 1 and 1 {condition} '''.format(condition=condition))
        
        total_class = 0
        total_present_students = 0

        for employee in employees:
            employee_attendance = []
            present_count = 0
            absent_count = 0
            student_total_class = 0
            for months in month_list[:len(nameMonth)]:
                
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                in_time = ''
                out_time = ''

                if datetime.strptime(str(months), '%d/%m/%Y').strftime('%a') == 'Sun':
                    attendance_status = 'SUN'
                elif checkHoliday(month_date,employee.organization_id):
                    attendance_status = checkHoliday(month_date,employee.organization_id)
                else:
                    total_class = total_class +1
                    student_total_class = student_total_class + 1

                    if checkEmployeeAttendance(month_date, employee.id):
                        attendance_status         = "P"
                        present_count = present_count +1
                        total_present_students = total_present_students +1
                        
                        punch_time = getEmployeePunchTime(month_date, employee.id)
                        in_time = punch_time['start_time']
                        out_time = punch_time['end_time']
                    else:
                        attendance_status         = "A"
                        absent_count = absent_count +1

                tmp               = {}
                tmp['attendance_status']   = attendance_status
                tmp['in_time']   = in_time
                tmp['out_time']   = out_time

                employee_attendance.append(tmp) 

            employee.attendances             = employee_attendance
            employee.present_count           = present_count
            employee.absent_count            = absent_count
            employee.attendance_percentage   = float((present_count / student_total_class ) * 100)
        
        context['total_class'] = total_class
        context['total_present_students'] = total_present_students
        if total_present_students == 0:
            context['attendance_percentage'] = 0
        else:
            context['attendance_percentage'] = float((total_present_students / total_class ) * 100)

        context['employees']     = employees
        context['page_title']   = "Employee Attendance Summary"
        template = 'attendance/ajax-employee-attendance-summary.html'
        return render(request, template, context)  
    


@login_required
def exportEmployeeAttendanceToXlsx(request, college_id, department, filter_date):

    if request.method == "GET":
        from datetime import date
        today   = date.today()
        year  = today.year
        month = today.month
        current_month = today.month
        date = today.day
        if filter_date !='':
            filter_date = filter_date.split('-')
            
            year = int(filter_date[1])
            month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date)
            else:
                nameMonth.append(month_date)
        
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = days_in_months(year,month)
        
        condition = ''
        filter_string = ''
        if int(college_id) > 0:
            condition += " and organization_id= "+college_id
            filter_string += getModelColumnById(SpOrganizations,college_id,'organization_name')
        if int(department) > 0:
            condition += " and department_id= "+department
            filter_string += ' - '+getModelColumnById(SpDepartments,department,'department_name')

        # students = TblStudents.objects.raw(''' SELECT id,first_name,middle_name,last_name,reg_no FROM tbl_students WHERE  is_registered = 1 and 1 {condition} '''.format(condition=condition))
        if int(department) > 0:
            employees = SpUsers.objects.raw(''' SELECT id,first_name,middle_name,last_name,emp_sap_id,organization_id FROM sp_users WHERE status = 1 and 1 {condition} '''.format(condition=condition))
        else:
            employees = []

        
        total_class = 0
        total_present_students = 0

        for employee in employees:
            student_attendance = []
            present_count = 0
            absent_count = 0
            student_total_class = 0
            for months in month_list[:len(nameMonth)]:
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                if datetime.strptime(str(months), '%d/%m/%Y').strftime('%a') == 'Sun':
                    attendance_status = 'SUN'
                elif checkHoliday(month_date,employee.organization_id):
                    attendance_status = checkHoliday(month_date,employee.organization_id)
                else:
                    student_total_class = student_total_class + 1
                    total_class = total_class + 1

                    if checkEmployeeAttendance(month_date, employee.id):
                        attendance_status         = "P"
                        present_count = present_count +1
                        total_present_students = total_present_students +1
                    else:
                        attendance_status         = "A"
                        absent_count = absent_count +1

                tmp               = {}
                tmp['attendance_status']   = attendance_status

                student_attendance.append(tmp) 

            employee.attendances =  student_attendance
            employee.present_count =  present_count
            employee.absent_count =  absent_count
            employee.attendance_percentage   = float((present_count / student_total_class ) * 100)

        if total_present_students == 0:
            attendance_percentage = 0
        else:
            attendance_percentage = float((total_present_students / total_class ) * 100)

        filter_string += ' - ('+str(round(attendance_percentage,2))+')%'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employee-attendance-list.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Summary Report'
    
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True

    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/sansthaa.png')
    img.height = 63
    img.width = 80
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)
    worksheet['A1'].alignment = wrapped_alignment

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = calendar.month_name[int(filter_date[0])] + ' '+filter_date[1]
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=12, bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    column_length = len(nameMonth) + 4
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = filter_string
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    row_num = 2
    # Define the titles for columns
    columns = ['S.No']
    columns += [ 'Employee name' ]
    for month_date in nameMonth:
        # month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')#.strftime('%Y-%m-%d')
        columns += [ month_date.strftime('%d %b') ]

    columns += [ 'Total Present' ]
    columns += [ 'Total Absent' ]
    

    # Assign the titles for each cell of the header
    header_column_counter = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = thin_border
        
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        if header_column_counter == 0:
            cell.alignment = centered_alignment

        elif header_column_counter == 1:
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = centered_alignment


        header_column_counter = header_column_counter + 1

    # Iterate through all movies
    for employee in employees:
        row_num += 1
        # Define the data for each cell in the row 
        column = [str(int(row_num)-2)]
        # name = ''
        # if student.first_name:
        #     name += student.first_name
        # if student.middle_name:
        #     name += " "+student.middle_name
        # if student.last_name:
        #     name += " "+student.last_name
        column += [ getUserName(employee.id) +"\n("+ employee.emp_sap_id+")" ]
        for attendance in employee.attendances:
            column += [ attendance['attendance_status'] ]
        column += [ str(employee.present_count)+'('+ str(round(employee.attendance_percentage,2)) +'%)' ]
        column += [ employee.absent_count ]
        
        content_column_counter = 0
        for col_num, cell_value in enumerate(column, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = header_font
            cell.border = thin_border
            
            if content_column_counter == 0:
                column_dimensions.width = 5

            elif content_column_counter == 1:
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 30
                cell.alignment = Alignment(horizontal='left',wrap_text=True)
            else:
                cell.alignment = centered_alignment

            if content_column_counter > 1 and content_column_counter < len(nameMonth)+2:
                if cell_value == "P":
                    cell.fill = PatternFill(start_color="7EC857", end_color="7EC857", fill_type = "solid")
                elif cell_value == "SUN":
                    cell.fill = PatternFill(start_color="fd8823", end_color="fd8823", fill_type = "solid")
                elif cell_value == "A":
                    cell.fill = PatternFill(start_color="FF4859", end_color="FF4859", fill_type = "solid")
                else:
                    cell.fill = PatternFill(start_color="c7fd2370", end_color="c7fd2370", fill_type = "solid")
            
            content_column_counter = content_column_counter + 1

    workbook.save(response)

    return response






@login_required
def attendanceReports(request):
    
    context = {}
    today = date.today().strftime('%Y-%m-%d')
    
    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(sp_user_attendance.attendance_date_time) = %s
     group by user_id ''',[today])
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')
    users = []
    today = date.today()

    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.role_name,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id])

            temp['date']=datetime.now().strftime('%d/%m/%Y')
            temp['id'] = user.user_id
            temp['name'] = start_attendance[0].name
            temp['role_name'] = start_attendance[0].role_name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['start_latitude'] = start_attendance[0].latitude
            temp['start_longitude'] = start_attendance[0].longitude
            temp['img'] = start_attendance[0].attendance_img
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
                temp['end_latitude'] = end_attendance[0].latitude
                temp['end_longitude'] = end_attendance[0].longitude
                temp['Eod'] = end_attendance[0].Eod
            else:
                temp['end_time'] = None
                temp['Eod'] = None

            now = str(today)
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()
            hours = convert(total_seconds)
            temp['working_hours'] = str(hours)

            distance_travelled  = SpUserTracking.objects.filter(user_id=start_attendance[0].user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
            if distance_travelled['distance_travelled__sum'] is not None:
                temp['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                temp['distance_travelled'] = 0.00

                temp['added_employee_count'] = SpUsers.objects.filter(created_by=start_attendance[0].user_id, created_at__icontains=today).count()
                temp['tagged_employee_count'] = 0
            
            users.append(temp)
            
   
            
    roles                       =  SpRoles.objects.filter(status=1).exclude(id=1)
    organization                =  SpOrganizations.objects.all()
    if roles:
        context['roles']            = roles
    else:
        context['roles']         =0
    context['employee']         = SpUsers.objects.filter(user_type=1,status=1).exclude(id=1)
    context['organization']     = organization
    context['users']            = users
    context['page_title']       = "Daily Attendance Report"
    template = 'attendance/daily-attendance-reports.html'
    return render(request, template,context)


def viewDetails(request,id):
    context = {}
    try:
        attendancedetail = SpUserAttendance.objects.get(user_id=id)
        attendancedetail.name = getUserName(attendancedetail.user_id)    
        attendancedetail.emp_id = getModelColumnById(SpUsers, attendancedetail.user_id,'emp_sap_id')    
    except ObjectDoesNotExist :
        attendancedetail=0
        attendancedetail.name=""
        attendancedetail.emp_id=""
    context['attendancedetail'] = attendancedetail
    template = 'attendance/view-detail.html'
    return render(request, template, context ) 


@login_required
def monthPerormanceReport(request,month,year):

    employees_list = []
    today   = datetime.today()  
    year  = int(year)
    month = int(month)

    month_list = days_in_months(year,month)
    
    employees = SpUsers.objects.filter(status=1).exclude(id=1)
    for employee in employees:
        employees_dict = {}
        employee_name = getUserName(employee.id)
        employees_dict['Department']    = employee.department_name  
        employees_dict['EmployeeId']   = employee.emp_sap_id
        employees_dict['EmployeeName'] = employee_name 
                    
        total_present   = 0
        total_absent    = 0
        total_leave     = 0 
        total_holiday   = 0
        total_weekof    = 0
        total_halfleave = 0
        total_working_hours = 0
        emp_atten_detail_list = []
        for dates in month_list:
            datem = datetime.strptime(dates, "%d/%m/%Y").strftime("%Y-%m-%d")
            
            emp_atten_detail = {}
              
            # if datetime.strptime(dates, "%d/%m/%Y") <= datetime.today():         
            attendance_detail = SpUserAttendance.objects.filter(user_id=employee.id ,attendance_date_time__contains = datem)
            if attendance_detail:
                arrival_time = attendance_detail[0].start_time
                start = arrival_time.split(':')
                checkin_time = start[0]+':'+start[1]
                       
                if (len(attendance_detail) > 1):
                    today = date.today()
                    if arrival_time:
                        end    = attendance_detail[1].end_time
                        end = end.split(':')
                        dept_time = end[0]+':'+end[1]
                        
                        now = str(today)
                        start_datetime = now + ' '+arrival_time
                        if attendance_detail[1].end_time is None:
                            end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            end_datetime = now + ' '+attendance_detail[1].end_time
                        start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
                        end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
                        time_delta = (end_datetime - start_datetime)
                        total_seconds = time_delta.total_seconds()
                        hours = converts(total_seconds)
                        
                        total_working_hours+=hours
                else:
                    dept_time    = '--'  
                
                emp_atten_detail['Attendance Status'] = "P"
                emp_atten_detail['Arr Time']          = checkin_time
                emp_atten_detail['Dept Time']         = dept_time
                emp_atten_detail['Date']              = datetime.strptime(dates, "%d/%m/%Y").strftime("%d")
                total_present+=1


            elif checkEmpLeave(datem, employee.id):
                total_halfleave += checkEmpHalfLeave(datem, employee.id)
                data = checkEmpLeave(datem, employee.id)
                emp_atten_detail['Attendance Status'] = data['leave_type'] 
                total_leave += data['leave']    
                emp_atten_detail['Arr Time']          = 0
                emp_atten_detail['Dept Time']         = 0
                emp_atten_detail['Date']              = datetime.strptime(dates, "%d/%m/%Y").strftime("%d") 
                
            
                    
            elif checkHoliday(datem,employee.role_id):
                emp_atten_detail['Attendance Status'] = "HL"           

                emp_atten_detail['Arr Time']          = 0
                emp_atten_detail['Dept Time']         = 0
                emp_atten_detail['Date']              = datetime.strptime(dates, "%d/%m/%Y").strftime("%d")
                total_holiday+=1
            elif ceckWeekOfDay(datem,employee.id):
                emp_atten_detail['Attendance Status'] = "WO"           

                emp_atten_detail['Arr Time']          = 0
                emp_atten_detail['Dept Time']         = 0
                emp_atten_detail['Date']              = datetime.strptime(dates, "%d/%m/%Y").strftime("%d")
                total_weekof+=1
            else:
                emp_atten_detail['Attendance Status'] = "A"           

                emp_atten_detail['Arr Time']          = 0
                emp_atten_detail['Dept Time']         = 0
                emp_atten_detail['Date']              = datetime.strptime(dates, "%d/%m/%Y").strftime("%d")
                total_absent+=1
                
            emp_atten_detail_list.append(emp_atten_detail)  
        total_leave -= (total_halfleave * 0.5)     
        employees_dict['emp_atten_detail_list'] = emp_atten_detail_list
        employees_dict['total_present']     = total_present
        employees_dict['total_absent']      = total_absent
        employees_dict['total_leave']       = total_leave
        employees_dict['total_holiday']     = total_holiday
        employees_dict['total_weekof']      = total_weekof
        employees_dict['total_halfleave']   = total_halfleave
        employees_dict['total_working_hours']   = round(total_working_hours,2)
        employees_dict['total_paid_days']       = (total_present+total_leave+total_holiday+total_weekof+total_halfleave)
        employees_list.append(employees_dict)                        
        







    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filname = 'monthly-performance-report'+str(datetime.now().strftime('%m/%Y'))+'.xlsx'
    response['Content-Disposition'] = 'attachment; filename='+filname
       
    
    workbook = Workbook()
    
    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    



    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Performance Report For Month'

    row = []
    row += ['Performance Report For Month']
    date_count = 1
    for id, dates in  enumerate(month_list):
        if id == 2:
            row += ['Sakhi Mahila Milk Producer Company Limited']
        elif id == 1:
            
            row += [calendar.month_name[month] +' '+ str(year)]
           
        else:
            row += ['']
        date_count+=1

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        
        cell.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type = "solid")
        column_letter = get_column_letter(col_num)
        cell.border = black_border
        column_dimensions = worksheet.column_dimensions[column_letter]
        
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)        
    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=date_count)

    columns = []
    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment      
        cell.border = black_border
        
   
    # Define the titles for columns
    columns = []
    columns += [ 'Department' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'EmpCode' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Name' ]  
    columns += [ '' ]
    columns += [ '' ]  
    columns += [ 'Present' ] 
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'WO' ] 
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'HL' ] 
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Absent' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'Paid Leave' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ 'HalfDays' ]
    columns += [ '' ]
    columns += [ 'Paid Days' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
    columns += [ '' ]
   

   
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = black_border
    columns = []
    row_num = 4

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment 
        cell.border = black_border

    for emp in employees_list:
        row = []
        row_num+=1  
        row += [ emp['Department'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['EmployeeId'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['EmployeeName'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['total_present'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['total_weekof'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['total_holiday'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['total_absent'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ emp['total_leave'] ]
        row += [ '' ]
        row += [ '' ]
        row += [emp['total_halfleave']]
        row += [ '' ]
        row += [ emp['total_paid_days'] ]
        row += [ '' ]
        row += [ '' ]
        row += [ '' ]
        row += [ '' ]
        row += [ '' ]
        
        
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment 
            cell.border = black_border
            
        row_num+=1   
        rows = [] 
        for id,atten in enumerate(emp['emp_atten_detail_list']):
            if id == 0:
                rows += [ 'Date' ]
            rows += [ atten['Date'] ]
            
        for col_num, column_title in enumerate(rows, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.font = Font(size=12, color='ff0000', bold=True)
            cell.alignment = centered_alignment 
            cell.border = black_border
            
        row_num+=1
        rows = []
        for id,atten in enumerate(emp['emp_atten_detail_list']):
            if id == 0:
                rows += [ 'Arr Time' ]
            rows += [ atten['Arr Time'] ]
            
        for col_num, column_title in enumerate(rows, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment 
            cell.border = black_border
 
        row_num+=1
        rows = []
        for id,atten in enumerate(emp['emp_atten_detail_list']):
            if id == 0:
                rows += [ 'Dept Time' ]
            rows += [ atten['Dept Time'] ]
            
        for col_num, column_title in enumerate(rows, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment    
            cell.border = black_border
            
        row_num+=1
        rows = []
        for id,atten in enumerate(emp['emp_atten_detail_list']):
            if id == 0:
                rows += [ 'Status' ]
            rows += [ atten['Attendance Status'] ]
            
        for col_num, column_title in enumerate(rows, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            if column_title == 'A':
                cell.font = Font(size=12, color='ff0000', bold=True)  
            elif column_title == 'P':
                cell.font = Font(size=12,color='00FF00', bold=True)   
            else:
                cell.font = Font(size=12,color='000000', bold=True)      
            cell.alignment = centered_alignment    
            cell.border = black_border                  
 
        columns = []
        row_num += 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.border = black_border
            
    row = []
    row += ['']
    date_count = 1
    for id, dates in  enumerate(month_list):
        if id == 1:
            row += ['Sakhi Mahila Milk Producer Company Limited']
            
        else:
            row += ['']
        date_count+=1

    row_num += 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = black_border
        
        cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type = "solid")
        column_letter = get_column_letter(col_num)
        
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 25
    worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=date_count)

      
    workbook.save(response)

    return response
    

def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
      
    return "%d:%02d" % (hour, minutes)

#done
def checkEmpLeave(date, employee_id):
    # if SpUserLeaves.objects.filter(leave_from_date__day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%d'), leave_from_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),leave_from_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),leave_status = 3 ,user_id = employee_id).exists():
    #     employees = SpUserLeaves.objects.filter(leave_from_date__day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%d'), leave_from_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),leave_from_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),leave_status = 3 ,user_id = employee_id)
    if SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=employee_id ).exists():
        employees = SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=employee_id)    
        leave = 0
        leave_details = {}
        for employee in employees:
            
            start_date = employee.leave_from_date
            start_date = start_date.strftime('%Y-%m-%d')
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d')

            end_date   = employee.leave_to_date
            end_date = end_date.strftime('%Y-%m-%d')
            end_date = datetime.strptime(str(end_date), '%Y-%m-%d')
            current_date = datetime.strptime(str(date), '%Y-%m-%d')
            
            leave_type = getModelColumnById(SpLeaveTypes,employee.leave_type_id, 'alias')
            
            leave_details['leave_type'] = leave_type


            delta = end_date - start_date
                       
            if current_date>=start_date:
                if delta.days >= 0:
                    # leave += delta.days + 1
                    leave += 1
                    leave_details['leave'] = leave 
                 
        if leave > 0:
            return leave_details
        else:
            return False 
    else:
        return False
  


    
def checkEmpHalfLeave(date, employee_id):
    if SpUserLeaves.objects.filter(leave_from_date__day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%d'), leave_from_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),leave_from_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),leave_status = 3 ,user_id = employee_id).exists():
        employees = SpUserLeaves.objects.filter(leave_from_date__day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%d'), leave_from_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),leave_from_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),leave_status = 3 ,user_id = employee_id)
        halfleave = 0
        for employee in employees:
            if employee.is_first_half_day != 0 or employee.is_last_half_day != 0:
                if employee.is_first_half_day != 0 and employee.is_last_half_day != 0:
                    halfleave += employee.is_first_half_day +  employee.is_last_half_day
                elif employee.is_first_half_day !=0 :
                    halfleave += employee.is_first_half_day
                else:
                    halfleave += employee.is_last_half_day
            else:
                return False
        if halfleave > 0:
            return halfleave
        else:
            return False    
    else:
        return False 
                      
def checkHoliday(date,role_id):
    if SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).exists():
        holiday_id = SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).values_list('id',flat=True)
        holiday_ids = SpRoleEntityMapping.objects.filter(role_id = role_id,entity_type = 'holiday',entity_id__in = holiday_id).values_list('entity_id',flat=True)
        count = 0
        for holiday_id in holiday_ids:
            last_date = getModelColumnById(SpHolidays,holiday_id,'end_date')
            last_date = last_date.strftime('%Y-%m-%d')
            last_date = datetime.strptime(str(last_date), '%Y-%m-%d')
            from_date = datetime.strptime(str(date), '%Y-%m-%d')#.strftime('%Y-%m-%d')
            delta = last_date - from_date
            start_date = getModelColumnById(SpHolidays,holiday_id,'start_date')
            start_date = start_date.strftime('%Y-%m-%d')
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d')
            current_date = datetime.strptime(str(date), '%Y-%m-%d')
            if current_date>=start_date:
                if delta.days >= 0:
                    count+=1
        if count > 0:
            return 'HL'
        else:
            return False 
    else:
        return False 

def converts(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    min_to_hr = int(minutes) / 60
    seconds %= 60
    hours =  hour + min_to_hr
    return hours


def ceckWeekOfDay(date,user_id):
    # print(date)
    user_week_of_day = getModelColumnByColumnId(SpBasicDetails,'user_id',user_id,'week_of_day')
    if user_week_of_day:
        user_week_of_day = user_week_of_day.split(',')
        week_day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%A')
        if week_day in user_week_of_day:
            return True
        else:
            return False
