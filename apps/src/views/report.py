import sys
import os
import openpyxl
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password,check_password
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from rest_framework.response import Response
from django.db.models import Q
from utils import *
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.forms.models import model_to_dict
import time
import math 
from datetime import datetime, date
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage

# Create your views here.

@login_required
def attendanceReport(request):
    
    context = {}
    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(attendance_date_time) = CURDATE()
     group by user_id ''')
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')
    users = []
    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id])

            temp['name'] = start_attendance[0].name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['latitude'] = start_attendance[0].latitude
            temp['longitude'] = start_attendance[0].longitude
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
            else:
                temp['end_time'] = None

            now = datetime.now().strftime('%Y-%m-%d')
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()

            hours = math.floor(total_seconds / 3600)
            mins = math.floor((total_seconds - (hours * 3600)) / 60)

            temp['working_hours'] = str(hours)+'.'+str(mins)

            users.append(temp)

    context['users'] = users
    template = 'reports/attendance-report.html'
    return render(request, template, context)

@login_required
def ajaxAttendanceReport(request):
    if 'attendance_date' in request.GET and request.GET['attendance_date'] != "" :
            today                   = request.GET['attendance_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
    
    context = {}

    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(attendance_date_time) = %s
     group by user_id ''',[today])

    users = []
    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            print(user.user_id)
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id,today])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id,today])

            temp['name'] = start_attendance[0].name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['latitude'] = start_attendance[0].latitude
            temp['longitude'] = start_attendance[0].longitude
            temp['img'] = start_attendance[0].attendance_img
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
            else:
                temp['end_time'] = None

            now = today
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()
            hours = (total_seconds/60)/60

            temp['working_hours'] = hours

        
            users.append(temp)


    context['users'] = users
    template = 'reports/ajax-attendance-report.html'
    return render(request, template, context)
    
    


@login_required
def attendanceReports(request):
    
    context = {}
    today_attendance_users = SpUserAttendance.objects.raw(''' SELECT id,user_id FROM sp_user_attendance WHERE date(attendance_date_time) = CURDATE()
     group by user_id ''')
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')
    users = []
    today = date.today()
    if len(today_attendance_users):
        for user in today_attendance_users:
            temp = {}
            start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
            sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.role_name,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id LIMIT 1 ''',[user.user_id])
            end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
            sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
            dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
            FROM sp_user_attendance 
            left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
            left join sp_users on sp_users.id = sp_user_attendance.user_id 
            WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) = CURDATE() 
            order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id])

            temp['date']=datetime.now().strftime('%d/%m/%Y')
            temp['id'] = user.user_id
            temp['name'] = start_attendance[0].name
            temp['role_name'] = start_attendance[0].role_name
            temp['emp_sap_id'] = start_attendance[0].emp_sap_id
            temp['profile_image'] = start_attendance[0].profile_image
            temp['dis_sap_id'] = start_attendance[0].dis_sap_id
            temp['store_name'] = start_attendance[0].store_name
            temp['store_image'] = start_attendance[0].store_image
            temp['start_time'] = start_attendance[0].start_time
            temp['start_latitude'] = start_attendance[0].latitude
            temp['start_longitude'] = start_attendance[0].longitude
            
            if end_attendance:
                temp['end_time'] = end_attendance[0].end_time
                temp['end_latitude'] = end_attendance[0].latitude
                temp['end_longitude'] = end_attendance[0].longitude
                # temp['attendance_type'] = end_attendance[0].attendance_type
            else:
                temp['end_time'] = None

            now = str(today)
            start_datetime = now + ' '+start_attendance[0].start_time
            if temp['end_time'] is None:
                end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                end_datetime = now + ' '+end_attendance[0].end_time
            
            start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
            end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
            time_delta = (end_datetime - start_datetime)
            total_seconds = time_delta.total_seconds()
            hours = convert(total_seconds)
            temp['working_hours'] = str(hours)

            distance_travelled  = SpUserTracking.objects.filter(user_id=start_attendance[0].user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
            if distance_travelled['distance_travelled__sum'] is not None:
                temp['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                temp['distance_travelled'] = 0.00

            temp['added_employee_count'] = SpUsers.objects.filter(created_by=start_attendance[0].user_id, created_at__icontains=today).count()
            temp['tagged_employee_count'] = 0
            
            users.append(temp)
            
    roles                       =  SpRoles.objects.filter(status=1).all()
    organization                =  SpOrganizations.objects.all()
    context['roles']            = roles
    context['employee']         = SpUsers.objects.filter(user_type=1,status=1).exclude(id=1)
    context['organization']     = organization
    context['users']            = users
    context['page_title']       = "Monthly Attendance Report"
    template = 'reports/attendance-report/attendance-report.html'
    return render(request, template, context)

  
@login_required
def ajaxAttendanceReports(request):
    if 'attendance_start_date' in request.GET and request.GET['attendance_start_date'] != "" :
        attendance_start_date                   = datetime.strptime(str(request.GET['attendance_start_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
        attendance_end_date                   = datetime.strptime(str(request.GET['attendance_end_date']), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        attendance_start_date=attendance_end_date = date.today()

    condition=""
    if 'role_id' in request.GET and request.GET['role_id'] != ""  or 'organization_id' in request.GET and request.GET['organization_id'] != "" or 'employee_id' in request.GET and request.GET['employee_id'] != "":
        if 'role_id' in request.GET and request.GET['role_id'] != "" :
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.role_id="+request.GET['role_id']+ " and"
        
        if 'employee_id' in request.GET and request.GET['employee_id'] != "" :
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.id="+request.GET['employee_id']+ " and"
        
        if 'organization_id' in request.GET and request.GET['organization_id'] != "" :
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.organization_id="+request.GET['organization_id']+ " and"
        
        if 'role_id' in request.GET and request.GET['role_id'] != "" and 'employee_id' in request.GET and request.GET['employee_id'] != "":
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.id="+request.GET['employee_id']+ " and " + "sp_users.role_id="+request.GET['role_id']+ " and"
        
        if 'organization_id' in request.GET and request.GET['organization_id'] != "" and 'employee_id' in request.GET and request.GET['employee_id'] != "":
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.id="+request.GET['employee_id']+ " and " + "sp_users.organization_id="+request.GET['organization_id']+ " and"

        
        
        if 'organization_id' in request.GET and request.GET['organization_id'] != "" and 'role_id' in request.GET and request.GET['role_id'] != "":
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.organization_id="+request.GET['organization_id']+ " and " + "sp_users.role_id="+request.GET['role_id']+ " and"
    
        if 'organization_id' in request.GET and request.GET['organization_id'] != "" and 'employee_id' in request.GET and request.GET['employee_id'] != "" and 'role_id' in request.GET and request.GET['role_id'] != "":
            condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.id="+request.GET['employee_id']+ " and " + "sp_users.role_id="+request.GET['role_id']+ " and " + "sp_users.organization_id="+request.GET['organization_id']+ " and"
    
    else:
        condition=" where "
        
    context = {}

    start_date = date(int(datetime.strptime(str(request.GET['attendance_start_date']), '%d/%m/%Y').strftime('%Y')),int(datetime.strptime(str(request.GET['attendance_start_date']), '%d/%m/%Y').strftime('%m')),int(datetime.strptime(str(request.GET['attendance_start_date']), '%d/%m/%Y').strftime('%d')))
    end_date = date(int(datetime.strptime(str(request.GET['attendance_end_date']), '%d/%m/%Y').strftime('%Y')),int(datetime.strptime(str(request.GET['attendance_end_date']), '%d/%m/%Y').strftime('%m')),int(datetime.strptime(str(request.GET['attendance_end_date']), '%d/%m/%Y').strftime('%d')))
    
    users=[]
    while start_date <= end_date:
        today = start_date
        today_attendance_users = SpUserAttendance.objects.raw(''' SELECT sp_user_attendance.id,sp_user_attendance.user_id FROM sp_user_attendance {condition} date(attendance_date_time) = %s
            group by user_id '''.format(condition=condition),[today])
        
        if len(today_attendance_users):
            for user in today_attendance_users:
                temp = {}
                start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
                sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.role_name,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s  
                order by sp_user_attendance.id LIMIT 1 ''',[user.user_id,today])
                end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
                sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
                order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id,today])
               
                temp['date']=start_date.strftime('%d/%m/%Y')
                temp['id'] = user.user_id
                temp['name'] = start_attendance[0].name
                temp['role_name'] = start_attendance[0].role_name
                temp['emp_sap_id'] = start_attendance[0].emp_sap_id
                temp['profile_image'] = start_attendance[0].profile_image
                temp['dis_sap_id'] = start_attendance[0].dis_sap_id
                temp['store_name'] = start_attendance[0].store_name
                temp['store_image'] = start_attendance[0].store_image
                temp['start_time'] = start_attendance[0].start_time
                temp['start_latitude'] = start_attendance[0].latitude
                temp['start_longitude'] = start_attendance[0].longitude
                temp['img'] = start_attendance[0].attendance_img
                if end_attendance:
                    temp['end_time'] = end_attendance[0].end_time
                    temp['end_latitude'] = end_attendance[0].latitude
                    temp['end_longitude'] = end_attendance[0].longitude
                    # temp['attendance_type'] = end_attendance[0].attendance_type
                    temp['Eod'] = end_attendance[0].Eod
                else:
                    temp['end_time'] = None
                    temp['Eod'] = None

                now = str(today)
                start_datetime = now + ' '+start_attendance[0].start_time
                if temp['end_time'] is None:
                    end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    end_datetime = now + ' '+end_attendance[0].end_time
                
                start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
                end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
                time_delta = (end_datetime - start_datetime)
                total_seconds = time_delta.total_seconds()
                hours = convert(total_seconds)
                temp['working_hours'] = str(hours)

                distance_travelled  = SpUserTracking.objects.filter(user_id=start_attendance[0].user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
                if distance_travelled['distance_travelled__sum'] is not None:
                    temp['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    temp['distance_travelled'] = 0.00

                temp['added_employee_count'] = SpUsers.objects.filter(created_by=start_attendance[0].user_id, created_at__icontains=today).count()
                temp['tagged_employee_count'] = 0
                
                users.append(temp)
                
        
        start_date    = start_date + timedelta(days=1)
        
    context['users'] = users
    template = 'reports/attendance-report/ajax-attendance-report.html'
    return render(request, template, context)  

 
def convert(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
      
    return "%d:%02d" % (hour, minutes)
    # return "%d:%02d:%02d" % (hour, minutes, seconds)

@login_required
def exportAttendanceReport(request, columns,attendance_start_date, attendance_end_date,role_id):
    column_list = columns.split (",")
    

    attendance_start_date                   = datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%Y-%m-%d')
    attendance_end_date                   = datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%Y-%m-%d')
    

    condition=""
    if role_id != "0" :
        condition="left join sp_users on sp_users.id = sp_user_attendance.user_id where sp_users.role_id="+role_id+ " and"
    else:
        condition=" where "
    context = {}
    
    start_date = date(int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%Y')),int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%m')),int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%d')))
    end_date = date(int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%Y')),int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%m')),int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%d')))
    
    users=[]
    while start_date <= end_date:
        today = start_date
        today_attendance_users = SpUserAttendance.objects.raw(''' SELECT sp_user_attendance.id,sp_user_attendance.user_id FROM sp_user_attendance {condition} date(attendance_date_time) = %s
            group by user_id '''.format(condition=condition),[today])
        
        if len(today_attendance_users):
            for user in today_attendance_users:
                temp = {}
                start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
                sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.role_name,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s  
                order by sp_user_attendance.id LIMIT 1 ''',[user.user_id,today])
                end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
                sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
                order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id,today])
               
                temp['date']=start_date.strftime('%d/%m/%Y')
                temp['id'] = user.user_id
                temp['name'] = start_attendance[0].name
                temp['role_name'] = start_attendance[0].role_name
                temp['emp_sap_id'] = start_attendance[0].emp_sap_id
                temp['profile_image'] = start_attendance[0].profile_image
                temp['dis_sap_id'] = start_attendance[0].dis_sap_id
                temp['store_name'] = start_attendance[0].store_name
                temp['store_image'] = start_attendance[0].store_image
                temp['start_time'] = start_attendance[0].start_time
                temp['start_latitude'] = start_attendance[0].latitude
                temp['start_longitude'] = start_attendance[0].longitude
               
                if end_attendance:
                    temp['end_time'] = end_attendance[0].end_time
                    temp['end_latitude'] = end_attendance[0].latitude
                    temp['end_longitude'] = end_attendance[0].longitude
                    #temp['attendance_type'] = end_attendance[0].attendance_type
                else:
                    temp['end_time'] = None

                now = str(today)
                start_datetime = now + ' '+start_attendance[0].start_time
                if temp['end_time'] is None:
                    end_datetime = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    end_datetime = now + ' '+end_attendance[0].end_time
                
                start_datetime = datetime.strptime(start_datetime, '%Y-%m-%d %H:%M:%S')
                end_datetime = datetime.strptime(end_datetime, '%Y-%m-%d %H:%M:%S')
                time_delta = (end_datetime - start_datetime)
                total_seconds = time_delta.total_seconds()
                hours = convert(total_seconds)
                temp['working_hours'] = str(hours)

                distance_travelled  = SpUserTracking.objects.filter(user_id=start_attendance[0].user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
                if distance_travelled['distance_travelled__sum'] is not None:
                    temp['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    temp['distance_travelled'] = 0.00

                temp['added_employee_count'] = SpUsers.objects.filter(created_by=start_attendance[0].user_id, created_at__icontains=today).count()
                temp['tagged_employee_count'] = 0
                
                users.append(temp)
                
        
        start_date    = start_date + timedelta(days=1)
        
        print(users)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=attendance-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='center')
        border_bottom = Border(
            bottom=Side(border_style='medium', color='FF000000'),
        )
        wrapped_alignment = Alignment(
            vertical='top',
            wrap_text=True
        )

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Attendance Report'
        
        # Define the titles for columns
        columns = []

        
        columns += [ 'Date' ]
            
        
        columns += [ 'Employee Name' ]
            
        
        columns += [ 'Employee Code' ]

        
        columns += [ 'Day Start Time' ]
        
        
        columns += [ 'Day End Time' ] 

        
        columns += [ 'Total working hours' ]
            
        
        # columns += [ 'Total Covered Distance (in Km)' ]
                
        
        # columns += [ 'No. of Tagged/Add User' ]

        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 20

        # Iterate through all movies
        for user in users:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            
            row += [ user['date'] ]
                
            
            row += [ user['name'] +"("+user['role_name']+")"  ]
                        
            
            row += [ user['emp_sap_id'] ]

            
            row += [ user['start_time'] ]
            
           
            if user['end_time']:
                row += [ user['end_time'] ]
            else:
                row += [ '' ]        

            
            row += [ user['working_hours'] ]   
            
            
            # row += [ user['distance_travelled'] ]   
            
            
            # row += [ str(user['added_employee_count'])+ " / "+  str(user['tagged_employee_count']) ]   
            
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment

        workbook.save(response)

    return response
    
@login_required
def userGeoAttendance(request):
    context = {}
  
    try:
        user_coordinates = SpUsers.objects.get(id=request.GET['id'])
    except SpUserAttendanceLocations.DoesNotExist:
        user_coordinates = None
    if user_coordinates is not None:
        user_coordinates.latitude=request.GET['lat']
        user_coordinates.longitude=request.GET['long']
        user_coordinates.type=request.GET['type']
        user_coordinates.date_time=request.GET['date_time']
    context['user_coordinates'] = user_coordinates
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'reports/attendance-report/user-geo-attendance-tracking.html'

    return render(request, template,context)


@login_required
def leaveReport(request):
    today = date.today()
    
    if request.user.role_id == 0:
        leaveReport = SpUserLeaves.objects.all().order_by('-id')

    else:
        leaveReport = SpUserLeaves.objects.raw('''SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
    left join sp_users on sp_users.id = sp_user_leaves.user_id 
    where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' order by id desc ''',[request.user.id])
    

    user_type = SpPermissionWorkflowRoles.objects.filter(sub_module_id=38,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context = {}

    user_ids = SpUserLeaves.objects.all().distinct().values_list('user_id',flat=True)
    users = SpUsers.objects.filter(id__in=user_ids).values('id','first_name','middle_name','last_name','emp_sap_id')

    context['leaveReport'] = leaveReport
    context['role_id'] = request.user.role_id
    context['level_id'] = level_id
    context['today_date'] = today.strftime("%d/%m/%Y")
    context['page_title'] = "Manage Leaves"
    context['users'] = users

    template = 'reports/attendance-report/leave-report.html'
    return render(request, template, context)


# ajax order list
@login_required
def ajaxLeaveReportLists(request):
    context = {}
    today = date.today()
    user_id = request.GET['user_id']
    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            context['leave_status'] = leave_status
            leaveReport = leaveReport.filter(leave_status=leave_status).order_by('-id')

        # if 'leave_from_date' in request.GET and request.GET['leave_from_date'] and request.GET['leave_to_date']:
        #     leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
        #     leave_to_date = datetime.strptime(request.GET['leave_to_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
            
        #     leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        # if request.GET['leave_from_date']:
        #     if request.GET['leave_to_date'] in today.strftime("%d/%m/%Y"):
        #         leave_from_date = datetime.strptime(request.GET['leave_from_date'], '%d/%m/%Y').strftime('%Y-%m-%d %H:%M:%S')
                
        #         leaveReport = leaveReport.filter(leave_from_date__range=(leave_from_date, leave_to_date))
        if user_id:
                leaveReport = leaveReport.filter(user_id=user_id).order_by('-id')
    else:
        condition = ''
        
        if 'leave_status' in request.GET and request.GET['leave_status'] != "":
            leave_status = request.GET['leave_status']
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status
        if user_id:
            condition += ' and sp_user_leaves.user_id = "%s"' % user_id
            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    

    user_type = SpRoleWorkflowPermissions.objects.filter(sub_module_id=8, permission_slug='add',
                                                             workflow_level_role_id=request.user.role_id).exclude(
            role_id=request.user.role_id).values('level_id').order_by('-id').first()
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context['leaveReport'] = leaveReport
    context['level_id'] = level_id

    context['role_id'] = request.user.role_id
    template = 'reports/attendance-report/ajax-leave-report-lists.html'
    return render(request, template, context)


@login_required
def leaveStatusDetails(request):
    leave_id = request.GET.get('leave_id')
    initiate_leave_details = SpUserLeaves.objects.get(id=leave_id)
    leave_details = SpApprovalStatus.objects.filter(row_id=leave_id, model_name='SpUserLeaves', status=1).values(
        'final_status_user_id').distinct().values('final_status_user_name', 'final_update_date_time', 'level_id')

    context = {}
    context['initiate_leave_details'] = initiate_leave_details
    context['leave_details'] = leave_details
    template = 'reports/attendance-report/leave-status-details.html'

    return render(request, template, context)


@login_required
def leaveExportToXlsx(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)

        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=leave-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Leave-reports'

    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += ['Name']

    if 'leave_from_date' in column_list:
        columns += ['Leave Apply From Date']

    if 'leave_to_date' in column_list:
        columns += ['Leave Apply To Date']

    if 'status' in column_list:
        columns += ['Status']

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    for results in leaveReport:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        if 'user_name' in column_list:
            row += [results.user_name]

        if 'leave_from_date' in column_list:
            if None in [results.leave_from_date]:
                leave_from_date = ['']
            else:
                leave_from_date = [results.leave_from_date]
            row += leave_from_date

        if 'leave_to_date' in column_list:
            if None in [results.leave_to_date]:
                leave_to_date = ['']
            else:
                leave_to_date = [results.leave_to_date]
            row += leave_to_date

        if 'status' in column_list:
            if results.leave_status == 1:
                status = 'Pending'
                row += [status]
            elif results.leave_status == 2:
                status = 'Forwarded'
                row += [status]
            elif results.leave_status == 3:
                status = 'Approved'
                row += [status]
            elif results.leave_status == 4:
                status = 'Declined'
                row += [status]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
    workbook.save(response)

    return response


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


# Automaticly downloads to PDF file
@login_required
def leaveExportToPDF(request, columns, userId,leave_status):
    column_list = columns.split(",")

    leaveReport = SpUserLeaves.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        leaveReport = SpUserLeaves.objects.raw("""SELECT * FROM sp_user_leaves WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if leave_status != "" and leave_status != "0":
            condition += ' and sp_user_leaves.leave_status = "%s"' % leave_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_leaves.user_id = "%s"' % userId

            

        query = """SELECT sp_user_leaves.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_leaves left join sp_approval_status on sp_approval_status.row_id = sp_user_leaves.id
        left join sp_users on sp_users.id = sp_user_leaves.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserLeaves' %s order by id desc """ % (request.user.id,condition)

        leaveReport = SpUserLeaves.objects.raw(query)
    
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('reports/attendance-report/leave_pdf_template.html',
                        {'leaveReport': leaveReport, 'url': baseurl, 'columns': column_list,
                         'columns_length': len(column_list)})
    
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'leave-report.pdf'
    content = "attachment; filename=%s" % filename
    response['Content-Disposition'] = content
    return response


@login_required
def editLeaveStatus(request):
    id = request.GET.get('id')
    context = {}
    context['leaveData'] = SpUserLeaves.objects.get(id=id)
    template = 'reports/attendance-report/edit-leave-status.html'
    return render(request, template, context)

@login_required
def updateLeaveRemark(request): 
    context = {}
    context['level_id']     = request.GET.get('level_id')
    context['leave_status'] = request.GET.get('leave_status')
    template = 'reports/attendance-report/update-leave-remark.html'
    return render(request, template, context)

#update order status
@login_required
def updateLeaveStatus(request):
    response = {}
    leave_id        = request.POST.getlist('leave_id[]')
    level_id        = request.POST['level_id']
    leave_status    = request.POST['leave_status']
    if request.user.role_id == 0:
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id)
            if approvals_request:
                for approval in approvals_request:
                    approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                    approval_data.level_id                  = leave_status
                    if leave_status == '2':
                        approval_data.level                    = 'Forward'
                    elif leave_status == '3':
                        approval_data.level                    = 'Approve'
                    elif leave_status == '4':
                        approval_data.level                    = 'Declined'         
                    approval_data.status                    = 1
                    approval_data.final_status_user_id      = request.user.id
                    approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    approval_data.save()

                user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
                if user_level_approval_count == 0:
                    leave                   = SpUserLeaves.objects.get(id=leave)   
                    leave.leave_status      = leave_status
                    if request.POST['remark']:
                        leave.remark      = request.POST['remark']
                    leave.save()
            else:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                leave.save()
                today   = date.today()
                 
    
    else:    
        for leave in leave_id:
            approvals_request = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', role_id=request.user.role_id, level_id=level_id)
            for approval in approvals_request:
                approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
                approval_data.status                    = 1
                approval_data.final_status_user_id      = request.user.id
                approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                approval_data.save()

            user_level_approval_count = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', level_id=level_id, status=0).count()
            if user_level_approval_count == 0:
                leave                   = SpUserLeaves.objects.get(id=leave)   
                leave.leave_status      = leave_status
                if request.POST['remark']:
                    leave.remark      = request.POST['remark']
                leave.save()   

    
    if leave_status == '2':
        for leave in leave_id:
            approvals_requests = SpApprovalStatus.objects.filter(row_id=leave, model_name='SpUserLeaves', status=0)
            if approvals_requests:
                for approval in approvals_requests:
                    notification                        = SpUserNotifications()
                    notification.row_id                 = approval.row_id
                    notification.user_id                = approval.user_id
                    notification.model_name             = 'SpUserLeaves'
                    notification.notification           = 'leave '+approval.level+' Request has been sent.'
                    notification.is_read                = 0
                    notification.created_by_user_id     = request.user.id
                    notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    notification.save()

    if leave_status == '2':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been forwarded'
            activity    = 'Leave Request has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)
            
            message_title = "Leave request forwarded"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been forwarded  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------# 
    elif leave_status == '3':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been approved'
            activity    = 'Leave Request has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            #update user leave count
            user_basic_details              = SpBasicDetails.objects.get(user_id=user_id)
            leave_count                     = int(user_basic_details.leave_count)-1   
            user_basic_details.leave_count  = leave_count
            user_basic_details.save()

            message_title = "Leave request approved"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been approved  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#    

    elif leave_status == '4':
        for leave in leave_id:
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = 'Leave Request has been declined'
            activity    = 'Leave Request has been declined by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
            if request.POST['remark']:
                activity += '. '+request.POST['remark']
            saveActivity('User Management', 'Leave', heading, activity, request.user.id, user_name, 'declined.svg', '1', 'web.png')

            #-----------------------------notify android block-------------------------------#
            user_id = getModelColumnById(SpUserLeaves,leave,'user_id')
            user_role = getModelColumnById(SpUsers,user_id,'role_name')
            userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
            employee_name = getUserName(user_id)

            message_title = "Leave request declined"
            message_body = "A Leave request("+employee_name+" - "+user_role+") has been declined  by "+user_name
            if request.POST['remark']:
                message_body += '. '+request.POST['remark']
            notification_image = ""

            if userFirebaseToken is not None and userFirebaseToken != "" :
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['image'] = notification_image
                send_android_notification(message_title,message_body,data_message,registration_ids)
                #-----------------------------notify android block-------------------------------#

            #-----------------------------save notification block----------------------------#
            saveNotification(leave,'SpUserLeaves','User Management','Leave request declined',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
            #-----------------------------save notification block----------------------------#
            employee_role = getModelColumnById(SpUsers, user_id, 'role_id')
            if employee_role == 5:
                user_area_allocation = SpUserAreaAllocations.objects.filter(user_id=user_id).values('town_id')
                user_list = []   
                for area_allocation in user_area_allocation:
                    operational_user_list  = SpUserAreaAllocations.objects.raw(''' SELECT sp_user_area_allocations.id, sp_user_area_allocations.town_name, sp_users.id as user_id, sp_users.first_name, sp_users.middle_name, sp_users.role_id, sp_users.user_type, sp_users.reporting_to_id, sp_users.reporting_to_name, sp_users.primary_contact_number, sp_users.store_name, sp_users.last_name, sp_users.is_tagged FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id WHERE sp_user_area_allocations.town_id=%s and sp_users.role_id=%s and sp_users.status=%s order by sp_users.first_name asc ''', [area_allocation['town_id'], 4, 1])
                    if operational_user_list: 
                        for operational_user in operational_user_list:
                            user_list.append(operational_user.user_id)
                            
                if len(user_list) > 0:
                    for tse in user_list:
                        userFirebaseToken = getModelColumnById(SpUsers,tse,'firebase_token')
                        employee_name = getUserName(tse)
                        if userFirebaseToken is not None and userFirebaseToken != "" :
                            registration_ids = []
                            registration_ids.append(userFirebaseToken)
                            data_message = {}
                            data_message['id'] = 1
                            data_message['status'] = 'notification'
                            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                            data_message['image'] = notification_image
                            send_android_notification(message_title,message_body,data_message,registration_ids)
                        #-----------------------------save notification block----------------------------#
                        saveNotification(leave,'SpUserLeaves','User Management','Leave request approved',message_title,message_body,notification_image,request.user.id,user_name,tse,employee_name,'profile.png',2,'app.png',1,1)
                        #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Leave status has been updated successfully."
    return JsonResponse(response)
    
@login_required
def employeeReport(request):
    
    context = {}
    today = date.today()
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today.strftime("%Y-%m-%d")).count()
        user.tagged_employee_count = 0
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    context['users'] = users
    context['page_title'] = "Employee Tagging Report"
    template = 'reports/attendance-report/employee-report.html'
    return render(request, template, context)


@login_required
def ajaxEmployeeReport(request):
    if 'search_date' in request.GET and request.GET['search_date'] != "" :
        today                   = request.GET['search_date']
        today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()
        today                   = today.strftime("%Y-%m-%d")
        
    context = {}
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today).count()
        user.tagged_employee_count = 0
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    context['users'] = users
    template = 'reports/attendance-report/ajax-employee-report.html'
    return render(request, template, context)

@login_required
def exportEmployeeReport(request, columns, search_date):
    column_list = columns.split (",")
    if search_date != "" :
        today                   = search_date
    else:
        today                   = date.today()
        today                   = today.strftime("%Y-%m-%d")
        
    context = {}
    users = SpUsers.objects.raw(''' SELECT id, first_name, middle_name, last_name, role_name, is_tagged, tagged_by, tagged_date, created_by FROM sp_users WHERE id!=%s and user_type=%s and (role_id=%s or role_id=%s)  ''', [1, 1, 4, 5])
    for user in users:
        name = user.first_name
        if user.middle_name:
            name += ' '+user.middle_name
        if user.last_name:
            name += ' '+user.last_name    
        user.name = name
        user.added_employee_count = SpUsers.objects.filter(created_by=user.id, created_at__icontains=today).count()
        user.tagged_employee_count = 0
    context['current_data'] = datetime.now().strftime('%d/%m/%Y')

    users = users
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employee-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employee Report'
    
    # Define the titles for columns
    columns = []

    if 'employee_name' in column_list:
        columns += [ 'Employee Name' ]

    if 'added_employee' in column_list:
        columns += [ 'Added Employee' ]
    
    if 'tagged_employee' in column_list:
        columns += [ 'Tagged Employee' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        if 'employee_name' in column_list:
            row += [ user.name ]

        if 'added_employee' in column_list:
            row += [ user.added_employee_count ]
        
        if 'tagged_employee' in column_list:
            row += [ user.tagged_employee_count ]  
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response        


@login_required
def incentiveReport(request):
    context = {}
    import datetime
    today = datetime.date.today()
    first = today.replace(day=1)
    lastMonth = first - datetime.timedelta(days=1)
    date = lastMonth.strftime('%Y-%m')
    superstockists = SpUserIncentive.objects.filter(payment_cycle=2,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=2,created_at__icontains=date).values_list('id',flat=True)
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
        
        
    context['current_date'] = datetime.date.today().strftime('%m/%Y')
    context['superstockists'] = superstockists
    context['superstockists_distinct_slab'] = superstockists_distinct_slab
    context['page_title'] = "Superstockist Incentive Report"
    template = 'reports/attendance-report/incentive-report.html'
    return render(request, template, context)


@login_required
def ajaxIncentiveReport(request):
    import datetime
    if 'search_date' in request.GET and request.GET['search_date'] != "" :
        today                   = request.GET['search_date']
        date                   = datetime.datetime.strptime(str(today), '%m/%Y').strftime('%Y-%m')
    else:
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        
    context = {}
    if 'fortnight' in request.GET and request.GET['fortnight'] == "" :
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        date = lastMonth.strftime('%Y-%m')
    
    if 'fortnight' in request.GET and request.GET['fortnight'] != "" :
        if int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%m')) == 2:
            mid_date=14
        else:
            mid_date=15
        
        if request.GET['fortnight'] == "1":
            start_date = str(date)+"-01"
            end_date = str(date)+"-"+str(mid_date)
            date =  str(date)+"-"+str(mid_date)
            
        elif request.GET['fortnight'] == "2":
            start_date = str(date)+"-"+str(mid_date+1)
            end_date = str(date)+"-"+str(mid_date)
            last_date = len(days_in_months(int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%Y')), int(datetime.datetime.strptime(str(today), '%m/%Y').strftime('%m'))))
            date =  str(date)+"-"+str(last_date)
            
        payment_cycle = 1
    else:
        payment_cycle = 2
        
        # superstockists = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).all()
        # user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).values_list('id',flat=True)
        
    superstockists = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).values_list('id',flat=True)
    
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
        
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
        
        
    context['current_date'] = datetime.date.today().strftime('%m/%Y')
    context['superstockists'] = superstockists
    context['superstockists_distinct_slab'] = superstockists_distinct_slab
    template = 'reports/attendance-report/ajax-incentive-report.html'
    return render(request, template, context)


@login_required
def exportIncentiveReport(request, fortnight, search_date):
    
    import datetime
    if search_date != "" :
        today                   = search_date
        date                   = today
    else:
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        
    context = {}
    if fortnight == "NA" :
        today = datetime.date.today()
        first = today.replace(day=1)
        lastMonth = first - datetime.timedelta(days=1)
        date = lastMonth.strftime('%Y-%m')
    
    if fortnight != "NA" :
        if int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%m')) == 2:
            mid_date=14
        else:
            mid_date=15
        
        if fortnight == "1":
            start_date = str(date)+"-01"
            end_date = str(date)+"-"+str(mid_date)
            date =  str(date)+"-"+str(mid_date)
            
        elif fortnight == "2":
            start_date = str(date)+"-"+str(mid_date+1)
            end_date = str(date)+"-"+str(mid_date)
            last_date = len(days_in_months(int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%Y')), int(datetime.datetime.strptime(str(today), '%Y-%m').strftime('%m'))))
            date =  str(date)+"-"+str(last_date)
            
        payment_cycle = 1
    else:
        payment_cycle = 2
        
        # superstockists = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).all()
        # user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=1,created_at__gte=start_date,created_at__lte=end_date).values_list('id',flat=True)
        
    superstockists = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).all()
    user_incentive_ids = SpUserIncentive.objects.filter(payment_cycle=payment_cycle,created_at__icontains=date).values_list('id',flat=True)
    
    superstockists_distinct_slab = SpUserIncentiveDetails.objects.filter(user_incentive_id__in=user_incentive_ids).all()
        
    for superstockist in superstockists:
        ss_details = SpUsers.objects.get(id=superstockist.user_id)
        superstockist.emp_sap_id = ss_details.emp_sap_id
        superstockist.name = ss_details.first_name+" "+ss_details.middle_name+" "+ss_details.last_name
        superstockist.store_name = ss_details.store_name
        
        slab_amount_list = []
        for slab in superstockists_distinct_slab:
            key = "slab_"+str(slab.master_slab_id)
            slab_incentive = SpUserIncentiveDetails.objects.filter(user_incentive_id=superstockist.id,master_slab_id=slab.master_slab_id).first()
            if slab_incentive:
                slab_amount = slab_incentive.slab_amount
            else:
                slab_amount = 0
            slab_amount_list.append(slab_amount)
        superstockist.slab_amount_list = slab_amount_list
               
    for slab in superstockists_distinct_slab:
        product_class_id = getModelColumnById(SpSlabMasterList,slab.master_slab_id,'product_class_id')
        slab.slab_detail = getModelColumnById(SpProductClass,product_class_id,'product_class')+" ( "+ str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'more_than_quantity'))+" - "+str(getModelColumnById(SpSlabMasterList,slab.master_slab_id,'upto_quantity'))+" )"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=incentive-report.xlsx'.format(
        date=date
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Incentive Report'
    
    # Define the titles for columns
    columns = []
    
    
    columns += [ 'Sap Code SS' ]
    columns += [ 'SS Name' ]
    columns += [ 'SS Incentive' ]
    columns += [ 'Distributor Incentive' ]
    columns += [ 'Primary  TPT Amount' ]
    columns += [ 'Distributor TPT Amount' ]
    for slab in superstockists_distinct_slab:
        columns += [ slab.slab_detail ]
    columns += [ 'NET Amount' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for superstockist in superstockists:
        row_num += 1
        # Define the data for each cell in the row 
        row = []
        
        row += [ superstockist.emp_sap_id ]
        row += [ superstockist.name+" ( "+superstockist.store_name+" )" ]
        row += [ superstockist.ss_incentive ]
        row += [ superstockist.distributor_incentive ]
        row += [ superstockist.primary_transporter_amount ]
        row += [ superstockist.secondary_transporter_amount ]
        for slab_amount in superstockist.slab_amount_list:
            row += [ slab_amount ]
        row += [ superstockist.net_amount ]
        
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response        


@login_required
def getDataTotalTagged(request, user_id, check):
    page = request.GET.get('page')
    context = {}
    context['reporting_to'] = SpUsers.objects.filter(
        user_type=1).exclude(id=1).order_by('-id')
    #non operational
    count_outlet = 0
    page = request.GET.get('non_page')
    if check == '1':
        users = SpUsers.objects.all().filter(
            tagged_by=user_id, is_tagged=1).exclude(id=1).order_by('-id')
        if users:
            for user in users:
                count_outlet += 1
                try:
                    user_basic_details = SpBasicDetails.objects.get(
                        user_id=user.id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                try:
                    emp = SpUsers.objects.get(id=user_id)
                except SpBasicDetails.DoesNotExist:
                    emp = None
                user.tagged_by = emp.first_name + ' ' + emp.middle_name + ' ' + emp.last_name

                user.tagged_date = user.tagged_date.strftime('%d/%m/%Y')
                if user_basic_details.production_unit_id:
                    user.production_unit = getModelColumnById(
                        SpProductionUnit, user_basic_details.production_unit_id, 'production_unit_name')
                else:
                    user.production_unit = ''
            context['first_retailer'] = SpUsers.objects.filter(
                tagged_by=user_id, is_tagged=1).order_by('-id')[0]

    else:
        users = SpUsers.objects.all().filter(tagged_by=user_id, is_tagged=1,
                                             tagged_date__icontains=datetime.now().strftime('%Y-%m-%d')).exclude(id=1).order_by('-id')
        if users:
             for user in users:
                count_outlet += 1
                try:
                    user_basic_details = SpBasicDetails.objects.get(
                        user_id=user.id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                try:
                    emp = SpUsers.objects.get(id=user_id)
                except SpBasicDetails.DoesNotExist:
                    emp = None
                user.tagged_by = emp.first_name + ' ' + emp.middle_name + ' ' + emp.last_name

                user.tagged_date = user.tagged_date.strftime('%d/%m/%Y')
                if user_basic_details.production_unit_id:
                    user.production_unit = getModelColumnById(
                        SpProductionUnit, user_basic_details.production_unit_id, 'production_unit_name')
                else:
                    user.production_unit = ''
                context['first_retailer'] = SpUsers.objects.filter(
                    tagged_by=user_id, is_tagged=1, tagged_date__icontains=datetime.now().strftime('%Y-%m-%d')).order_by('-id')[0]
    context['count_outlet'] = count_outlet
    context['check'] = check
    context['today_date'] = datetime.now().strftime('%d/%m/%Y')
    paginator = Paginator(users, getConfigurationResult('page_limit'))
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = math.ceil(
        paginator.count/getConfigurationResult('page_limit'))

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages % paginator.count
    if(temp > 0 and getConfigurationResult('page_limit') != paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages

    first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id!=%s order by id desc LIMIT 1 ''', [1, 'correspondence', 1])

    if first_employee:
        context['first_employee'] = first_employee[0]
        first_employee_permanent_address = SpAddresses.objects.get(
            user_id=first_employee[0].id, type='permanent')
    else:
        context['first_employee'] = []
        first_employee_permanent_address = None

    context['first_employee_permanent_address'] = first_employee_permanent_address
    context['total_distributor'] = SpUsers.objects.filter().count()
    context['total_super_stockist'] = SpUsers.objects.filter(
        is_super_stockist=1).count()
    context['total_retailer'] = SpUsers.objects.filter(role_id=10).count()

    context['total_tagged_distributor'] = SpUsers.objects.filter(
        is_distributor=1, is_tagged=1).count()
    context['total_tagged_super_stockist'] = SpUsers.objects.filter(
        is_super_stockist=1, is_tagged=1).count()
    context['total_tagged_retailer'] = SpUsers.objects.filter(
        role_id=10, is_tagged=1).count()

    town_data = []
    towns = SpTowns.objects.all()
    for town in towns:

        distributors = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as distributor_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_distributor = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]
        town.distributor_count = distributors.distributor_count

        super_stockist = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as super_stockist_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_super_stockist = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]
        town.super_stockist_count = super_stockist.super_stockist_count

        retailers = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as retailers_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_retailer = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]

        town.retailer_count = retailers.retailers_count

        town_data.append(town)

    ho_roles = SpRoles.objects.filter(status=1).exclude(
        id=8).exclude(id=9).exclude(id=10)
    context['ho_roles'] = ho_roles

    context['towns'] = town_data
    context['page_title'] = "Manage Users"
    template = 'reports/tagging-report.html'
    return render(request, template, context)


def taggingReport(request):
    check = '2'
    user_id = '0'
    context = {}
    context['reporting_to'] = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    #non operational
    count_outlet = 0
    page = request.GET.get('non_page')
    if check == '2':
        users = SpUsers.objects.all().filter(is_tagged=1).exclude(id=1).order_by('-id')
        if users:
            for user in users:
                count_outlet += 1
                try:
                    user_basic_details = SpBasicDetails.objects.get(
                        user_id=user.id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                try:
                    emp                 =       SpUsers.objects.get(id=user.tagged_by)
                    user.tagged_by      =       emp.first_name + ' ' + emp.middle_name + ' ' + emp.last_name
                    user.tagged_date    =       user.tagged_date.strftime('%d/%m/%Y')
                except SpBasicDetails.DoesNotExist:
                    emp = None
                user.production_unit = ''
            context['first_retailer'] = SpUsers.objects.filter(
                is_tagged=1).order_by('-id')[0]
    context['check'] = check
    context['count_outlet'] = count_outlet
    paginator = Paginator(users, getConfigurationResult('page_limit'))
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = math.ceil(
        paginator.count/getConfigurationResult('page_limit'))

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages % paginator.count
    if(temp > 0 and getConfigurationResult('page_limit') != paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages

    first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id!=%s order by id desc LIMIT 1 ''', [1, 'correspondence', 1])

    if first_employee:
        context['first_employee'] = first_employee[0]
        first_employee_permanent_address = SpAddresses.objects.get(
            user_id=first_employee[0].id, type='permanent')
    else:
        context['first_employee'] = []
        first_employee_permanent_address = None

    context['first_employee_permanent_address'] = first_employee_permanent_address
    context['total_distributor'] = SpUsers.objects.filter().count()
    context['total_super_stockist'] = SpUsers.objects.filter(
        is_super_stockist=1).count()
    context['total_retailer'] = SpUsers.objects.filter(role_id=10).count()

    context['total_tagged_distributor'] = SpUsers.objects.filter(
        is_distributor=1, is_tagged=1).count()
    context['total_tagged_super_stockist'] = SpUsers.objects.filter(
        is_super_stockist=1, is_tagged=1).count()
    context['total_tagged_retailer'] = SpUsers.objects.filter(
        role_id=10, is_tagged=1).count()

    town_data = []
    towns = SpTowns.objects.all()
    for town in towns:

        distributors = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as distributor_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_distributor = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]
        town.distributor_count = distributors.distributor_count

        super_stockist = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as super_stockist_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_super_stockist = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]
        town.super_stockist_count = super_stockist.super_stockist_count

        retailers = SpUserAreaAllocations.objects.raw('''SELECT sp_user_area_allocations.id, count(sp_user_area_allocations.id) as retailers_count
        FROM sp_user_area_allocations left join sp_users on sp_users.id = sp_user_area_allocations.user_id 
        where sp_users.is_retailer = %s and sp_user_area_allocations.town_id = %s ''', [1, town.id])[0]

        town.retailer_count = retailers.retailers_count

        town_data.append(town)

    ho_roles = SpRoles.objects.filter(status=1).exclude(
        id=8).exclude(id=9).exclude(id=10)
    context['ho_roles'] = ho_roles

    context['towns'] = town_data
    context['page_title'] = "Manage Users"
    template = 'reports/tagging-report.html'
    return render(request, template, context)


@login_required
def ajaxOutletTagginReport(request):
    page = request.GET.get('non_page')
    search = request.GET.get('search')
    emp_id = request.GET.get('emp_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    users = []
    users = SpUsers.objects.filter(user_type=3, is_tagged=1)
    if search:
        users = users.filter(store_name__icontains=search, is_tagged=1)
    if emp_id:
        users = users.filter(tagged_by=emp_id, is_tagged=1)
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        end_date = end_date.strftime('%Y-%m-%d')
        users = users.filter(tagged_date__range=[
                             start_date, end_date], is_tagged=1)

    count_outlet = 0
    for user in users:
        count_outlet += 1
        try:
            user_basic_details = SpBasicDetails.objects.get(user_id=user.id)
        except SpBasicDetails.DoesNotExist:
            user_basic_details = None
        user.production_unit = ''
        try:
            emp = SpUsers.objects.get(id=user.tagged_by)
        except SpBasicDetails.DoesNotExist:
            emp = None
        user.tagged_by = emp.first_name + ' ' + emp.middle_name + ' ' + emp.last_name
        user.tagged_date = user.tagged_date.strftime('%d/%m/%Y')
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = math.ceil(
        paginator.count/getConfigurationResult('page_limit'))

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages % paginator.count
    if(temp > 0 and getConfigurationResult('page_limit') != paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context = {}
    context['count_outlet'] = count_outlet
    context['scroll_check'] = 23
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'reports/ajax-tagging-report-list.html'
    return render(request, template, context)


@login_required
def exportTaggingReportToXlsx(request, columns, search, emp_id, start_date, end_date):
    column_list = columns.split(",")
    users = SpUsers.objects.all().filter(user_type=3, is_tagged=1)
    if search != '0':
        users = users.filter(store_name__icontains=search, is_tagged=1)
    if emp_id != '0':
        users = users.filter(tagged_by=emp_id, is_tagged=1)
    if start_date != '0' and end_date != '0':
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        end_date = end_date.strftime('%Y-%m-%d')
        users = users.filter(tagged_date__range=[
                             start_date, end_date], is_tagged=1)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=Outlet-tagging-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Outlet Tagging Report'

    # Define the titles for columns
    columns = []

    if 'non_store_name' in column_list:
        columns += ['Store Name']

    if 'non_role' in column_list:
        columns += ['Role']

    if 'non_contact_person' in column_list:
        columns += ['Contact Person']

    if 'non_contact_no' in column_list:
        columns += ['Contact No.']

    # columns += ['Tagged By']
    # columns += ['Tagged Date']

    # columns += [ 'Address' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf",
                                end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies

    for user in users:
        row_num += 1
        # Define the data for each cell in the row

        row = []
        if 'non_store_name' in column_list:
            row += [user.store_name]

        if 'non_role' in column_list:
            row += [user.role_name]

        if 'non_contact_person' in column_list:
            row += [user.first_name + ' ' +
                    user.middle_name + ' ' + user.last_name]

        if 'non_contact_no' in column_list:
            row += [user.primary_contact_number]

        users = SpUsers.objects.get(id=user.tagged_by)
        # row += [users.first_name + ' ' +
        #         users.middle_name + ' ' + users.last_name]
        # row += [user.tagged_date.strftime('%d/%m/%Y')]

        # row += [ organization.address + ', ' + organization.pincode ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response

#ajax non operational user list


@login_required
def ajaxTaggingReportList(request):
    page = request.GET.get('non_page')
    search = request.GET.get('search')
    emp_id = request.GET.get('emp_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    scroll_check = request.GET.get('scroll_check')
    users = []
    users = SpUsers.objects.filter(user_type=3, is_tagged=1)
    if search:
        users = users.filter(store_name__icontains=search, is_tagged=1)
    if emp_id:
        users = users.filter(tagged_by=emp_id, is_tagged=1)
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        end_date = end_date.strftime('%Y-%m-%d')
        users = users.filter(tagged_date__range=[
                             start_date, end_date], is_tagged=1)

    users = users.exclude(id=1).order_by('-id')
    for user in users:
        try:
            emp = SpUsers.objects.get(id=user.tagged_by)
        except SpBasicDetails.DoesNotExist:
            emp = None
        user.tagged_by = emp.first_name + ' ' + emp.middle_name + ' ' + emp.last_name
        user.tagged_date = user.tagged_date.strftime('%d/%m/%Y')
        user.outstanding_amount = SpBasicDetails.objects.filter(
            status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    if page is not None:
        page = page
    else:
        page = 1

    total_pages = math.ceil(
        paginator.count/getConfigurationResult('page_limit'))

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages % paginator.count
    if(temp > 0 and getConfigurationResult('page_limit') != paginator.count):
        total_pages = total_pages
    else:
        total_pages = total_pages

    context = {}
    context['non_operational_users'] = users
    context['scroll_check'] = scroll_check
    context['non_operational_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'reports/ajax-tagging-report-list.html'
    return render(request, template, context)


def locatesUsersOnMap(request):
    context = {}
    search = request.GET.get('search')
    emp_id = request.GET.get('emp_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    users               = []
    users               = SpUsers.objects.filter(is_tagged=1,latitude__isnull=False,longitude__isnull=False).exclude(user_type=1)
    if search:
        users = users.filter(store_name__icontains=search, is_tagged=1)
    if emp_id:
        users = users.filter(tagged_by=emp_id, is_tagged=1)
    if start_date and end_date:
        start_date = datetime.strptime(
            start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        end_date = end_date.strftime('%Y-%m-%d')
        users = users.filter(tagged_date__range=[
                             start_date, end_date], is_tagged=1)
    users_count = users.count()
    for user in users:
        last_order = SpOrders.objects.filter(
            user_id=user.id).values('created_at').first()
        shipping_address = SpAddresses.objects.filter(
            user_id=user.id, type="correspondence").values('address_line_1').first()
        if shipping_address:
            user.address = shipping_address['address_line_1']
        area_allocation = SpUserAreaAllocations.objects.filter(
            user_id=user.id).values('town_name', 'zone_name').first()
        if area_allocation:
            user.town_name = area_allocation['town_name']
            user.zone_name = area_allocation['zone_name']
        else:
            user.town_name = ''
            user.zone_name = ''
        if last_order:
            user.last_order_place_date = last_order['created_at']
        else:
            user.last_order_place_date = '-'
        if user.tagged_date:
            tagged_date = str(user.tagged_date).replace('+00:00', '')
            user.tagged_date = datetime.strptime(
                str(tagged_date), '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y | %I:%M:%p')
        else:
            user.tagged_date = '-'
        if user.tagged_by:
            user.tagged_by = getUserName(user.tagged_by)
        else:
            user.tagged_by = '-'
            
    context['users']        = users
    context['tagged']       = 1
    context['users_count']  = users_count
    template = 'user-management/locate-user-on-map.html'
    return render(request, template, context)

def userNotification(request):
    today   = date.today()
    context = {}
    # attendance['user_id']
    user_attendance = SpUserAttendance.objects.filter(attendance_date_time__icontains=today.strftime("%Y-%m-%d")).order_by('user_id').values('user_id').distinct()
    for attendance in user_attendance:
        start_day_count = SpUserAttendance.objects.filter(attendance_date_time__icontains=today.strftime("%Y-%m-%d"), user_id=attendance['user_id'], start_time__isnull=False, end_time__isnull=True).count()
        end_day_count   = SpUserAttendance.objects.filter(attendance_date_time__icontains=today.strftime("%Y-%m-%d"), user_id=attendance['user_id'], start_time__isnull=True, end_time__isnull=False).count()
        if start_day_count != end_day_count:
            user_tracking = SpUserTracking.objects.filter(user_id=attendance['user_id'], created_at__icontains=today.strftime("%Y-%m-%d")).last()
            user_log = SpUserLocationLogs.objects.filter(user_id=attendance['user_id'] , created_at__icontains = today.strftime("%Y-%m-%d")).last()
            user_logs = SpUserLocationLogs.objects.filter(user_id=attendance['user_id'] , created_at__icontains = today.strftime("%Y-%m-%d"),status = 0).last()
            
            #-----------------------------notify android block-------------------------------#
            userFirebaseToken   = getModelColumnById(SpUsers,attendance['user_id'],'firebase_token')
            employee_name       = getUserName(attendance['user_id'])
            message_title = ""
            message_body = ""
            notification_image = ""
            if userFirebaseToken is not None and userFirebaseToken != "" :
                if user_tracking and user_log:
                    if user_tracking.created_at > user_log.created_at:
                        last_date_time = user_tracking.created_at.strftime('%d-%m-%Y %I:%M %p')
                    if user_log.created_at > user_tracking.created_at:
                        last_date_time = user_log.created_at.strftime('%d-%m-%Y %I:%M %p')
                    else:
                        last_date_time = user_log.created_at.strftime('%d-%m-%Y %I:%M %p')
                    
                elif user_tracking:
                    last_date_time = user_tracking.created_at.strftime('%d-%m-%Y %I:%M %p')
                elif user_log:
                    last_date_time = user_log.created_at.strftime('%d-%m-%Y %I:%M %p')
                else:
                    last_date_time =  ""
                if user_logs:
                    last_logs_time = user_logs.created_at.strftime('%d-%m-%Y %I:%M %p')
                else:
                    last_logs_time = ''
                registration_ids = []
                registration_ids.append(userFirebaseToken)
                data_message = {}
                data_message['id'] = 1
                data_message['status'] = 'notification'
                data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                data_message['last_date_time'] = last_date_time
                data_message['last_logs_time'] = last_logs_time
                data_message['image'] = notification_image
                result = send_android_notification(message_title,message_body,data_message,registration_ids)
                
                #-----------------------------notify android block-------------------------------#
    context['response_codes']    = 'success'
    # context[user_attendance]    = user_attendance
    return HttpResponse(user_attendance, status='200') 



@login_required
def vlcGeoTagged(request):
    context = {}
    
    try:
        user_coordinates = SpVlc.objects.get(id=request.GET['id'])
    except SpVlc.DoesNotExist:
        user_coordinates = None
  
    context['user_coordinates'] = user_coordinates
    context['google_app_key']   = 'AIzaSyCNCcTQduJoNRebWEf7zgqlpe1YJibSuGI'
    template = 'reports/vlc-geo-tagged.html'

    return render(request, template,context)



@login_required
def vlcTaggingReport(request):
    vlces       = SpVlc.objects.all().order_by('-created_at')
    first_retailer       = SpVlc.objects.all().first()
    for emp in vlces:
        emp.tagged_by = getUserName(emp.created_by)
    vlces_count = SpVlc.objects.all().count()
    context = {}
    context['non_operational_users'] = vlces
    context['count_outlet'] = vlces_count
    context['first_retailer'] = first_retailer
    context['reporting_to'] = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    template = 'reports/vlc-tagging-report.html'
    return render(request, template, context)

def vlcShortDetail(request,user_id):
    vlces  = SpVlc.objects.get(id = user_id)
   
    context = {}
    context['user'] = vlces
    template = 'reports/vlc-short-details.html'
    return render(request, template, context)



@login_required
def ajaxVlcTagginReport(request):
    page = request.GET.get('non_page')
    search = request.GET.get('search')
    emp_id = request.GET.get('emp_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    vlces       = SpVlc.objects.all().order_by('-created_at')
    if search:
        vlces = vlces.filter(Q(first_name__icontains = search) or Q(last_name__icontains = search))
    if emp_id:
        vlces = vlces.filter(created_by=emp_id)
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        start_date = start_date.strftime('%Y-%m-%d')
        end_date = end_date.strftime('%Y-%m-%d')
        vlces = vlces.filter(created_at__range=[start_date,end_date])
 
    for emp in vlces:
        emp.tagged_by = getUserName(emp.created_by)
        
    vlces_count  = vlces.count()
    context = {}
    context['non_operational_users'] = vlces
    context['count_outlet'] = vlces_count
    context['reporting_to'] = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    template = 'reports/ajax-vlc-tagging.html'
    return render(request, template, context)

 


def locatesVlcOnMap(request):
    context = {}
   
    search = request.GET.get('search')
    emp_id = request.GET.get('emp_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    vlces       = SpVlc.objects.all().order_by('-created_at')
    if search:
        vlces = vlces.filter(Q(first_name__icontains = search) or Q(last_name__icontains = search))
    if emp_id:
        vlces = vlces.filter(created_by=emp_id)
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        start_date = start_date.strftime('%Y-%m-%d')
        end_date = end_date.strftime('%Y-%m-%d')
        vlces = vlces.filter(created_at__range=[start_date,end_date])
 
    for emp in vlces:
        emp.tagged_by = getUserName(emp.created_by)
    vlces_count  = vlces.count()      
    context['users']        = vlces
    context['tagged']       = 1
    context['users_count']  = vlces_count
    template = 'reports/locate-vlc-on-map.html'
    return render(request, template, context)



@login_required
def exportVlcTaggingReportToXlsx(request, columns, search, emp_id, start_date, end_date):
    column_list = columns.split(",")
    vlces       = SpVlc.objects.all().order_by('-created_at')
    if search != '0':
        vlces = vlces.filter(Q(first_name__icontains = search) or Q(last_name__icontains = search))
    if emp_id != '0':
        vlces = vlces.filter(created_by=emp_id)
    if start_date != '0' and end_date != '0':
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        end_date = end_date + timedelta(days=1)
        start_date = start_date.strftime('%Y-%m-%d')
        end_date = end_date.strftime('%Y-%m-%d')
        vlces = vlces.filter(created_at__range=[start_date,end_date])
    for emp in vlces:
        emp.tagged_by = getUserName(emp.created_by)
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=vlc-tagging-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'VLC Tagging Report'

    # Define the titles for columns
    columns = []

    if 'non_store_name' in column_list:
        columns += ['VLC Name']

    if 'non_role' in column_list:
        columns += ['VLC Code']

    if 'non_contact_no' in column_list:
        columns += ['Contact No.']

    columns += ['Organization Name']
    columns += ['Production Unit']
    columns += ['Vehicle']
    columns += ['Address']

    columns += ['Tagged By']
    columns += ['Tagged Date']

    # columns += [ 'Address' ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf",
                                end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies

    for user in vlces:
        row_num += 1
        # Define the data for each cell in the row

        row = []
        if 'non_store_name' in column_list:
            row += [user.first_name + ' ' + user.last_name]

        if 'non_role' in column_list:
            row += [user.vlc_code]

        if 'non_contact_no' in column_list:
            row += [user.contact_number]
        row += [user.organization_name]
        row += [user.production_unit_name]
        row += [user.vehicle_number]
        row += [user.address]

        row += [user.tagged_by]
        row += [user.created_at.strftime('%d/%m/%Y')]

        # row += [ organization.address + ', ' + organization.pincode ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response
    
    
#update user status
@login_required
def updateVlcStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpVlc.objects.get(id=id)
            data.status = is_active
            data.save()

           
                
            # user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            # heading     = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status
            # activity    = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            # saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
    return JsonResponse(response)
  

def ceckWeekOfDay(date,user_id):
    # print(date)
    user_week_of_day = getModelColumnByColumnId(SpBasicDetails,'user_id',user_id,'week_of_day')
    if user_week_of_day:
        user_week_of_day = user_week_of_day.split(',')
        week_day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%A')
        if week_day in user_week_of_day:
            return True
        else:
            return False

@login_required
def staffAttendanceSummary(request):
    today   = datetime.today()  
    year  = today.year
    month = today.month
    date = today.day
    
    month_list = days_in_months(year,month)
    
    context = {}
    context['today_date']                   = today.strftime("%m/%Y")
    nameMonth = []
    for month_date in month_list:
        month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
        if int(date) >= int(month_date.strftime('%d')):
            nameMonth.append(month_date.strftime('%d %b %a'))

    context['month_list']                   = nameMonth

    staffs = SpUsers.objects.raw(" SELECT DISTINCT sp_users.id, sp_users.* FROM sp_users LEFT JOIN sp_user_attendance on sp_user_attendance.user_id = sp_users.id WHERE sp_users.status = 1 AND sp_users.role_id !=0 AND sp_users.user_type = 1 ORDER BY sp_users.id DESC ")
    for staff in staffs:
        staff.attendance = SpUserAttendance.objects.filter(user_id=staff.id,attendance_date_time__icontains=date).order_by('-id')
    
    total_class = 0
    total_present_staffs = 0

    for staff in staffs:
        staff_attendance = []
        present_count = 0
        absent_count = 0
        staff_total_class = 0
        for months in month_list[:len(nameMonth)]:
            
            month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
            in_time = ''
            out_time = ''
            if checkAttendance(month_date, staff.id):
                attendance_status         = "P"
                present_count = present_count +1
                total_present_staffs = total_present_staffs +1
                punch_time = getPunchTime(month_date, staff.id)
                in_time = punch_time['start_time']
                out_time = punch_time['end_time']
            elif checkLeave(month_date, staff.id):
                attendance_status = checkLeave(month_date, staff.id)
                absent_count = absent_count +1
            elif checkHoliday(month_date,staff.role_id):
                attendance_status = checkHoliday(month_date,staff.role_id)
            elif ceckWeekOfDay(month_date,staff.id):
                attendance_status = 'WO'
                present_count = present_count +1
            else:
                total_class = total_class +1
                # staff_total_class = staff_total_class + 1
                
                attendance_status         = "A"
                absent_count = absent_count +1
            staff_total_class = staff_total_class + 1
            tmp               = {}
            tmp['attendance_status']   = attendance_status
            tmp['in_time']   = in_time
            tmp['out_time']   = out_time

            staff_attendance.append(tmp) 

        staff.father_name             = getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'father_name')
        staff.date_of_joining         = datetime.strptime(str(getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'date_of_joining')), '%Y-%m-%d').strftime('%Y-%m-%d')
        date_of_joining_str = getModelColumnByColumnId(SpBasicDetails, 'user_id', staff.id, 'date_of_joining')

        if date_of_joining_str is not None:
            try:
                date_of_joining = datetime.strptime(str(date_of_joining_str), '%Y-%m-%d').strftime('%Y-%m-%d')
            except ValueError:
                date_of_joining = None
        else:
            
            date_of_joining = None

       
        staff.attendances             = staff_attendance
        staff.present_count           = present_count
        staff.absent_count            = absent_count
        if present_count == 0 or staff_total_class == 0:
            staff.attendance_percentage = 0
        else:
            staff.attendance_percentage   = float((present_count / staff_total_class ) * 100)
    
    context['total_class'] = total_class
    context['total_present_staffs'] = total_present_staffs
    if total_present_staffs == 0 or total_class == 0:
        context['attendance_percentage'] = 0
    else:
        context['attendance_percentage'] = float((total_present_staffs / total_class ) * 100)

    context['staffs'] = staffs
    context['leave_types'] = SpLeaveTypes.objects.values_list('alias', flat=True)
    context['dateForMonthlyAttendance'] = str(today.strftime("%m/%Y")).replace('/','-')
    
    context['page_title'] = "Monthly Attendance Report"
    template = 'reports/staff/staff-attendance-summary.html'
    return render(request, template, context)

# ajax List View
@login_required
def filterStaffAttendanceSummary(request):
    if request.method == "GET":
        today   = datetime.today()  
        year  = today.year
        month = today.month
        current_month = today.month

        date = today.day
        if request.GET['date'] !='':
            filter_date = str(request.GET['date']).split('/')
            if len(filter_date) >2:
                year = int(filter_date[2])
                month = int(filter_date[1])
            else:
                year = int(filter_date[1])
                month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date.strftime('%d %b %a'))
            else:
                nameMonth.append(month_date.strftime('%d %b %a'))
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = nameMonth
        
        condition = ''
        if 'date' in request.GET or  request.GET['date']:
            Date = request.GET['date']
            date_month = str(Date.split('/')[0])
        if 'date' in request.GET or  request.GET['date']:
            staffs = SpUsers.objects.raw(" SELECT DISTINCT sp_users.id, sp_users.* FROM sp_users LEFT JOIN sp_user_attendance on sp_user_attendance.user_id = sp_users.id WHERE sp_users.status = 1 AND sp_users.role_id !=0 ORDER BY sp_users.id DESC ")
            
        for staff in staffs:
            staff.attendance = SpUserAttendance.objects.filter(user_id=staff.id,created_at__icontains=date).order_by('-id')
        
        total_class = 0
        total_present_staffs = 0
        for staff in staffs:
            staff.attendance = SpUserAttendance.objects.filter(user_id=staff.id,attendance_date_time__icontains=date).order_by('-id')

        for staff in staffs:
            staff_attendance = []
            present_count = 0
            absent_count = 0
            staff_total_class = 0
            for months in month_list[:len(nameMonth)]:
                
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                in_time = ''
                out_time = ''
                if checkAttendance(month_date, staff.id):
                    attendance_status         = "P"
                    present_count = present_count +1
                    total_present_staffs = total_present_staffs +1
                    punch_time = getPunchTime(month_date, staff.id)
                    in_time = punch_time['start_time']
                    out_time = punch_time['end_time']
                elif checkLeave(month_date, staff.id):
                    attendance_status = checkLeave(month_date, staff.id)
                    absent_count = absent_count +1
                elif checkHoliday(month_date,staff.role_id):
                    attendance_status = checkHoliday(month_date,staff.role_id)
                elif ceckWeekOfDay(month_date,staff.id):
                    attendance_status = 'WO'
                    present_count = present_count +1
                else:
                    total_class = total_class +1
                    # staff_total_class = staff_total_class + 1
                    
                    attendance_status         = "A"
                    absent_count = absent_count +1
                staff_total_class = staff_total_class + 1
                tmp               = {}
                tmp['attendance_status']   = attendance_status
                tmp['in_time']   = in_time
                tmp['out_time']   = out_time

                staff_attendance.append(tmp) 

            staff.father_name             = getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'father_name')
            #staff.date_of_joining         = datetime.strptime(str(getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'date_of_joining')), '%Y-%m-%d').strftime("%d/%m/%Y")
            date_of_joining_str = getModelColumnByColumnId(SpBasicDetails, 'user_id', staff.id, 'date_of_joining')

            if date_of_joining_str is not None:
                try:
                    date_of_joining = datetime.strptime(str(date_of_joining_str), '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    date_of_joining = None
            else:
                
                date_of_joining = None
            staff.attendances             = staff_attendance
            staff.present_count           = present_count
            staff.absent_count            = absent_count
            if present_count == 0 or staff_total_class == 0:
                staff.attendance_percentage = 0
            else:
                staff.attendance_percentage   = float((present_count / staff_total_class ) * 100)

        context['total_class'] = total_class
        context['total_present_staffs'] = total_present_staffs
        if total_present_staffs == 0 or total_class == 0:
            context['attendance_percentage'] = 0
        else:
            context['attendance_percentage'] = float((total_present_staffs / total_class ) * 100)

        context['staffs'] = staffs
        context['leave_types'] = SpLeaveTypes.objects.values_list('alias', flat=True)
        context['dateForMonthlyAttendance'] = str(request.GET['date']).replace('/','-')
        context['page_title']                   = "Attendance Summary"
        template = 'reports/staff/ajax-attendance-summary.html'
        return render(request, template, context)  


@login_required
def exportToXlsx(request,  filter_date):

    if request.method == "GET":
        today   = datetime.today()  
        year  = today.year
        month = today.month
        current_month = today.month

        date = today.day
        if filter_date !='':
            filter_date = str(filter_date).split('-')
            
            year = int(filter_date[1])
            month = int(filter_date[0])
            month_list = days_in_months(year,month)
        else:
            month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
                if int(date) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date.strftime('%d %b %a' ))
            else:
                nameMonth.append(month_date.strftime('%d %b %a'))
        context = {}
        context['today_date']                   = today.strftime("%m/%Y")
        context['month_list']                   = nameMonth
        
        condition = ''
        staffs = SpUsers.objects.raw(" SELECT DISTINCT sp_users.id, sp_users.* FROM sp_users LEFT JOIN sp_user_attendance on sp_user_attendance.user_id = sp_users.id WHERE sp_users.status = 1 AND sp_users.role_id !=0  ORDER BY sp_users.id DESC ")
            
        for staff in staffs:
            staff.attendance = SpUserAttendance.objects.filter(user_id=staff.id,created_at__icontains=date).order_by('-id')
        
        total_class = 0
        total_present_staffs = 0
        for staff in staffs:
            staff.attendance = SpUserAttendance.objects.filter(user_id=staff.id,attendance_date_time__icontains=date).order_by('-id')

        for staff in staffs:
            staff_attendance = []
            present_count = 0
            absent_count = 0
            staff_total_class = 0
            for months in month_list[:len(nameMonth)]:
                
                month_date                = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                in_time = ''
                out_time = ''
                if checkAttendance(month_date, staff.id):
                    attendance_status         = "P"
                    present_count = present_count +1
                    total_present_staffs = total_present_staffs +1
                    punch_time = getPunchTime(month_date, staff.id)
                    in_time = punch_time['start_time']
                    out_time = punch_time['end_time']
                elif checkLeave(month_date, staff.id):
                    attendance_status = checkLeave(month_date, staff.id)
                    absent_count = absent_count +1
                elif checkHoliday(month_date,staff.role_id):
                    attendance_status = checkHoliday(month_date,staff.role_id)
                elif ceckWeekOfDay(month_date,staff.id):
                    attendance_status = 'WO'
                    present_count = present_count +1
                else:
                    total_class = total_class +1
                    
                    
                    attendance_status         = "A"
                    absent_count = absent_count +1
                staff_total_class = staff_total_class + 1
                tmp               = {}
                tmp['attendance_status']   = attendance_status
                tmp['in_time']   = in_time
                tmp['out_time']   = out_time

                staff_attendance.append(tmp) 

            staff.father_name             = getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'father_name')
            #staff.date_of_joining         = datetime.strptime(str(getModelColumnByColumnId(SpBasicDetails,'user_id',staff.id,'date_of_joining')), '%Y-%m-%d').strftime("%d/%m/%Y")
            date_of_joining_str = getModelColumnByColumnId(SpBasicDetails, 'user_id', staff.id, 'date_of_joining')

            if date_of_joining_str is not None:
                try:
                    date_of_joining = datetime.strptime(str(date_of_joining_str), '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    date_of_joining = None
            else:
                
                date_of_joining = None
            staff.attendances             = staff_attendance
            staff.present_count           = present_count
            staff.absent_count            = absent_count
            if present_count == 0:
                staff.attendance_percentage   = 0
            else:
                staff.attendance_percentage   = float((present_count / staff_total_class ) * 100)

        context['total_class'] = total_class
        context['total_present_staffs'] = total_present_staffs
        if total_present_staffs == 0:
            context['attendance_percentage'] = 0
        else:
            context['attendance_percentage'] = float((total_present_staffs / total_class ) * 100)

    
    leave_types = SpLeaveTypes.objects.values_list('alias', flat=True)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=staff-attendance-list.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style="thin", color="303030") 
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='center',
        horizontal='center',
        wrap_text=True
    )
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))


    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Attendance Summary Report'
    
    worksheet.merge_cells('A1:A1') 
    
    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True

    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
    img.height = 50
    img.width = 60
    img.alignment = centered_alignment
    img.anchor = 'A1'
    worksheet.add_image(img)
    worksheet['A1'].alignment = wrapped_alignment

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    cell = worksheet.cell(row=1, column=2)  
    cell.value = calendar.month_name[int(filter_date[0])] + ' '+filter_date[1]
    cell.font = header_font
    cell.alignment = wrapped_alignment
    cell.border = black_border
    cell.font = Font(size=15, bold=True)
    cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    column_length = len(nameMonth) + 8
    
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=3).value = "Attendance Summary Report"
    worksheet.cell(row=1, column=3).font = header_font
    worksheet.cell(row=1, column=3).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=3).font = Font(size=20, bold=True)
    worksheet.cell(row=1, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type = "solid")

    row_num = 2
    # Define the titles for columns
    columns = ['S.No']
    columns += [ 'Employee No.' ]
    columns += [ 'Associate Name' ]
    columns += [ 'Father Name ' ]
    columns += [ 'Department' ]
    #columns += [ 'DOJ' ]
    for month_date in nameMonth:
       
        columns += [ month_date ]

    columns += [ 'Total Present' ]
    columns += [ 'Total Absent' ]
    

    # Assign the titles for each cell of the header
    header_column_counter = 0
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = thin_border
        
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        if header_column_counter == 0:
            cell.alignment = centered_alignment

        elif header_column_counter == 1:
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 30
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = centered_alignment


        header_column_counter = header_column_counter + 1

    # Iterate through all movies
    for staff in staffs:
        row_num += 1
        # Define the data for each cell in the row 
        column = [str(int(row_num)-2)]
        name = ''
        if staff.first_name:
            name += staff.first_name
        if staff.middle_name:
            name += " "+staff.middle_name
        if staff.last_name:
            name += " "+staff.last_name
        column += [ staff.emp_sap_id ]
        column += [ name +"\n("+ staff.role_name+")" ]
        column += [ staff.father_name ]
        column += [ staff.department_name ]
        #column += [ staff.date_of_joining ]
        for attendance in staff.attendances:
            column += [ attendance['attendance_status'] ]
        column += [ str(staff.present_count)+'('+ str(round(staff.attendance_percentage,2)) +'%)' ]
        column += [ staff.absent_count ]
        
        content_column_counter = 0
        for col_num, cell_value in enumerate(column, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.font = header_font
            cell.border = thin_border
            
            if content_column_counter == 0:
                column_dimensions.width = 5

            elif content_column_counter <= 5 or content_column_counter > (len(nameMonth)+5):
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = 20
                cell.alignment = Alignment(horizontal='left',wrap_text=True)
            else:
                cell.alignment = centered_alignment

            if content_column_counter > 5 and content_column_counter < len(nameMonth)+6:
                if cell_value == "P":
                    cell.fill = PatternFill(start_color="7EC857", end_color="7EC857", fill_type = "solid")
                elif cell_value == 'Week Of' or cell_value == 'HL':
                    cell.fill = PatternFill(start_color="B7F54C", end_color="B7F54C", fill_type = "solid")
                elif cell_value == "A":
                    cell.fill = PatternFill(start_color="FF4859", end_color="FF4859", fill_type = "solid")
                elif cell_value in leave_types :
                    cell.fill = PatternFill(start_color="CF3847", end_color="CF3847", fill_type = "solid")
                else:
                    cell.fill = PatternFill(start_color="B7F54C", end_color="B7F54C", fill_type = "solid")
            
            content_column_counter = content_column_counter + 1

    workbook.save(response)

    return response

#done
def checkLeave(date, user_id):
    try:
        leaves = SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=user_id).first()
    except SpUserLeaves.DoesNotExist:
        leaves = None
    if leaves:
        return getModelColumnById(SpLeaveTypes, leaves.leave_type_id, 'alias')
    else:
        return False
#done

def checkHoliday(date,role_id):
    if SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).exists():
        holiday_id = SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).values_list('id',flat=True)
        holiday_ids = SpRoleEntityMapping.objects.filter(role_id = role_id,entity_type = 'holiday',entity_id__in = holiday_id).values_list('entity_id',flat=True)
        count = 0
        for holiday_id in holiday_ids:
            last_date = getModelColumnById(SpHolidays,holiday_id,'end_date')
            last_date = last_date.strftime('%Y-%m-%d')
            last_date = datetime.strptime(str(last_date), '%Y-%m-%d')
            from_date = datetime.strptime(str(date), '%Y-%m-%d')#.strftime('%Y-%m-%d')
            delta = last_date - from_date
            start_date = getModelColumnById(SpHolidays,holiday_id,'start_date')
            start_date = start_date.strftime('%Y-%m-%d')
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d')
            current_date = datetime.strptime(str(date), '%Y-%m-%d')
            if current_date>=start_date:
                if delta.days >= 0:
                    count+=1
        if count > 0:
            return 'HL'
        else:
            return False 
    else:
        return False 

#done
def checkAttendance(date,user_id):
    if SpUserAttendance.objects.filter(user_id=user_id,attendance_date_time__contains=date).exists():
        return True
    else:
        return False
#done



def getPunchTime(date, user_id):
    context = {}
    first_record = SpUserAttendance.objects.filter(user_id=user_id,created_at__contains=date).first()
    if first_record is None:
        context['start_time'] = ""
    else:
        context['start_time'] = first_record.attendance_date_time

    last_record = SpUserAttendance.objects.filter(user_id=user_id,created_at__contains=date).last()
    if last_record is None:
        context['end_time'] = ""
    else:
        context['end_time'] = last_record.end_time
    return context
  
@login_required
def staffMonthlyAttendance(request, user_id, MonthDate):
    if datetime.now().strftime('%m-%Y') == MonthDate:
        attendance_start_date                 = datetime.today().replace(day=1).strftime('%Y-%m-%d')
        attendance_end_date                   = datetime.now().strftime('%Y-%m-%d')
    else:
        MonthDate                               = "01-"+ str(MonthDate)
        mD = str(MonthDate).split('-')
        nDays=numberOfDays(int(mD[2]), int(mD[1]))
        attendance_start_date                   = datetime.strptime(str(MonthDate), '%d-%m-%Y').replace(day=1).strftime('%Y-%m-%d')
        attendance_end_date                     = datetime.strptime(str(MonthDate), '%d-%m-%Y').strftime('%Y-%m')+"-"+str(nDays)
    condition=""
    if user_id != "" :
        condition=" where sp_user_attendance.user_id="+user_id+ " and"
    else:
        condition=" where "
    context = {}

    start_date = date(int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%Y')),int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%m')),int(datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%d')))
    end_date = date(int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%Y')),int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%m')),int(datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%d')))
    
    users=[]
    while start_date <= end_date:
        today = start_date
        today_attendance_users = SpUserAttendance.objects.raw(''' SELECT sp_user_attendance.id,sp_user_attendance.user_id FROM sp_user_attendance {condition} date(attendance_date_time) = %s
            group by user_id '''.format(condition=condition),[today])
        
        if len(today_attendance_users):
            for user in today_attendance_users:
                temp = {}
                start_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, CONCAT(sp_users.first_name," ",
                sp_users.middle_name," ",sp_users.last_name) as name,sp_users.emp_sap_id,sp_users.role_name,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.start_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s  
                order by sp_user_attendance.id LIMIT 1 ''',[user.user_id,today])
                end_attendance =  SpUserAttendance.objects.raw('''SELECT sp_user_attendance.*, sp_users.first_name,
                sp_users.middle_name,sp_users.last_name,sp_users.emp_sap_id,sp_users.profile_image,
                dis_ss.emp_sap_id as dis_sap_id,dis_ss.store_name,dis_ss.store_image
                FROM sp_user_attendance 
                left join sp_users as dis_ss on dis_ss.id = sp_user_attendance.dis_ss_id  
                left join sp_users on sp_users.id = sp_user_attendance.user_id 
                WHERE sp_user_attendance.end_time IS NOT NULL and sp_user_attendance.user_id = %s and date(sp_user_attendance.attendance_date_time) =  %s 
                order by sp_user_attendance.id desc LIMIT 1 ''',[user.user_id,today])
               
                temp['date']=start_date.strftime('%d/%m/%Y')
                temp['id'] = user.user_id
                temp['name'] = start_attendance[0].name
                temp['role_name'] = start_attendance[0].role_name
                temp['emp_sap_id'] = start_attendance[0].emp_sap_id
                temp['profile_image'] = start_attendance[0].profile_image
                temp['dis_sap_id'] = start_attendance[0].dis_sap_id
                temp['store_name'] = start_attendance[0].store_name
                temp['store_image'] = start_attendance[0].store_image
                temp['start_time'] = start_attendance[0].start_time
                temp['start_latitude'] = start_attendance[0].latitude
                temp['start_longitude'] = start_attendance[0].longitude
                temp['img'] = start_attendance[0].attendance_img
                if end_attendance:
                    temp['end_time'] = end_attendance[0].end_time
                    temp['end_latitude'] = end_attendance[0].latitude
                    temp['end_longitude'] = end_attendance[0].longitude
                    temp['attendance_type'] = end_attendance[0].attendance_type
                else:
                    temp['end_time'] = None

                now = str(today)
                start_time = now + ' '+start_attendance[0].start_time
                if temp['end_time'] is None:
                    end_time = now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                else:
                    end_time = now + ' '+end_attendance[0].end_time
                
                start_time = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
                end_time = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
                time_delta = (end_time - start_time)
                total_seconds = time_delta.total_seconds()
                hours = convert(total_seconds)
                temp['working_hours'] = str(hours)

                distance_travelled  = SpUserTracking.objects.filter(user_id=start_attendance[0].user_id, created_at__icontains=today).aggregate(Sum('distance_travelled'))
                if distance_travelled['distance_travelled__sum'] is not None:
                    temp['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    temp['distance_travelled'] = 0.00

                temp['added_employee_count'] = SpUsers.objects.filter(created_by=start_attendance[0].user_id, created_at__icontains=today).count()
                temp['tagged_employee_count'] = 0
                
                users.append(temp)
                
        
        start_date    = start_date + timedelta(days=1)
        
    context['attendance_start_date'] =  datetime.strptime(str(attendance_start_date), '%Y-%m-%d').strftime('%d/%m/%Y')
    context['attendance_end_date'] = datetime.strptime(str(attendance_end_date), '%Y-%m-%d').strftime('%d/%m/%Y')
    roles =  SpRoles.objects.filter(status=1).all()
    context['roles'] = roles
    context['users'] = users
    context['page_title'] = "Monthly Attendance Report"
    template = 'reports/attendance-report/attendance-report.html'
    return render(request, template, context)  
 

 
 
 
 
 
def regularizationReport(request):
    today = date.today()
    
    if request.user.role_id == 0:
        regularizationReport = SpUserRegularization.objects.all().order_by('-id')
    else:
        regularizationReport = SpUserRegularization.objects.raw('''SELECT sp_user_regularization.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
    FROM sp_user_regularization left join sp_approval_status on sp_approval_status.row_id = sp_user_regularization.id
    left join sp_users on sp_users.id = sp_user_regularization.user_id 
    where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserRegularization' order by id desc ''',[request.user.id])
    

    user_type = SpPermissionWorkflowRoles.objects.filter(sub_module_id=38,  workflow_level_role_id=request.user.role_id).exclude(level_id=1).values('level_id').order_by('-id').first()  
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
    context = {}

    user_ids    = SpUserRegularization.objects.all().distinct().values_list('user_id',flat=True)
    users       = SpUsers.objects.filter(status = 1).exclude(id=1).values('id','first_name','middle_name','last_name','emp_sap_id')

    context['regularizationReport'] = regularizationReport
    context['role_id']              = request.user.role_id
    context['level_id']             = level_id
    context['today_date']           = today.strftime("%d/%m/%Y")
    context['page_title']           = "Regularization Request"
    context['users']                = users

    template = 'reports/regularization/regularization-report.html'
    return render(request, template, context)


# ajax order list

def ajaxRegularizationReportLists(request):
    context = {}
    today   = date.today()
    user_id = request.GET['user_id']
    regularizationReport = SpUserRegularization.objects.all().order_by('-id')
    if request.user.role_id == 0:
        if 'regularization_status' in request.GET and request.GET['regularization_status'] != "":
            regularization_status = request.GET['regularization_status']
            context['regularization_status'] = regularization_status
            regularizationReport = regularizationReport.filter(regularization_status=regularization_status).order_by('-id')
        if user_id:
                regularizationReport = regularizationReport.filter(user_id=user_id).order_by('-id')
    else:
        condition = ''
        
        if 'regularization_status' in request.GET and request.GET['regularization_status'] != "":
            regularization_status = request.GET['regularization_status']
            condition += ' and sp_user_regularization.regularization_status = "%s"' % regularization_status
        if user_id:
            condition += ' and sp_user_regularization.user_id = "%s"' % user_id
            
        query = """SELECT sp_user_regularization.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_regularization left join sp_approval_status on sp_approval_status.row_id = sp_user_regularization.id
        left join sp_users on sp_users.id = sp_user_regularization.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserRegularization' %s order by id desc """ % (request.user.id,condition)

        regularizationReport = SpUserRegularization.objects.raw(query)

    user_type = SpRoleWorkflowPermissions.objects.filter(sub_module_id=50, permission_slug='add', workflow_level_role_id=request.user.role_id).exclude(role_id=request.user.role_id).values('level_id').order_by('-id').first()
    if user_type:
        level_id = user_type['level_id']
    else:
        level_id = '0'
        
    context['regularizationReport'] = regularizationReport
    context['level_id']             = level_id
    context['role_id']              = request.user.role_id
    template = 'reports/regularization/ajax-regularization-list.html'
    return render(request, template, context)

def updateRegularizationRemark(request): 
    context = {}
    context['level_id']                 = request.GET.get('level_id')
    context['regularization_status']    = request.GET.get('regularization_status')
    template = 'reports/regularization/update-regularization-remark.html'
    return render(request, template, context)

# #update regularization status
@login_required
def updateRegularizationStatus(request):
    response = {}
    level_id                    = request.POST.get('level_id')
    regularization_id = level_id
    regularization_status       = request.POST.get('regularization_status')
    # if request.user.role_id == 0:
    #     approvals_request = SpApprovalStatus.objects.filter(row_id=regularization, model_name='SpUserRegularization', level_id=level_id)
    #     if approvals_request:
    #         for approval in approvals_request:
    #             approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
    #             approval_data.level_id                  = regularization_status
    #             # if regularization_status == '2':
    #             #     approval_data.level                    = 'Forward'
    #             if regularization_status == '3':
    #                 approval_data.level                    = 'Approve'
    #             elif regularization_status == '4':
    #                 approval_data.level                    = 'Declined'         
    #             approval_data.status                    = 1
    #             approval_data.final_status_user_id      = request.user.id
    #             approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    #             approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #             approval_data.save()

    #         user_level_approval_count = SpApprovalStatus.objects.filter(row_id=regularization, model_name='SpUserRegularization', level_id=level_id, status=0).count()
    #         if user_level_approval_count == 0:
    #             regularization                   = SpUserRegularization.objects.get(id=regularization)   
    #             regularization.regularization_status      = regularization_status
    #             if request.POST['remark']:
    #                 regularization.remark      = request.POST['remark']
    #             regularization.save()
    #         else:
    #             regularization                   = SpUserRegularization.objects.get(id=regularization)   
    #             regularization.regularization_status      = regularization_status
    #             if request.POST['remark']:
    #                 regularization.remark      = request.POST['remark']
    #             regularization.updated_at        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
    #             regularization.save()
    #             today   = date.today()
                 
    
    # else:    
       
    #     approvals_request = SpApprovalStatus.objects.filter(row_id=regularization, model_name='SpUserRegularization', role_id=request.user.role_id, level_id=level_id)
    #     for approval in approvals_request:
    #         approval_data                           = SpApprovalStatus.objects.get(id=approval.id)
    #         approval_data.status                    = 1
    #         approval_data.final_status_user_id      = request.user.id
    #         approval_data.final_status_user_name    = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    #         approval_data.final_update_date_time    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #         approval_data.save()

    # user_level_approval_count = SpApprovalStatus.objects.filter(row_id=regularization, model_name='SpUserRegularization', level_id=level_id, status=0).count()
    # if user_level_approval_count == 0:
    regularization                   = SpUserRegularization.objects.get(id=regularization_id)   
    regularization.regularization_status      = regularization_status
    if request.POST['remark']:
        regularization.remark      = request.POST['remark']
    regularization.save()   

    
    # if regularization_status == '2':
    #     for regularization in regularization_id:
    #         approvals_requests = SpApprovalStatus.objects.filter(row_id=regularization, model_name='SpUserRegularization', status=0)
    #         if approvals_requests:
    #             for approval in approvals_requests:
    #                 notification                        = SpUserNotifications()
    #                 notification.row_id                 = approval.row_id
    #                 notification.user_id                = approval.user_id
    #                 notification.model_name             = 'SpUserRegularization'
    #                 notification.notification           = 'regularization '+approval.level+' Request has been sent.'
    #                 notification.is_read                = 0
    #                 notification.created_by_user_id     = request.user.id
    #                 notification.created_by_user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    #                 notification.save()

    # if regularization_status == '2':
    #     for regularization in regularization_id:
    #         user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
    #         heading     = 'Regularization Request has been forwarded'
    #         activity    = 'Regularization Request has been forwarded by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
    #         if request.POST['remark']:
    #             activity += '. '+request.POST['remark']
    #         saveActivity('User Management', 'Regularization', heading, activity, request.user.id, user_name, 'forwaord.png', '1', 'web.png')

    #         #-----------------------------notify android block-------------------------------#
    #         user_id = getModelColumnById(SpUserRegularization,regularization,'user_id')
    #         user_role = getModelColumnById(SpUsers,user_id,'role_name')
    #         userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
    #         employee_name = getUserName(user_id)
            
    #         message_title = "Regularization request forwarded"
    #         message_body = "A Regularization request("+employee_name+" - "+user_role+") has been forwarded  by "+user_name
    #         if request.POST['remark']:
    #             message_body += '. '+request.POST['remark']
    #         notification_image = ""

    #         if userFirebaseToken is not None and userFirebaseToken != "" :
    #             registration_ids = []
    #             registration_ids.append(userFirebaseToken)
    #             data_message = {}
    #             data_message['id'] = 1
    #             data_message['status'] = 'notification'
    #             data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
    #             data_message['image'] = notification_image
    #             send_android_notification(message_title,message_body,data_message,registration_ids)
    #             #-----------------------------notify android block-------------------------------#

    #         #-----------------------------save notification block----------------------------#
    #         saveNotification(regularization,'SpUserRegularization','User Management','Regularization request forwarded',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
    #         #-----------------------------save notification block----------------------------#
            
    if regularization_status == '3':
        
        saveAttendanceData(regularization_id)
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'Regularization Request has been approved'
        activity    = 'Regularization Request has been approved by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        if request.POST['remark']:
            activity += '. '+request.POST['remark']
        saveActivity('User Management', 'Regularization', heading, activity, request.user.id, user_name, 'approved.svg', '1', 'web.png')

        #-----------------------------notify android block-------------------------------#
        user_id = getModelColumnById(SpUserRegularization,regularization_id,'user_id')
        user_role = getModelColumnById(SpUsers,user_id,'role_name')
        userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
        employee_name = getUserName(user_id)

        message_title = "Regularization request approved"
        message_body = "A Regularization request("+employee_name+" - "+user_role+") has been approved  by "+user_name
        if request.POST['remark']:
            message_body += '. '+request.POST['remark']
        notification_image = ""

        if userFirebaseToken is not None and userFirebaseToken != "" :
            registration_ids = []
            registration_ids.append(userFirebaseToken)
            data_message = {}
            data_message['id'] = 1
            data_message['status'] = 'notification'
            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
            data_message['image'] = notification_image
            send_android_notification(message_title,message_body,data_message,registration_ids)
            #-----------------------------notify android block-------------------------------#

        #-----------------------------save notification block----------------------------#
        saveNotification(regularization_id,'SpUserRegularization','User Management','Regularization request approved',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
        #-----------------------------save notification block----------------------------#

    elif regularization_status == '4':
        
        user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
        heading     = 'Regularization Request has been declined'
        activity    = 'Regularization Request has been declined by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
        if request.POST['remark']:
            activity += '. '+request.POST['remark']
        saveActivity('User Management', 'regularization', heading, activity, request.user.id, user_name, 'declined.svg', '1', 'web.png')

        #-----------------------------notify android block-------------------------------#
        user_id = getModelColumnById(SpUserRegularization,regularization_id,'user_id')
        user_role = getModelColumnById(SpUsers,user_id,'role_name')
        userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
        employee_name = getUserName(user_id)

        message_title = "Regularization request declined"
        message_body = "A Regularization request("+employee_name+" - "+user_role+") has been declined  by "+user_name
        if request.POST['remark']:
            message_body += '. '+request.POST['remark']
        notification_image = ""

        if userFirebaseToken is not None and userFirebaseToken != "" :
            registration_ids = []
            registration_ids.append(userFirebaseToken)
            data_message = {}
            data_message['id'] = 1
            data_message['status'] = 'notification'
            data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
            data_message['image'] = notification_image
            send_android_notification(message_title,message_body,data_message,registration_ids)
            #-----------------------------notify android block-------------------------------#

        #-----------------------------save notification block----------------------------#
        saveNotification(regularization_id,'SpUserRegularization','User Management','Regularization request declined',message_title,message_body,notification_image,request.user.id,user_name,user_id,employee_name,'profile.png',2,'app.png',1,1)
        #-----------------------------save notification block----------------------------#

    response['error'] = False
    response['message'] = "Regularization request status has been updated successfully."
    return JsonResponse(response)


def regularizationExportToXlsx(request, columns, userId, regularization_status):
    column_list = columns.split(",")

    regularizationReport = SpUserRegularization.objects.all().order_by('-id')
    if request.user.role_id == 0:
        condition = ''
        if regularization_status != "" and regularization_status != "0":
            condition += ' and regularization_status = "%s"' % regularization_status

        if userId != "" and userId != "0":
            condition += ' and user_id = "%s"' % userId
        
        regularizationReport = SpUserRegularization.objects.raw("""SELECT * FROM sp_user_regularization WHERE 1 {condition}  order by id desc """.format(condition=condition))
    else:
        condition = ''
        if regularization_status != "" and regularization_status != "0":
            condition += ' and sp_user_regularization.regularization_status = "%s"' % regularization_status

        if userId != "" and userId != "0":
            condition += ' and sp_user_regularization.user_id = "%s"' % userId

            

        query = """SELECT sp_user_regularization.*, sp_approval_status.level_id, sp_approval_status.level, sp_approval_status.status, sp_approval_status.final_status_user_id, sp_approval_status.final_status_user_name, sp_users.store_name
        FROM sp_user_regularization left join sp_approval_status on sp_approval_status.row_id = sp_user_regularization.id
        left join sp_users on sp_users.id = sp_user_regularization.user_id 
        where  sp_approval_status.user_id = %s and sp_approval_status.model_name = 'SpUserRegularization' %s order by id desc """ % (request.user.id,condition)

        regularizationReport = SpUserRegularization.objects.raw(query)

        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=regularization-report.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Regularization-reports'

    # Define the titles for columns
    columns = []

    if 'user_name' in column_list:
        columns += ['User name']

    if 'regularization_type' in column_list:
        columns += ['Regularization Type']

    if 'regularization_from_date' in column_list:
        columns += ['Date From']

    if 'regularization_to_date' in column_list:
        columns += ['Date To']

    if 'mobile_no' in column_list:
        columns += ['Mobile No.']

    if 'place' in column_list:
        columns += ['Place']

    if 'reason' in column_list:
        columns += ['Reason for Leave']

    # if 'manager' in column_list:
    #     columns += ['Manager']

    # if 'hod' in column_list:
    #     columns += ['HOD']

    if 'status' in column_list:
        columns += ['Status']

    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 23

    for results in regularizationReport:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        if 'user_name' in column_list:
            row += [results.user_name]

        if 'regularization_type' in column_list:
            row += [results.regularization_type_name]

        if 'regularization_from_date' in column_list:
            if results.from_date and results.from_time:
                row += [datetime.strptime(str(results.from_date), '%Y-%m-%d').strftime('%d/%m/%Y')+' '+str(datetime.strptime(str(results.from_time), '%H:%M').strftime('%I:%M'))]
            elif results.from_date:
                row += [datetime.strptime(str(results.from_date), '%Y-%m-%d').strftime('%d/%m/%Y')]
            else:
                row += ['']
        if 'regularization_to_date' in column_list:
            if results.to_date and results.to_time:
                row += [datetime.strptime(str(results.to_date), '%Y-%m-%d').strftime('%d/%m/%Y')+' '+str(datetime.strptime(str(results.to_time), '%H:%M').strftime('%I:%M'))]
            elif results.to_date:
                row += [datetime.strptime(str(results.to_date), '%Y-%m-%d').strftime('%d/%m/%Y')]
            else:
                row += ['']

        if 'mobile_no' in column_list:
            row += [results.mobile_no]

        if 'place' in column_list:
            row += [results.place]

        if 'reason' in column_list:
            row += [results.reason_for_leave]

        # if 'manager' in column_list:
        #     row += [results.manager]

        # if 'hod' in column_list:
        #     row += [results.hod]

        if 'status' in column_list:
            if results.regularization_status	 == 1:
                status = 'Initiated'
            elif results.regularization_status	 == 2:
                status = 'Forwarded'
            elif results.regularization_status	 == 3:
                status = 'Approved'
            elif results.regularization_status	 == 4:
                status = 'Declined'
            
            row += [status]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border
    workbook.save(response)

    return response    


