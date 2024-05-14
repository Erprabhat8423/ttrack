import sys
import shutil, os
import time,json
from django.shortcuts import get_object_or_404, render, redirect
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.db import connection
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound,Http404
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.apps import apps
from ..models import *
from django.db.models import Q
from utils import *
from datetime import datetime
import openpyxl
from decimal import Decimal
from datetime import timedelta,date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.core import serializers
from math import sin, cos, sqrt, atan2, radians
from io import BytesIO
from django.template.loader import get_template
from django.views import View
from xhtml2pdf import pisa
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.hashers import make_password

from PIL import Image

# Create your views here.


def mapUserLeavess(request,role_id):
    user_id = SpUsers.objects.filter(role_id  = role_id,status = 1)
    for i in user_id:
        mapUserLeaves(role_id,i.id)
    
    
# User List View
@login_required
def index(request):
    #employee
    page = request.GET.get('employee_page')
    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    all_users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1
    
    total_pages = int(paginator.count/getConfigurationResult('page_limit')) 

    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages
    context = {}   
    context['all_users'] = all_users
    context['employee_users'] = users
    context['userscount'] = all_users.count()
    context['employee_total_pages'] = total_pages

    first_employee_id = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id').first()

    if first_employee_id is not None:
        # first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
        # ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
        # FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
        # left join sp_addresses on sp_addresses.user_id = sp_users.id  
        # where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id=%s order by id desc LIMIT 1 ''',[1,'correspondence', first_employee_id.id])
        first_employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type ,sp_basic_details.gender, sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
        ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id left join contract_type on contract_type.id=sp_basic_details.contract_type
        left join sp_addresses on sp_addresses.user_id = sp_users.id where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id=%s order by id desc LIMIT 1 ''',[1,'correspondence', first_employee_id.id])
        
        if first_employee :
            context['first_employee'] = first_employee[0]
            try:
                first_employee_permanent_address = SpAddresses.objects.get(user_id=first_employee[0].id,type='permanent')
            except SpAddresses.DoesNotExist:
                first_employee_permanent_address = None
            try:
                contacts = SpContactNumbers.objects.filter(user_id=first_employee[0].id).first()
            except SpContactNumbers.DoesNotExist:
                contacts = None
            context['contacts'] = contacts
            
        else : 
            context['first_employee'] = []
            first_employee_permanent_address = None
            context['contacts'] = None
        
        single = SpUsers.objects.filter(id = first_employee_id.id).first()
        if single and single.web_auth_token:
            device_data = {item.split(":")[0].strip().capitalize(): item.split(":")[1].strip() for item in str(single.web_auth_token).split(",")}
    
            context['device_data'] =  device_data
        else:
            context['device_data'] =  None

        context['first_employee_permanent_address'] = first_employee_permanent_address
  
    if request.user.role_id == 0:
        organizations = SpOrganizations.objects.all()
        departments = SpDepartments.objects.all()
        roles = SpRoles.objects.all()
    else:
        organizations = SpOrganizations.objects.all()
        departments = SpDepartments.objects.all()
        roles = SpRoles.objects.all()
        
    town_data = []
    towns = SpTowns.objects.all()
    context['jobs']  = ContractType.objects.all()
    context['organizations']  = organizations
    context['departments']  = SpDepartments.objects.all()
    context['roles']  = roles

    context['towns'] = town_data
    context['page_title'] = "Manage Employees"
    template = 'user-management/index.html'
    return render(request, template, context)

#ajax operational user list
@login_required
def ajaxOperationalUsersList(request):
    page = request.GET.get('page')

    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['users'] = users
    context['total_pages'] = total_pages
    context['page_limit'] = getConfigurationResult('page_limit')
    template = 'user-management/ajax-operational-users-list.html'
    return render(request, template, context)

#ajax non operational user list
@login_required
def ajaxNonOperationalUsersList(request):
    page = request.GET.get('non_page')

    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.filter(status=1, user_id=user.id).values('outstanding_amount').first()

    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1

    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['non_operational_users'] = users
    context['non_operational_total_pages'] = total_pages
    template = 'user-management/ajax-non-operational-users-list.html'
    return render(request, template, context)

@login_required
def ajaxEmployeeUsersList(request):
    page = request.GET.get('employee_page')
    search = request.GET.get('search')
    jobs = request.GET.get('jobs')
    departments = request.GET.get('depts')
    roles = request.GET.get('roles')


    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    if search:
        users = users.filter(id = search)
    if len(departments) > 0:
        users = users.filter(department_id = departments)
    if len(roles) > 0:
        users = users.filter(role_id = roles)
    
    if len(jobs) > 0:
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type = jobs):
                users = users.exclude(id = user.id)

    users_count = users.count()
    paginator = Paginator(users, getConfigurationResult('page_limit'))

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)  
    if page is not None:
           page = page
    else:
           page = 1

    total_pages = int(paginator.count/getConfigurationResult('page_limit'))  
    
    if(paginator.count == 0):
        paginator.count = 1
        
    temp = total_pages%paginator.count
    if(temp > 0 and getConfigurationResult('page_limit')!= paginator.count):
        total_pages = total_pages+1
    else:
        total_pages = total_pages

    context = {}
    context['employee_users'] = users
    context['userscount'] = users_count
    context['employee_total_pages'] = total_pages
    context['page_loading_type'] = request.GET.get('page_loading_type')
    template = 'user-management/ajax-employee-users-list.html'
    
    return render(request, template, context)



@login_required
def userGeoTagged(request):
    context = {}
  
    try:
        user_coordinates = SpUsers.objects.get(id=request.GET['id'])
    except SpUserAttendanceLocations.DoesNotExist:
        user_coordinates = None
           
    context['user_coordinates'] = user_coordinates
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-geo-tagged.html'

    return render(request, template,context)

# User basic details View
@login_required
def addUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.filter().order_by('country')
    country_codes   = SpCountryCodes.objects.filter(status=1)

    context = {}
    context['contact_types']    = contact_types
    context['countries']        = countries
    context['country_codes']    = country_codes

    template = 'user-management/add-user-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password    

            error_count = 0
            if request.POST['last_user_id'] != '' and request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            elif request.POST['official_email'] !='':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            else:
                user_exists = 0

            if user_exists:
                error_count = 1
                error_response['official_email_error'] = "Email already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if bool(request.FILES.get('store_image', False)) == True:
                    if request.POST['previous_store_image'] != '':
                        deleteMediaFile(request.POST['previous_store_image'])
                    uploaded_store_image = request.FILES['store_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    store_image_name = uploaded_store_image.name
                    temp = store_image_name.split('.')
                    store_image_name = 'store_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    store_image = storage.save(store_image_name, uploaded_store_image)
                    store_image = storage.url(store_image)

                else:
                    if request.POST['previous_store_image'] != '':
                        store_image = request.POST['previous_store_image'] 
                    else:
                        store_image = None    
                
                if bool(request.FILES.get('profile_image', False)) == True:
                    if request.POST['previous_profile_image'] != '':
                        deleteMediaFile(request.POST['previous_profile_image'])
                    uploaded_profile_image = request.FILES['profile_image']
                    storage = FileSystemStorage()
                    timestamp = int(time.time())
                    profile_image_name = uploaded_profile_image.name
                    temp = profile_image_name.split('.')
                    profile_image_name = 'profile_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                    
                    profile_image = storage.save(profile_image_name, uploaded_profile_image)
                    profile_image = storage.url(profile_image)
                else:
                    if request.POST['previous_profile_image'] != '':
                        profile_image = request.POST['previous_profile_image'] 
                    else:
                        profile_image = None        
                
                if request.POST['last_user_id'] != '':
                    SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                    SpContactPersons.objects.filter(user_id=request.POST['last_user_id']).delete()

                    user = SpUsers.objects.get(id=request.POST['last_user_id'])
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = request.POST['last_user_id']
                    user_inserted = 1
                else:
                    user = SpUsers()
                    user.store_name     = request.POST['store_name']
                    user.salutation     = request.POST['salutation']
                    user.store_image    = store_image
                    user.first_name     = request.POST['first_name']
                    user.middle_name    = request.POST['middle_name']
                    user.profile_image  = profile_image
                    user.last_name      = request.POST['last_name']
                    user.official_email = request.POST['official_email']
                    user.self_owned     = request.POST['self_owned']
                    user.password       = make_password(str(password))
                    user.plain_password = str(password)
                    if request.POST['store_name']:
                        user.user_type = 2
                    user.save()
                    last_user_id = user.id
                    user_inserted = 0
                    #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    user_contact_no         = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

            
                contact_person_names = request.POST.getlist('contact_person_name[]')
                designations = request.POST.getlist('designation[]')
                contact_numbers = request.POST.getlist('contact_number[]')

                for id, val in enumerate(contact_person_names):
                    user_contact_person         = SpContactPersons()
                    user_contact_person.user_id = last_user_id
                    if contact_person_names[id]!='':
                        user_contact_person.contact_person_name = contact_person_names[id]
                    if designations[id]!='':
                        user_contact_person.designation         = designations[id]
                    if contact_numbers[id]!='':
                        user_contact_person.contact_number      = contact_numbers[id]      
                    user_contact_person.save()

                if request.POST['last_user_id'] != '':
                    basic = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save()
                else:
                    basic = SpBasicDetails()
                    basic.user_id       = last_user_id
                    basic.father_name   = request.POST['father_name']
                    basic.mother_name   = request.POST['mother_name']
                    basic.gender        = request.POST['user_gender']
                    basic.date_of_birth = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    basic.blood_group   = request.POST['blood_group']
                    basic.save() 

                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(SpCountries, request.POST['store_country_id'],'country')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(SpStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(SpCities, request.POST['store_city_id'],'city')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id           = last_user_id
                permanent.type              = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(SpCountries, request.POST['permanent_country_id'],'country')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(SpStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(SpCities, request.POST['permanent_city_id'],'city')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()

                oganizations    = SpOrganizations.objects.filter(status=1)
                working_shifts  = TblClWorkingShifts.objects.all()
                zones           = SpZones.objects.all()
                routes          = SpRoutes.objects.all()
                user_details    = SpUsers.objects.get(id=last_user_id)

                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=last_user_id)
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None
                
                try:
                    user_basic_details = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    user_basic_details = None
                
                try:
                    departments = SpDepartments.objects.filter(organization_id=user_details.organization_id)
                except SpDepartments.DoesNotExist:
                    departments = None

                try:
                    roles = SpRoles.objects.filter(department_id=user_details.department_id)#.filter(id__in=[8,9])
                except SpRoles.DoesNotExist:
                    roles = None
                
                if user_area_allocations is None:
                    towns = None
                else:
                    towns = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)

                if user_inserted == 0:
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' added by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
                else:
                    user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                    heading     = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated'
                    activity    = request.POST['first_name']+' '+request.POST['middle_name']+' '+request.POST['last_name']+' '+' updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                    
                    saveActivity('Users Management', 'User', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
                
                context = {}
                context['oganizations']             = oganizations
                context['working_shifts']           = working_shifts
                context['zones']                    = zones
                context['routes']                   = routes
                context['user_details']             = user_details
                context['user_area_allocations']     = user_area_allocations
                context['user_basic_details']       = user_basic_details
                context['departments']              = departments
                context['roles']                    = roles
                context['towns']                    = towns
                context['last_user_id']             = last_user_id
                
                template = 'user-management/add-user-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template, context)

# Edit User basic details View
@login_required
def editUserBasicDetail(request):
    contact_types   = SpContactTypes.objects.filter(status=1)
    countries       = SpCountries.objects.filter().order_by('country')
    country_codes   = SpCountryCodes.objects.filter(status=1)

    try:
        user_basic_details = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    except SpBasicDetails.DoesNotExist:
        user_basic_details = None

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_contact_details    = SpContactNumbers.objects.filter(user_id=request.GET['last_user_id'])
    user_basic_details      = user_basic_details
    user_store_address      = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='correspondence').first()
    user_permanent_address  = SpAddresses.objects.filter(user_id=request.GET['last_user_id']).filter(type='permanent').first()
    user_contact_person     = SpContactPersons.objects.filter(user_id=request.GET['last_user_id'])

    if user_store_address:
        store_states = SpStates.objects.filter(country_id=user_store_address.country_id)
        store_cities = SpCities.objects.filter(state_id=user_store_address.state_id)
        if user_permanent_address:
            permanent_states = SpStates.objects.filter(country_id=user_permanent_address.country_id)
            permanent_cities = SpCities.objects.filter(state_id=user_permanent_address.state_id)
        else:
            permanent_states = None
            permanent_cities = None    
    else:
        store_states = None
        store_cities = None
        permanent_states = None
        permanent_cities = None

    context = {}
    context['contact_types']            = contact_types
    context['countries']                = countries
    context['country_codes']            = country_codes
    context['user_details']             = user_details
    context['user_contact_details']     = user_contact_details
    context['user_basic_details']       = user_basic_details
    context['user_store_address']       = user_store_address
    context['user_permanent_address']   = user_permanent_address
    context['user_contact_person']      = user_contact_person
    context['store_states']             = store_states
    context['store_cities']             = store_cities
    context['permanent_states']         = permanent_states
    context['permanent_cities']         = permanent_cities
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-basic-detail.html'
    
    return render(request, template, context)    
    
    
    


# User offical details View
# @login_required
# def addUserOfficalDetail(request):
#     template = 'user-management/add-user-offical-detail.html'
#     response = {}
#     error_response = {}
#     if request.method == "POST":
#         try:
#             error_count = 0
#             if request.POST['last_user_id'] != '':
#                 emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
#             else:
#                 emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
#             if emp_sap_id_exists:
#                 error_count = 1
#                 error_response['emp_sap_id_error'] = "SAP ID already exists"

#             if(error_count > 0):
#                 response['error'] = True
#                 response['message'] = error_response

#                 return JsonResponse(response)
#             else:
#                 user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

#                 # update user role & map product
#                 if user_data.role_id is None :
#                     updateUserRole(user_data.id,request)
#                     mapProductToUser(user_data.id)
#                 if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
#                     updateUserRole(user_data.id,request)

#                 user_data.organization_id       = request.POST['organization_id']
#                 user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
#                 user_data.department_id         = request.POST['department_id']
#                 user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
#                 user_data.role_id               = request.POST['role_id']
#                 user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
#                 user_data.emp_sap_id            = request.POST['emp_sap_id']
#                 if request.POST['role_id'] == '8':
#                     user_data.is_distributor = 1
#                     user_data.is_super_stockist = 0
#                 else:
#                     user_data.is_distributor = 0
#                     user_data.is_super_stockist = 1    
#                 user_data.save()

                

#                 try:
#                     user_basic_detail = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
#                 except SpBasicDetails.DoesNotExist:
#                     user_basic_detail = None

#                 if user_basic_detail is None:        
#                     user_basic_details                  = SpBasicDetails()
#                 else:
#                     user_basic_details                  = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])

#                 user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
#                 user_basic_details.pan_number           = request.POST['pan_number']
#                 user_basic_details.cin                  = request.POST['cin']
#                 user_basic_details.gstin                = request.POST['gstin']
#                 user_basic_details.fssai                = request.POST['fssai']
#                 user_basic_details.working_shift_id     = request.POST['working_shift_id']
#                 user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
#                 user_basic_details.order_timing         = request.POST['order_timing']
#                 user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
#                 user_basic_details.outstanding_amount   = request.POST['outstanding_amount']
#                 user_basic_details.security_amount      = request.POST['security_amount']
#                 user_basic_details.opening_crates       = request.POST['opening_crates']
#                 user_basic_details.save()
                
#                 try:
#                     user_area_allocations = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])
#                 except SpUserAreaAllocations.DoesNotExist:
#                     user_area_allocations = None

#                 if user_area_allocations is None:        
#                     area_allocation = SpUserAreaAllocations()
#                 else:
#                     area_allocation = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])

#                 area_allocation.user_id     = request.POST['last_user_id']
#                 area_allocation.state_id     = getModelColumnById(SpZones,request.POST['zone_id'],'state_id')
#                 area_allocation.state_name   = getModelColumnById(SpZones,request.POST['zone_id'],'state_name')
#                 area_allocation.zone_id     = request.POST['zone_id']
#                 area_allocation.zone_name   = getModelColumnById(SpZones,request.POST['zone_id'],'zone')
#                 area_allocation.town_id     = request.POST['town_id']
#                 area_allocation.town_name   = getModelColumnById(SpTowns,request.POST['town_id'],'town')
#                 area_allocation.route_id    = request.POST['route_id']
#                 area_allocation.route_name  = getModelColumnById(SpRoutes,request.POST['route_id'],'route')
#                 area_allocation.save()
                
#                 try:
#                     user_variants = SpUserProductVariants.objects.filter(user_id=request.POST['last_user_id'])
#                 except SpUserProductVariants.DoesNotExist:
#                     user_variants = None

#                 user_details = SpUsers.objects.filter(id=request.POST['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]
                    
#                 context                  = {}
#                 context['last_user_id']  = request.POST['last_user_id']
#                 context['user_variants'] = user_variants
#                 context['user_details']  = user_details
#                 template                 = 'user-management/add-user-product-detail.html'
#                 return render(request, template, context)
#         except Exception as e:
#             response['error'] = True
#             response['message'] = e
#             return HttpResponse(e)
#     return render(request, template)

# User offical details View
@login_required
def addUserOfficalDetail(request):
    template = 'user-management/add-user-offical-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
            else:
                emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "SAP ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role & map product
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                    mapProductToUser(user_data.id)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                if request.POST['role_id'] == '8':
                    user_data.is_distributor = 1
                    user_data.is_super_stockist = 0
                else:
                    user_data.is_distributor = 0
                    user_data.is_super_stockist = 1    
                user_data.save()

                

                try:
                    user_basic_detail = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                except SpBasicDetails.DoesNotExist:
                    user_basic_detail = None

                if user_basic_detail is None:        
                    user_basic_details                  = SpBasicDetails()
                else:
                    user_basic_details                  = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])

                user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                user_basic_details.pan_number           = request.POST['pan_number']
                user_basic_details.cin                  = request.POST['cin']
                user_basic_details.gstin                = request.POST['gstin']
                user_basic_details.fssai                = request.POST['fssai']
                user_basic_details.working_shift_id     = request.POST['working_shift_id']
                user_basic_details.working_shift_name   = getModelColumnById(SpWorkingShifts,request.POST['working_shift_id'],'working_shift')
                user_basic_details.order_timing         = request.POST['order_timing']
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_basic_details.outstanding_amount   = request.POST['outstanding_amount']
                user_basic_details.security_amount      = request.POST['security_amount']
                user_basic_details.opening_crates       = request.POST['opening_crates']
                user_basic_details.save()
                
                try:
                    user_area_allocations = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])
                except SpUserAreaAllocations.DoesNotExist:
                    user_area_allocations = None

                if user_area_allocations is None:        
                    area_allocation = SpUserAreaAllocations()
                else:
                    area_allocation = SpUserAreaAllocations.objects.get(user_id=request.POST['last_user_id'])

                area_allocation.user_id     = request.POST['last_user_id']
                area_allocation.state_id     = getModelColumnById(SpZones,request.POST['zone_id'],'state_id')
                area_allocation.state_name   = getModelColumnById(SpZones,request.POST['zone_id'],'state_name')
                area_allocation.zone_id     = request.POST['zone_id']
                area_allocation.zone_name   = getModelColumnById(SpZones,request.POST['zone_id'],'zone')
                area_allocation.town_id     = request.POST['town_id']
                area_allocation.town_name   = getModelColumnById(SpTowns,request.POST['town_id'],'town')
                area_allocation.route_id    = request.POST['route_id']
                area_allocation.route_name  = getModelColumnById(SpRoutes,request.POST['route_id'],'route')
                area_allocation.save()
                
                try:
                    user_variants = SpUserProductVariants.objects.filter(user_id=request.POST['last_user_id'])
                except SpUserProductVariants.DoesNotExist:
                    user_variants = None

                user_details = SpUsers.objects.filter(id=request.POST['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]
                    
                context                  = {}
                context['last_user_id']  = request.POST['last_user_id']
                context['user_variants'] = user_variants
                context['user_details']  = user_details
                template                 = 'user-management/add-user-product-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)

# Edit User basic details View
@login_required
def editUserOfficalDetail(request):
    oganizations    = SpOrganizations.objects.filter(status=1)
    working_shifts  = TblClWorkingShifts.objects.all()
    zones           = SpZones.objects.all()
    routes          = SpRoutes.objects.all()

    user_details            = SpUsers.objects.get(id=request.GET['last_user_id'])
    user_area_allocations   = SpUserAreaAllocations.objects.get(user_id=request.GET['last_user_id'])
    user_basic_details      = SpBasicDetails.objects.get(user_id=request.GET['last_user_id'])
    
    departments     = SpDepartments.objects.filter(organization_id=user_details.organization_id)
    roles           = SpRoles.objects.filter(department_id=user_details.department_id)#.filter(id__in=[8,9])
    towns           = SpTowns.objects.filter(zone_id=user_area_allocations.zone_id)
    
    context = {}
    context['oganizations']             = oganizations
    context['working_shifts']           = working_shifts
    context['zones']                    = zones
    context['routes']                   = routes
    context['user_details']             = user_details
    context['user_area_allocations']     = user_area_allocations
    context['user_basic_details']       = user_basic_details
    context['departments']              = departments
    context['roles']                    = roles
    context['towns']                    = towns
    context['last_user_id']             = request.GET['last_user_id']
    template = 'user-management/add-user-offical-detail.html'
    
    return render(request, template, context)      

# User product details View
@login_required
def addUserProductDetail(request):
    context = {}
    template = 'user-management/add-user-product-detail.html'
    if request.method == "POST":
        try:
            user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
        except SpUserDocuments.DoesNotExist:
            user_documents = None
        context['user_documents']   = user_documents
        context['last_user_id']     = request.POST['last_user_id']
        template = 'user-management/add-user-document-detail.html'
    return render(request, template, context)

# User product details View
@login_required
def editUserProductDetail(request):
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=request.GET['last_user_id'])
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    user_details = SpUsers.objects.filter(id=request.GET['last_user_id']).values('is_distributor', 'is_super_stockist', 'user_type')[0]

    context = {}
    context['last_user_id']  = request.GET['last_user_id']
    context['user_variants'] = user_variants
    context['user_details']  = user_details
    template = 'user-management/add-user-product-detail.html'
    return render(request, template, context) 

# User document details View
@login_required
def addUserDocumentDetail(request):
    template = 'user-management/add-user-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                if request.POST['previous_aadhaar_card'] != '':
                        deleteMediaFile(request.POST['previous_aadhaar_card'])
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                aadhaar_card_name = uploaded_aadhaar_card.name
                temp = aadhaar_card_name.split('.')
                aadhaar_card_name = 'aadhaar_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                aadhaar_card = storage.save(aadhaar_card_name, uploaded_aadhaar_card)
                aadhaar_card = storage.url(aadhaar_card)
            else:
                if request.POST['previous_aadhaar_card'] != '':
                        aadhaar_card = request.POST['previous_aadhaar_card'] 
                else:
                    aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:
                if request.POST['previous_pan_card'] != '':
                        deleteMediaFile(request.POST['previous_pan_card'])        
                uploaded_pan_card = request.FILES['pan_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                pan_card_name = uploaded_pan_card.name
                temp = pan_card_name.split('.')
                pan_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                pan_card = storage.save(pan_card_name, uploaded_pan_card)
                pan_card = storage.url(pan_card)
            else:
                if request.POST['previous_pan_card'] != '':
                        pan_card = request.POST['previous_pan_card'] 
                else:
                    pan_card = None

            if bool(request.FILES.get('cin', False)) == True:
                if request.POST['previous_cin'] != '':
                        deleteMediaFile(request.POST['previous_cin'])
                uploaded_cin = request.FILES['cin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                cin_name = uploaded_cin.name
                temp = cin_name.split('.')
                cin_name = 'cin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                cin = storage.save(cin_name, uploaded_cin)
                cin = storage.url(cin)
            else:
                if request.POST['previous_cin'] != '':
                        cin = request.POST['previous_cin'] 
                else:
                    cin = None

            if bool(request.FILES.get('gstin', False)) == True:
                if request.POST['previous_gstin'] != '':
                        deleteMediaFile(request.POST['previous_gstin'])
                uploaded_gstin = request.FILES['gstin']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                gstin_name = uploaded_gstin.name
                temp = gstin_name.split('.')
                gstin_name = 'gstin_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                gstin = storage.save(gstin_name, uploaded_gstin)
                gstin = storage.url(gstin)
            else:
                if request.POST['previous_gstin'] != '':
                        gstin = request.POST['previous_gstin'] 
                else:
                    gstin = None

            if bool(request.FILES.get('fssai', False)) == True:
                if request.POST['previous_fssai'] != '':
                        deleteMediaFile(request.POST['previous_fssai'])
                uploaded_fssai = request.FILES['fssai']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                fssai_name = uploaded_fssai.name
                temp = fssai_name.split('.')
                fssai_name = 'fssai_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                fssai = storage.save(fssai_name, uploaded_fssai)
                fssai = storage.url(fssai)
            else:
                if request.POST['previous_fssai'] != '':
                        fssai = request.POST['previous_fssai'] 
                else:
                    fssai = None        
            
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None

            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.aadhaar_card  = aadhaar_card
                documents.pan_card      = pan_card
                documents.cin           = cin
                documents.gstin         = gstin
                documents.fssai         = fssai
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error']            = True
            response['message']          = e
            return HttpResponse(e)
    return render(request, template)
    

# Employee Payroll Master
@login_required
def saveEmployeePayrollMaster(request):
    if request.method == "POST":
        emp_ctc         = request.POST['emp_ctc']
        gross_salary    = request.POST['gross_salary']
        emp_hra         = request.POST['emp_hra']
        emp_ta          = request.POST['emp_ta']
        emp_da          = request.POST['emp_da']
        emp_pf          = request.POST['emp_pf']
        emp_tds         = request.POST['emp_tds']
        employee_id     = request.POST['employee_id']
        
        emp_payroll                 = SpEmployeePayrollMaster()
        emp_payroll.user_id         = employee_id
        emp_payroll.emp_ctc         = emp_ctc
        emp_payroll.gross_salary    = gross_salary
        emp_payroll.emp_hra         = emp_hra
        emp_payroll.emp_ta          = emp_ta
        emp_payroll.emp_da          = emp_da
        emp_payroll.emp_tds         = emp_tds
        emp_payroll.emp_pf          = emp_pf
        emp_payroll.save()
        
        message  = "Employee Payroll Added Successfully"
        response = {}
        response['error'] = False
        response['message'] = message
        return JsonResponse(response)

@login_required
def updateEmployeePayrollMaster(request):
    if request.method == "POST":
        emp_ctc         = request.POST['emp_ctc']
        gross_salary    = request.POST['gross_salary']
        emp_hra         = request.POST['emp_hra']
        emp_ta          = request.POST['emp_ta']
        emp_da          = request.POST['emp_da']
        emp_pf          = request.POST['emp_pf']
        emp_tds         = request.POST['emp_tds']
        employee_id     = request.POST['employee_id']
        
        emp_payroll                 = SpEmployeePayrollMaster.objects.get(user_id = employee_id)
        emp_payroll.user_id         = employee_id
        emp_payroll.emp_ctc         = emp_ctc
        emp_payroll.gross_salary    = gross_salary
        emp_payroll.emp_hra         = emp_hra
        emp_payroll.emp_ta          = emp_ta
        emp_payroll.emp_da          = emp_da
        emp_payroll.emp_tds         = emp_tds
        emp_payroll.emp_pf          = emp_pf
        emp_payroll.save()
        
        message  = "Employee Payroll Updated Successfully"
        response = {}
        response['error'] = False
        response['message'] = message
        return JsonResponse(response)



# Employee basic details View
@login_required
def addEmployeeBasicDetail(request):
    contact_types = SpContactTypes.objects.filter(status=1)
    countries = TblCountry.objects.filter().order_by('country_name')
    country_codes   = SpCountryCodes.objects.filter(status=1)
    #print(countries)
    context = {}
    context['contact_types'] = contact_types
    context['countries']     = countries
    context['country_codes'] = country_codes
    template = 'user-management/add-employee-basic-detail.html'
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            password = '123456'
            user_context = {}
            user_context['first_name']      = request.POST['first_name']
            user_context['middle_name']     = request.POST['middle_name']
            user_context['last_name']       = request.POST['last_name']
            user_context['official_email']  = request.POST['official_email']
            user_context['password']        = password
            
            error_count = 0 
            user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
           
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
                return JsonResponse(response)
            else:
                
                user = SpUsers()
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.user_type = 1
                user.password       = make_password(str(password))
                user.plain_password = str(password)
                user.save()

                if request.POST['previous_profile_image']:
                    profile_image = request.POST['previous_profile_image']
                    im = Image.open(BytesIO(base64.b64decode(profile_image.split(",")[1])))
                    im.save("media/profileImage/employee_photo_"+str(user.id)+".png", 'PNG')
                    filePath="media/profileImage/employee_photo_" +str(user.id) + ".png"
                    profile_image = filePath 
                else:
                    profile_image = None
                

                user.profile_image = profile_image
                user.save()
                last_user_id = user.id
                #sendEmail(request, 'user-management/email.html', user_context, 'Welcome to Sales Port', request.POST['official_email'])
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                contact_nos         = request.POST.getlist('contact_no[]')
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()

                basic = SpBasicDetails()
                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

            
                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(TblCountry, request.POST['store_country_id'],'country_name')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(TblStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(TblNewDistrict, request.POST['store_city_id'],'district_name')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(TblCountry, request.POST['permanent_country_id'],'country_name')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(TblStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(TblNewDistrict, request.POST['permanent_city_id'],'district_name')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()


                oganizations                = SpOrganizations.objects.filter(status=1)
                working_shifts              = TblClWorkingShifts.objects.all()
                # Emp Sap Id Code -----------------
                last_user_id_id = SpUsers.objects.filter().last()
                last_user_id_num = last_user_id_id.id
                incremented_id = last_user_id_num + 1
                formatted_number = str(incremented_id).zfill(4)
                
                user_name_slag = Configuration.objects.filter().first()
                genrated_emp_id = f"{user_name_slag.user_name}{formatted_number}"
                # --------------------------
                context                     = {}
                context['banks'] =   SpBanks.objects.all()
                context['genrated_emp_id'] = genrated_emp_id
                contract_type    = ContractType.objects.all()
                context['working_shifts']   = working_shifts
                context['contract_type']    = contract_type
                context['oganizations']     = oganizations
                
                context['last_user_id']     = last_user_id

                template = 'user-management/add-employee-offical-detail.html'
                return render(request, template, context)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))
    return render(request, template, context)



    
def mapUserLeaves(role_id,user_id):
    try:
        leave_policy_id = SpRoleEntityMapping.objects.get(role_id = role_id , entity_type = "leave_policy")
        leave_policy_id = leave_policy_id.entity_id
    except SpRoleEntityMapping.DoesNotExist:
        leave_policy_id = None
    if leave_policy_id:
        leave_policy_dettail = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id)
        for policy_detail in leave_policy_dettail:
            leave_polcy_ledger = SpUserLeavePolicyLedger()
            leave_polcy_ledger.user_id = user_id
            leave_polcy_ledger.leave_policy_id = leave_policy_id
            leave_polcy_ledger.leave_type_id = policy_detail.leave_type_id
            current_month = datetime.today().strftime('%m')
            if int(current_month) == 1:
                month_leave_count = policy_detail.year_leave_count / 12
                leave_polcy_ledger.year_leave_count = policy_detail.year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
            else:   
                current_month = 12 - int(current_month)
                sub_leave_count = policy_detail.year_leave_count / 12
                year_leave_count = sub_leave_count*current_month
                month_leave_count = year_leave_count  / current_month
                leave_polcy_ledger.year_leave_count = year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
            
            leave_polcy_ledger.consecutive_leave = policy_detail.consecutive_leave 
            leave_polcy_ledger.save()
  


# @login_required
# def addEmployeeOfficalDetail(request):
#     template = 'user-management/add-employee-offical-detail.html'
    
#     response = {}
#     error_response = {}

#     if request.method == "POST":
#         try:
#             error_count = 0
#             emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
#             if emp_sap_id_exists:
#                 error_count = 1
#                 error_response['emp_sap_id_error'] = "Employee ID already exists"

#             if(error_count > 0):
#                 response['error'] = True
#                 response['message'] = error_response

#                 return JsonResponse(response)
#             else: 
#                 user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

#                 # update user role
#                 if user_data.role_id is None :
#                     updateUserRole(user_data.id,request)
#                 if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
#                     updateUserRole(user_data.id,request)

#                 user_data.organization_id       = request.POST['organization_id']
                
#                 user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
#                 user_data.department_id         = request.POST['department_id']
#                 user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
#                 user_data.role_id               = request.POST['role_id']
#                 user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
#                 user_data.emp_sap_id            = request.POST['emp_sap_id']
#                 #user_data.login_status         = request.POST['login_status']
                
#                 if request.POST['reporting_to_id']:
#                     user_data.reporting_to_id   = request.POST['reporting_to_id']
#                     user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

#                 user_data.save()
#                 user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                
#                 user_basic_details.contract_type          = request.POST['contract_type']
#                 user_basic_details.is_esic                = request.POST['is_esic']
               
#                 user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                
#                 user_basic_details.week_of_day                = request.POST['week_of_day']
               
#                 user_basic_details.save()
                
#                 role_id = request.POST['role_id']
#                 if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
#                     mapUserLeaves(role_id,request.POST['last_user_id'])
                
                
#                 role_id = request.POST['role_id']
#                 if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
#                     mapUserLeaves(role_id,request.POST['last_user_id'])
                    
                    
#                 if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
#                     banks = SpBankDetails()
#                     banks.user_id = request.POST['last_user_id']
#                     banks.bank_id = request.POST['bank_name']
#                     banks.bank_name = getModelColumnById(
#                         SpBanks, request.POST['bank_name'], 'bank_name')
#                     banks.bank_account_no = request.POST['bank_account_no']
#                     banks.ifsc_code = request.POST['ifsc_code']
#                     banks.bank_address = request.POST['bank_address']
#                     banks.save()
                

#                 context = {}
#                 context['last_user_id'] = request.POST['last_user_id']
#                 context['working_shifts'] = TblClWorkingShifts.objects.all()
                
#                 distributors = 0
                
    
#                 contract_type    = ContractType.objects.all()
#                 context['contract_type']    = contract_type
             
#                 if distributors:
#                     context['distributors'] = distributors
#                 else:
#                     context['distributors'] = None
#                 template = 'user-management/add-employee-biometric-details.html'
#                 return render(request, template, context)
#         except Exception as e:
#             response['error'] = True
#             response['message'] = e
#             return HttpResponse(e)
#     return render(request, template,response)


@login_required
def addEmployeeOfficalDetail(request):
    template = 'user-management/add-employee-offical-detail.html'
    
    response = {}
    error_response = {}
    
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['emp_sap_id_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response
                return JsonResponse(response)
            else: 
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)

                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name') 
                user_data.emp_sap_id            = request.POST['emp_sap_id']
                #user_data.login_status         = request.POST['login_status']
                
                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()
                user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                
                user_basic_details.contract_type          = request.POST['contract_type']
                user_basic_details.is_esic                = request.POST['is_esic']
               
                user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                
                user_basic_details.week_of_day                = request.POST['week_of_day']

               
                user_basic_details.save()
                
                role_id = request.POST['role_id']
                if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
                    mapUserLeaves(role_id,request.POST['last_user_id'])
                
                
                role_id = request.POST['role_id']
                if SpRoleEntityMapping.objects.filter(role_id = role_id , entity_type = "leave_policy").exists():
                    mapUserLeaves(role_id,request.POST['last_user_id'])
                    
                    
                if request.POST['bank_name'] != '' and request.POST['bank_account_no']:
                    banks = SpBankDetails()
                    banks.user_id = request.POST['last_user_id']
                    banks.bank_id = request.POST['bank_name']
                    banks.bank_name = getModelColumnById(SpBanks, request.POST['bank_name'], 'bank_name')
                    banks.bank_account_no = request.POST['bank_account_no']
                    banks.ifsc_code = request.POST['ifsc_code']
                    banks.bank_address = request.POST['bank_address']
                    banks.save()
                

                context = {}
                context['last_user_id'] = request.POST['last_user_id']
                context['working_shifts'] = TblClWorkingShifts.objects.all()
                
                distributors = 0
                
                if distributors:
                    context['distributors'] = distributors
                else:
                    context['distributors'] = None
                template = 'user-management/add-employee-biometric-details.html'
                return render(request, template, context)
        except Exception as e:
            response = {}
            response['error'] = True
            response['message'] = str(e)
            return JsonResponse(response)
    return render(request, template,response)
# Employee attendance details View
@login_required
def addEmployeeAttendanceDetail(request):
    if request.method == "POST":

        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = request.POST['last_user_id']
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()

        context = {}
        context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-document-detail.html'
        return render(request, template, context)
    else:
        context = {}
        distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
            from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        if distributors:
            context['distributors'] = distributors
        else:
            context['distributors'] = None
            context['last_user_id'] = request.POST['last_user_id']
        template = 'user-management/add-employee-attendance.html'
        return render(request, template, context)

    

# Employee document details View
@login_required
def addEmployeeDocumentDetail(request):
    template = 'user-management/add-employee-document-detail.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('aadhaar_card', False)) == True:
                uploaded_aadhaar_card = request.FILES['aadhaar_card']
                aadhaar = FileSystemStorage()
                aadhaar_card = aadhaar.save(uploaded_aadhaar_card.name, uploaded_aadhaar_card)
                aadhaar_card = aadhaar.url(aadhaar_card)
            else:
                aadhaar_card = None
                
            if bool(request.FILES.get('pan_card', False)) == True:        
                uploaded_pan_card = request.FILES['pan_card']
                pan = FileSystemStorage()
                pan_card = pan.save(uploaded_pan_card.name, uploaded_pan_card)
                pan_card = pan.url(pan_card)
            else:
                pan_card = None

            documents               = SpUserDocuments()
            documents.user_id       = request.POST['last_user_id']
            documents.aadhaar_card  = aadhaar_card
            documents.pan_card      = pan_card
            documents.save()
            
            #-----------------------------notify android block-------------------------------#
            organization_id = getModelColumnById(SpUsers, request.POST['last_user_id'], 'organization_id')
            all_users = SpUsers.objects.filter(organization_id = organization_id).exclude(id=request.POST['last_user_id']).values_list("id", flat=True)
            for each_user in all_users:
                user_id = each_user
                if SpUsers.objects.filter(id=user_id, firebase_token__isnull=False).exists():
                    userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
                    employee_name = getUserName(request.user.id)
    
                    message_title = document.document_name+" document scan request."
                    message_body =  document.document_name+" document scan request has been generated by "+employee_name
                    notification_image = ""
        
                    if userFirebaseToken is not None and userFirebaseToken != "" :
                        registration_ids = []
                        registration_ids.append(userFirebaseToken)
                        data_message = {}
                        data_message['id'] = 1
                        data_message['status'] = 'notification'
                        data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image'] = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

            response['error'] = False
            response['message'] = "Record has been updated successfully."

            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)


#csrf_exempt
@login_required
def addEmployeePhoto(request):
    context = {}
    return render(request,"user-management/add-employee-photo.html",context)



@login_required
def editEmployeeBasicDetail(request, employee_id):
    response = {}
    error_response = {}
    if request.method == "POST":
        try:
            error_count = 0
            if request.POST['last_user_id'] != '':
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exclude(id=request.POST['last_user_id']).exists()
            else:
                user_exists = SpUsers.objects.filter(official_email=request.POST['official_email']).exists()
  
            if user_exists:
                error_count = 1
                error_response['emailId_error'] = "Email already exists"
            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:
                if request.POST['previous_profile_image']:
                    if not SpUsers.objects.filter(id = request.POST['last_user_id'], profile_image = request.POST['previous_profile_image']).exists():
                        profile_image = request.POST['previous_profile_image']
                        im = Image.open(
                            BytesIO(base64.b64decode(profile_image.split(",")[1])))
                        im.save("media/profileImage/employee_photo_" +
                                str(employee_id)+".png", 'PNG')
                        filePath = "media/profileImage/employee_photo_" + \
                            str(employee_id) + ".png"
                        profile_image = filePath
                    else:
                        profile_image = request.POST['previous_profile_image']
                else:
                    profile_image = None

                SpAddresses.objects.filter(user_id=request.POST['last_user_id']).delete()
                SpContactNumbers.objects.filter(user_id=request.POST['last_user_id']).delete()
                contact_nos         = request.POST.getlist('contact_no[]')
                user = SpUsers.objects.get(id=request.POST['last_user_id'])
                user.salutation = request.POST['salutation']
                user.first_name = request.POST['first_name']
                user.middle_name = request.POST['middle_name']
                user.profile_image = profile_image
                user.last_name = request.POST['last_name']
                user.official_email = request.POST['official_email']
                user.primary_contact_number = contact_nos[0]
                user.save()
                last_user_id = request.POST['last_user_id']

            
                country_codes       = request.POST.getlist('country_code[]') 
                contact_person_name = request.POST.getlist('contact_person_name[]')
                contact_types       = request.POST.getlist('contact_type[]')
                
                is_primary          = request.POST.getlist('primary_contact[]')

                for id, val in enumerate(contact_nos):
                    if is_primary[id] == '1':
                        user_data = SpUsers.objects.get(id=last_user_id)
                        user_data.primary_contact_number = contact_nos[id]
                        user_data.save()

                    user_contact_no = SpContactNumbers()
                    user_contact_no.user_id = last_user_id
                    if country_codes[id]!='':
                        user_contact_no.country_code = country_codes[id]
                    if contact_types[id]!='':    
                        user_contact_no.contact_type = contact_types[id]
                        user_contact_no.contact_type_name = getModelColumnById(SpContactTypes,contact_types[id],'contact_type')
                    if contact_nos[id]!='':    
                        user_contact_no.contact_number = contact_nos[id]
                    if is_primary[id]!='':    
                        user_contact_no.is_primary = is_primary[id]
                    user_contact_no.save()
                try:
                    basic   = SpBasicDetails.objects.get(user_id=last_user_id)
                except SpBasicDetails.DoesNotExist:
                    basic  = None
                if basic:
                    basic                   = SpBasicDetails.objects.get(user_id=last_user_id)
                else:
                    basic                   = SpBasicDetails()

                basic.user_id               = last_user_id
                basic.father_name           = request.POST['father_name']
                basic.mother_name           = request.POST['mother_name']
                basic.gender                = request.POST['user_gender']
                basic.date_of_birth         = datetime.strptime(request.POST['date_of_birth'], '%d/%m/%Y').strftime('%Y-%m-%d')
                basic.blood_group           = request.POST['blood_group']
                basic.save()

                correspondence = SpAddresses()
                correspondence.user_id          = last_user_id
                correspondence.type             = 'correspondence'
                correspondence.address_line_1   = request.POST['store_address_line_1']
                correspondence.address_line_2   = request.POST['store_address_line_2']
                correspondence.country_id       = request.POST['store_country_id']
                correspondence.country_name     = getModelColumnById(TblCountry, request.POST['store_country_id'],'country_name')
                correspondence.state_id         = request.POST['store_state_id']
                correspondence.state_name       = getModelColumnById(TblStates, request.POST['store_state_id'],'state')
                correspondence.city_id          = request.POST['store_city_id']
                correspondence.city_name        = getModelColumnById(TblNewDistrict, request.POST['store_city_id'],'district_name')
                correspondence.pincode          = request.POST['store_pincode']
                correspondence.save()

                permanent = SpAddresses()
                permanent.user_id = last_user_id
                permanent.type = 'permanent'
                permanent.address_line_1    = request.POST['permanent_address_line_1']
                permanent.address_line_2    = request.POST['permanent_address_line_2']
                permanent.country_id        = request.POST['permanent_country_id']
                permanent.country_name      = getModelColumnById(TblCountry, request.POST['permanent_country_id'],'country_name')
                permanent.state_id          = request.POST['permanent_state_id']
                permanent.state_name        = getModelColumnById(TblStates, request.POST['permanent_state_id'],'state')
                permanent.city_id           = request.POST['permanent_city_id']
                permanent.city_name         = getModelColumnById(TblNewDistrict, request.POST['permanent_city_id'],'district_name')
                permanent.pincode           = request.POST['permanent_pincode']
                permanent.save()
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,request.POST['last_user_id'],'firebase_token')
                employee_name = getUserName(request.POST['last_user_id'])

                message_title = "Profile updated"
                message_body = "Your profile has been updated by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(request.POST['last_user_id'],'SpUsers','User Management','Profile updated',message_title,message_body,notification_image,request.user.id,user_name,request.POST['last_user_id'],employee_name,'profile.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['error'] = False
                response['last_user_id'] = last_user_id
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(str(e))

    else : 
        contact_types = SpContactTypes.objects.filter(status=1)
        countries = TblCountry.objects.filter().order_by('country_name')
        country_codes   = SpCountryCodes.objects.filter(status=1)

        context = {}
        context['contact_types'] = contact_types
        context['countries']     = countries
        context['country_codes'] = country_codes

        employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
        sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
        FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
        where sp_users.id = %s''',[employee_id])
        
        try:
            employee_correspondence_address = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
        except SpAddresses.DoesNotExist:
            employee_correspondence_address = None

        try:
            employee_permanent_address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
        except SpAddresses.DoesNotExist:
            employee_permanent_address = None

        try:
            employee_contact_numbers = SpContactNumbers.objects.filter(user_id=employee_id)
        except SpContactNumbers.DoesNotExist:
            employee_contact_numbers = None    

        employee_user_allocation = None

        try:
            if employee_correspondence_address is not None:
                employee_store_states = TblStates.objects.filter(country_id=employee_correspondence_address.country_id).order_by('state')
            else:
                employee_store_states = None    
        except TblStates.DoesNotExist:
            employee_store_states = None    

        try:
            if employee_correspondence_address is not None:
                employee_store_cities = TblNewDistrict.objects.filter(state_id=employee_correspondence_address.state_id).order_by('district_name')
            else:
                employee_store_cities = None
        except TblNewDistrict.DoesNotExist:
            employee_store_cities = None 
        
        try:
            if employee_permanent_address is not None:
                employee_permanent_states = TblStates.objects.filter(country_id=employee_permanent_address.country_id).order_by('state')
            else:
                employee_permanent_states = None    
        except TblStates.DoesNotExist:
            employee_permanent_states = None

        try:
            if employee_permanent_address is not None:
                employee_permanent_cities = TblNewDistrict.objects.filter(state_id=employee_permanent_address.state_id).order_by('district_name')
            else:
                employee_permanent_cities = None    
        except TblNewDistrict.DoesNotExist:
            employee_permanent_cities = None
        try:
            employee_contact_numbers = SpContactNumbers.objects.filter(user_id=employee_id)
        except SpContactNumbers.DoesNotExist:
            employee_contact_numbers = None   

        context['user_contacts']                    = employee_contact_numbers



        if employee:
            context['employee'] = employee[0]
            context['employee_correspondence_address']  = employee_correspondence_address
            context['employee_permanent_address']       = employee_permanent_address
            context['user_contacts']                    = employee_contact_numbers
            context['user_areas']                       = employee_user_allocation
            context['store_states']                     = employee_store_states
            context['store_cities']                     = employee_store_cities
            context['permanent_states']                 = employee_permanent_states
            context['permanent_cities']                 = employee_permanent_cities
            context['last_user_id']                     = employee_id
            try:
                user_documents                              = SpUserDocuments.objects.get(user_id=employee_id)
                context['user_documents'] = user_documents
            except SpUserDocuments.DoesNotExist:
                context['user_documents'] = None

            context['user_attendance_locations'] = None
            template = 'user-management/edit-employee/employee-basic-detail.html'
            return render(request, template, context)
        else:
            return HttpResponse('Employee not found')


@login_required
def editEmployeeOfficalDetail(request,employee_id):
    error_response = {}
    response = {}
    if request.method == "POST":
        try:
            error_count = 0
            emp_sap_id_exists = SpUsers.objects.filter(emp_sap_id=request.POST['emp_sap_id']).exclude(id=request.POST['last_user_id']).exists()
  
            if emp_sap_id_exists:
                error_count = 1
                error_response['SAPID_error'] = "Employee ID already exists"

            if(error_count > 0):
                response['error'] = True
                response['message'] = error_response

                return JsonResponse(response)
            else:  
                user_data                       = SpUsers.objects.get(id=request.POST['last_user_id'])

                # update user role
                if user_data.role_id is None :
                    updateUserRole(user_data.id,request)
                if user_data.role_id is not None and user_data.role_id !=  int(request.POST['role_id']) :
                    updateUserRole(user_data.id,request)
                    
                user_data.organization_id       = request.POST['organization_id']
                
                user_data.organization_name     = getModelColumnById(SpOrganizations,request.POST['organization_id'],'organization_name')
                user_data.department_id         = request.POST['department_id']
                user_data.department_name       = getModelColumnById(SpDepartments,request.POST['department_id'],'department_name')
                user_data.role_id               = request.POST['role_id']
                user_data.role_name             = getModelColumnById(SpRoles,request.POST['role_id'],'role_name')
                user_data.emp_sap_id            = request.POST['emp_sap_id']  
                #user_data.login_status          = request.POST['login_status']

                if request.POST['reporting_to_id']:
                    user_data.reporting_to_id   = request.POST['reporting_to_id']
                    user_data.reporting_to_name = getModelColumnById(SpUsers,request.POST['reporting_to_id'],'first_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'middle_name')+' '+getModelColumnById(SpUsers,request.POST['reporting_to_id'],'last_name') 

                user_data.save()

                if request.POST['last_user_id'] != '':
                    user_basic_details                      = SpBasicDetails.objects.get(user_id=request.POST['last_user_id'])
                    # user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    # user_basic_details.pan_number           = request.POST['pan_number']
                    # user_basic_details.working_shift_id        = request.POST['working_shift_id']
                    user_basic_details.contract_type          = request.POST['contract_type']
                    user_basic_details.is_esic                = request.POST['is_esic']
                    # user_basic_details.geofencing        = request.POST['geofencing_type']
                    # user_basic_details.working_shift_name   = getModelColumnById(TblClWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    user_basic_details.save()
                else:
                    user_basic_details                      = SpBasicDetails()
                    # user_basic_details.aadhaar_nubmer       = request.POST['aadhaar_nubmer']
                    # user_basic_details.pan_number           = request.POST['pan_number']
                    # user_basic_details.working_shift_id        = request.POST['working_shift_id']
                    user_basic_details.contract_type        = request.POST['contract_type']
                    # user_basic_details.working_shift_name   = getModelColumnById(TblClWorkingShifts,request.POST['working_shift_id'],'working_shift')
                    user_basic_details.date_of_joining      = datetime.strptime(request.POST['date_of_joining'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    
                # if request.POST['total_leave']!='' and request.POST['total_leave']:
                       
                #     user_basic_details.leave_count  = request.POST['total_leave']
                # else:
                #     user_basic_details.leave_count = 0 
                user_basic_details.working_state_name       =  None
             
                user_basic_details.working_location        = None

                user_basic_details.pf_no       =  0
             
                user_basic_details.uan         = 0

                user_basic_details.esi_no       =  0
             
                user_basic_details.pan_no        = 0

                user_basic_details.esi_no       =  0
             
                user_basic_details.adhaar_no        = 0
                
                user_basic_details.week_of_day= ','.join([str(elem) for elem in request.POST.getlist('week_of_day')]) 
                
                user_basic_details.save()
                    
                # TblClAllocatedShifts.objects.filter(user_id = request.POST['last_user_id']).delete()
                # for each_shift in request.POST.getlist('working_shift_id[]'):
                #     working_shift = TblClAllocatedShifts()
                #     working_shift.user_id = request.POST['last_user_id']
                #     working_shift.working_shift_id = each_shift
                #     working_shift.save()
                    
               
                try:
                    banks                   =   SpBankDetails.objects.get(user_id = request.POST['last_user_id'])
                    banks.bank_id           =   request.POST['bank_name']
                    banks.bank_name         =   getModelColumnById(SpBanks,request.POST['bank_name'],'bank_name')
                    banks.bank_account_no   =   request.POST['bank_account_no']
                    banks.ifsc_code         =   request.POST['ifsc_code']
                    banks.bank_address      =   request.POST['bank_address']
                    banks.save()
                except SpBankDetails.DoesNotExist:
                    
                    banks                   =   SpBankDetails()
                    banks.user_id           =   request.POST['last_user_id']
                    banks.bank_id           =   request.POST['bank_name']
                    banks.bank_name         =   getModelColumnById(SpBanks,request.POST['bank_name'],'bank_name')
                    banks.bank_account_no   =   request.POST['bank_account_no']
                    banks.ifsc_code         =   request.POST['ifsc_code']
                    banks.bank_address      =   request.POST['bank_address']
                    banks.save()
                
                
                response = {}
                response['error'] = False
                response['last_user_id'] = request.POST['last_user_id']
                return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    else:
        
        template = 'user-management/edit-employee/employee-offical-detail.html'
        context = {}

        employee_details = SpUsers.objects.get(id=employee_id)
        employee_area_allocations = None
        
        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        
        employee_basic_details = SpBasicDetails.objects.filter(user_id=employee_id).first()


        
        if employee_basic_details.week_of_day:
            week_of_day = employee_basic_details.week_of_day.split(',')
        else:
            week_of_day = None
        # working_shifts = TblClWorkingShifts.objects.all()
        departments = SpDepartments.objects.filter(organization_id=employee_details.organization_id)
        try:
           roles = SpRoles.objects.filter(department_id=employee_details.department_id)
        except:
            roles=None

        try:
            if employee_basic_details is not None:
                working_state_name = TblStates.objects.filter(country_id=1).order_by('state')
            else:
                employee_basic_details = None    
        except TblStates.DoesNotExist:
               working_state_name = None

        try:
            if employee_basic_details is not None:
                working_location = TblNewDistrict.objects.filter(state_id=employee_basic_details.working_state_name).order_by('district_name')
            else:
                employee_basic_details = None    
        except TblNewDistrict.DoesNotExist:
            working_location = None


        try:
            reporting_role = SpRoles.objects.get(id = employee_details.role_id)
        except:
            reporting_role = None

        try:
            if reporting_role is not None:
                reporting_users = SpUsers.objects.filter(role_id=reporting_role.reporting_role_id)
            else:
                reporting_users = None
        except SpUsers.DoesNotExist:
            reporting_users = None
                 
        oganizations = SpOrganizations.objects.filter(status=1)

        contract_type   = ContractType.objects.all()

        try:
            employee_bank_details = SpBankDetails.objects.get(user_id=employee_id)
        except SpBankDetails.DoesNotExist:
            employee_bank_details = None
        banks                                   = SpBanks.objects.all().order_by('bank_name')
        # allocat_shift                           = TblClAllocatedShifts.objects.filter(user_id=employee_id).values_list("working_shift_id", flat=True)

        context['banks']                        = banks
        context['employee_bank_details']        = employee_bank_details
        context['oganizations']                 = oganizations
        # context['working_shifts']               = working_shifts
        # context['allocat_shifts']               = allocat_shift
        context['contract_type']                = contract_type
        context['week_of_days']                 = week_of_day
        context['store_states']                 = working_state_name
        context['store_cities']                 = working_location
        context['employee_details']             = employee_details
        context['employee_area_allocations']    = employee_area_allocations
        context['employee_basic_details']       = employee_basic_details
        context['departments']                  = departments
        context['roles']                        = roles
        context['reporting_users']              = reporting_users 
        context['last_user_id']                 = employee_id
        
        return render(request, template,context)




# Employee attendance details View
@login_required
def editEmployeeAttendanceDetail(request,employee_id):
    if request.method == "POST":
        SpUserAttendanceLocations.objects.filter(user_id=request.POST['last_user_id']).delete()
        distributors    = request.POST.getlist('distributor_ss_id[]')
        periphery    = request.POST.getlist('periphery[]')
        timing    = request.POST.getlist('timing[]')
        for id, val in enumerate(distributors):
            
            if distributors[id] != '':
                user_attendance_location = SpUserAttendanceLocations()
                user_attendance_location.user_id = employee_id
                user_attendance_location.attendance_config_id    =   1
                user_attendance_location.distributor_ss_id = distributors[id]
                user_attendance_location.distributor_ss_name     = getModelColumnById(SpUsers,distributors[id],'first_name')
                user_attendance_location.periphery = periphery[id]
                user_attendance_location.timing = timing[id]
                user_attendance_location.status = 1
                user_attendance_location.save()
                
        response = {}
        response['error'] = False
        response['last_user_id'] = request.POST['last_user_id']
        return JsonResponse(response)
    else:
        context = {}
        # user_attendance_locations = SpUserAttendanceLocations.objects.filter(user_id=employee_id)
        # if user_attendance_locations:
        #     user_attendance_locations = user_attendance_locations
        # else:
        #     user_attendance_locations = None
        # context['user_attendance_locations'] = user_attendance_locations
        # distributors = SpUsers.objects.raw(''' select id,first_name, middle_name, last_name 
        #     from sp_users where is_distributor = %s or is_super_stockist = %s ''',[1,1])
        # if distributors:
        #     context['distributors'] = distributors
        # else:
        #     context['distributors'] = None

        # context['last_user_id'] = employee_id
        # context['user_attendance_locations'] = user_attendance_locations
        template = 'user-management/edit-employee/employee-biometric-details.html'
        return render(request, template, context)

# Employee attendance details View
@login_required
def editEmployeeBiometricDetails(request,employee_id):
    if request.method == "POST":
        user                = SpUsers.objects.get(id=request.POST['last_user_id'])
        
        user.periphery      = request.POST['periphery']
        user.save()
        if request.POST['geofencing_type'] != "":
            SpBasicDetails.objects.filter(user_id=employee_id).update(geofencing=request.POST['geofencing_type'])

        for each_shift in request.POST.getlist('working_shift_id[]'):
            SpBasicDetails.objects.filter(user_id=request.POST['last_user_id']).update(working_shift_id=each_shift)
        response = {}
        response['error'] = False
        response['last_user_id'] = request.POST['last_user_id']
        return JsonResponse(response)
    else:
        context = {}
        user_attendance_locations = SpUsers.objects.get(id=employee_id)
        
        context['user_attendance_locations'] = user_attendance_locations
        working_shifts         = TblClWorkingShifts.objects.all()
        allocat_shift          = TblClAllocatedShifts.objects.filter(user_id=employee_id).values_list("working_shift_id", flat=True)
        user                   = SpUsers.objects.get(id=employee_id)
        employee_basic_details = SpBasicDetails.objects.get(user_id=employee_id)
        context['working_shifts']    = working_shifts
        context['allocat_shifts']    = allocat_shift
        
        context['last_user_id']      = employee_id
        context['user']              = user
        context['employee_basic_details'] = employee_basic_details
        context['last_user_id'] = employee_id
        template = 'user-management/edit-employee/employee-biometric-details.html'
        return render(request, template, context)

@login_required
def editEmployeeDocumentDetail(request,employee_id):
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('qatar_id', False)) == True:
                if request.POST['previous_qatar_id']:
                    deleteMediaFile(request.POST['previous_qatar_id'])
                uploaded_qatar_id = request.FILES['qatar_id']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                qatar_id_name = uploaded_qatar_id.name
                temp = qatar_id_name.split('.')
                qatar_id_name = 'qatar_id_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                qatar_id = storage.save(qatar_id_name, uploaded_qatar_id)
                qatar_id = storage.url(qatar_id)
            else:
                if request.POST['previous_qatar_id'] != '':
                    qatar_id = request.POST['previous_qatar_id'] 
                else:
                    qatar_id = None
                
            if bool(request.FILES.get('passport_card', False)) == True:
                if request.POST['previous_passport_card']:
                    deleteMediaFile(request.POST['previous_passport_card'])  
                uploaded_passport_card = request.FILES['passport_card']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                passport_card_name = uploaded_passport_card.name
                temp = passport_card_name.split('.')
                passport_card_name = 'pan_card_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                passport_card = storage.save(passport_card_name, uploaded_passport_card)
                passport_card = storage.url(passport_card)
            else:
                if request.POST['previous_passport_card'] != '':
                    passport_card = request.POST['previous_passport_card'] 
                else:
                    passport_card = None

            if bool(request.FILES.get('resume', False)) == True:
                if request.POST['previous_resume']:
                    deleteMediaFile(request.POST['previous_resume'])  
                uploaded_resume= request.FILES['resume']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                resume_name = uploaded_resume.name
                temp = resume_name.split('.')
                resume_name = 'resume_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                resume = storage.save(resume_name, uploaded_resume)
                resume = storage.url(resume)
            else:
                if request.POST['previous_resume'] != '':
                    resume = request.POST['previous_resume'] 
                else:
                    resume = None
            
            if bool(request.FILES.get('educationaldoc', False)) == True:
                if request.POST['previous_educationaldoc']:
                    deleteMediaFile(request.POST['previous_educationaldoc'])  
                uploaded_educationaldoc = request.FILES['educationaldoc']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                educationaldoc_name = uploaded_educationaldoc.name
                temp = educationaldoc_name.split('.')
                educationaldoc_name = 'educationaldoc_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                educationaldoc = storage.save(educationaldoc_name, uploaded_educationaldoc)
                educationaldoc = storage.url(educationaldoc)
            else:
                if request.POST['previous_educationaldoc'] != '':
                    educationaldoc = request.POST['previous_educationaldoc'] 
                else:
                    educationaldoc = None
                    
            if bool(request.FILES.get('offerletter', False)) == True:
                if request.POST['previous_offerletter']:
                    deleteMediaFile(request.POST['previous_offerletter'])  
                uploaded_offerletter = request.FILES['offerletter']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                offerletter_name = uploaded_offerletter.name
                temp = offerletter_name.split('.')
                offerletter_name = 'offerletter_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                offerletter = storage.save(offerletter_name, uploaded_offerletter)
                offerletter = storage.url(offerletter)
            else:
                if request.POST['previous_offerletter'] != '':
                    offerletter = request.POST['previous_offerletter'] 
                else:
                    offerletter = None
                    
            if bool(request.FILES.get('visaletter', False)) == True:
                if request.POST['previous_visaletter']:
                    deleteMediaFile(request.POST['previous_visaletter'])  
                uploaded_visaletter = request.FILES['visaletter']
                storage = FileSystemStorage()
                timestamp = int(time.time())
                visaletter_name = uploaded_visaletter.name
                temp = visaletter_name.split('.')
                visaletter_name = 'visaletter_'+str(timestamp)+"."+temp[(len(temp) - 1)]
                
                visaletter = storage.save(visaletter_name, uploaded_visaletter)
                visaletter = storage.url(visaletter)
            else:
                if request.POST['previous_visaletter'] != '':
                    visaletter = request.POST['previous_visaletter'] 
                else:
                    visaletter = None
                    
           
            try:
                user_documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
            except SpUserDocuments.DoesNotExist:
                user_documents = None
            
            if user_documents is None:        
                documents = SpUserDocuments()
                documents.user_id       = request.POST['last_user_id']
                documents.qatar_id          = qatar_id
                documents.passport_card     = passport_card
                documents.resume            =  resume
                documents.educationaldoc    = educationaldoc
                documents.offerletter       = offerletter
                documents.visaletter       = visaletter
                if request.POST['qatar_id_expairy']:
                    documents.qatar_id_expairy            =  datetime.strptime(request.POST['qatar_id_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['visa_letter_expairy']:
                    documents.visaletter_expairy            =  datetime.strptime(request.POST['visa_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['offer_letter_expairy']:
                    documents.offerletter_expairy            =  datetime.strptime(request.POST['offer_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['educational_documents_expairy']:
                    documents.educationaldoc_expairy            =  datetime.strptime(request.POST['educational_documents_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['resume_expairy']:
                    documents.resume_expairy            =  datetime.strptime(request.POST['resume_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['passport_expairy']:
                    documents.passport_card_expairy            =  datetime.strptime(request.POST['passport_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                documents.save()
                response['error'] = False
                response['message'] = "Record has been saved successfully."
            else:
                documents = SpUserDocuments.objects.get(user_id=request.POST['last_user_id'])
                documents.user_id       = request.POST['last_user_id']
                documents.qatar_id          = qatar_id
                documents.passport_card     = passport_card
                documents.resume            =  resume
                documents.educationaldoc    = educationaldoc
                documents.offerletter       = offerletter
                documents.visaletter       = visaletter
                if request.POST['qatar_id_expairy']:
                    documents.qatar_id_expairy            =  datetime.strptime(request.POST['qatar_id_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['visa_letter_expairy']:
                    documents.visaletter_expairy            =  datetime.strptime(request.POST['visa_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['offer_letter_expairy']:
                    documents.offerletter_expairy            =  datetime.strptime(request.POST['offer_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['educational_documents_expairy']:
                    documents.educationaldoc_expairy            =  datetime.strptime(request.POST['educational_documents_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['resume_expairy']:
                    documents.resume_expairy            =  datetime.strptime(request.POST['resume_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                if request.POST['passport_expairy']:
                    documents.passport_card_expairy            =  datetime.strptime(request.POST['passport_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
                documents.save()

                response['error'] = False
                response['message'] = "Record has been updated successfully."
            
            return JsonResponse(response)

        except Exception as e:
            return HttpResponse(e)
    else:
        try:
            employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
        except SpUserDocuments.DoesNotExist:
            employee_documents = None

        template = 'user-management/edit-employee/employee-document-detail.html'
        response = {}
        response['last_user_id'] = employee_id
        response['employee_documents'] = employee_documents
        return render(request, template,response)




@login_required
def userShortDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.gender,
    sp_basic_details.outstanding_amount,sp_basic_details.opening_crates,
    sp_addresses.address_line_1,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,
    sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    left join sp_addresses on sp_addresses.user_id = sp_users.id   
    where sp_addresses.type=%s  and sp_users.id = %s ''',['correspondence',user_id])
    if user :
        context['user'] = user[0]
        context['contact_persons'] = SpContactPersons.objects.filter(user_id=user_id)
    else : 
        context['user'] = []
      
    template = 'user-management/user-short-details.html'
    return render(request, template,context)

@login_required
def userDetail(request,user_id):
    context = {}
    user = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,
    sp_basic_details.gender,sp_basic_details.personal_email,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.date_of_joining,sp_basic_details.working_shift_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id 
    where sp_users.id = %s''',[user_id])[0]
    
    try:
        documents = SpUserDocuments.objects.get(user_id=user_id)
    except SpUserDocuments.DoesNotExist:
        documents = None

    try:
        allocation = SpUserAreaAllocations.objects.get(user_id=user_id)
    except SpUserAreaAllocations.DoesNotExist:
        allocation = None
            
    try:
        user_variants = SpUserProductVariants.objects.filter(user_id=user_id)
    except SpUserProductVariants.DoesNotExist:
        user_variants = None

    try:
        permanent_address = SpAddresses.objects.get(user_id=user_id,type='permanent')
    except SpAddresses.DoesNotExist:
        permanent_address = None    

    context['user'] = user
    context['user_correspondence_address'] = SpAddresses.objects.get(user_id=user_id,type='correspondence')
    context['user_permanent_address']      = permanent_address
    context['contact_persons']             = SpContactPersons.objects.filter(user_id=user_id)
    context['user_contacts']               = SpContactNumbers.objects.filter(user_id=user_id)
    context['area_allocated']              = allocation 
    context['user_documents']              = documents
    context['user_variants']               = user_variants
    context['user_attendance_locations']   = SpUserAttendanceLocations.objects.filter(user_id=user_id,status=1)
    template = 'user-management/user-details.html'

    return render(request, template,context)


@login_required
def employeeShortDetail(request,employee_id):
    context = {}
    employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,contract_type.contract_type,sp_basic_details.gender,sp_basic_details.father_name,sp_basic_details.mother_name,sp_basic_details.date_of_birth, sp_addresses.address_line_1
    ,sp_addresses.address_line_2, sp_addresses.country_name, sp_addresses.state_name,sp_addresses.city_name,sp_addresses.pincode
    FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    left join contract_type on contract_type.id=sp_basic_details.contract_type
    left join sp_addresses on sp_addresses.user_id = sp_users.id  
    where sp_users.user_type = %s and sp_addresses.type=%s and sp_users.id = %s ''',[1,'correspondence',employee_id])
    if employee :
        context['employee'] = employee[0]
    else : 
        context['employee'] = []
    
    single = SpUsers.objects.filter(id = employee_id).first()
    if single and single.web_auth_token:
        # device_data = dict(item.split(":") for item in str(single.web_auth_token).split(","))
        device_data = {item.split(":")[0].strip().capitalize(): item.split(":")[1].strip() for item in str(single.web_auth_token).split(",")}

        context['device_data'] =  device_data
    else:
        context['device_data'] =  None

    try:
        address = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    except SpAddresses.DoesNotExist:
        address = None

    context['employee_permanent_address'] = address    
    template = 'user-management/employee-short-details.html'
    return render(request, template,context)




@login_required
def employeeDetail(request,employee_id):
    context = {}
    context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,sp_basic_details.geofencing,
    sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.leave_count,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    where sp_users.id = %s''',[employee_id])[0]

    context['contractval'] = contractval = SpBasicDetails.objects.raw('''SELECT sp_basic_details.*,contract_type.id,contract_type.contract_type as contract_name 
    FROM sp_basic_details left join contract_type on contract_type.id = sp_basic_details.contract_type
    where sp_basic_details.user_id  = %s''',[employee_id])
    
    context['workingshift'] = workingshift = TblClAllocatedShifts.objects.raw('''SELECT tbl_cl_allocated_shifts.*,tbl_cl_working_shifts.id, tbl_cl_working_shifts.working_shift ,tbl_cl_working_shifts.start_timing,tbl_cl_working_shifts.end_timing 
    FROM tbl_cl_allocated_shifts left join tbl_cl_working_shifts on tbl_cl_working_shifts.id = tbl_cl_allocated_shifts.working_shift_id
    where tbl_cl_allocated_shifts.user_id  = %s''',[employee_id])
    hours_sum = timedelta(days=0,seconds=0,minutes=0,hours=0)
    for workingshifts in workingshift:
        FMT = '%H:%M:%S'
        starttime = workingshifts.start_timing.strftime("%H:%M:%S")
        endtime = workingshifts.end_timing.strftime("%H:%M:%S")
        tdelta = datetime.strptime(endtime, FMT) - datetime.strptime(starttime,FMT)
        hours_sum += tdelta
    totalhours = str(hours_sum).split(', ')

    try:
        documents = TblClEmployeeFolderFiles.objects.filter(employee_id=employee_id)
    except TblClEmployeeDocuments.DoesNotExist:
        documents = None

    try:
        user_contacts = SpContactNumbers.objects.filter(user_id=employee_id)
    except SpContactNumbers.DoesNotExist:
        user_contacts = None

    try:
        bank_details = SpBankDetails.objects.get(user_id=employee_id)
    except SpBankDetails.DoesNotExist:
        bank_details = None
    try:
        contract_type = SpBasicDetails.objects.get(user_id=employee_id)
    except SpBasicDetails.DoesNotExist:
        contract_type = None

    try:
        contracttype =  ContractType.objects.get(id=contract_type.contract_type)
    except ContractType.DoesNotExist:
         contracttype = None
         
    leave_ledger = SpUserLeavePolicyLedger.objects.filter(user_id = employee_id)
    for leave in leave_ledger:
        leave.leave_policy_name = getModelColumnById(SpLeavePolicies,leave.leave_policy_id,'leave_policy')
        leave.laave_type_name = getModelColumnById(SpLeaveTypes,leave.leave_type_id,'leave_type')
        leave.month_leave_counts = round(leave.month_leave_count,1)
        
    try:
        emp_payroll = SpPayrollMaster.objects.get(user_id=employee_id)
    except SpPayrollMaster.DoesNotExist:
        emp_payroll = None

    context['contract_type']                     = contracttype
    context['basic_data']                       = contract_type
    context['bank_details']                      = bank_details
    context['working_shift']                     = workingshift
    if len(totalhours) > 0:
        context['totalhours']                        = totalhours
    else:
        context['totalhours']                        = '-'
    try:
        employee_documents = SpUserDocuments.objects.get(user_id=employee_id)
    except SpUserDocuments.DoesNotExist:
        employee_documents = None
    
    context['employee_correspondence_address']  = SpAddresses.objects.get(user_id=employee_id,type='correspondence')
    context['employee_permanent_address']       = SpAddresses.objects.get(user_id=employee_id,type='permanent')
    context['employee_id']                      = employee_id
    context['emp_payroll']                      = emp_payroll
    context['user_contacts']                    = user_contacts
    context['user_areas']                       = None
    context['user_documents']                   = documents
    context['leave_ledger']                     = leave_ledger
    context['employee_documents']               = employee_documents
    context['user_coordinates']                 = None
    context['user_attendance_locations']        = None
    context['google_app_key']                   = getConfigurationResult('google_app_key')
    context['currency_code']                    = SpCurrencyCode.objects.all()
    context['fixed_type']                       = SpSalaryHead.objects.filter(salary_head_type = 1)
    context['fixed_type_len']                   = SpSalaryHead.objects.filter(salary_head_type = 1).count()
    context['addition_type']                       = SpSalaryHead.objects.filter(salary_head_type = 2)
    
    user_salary_details = SpUserSalarySlip.objects.filter(user_id = employee_id).first()
    if user_salary_details:
        fixed_pay_type_ids_str = user_salary_details.fixed_pay_type_ids

        fixed_pay_type_ids_list = fixed_pay_type_ids_str.split(',')

        list_length = len(fixed_pay_type_ids_list)
        
        count_list = list(range(list_length))
        count_list_str = ','.join(map(str, count_list))
        fixed_pay_type_ids_list = user_salary_details.fixed_pay_type_ids.split(',')
        fixed_pay_per_val_list = user_salary_details.fixed_pay_per_val.split(',')
        fixed_pay_converted_val_list = user_salary_details.fixed_pay_converted_val.split(',')
        fixed_pay_currency_list = user_salary_details.fixed_pay_currency.split(',')
        if user_salary_details.additional_type:
            additional_type_list = user_salary_details.additional_type.split(',')
            
        fixed_pay_type_ids_list = list(map(int, fixed_pay_type_ids_list))
        fixed_pay_per_val_list = list(map(int, fixed_pay_per_val_list))
        fixed_pay_converted_val_list = list(map(int, fixed_pay_converted_val_list))
        fixed_pay_currency_list = list(map(int,fixed_pay_currency_list))
        if user_salary_details.additional_type:
            additional_type_list = list(map(int, additional_type_list))
        
        combined_list = []
        for i in range(len(fixed_pay_type_ids_list)):
            combined_list.append({
                'type_id': fixed_pay_type_ids_list[i],
                'per_val': fixed_pay_per_val_list[i],
                'converted_val': fixed_pay_converted_val_list[i],
                'fixed_pay_currency':fixed_pay_currency_list[i]
            })
        if user_salary_details.additional_type:
            additional_combined_list = []
            for i in range(len(additional_type_list)):
                additional_combined_list.append({
                    'additional_type_list': additional_type_list[i],
                })
        
    
        context['fixed_pay_type_ids_count_list']  = count_list_str,
        context['user_salary_details']      = user_salary_details
        context['combined_list']    = combined_list
        context['len_combined_list'] = len(combined_list)
        if user_salary_details.additional_type:
            context['additional_combined_list'] = additional_combined_list
    template = 'user-management/employee-details.html'

    return render(request, template,context)



def getGroupedTownOptions(request):
    options = ''
    zone_ids = request.POST['zone_ids'].split(',')
    zones = SpZones.objects.raw(''' select * from sp_zones where id in %s ''',[zone_ids])
    for zone in zones:
        towns = SpTowns.objects.filter(zone_id=zone.id)
        if towns:
            options += '<optgroup label="' + zone.zone + '">'
            for town in towns : 
                options += "<option value="+str(town.id)+">"+town.town+"</option>"
            options += '</optgroup>'
    
    return HttpResponse(options)

def getReportingUserOptions(request, role_id):
    reporting_role = SpRoles.objects.get(id=role_id)
    options = '<option value="">Select Reporting to User*</option>'
    reporting_users = SpUsers.objects.raw(''' select id,first_name,last_name from sp_users where role_id = %s and user_type = %s ''',[reporting_role.reporting_role_id,1])
    for reporting_user in reporting_users:
        options += "<option value="+str(reporting_user.id)+">"+reporting_user.first_name+" "+reporting_user.last_name+"</option>"
    
    return HttpResponse(options)

#update user status
@login_required
def updateUserStatus(request):

    response = {}
    if request.method == "POST":
        try:
            id = request.POST.get('id')
            is_active = request.POST.get('is_active')

            data = SpUsers.objects.get(id=id)
            data.status = is_active
            data.save()

            if is_active == '1':
                status = 'Unblock'
            else:
                status = 'Block'
                
            user_name   = request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
            heading     = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status
            activity    = getModelColumnById(SpUsers, id, 'first_name')+' '+getModelColumnById(SpUsers, id, 'middle_name')+' '+getModelColumnById(SpUsers, id, 'last_name')+' '+status+' by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'icon', '1', 'platform_icon')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#update user status
@login_required
def updateUserVariantPrice(request):

    response = {}
    if request.method == "POST":
        try:
            id              = request.POST.get('id')
            price           = request.POST.get('price')
            user_type       = request.POST.get('user_type')
            is_distributor  = request.POST.get('is_distributor')
            
            data = SpUserProductVariants.objects.get(id=id)
            if user_type == '2':
                if is_distributor == '1':
                    data.sp_distributor = price
                    previous_price = data.sp_distributor
                else:
                    data.sp_superstockist = price
                    previous_price = data.sp_superstockist    
            else:
                data.sp_employee = price
                previous_price = data.sp_employee 
    
            data.save()
            
            response['error'] = False
            response['message'] = "Record has been updated successfully."
            response['id'] = id
            response['price'] = price
            response['user_type'] = user_type
            response['is_distributor'] = is_distributor
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

#export to excel operational user list
@login_required
def exportOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        try:
            user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)
        except SpBasicDetails.DoesNotExist:
            user.outstanding_amount = None
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Operational User'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]

    if 'outstanding_amount' in column_list:
        columns += [ 'Outstanding Amount' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 
        if user.outstanding_amount:
            outstanding_amount = user.outstanding_amount.outstanding_amount 
        else:
            outstanding_amount = ''   
        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]

        if 'outstanding_amount' in column_list:
            row += [ outstanding_amount ]           
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

def render_to_pdf(template_src, context_dict={}):
	template = get_template(template_src)
	html  = template.render(context_dict)
	result = BytesIO()
	pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
	if not pdf.err:
		return HttpResponse(result.getvalue(), content_type='application/pdf')
	return None


#export to pdf operational user list
@login_required
def exportOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=2).exclude(id=1).order_by('-id')
    for user in users :
        user.outstanding_amount = SpBasicDetails.objects.get(status=1, user_id=user.id)

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 

#export to excel non-operational user list
@login_required
def exportNonOperationalUserToXlsx(request, columns):
    column_list = columns.split (",")
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=non-operatonal-users.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Non-Operational-Users'
    
    # Define the titles for columns
    columns = []

    if 'store_name' in column_list:
        columns += [ 'Store Name' ]

    if 'role' in column_list:
        columns += [ 'Role' ]
    
    if 'contact_person' in column_list:
        columns += [ 'Contact Person' ] 

    if 'contact_no' in column_list:
        columns += [ 'Contact No.' ]    

        # columns += [ 'Address' ] 

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        if 'store_name' in column_list:
            row += [ user.store_name ]

        if 'role' in column_list:
            row += [ user.role_name ]
        
        if 'contact_person' in column_list:
            row += [ user.first_name + ' ' + user.middle_name + ' ' + user.last_name ] 

        if 'contact_no' in column_list:
            row += [ user.primary_contact_number ]         
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response 

#export to pdf non operational user list
@login_required
def exportNonOperationalUserToPdf(request, columns):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=3).exclude(id=1).order_by('-id')

    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/non_operational_user_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Non-Operational-User.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response 




#export to excel employee list
@login_required
def exportEmployeeToXlsx(request, columns, search, jobs, depts, roles):
    column_list = columns.split (",")
    users = SpUsers.objects.filter(user_type=1).exclude(id=1).order_by('-id')
    
    for user in users:
        user_basic_details = SpBasicDetails.objects.get(user_id=user.id)
        if user_basic_details:
            if user_basic_details.father_name:    
                user.father_name  = user_basic_details.father_name
            else:
                user.father_name  = ''
            
            if user_basic_details.mother_name:
                user.mother_name = user_basic_details.mother_name
            else:
                user.mother_name = ''
            
            if user_basic_details.date_of_birth:
                user.date_of_birth = user_basic_details.date_of_birth
            else:
                user.date_of_birth = ''

            if user_basic_details.gender:
                user.gender = user_basic_details.gender
            else:
                user.gender = ''
            if user_basic_details.date_of_joining:
                user.date_of_joining = user_basic_details.date_of_joining
            else:
                user.date_of_joining = ''
            if user_basic_details.week_of_day:
                user.week_of_day = user_basic_details.week_of_day
            else:
                user.week_of_day = ''
        else:
            user.father_name  = ''
            user.mother_name = ''
            user.gender = ''
            user.week_of_day = ''
            user.date_of_joining = ''
        user_bank_details = SpBankDetails.objects.filter(user_id=user.id).first()
        if user_bank_details:
            if user_bank_details.bank_name:
                user.p_bank_name = user_bank_details.bank_name
            else:
                user.p_bank_name = ''
            if user_bank_details.bank_account_no:
                user.bank_account_no = user_bank_details.bank_account_no
            else:
                user.bank_account_no = ''
            if user_bank_details.ifsc_code:
                user.ifsc_code = user_bank_details.ifsc_code
            else:
                user.ifsc_code = ''
            if user_bank_details.bank_address:
                user.bank_address = user_bank_details.bank_address
            else:
                user.bank_address = ''
        else:
            user.p_bank_name = ''
            user.bank_account_no = ''
            user.ifsc_code = ''
            user.bank_address = ''
        ccr_user_addresses = SpAddresses.objects.filter(user_id=user.id,type="correspondence").first()
        if ccr_user_addresses:
            if ccr_user_addresses.address_line_1:
                user.cr_address_line_1 = ccr_user_addresses.address_line_1
            else:
                user.cr_address_line_1 = ''
            if ccr_user_addresses.address_line_2:
                user.cr_address_line_2 = ccr_user_addresses.address_line_2
            else:
                user.cr_address_line_2 = ''
                
            if ccr_user_addresses.country_name:
                user.cr_country_name = ccr_user_addresses.country_name
            else:
                user.cr_country_name = ''
                
            if ccr_user_addresses.state_name:
                user.cr_state_name = ccr_user_addresses.state_name
            else:
                user.cr_state_name = ''
                
            if ccr_user_addresses.city_name:
                user.cr_city_name = ccr_user_addresses.city_name
            else:
                user.cr_city_name = ''
            if ccr_user_addresses.pincode:
                user.cr_pincode = ccr_user_addresses.pincode
            else:
                user.cr_pincode = ''
        else:
            user.cr_address_line_1 = ''
            user.cr_address_line_2 = ''
            user.cr_country_name = ''
            user.cr_state_name = ''
            user.cr_pincode = ''
            user.cr_city_name = ''
        ppr_user_addresses = SpAddresses.objects.filter(user_id=user.id,type="permanent").first()
        if ppr_user_addresses:
            if ppr_user_addresses.address_line_1:
                user.pr_address_line_1 = ppr_user_addresses.address_line_1
            else:
                user.pr_address_line_1 = ''
            if ppr_user_addresses.address_line_2:
                user.pr_address_line_2 = ppr_user_addresses.address_line_2
            else:
                user.pr_address_line_2 = ''
                
            if ppr_user_addresses.country_name:
                user.pr_country_name = ppr_user_addresses.country_name
            else:
                user.pr_country_name = ''
                
            if ppr_user_addresses.state_name:
                user.pr_state_name = ppr_user_addresses.state_name
            else:
                user.pr_state_name = ''
                
            if ppr_user_addresses.city_name:
                user.pr_city_name = ppr_user_addresses.city_name
            else:
                user.pr_city_name = ''
            if ppr_user_addresses.pincode:
                user.pr_pincode = ppr_user_addresses.pincode
            else:
                user.pr_pincode = ''
        else:
            user.pr_address_line_1 = ''
            user.pr_address_line_2 = ''
            user.pr_country_name = ''
            user.pr_state_name = ''
            user.pr_pincode = ''
            user.pr_city_name = ''
    jobss = jobs.split(",")
    deptss = depts.split(",")
    roless = roles.split(",")

    if search != '0':
        users = users.filter(id = search)
    if depts != '0':
        users = users.filter(department_id__in = deptss)
    if roles != '0':
        users = users.filter(role_id__in = roless)
    if jobs != '0':
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type__in = jobss):
                users = users.exclude(id = user.id)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employees'
    
    # Define the titles for columns
    columns = []
    columns += [ 'Employee ID' ]
    columns += [ 'Name' ]
    columns += ['Father Name']
    columns += ['Mother Name']
    columns += ['DOB']
    columns += ['Gender']
    columns += [ 'Contact' ]
    columns += ['Email']
    columns += ['Correspondence Address 1']
    columns += ['Correspondence Address 2']
    columns += ['Correspondence Country']
    columns += ['Correspondence Zone']
    columns += ['Correspondence District']
    columns += ['Correspondence Pin Code']
    columns += ['Permanent Address 1']
    columns += ['Permanent Address 2']
    columns += ['Permanent Country']
    columns += ['Permanent Zone']
    columns += ['Permanent District']
    columns += ['Permanent Pin Code']
    columns += ['Date Of Joining']
    columns += [ 'Role' ]    
    columns += ['Organization Name']
    columns += [ 'Department Name' ] 
    columns += ['Reporting to Name']
    columns += ['Week of Day']
    columns += ['Bank Name']
    columns += ['Bank Account No.']
    columns += ['IFSC Code']
    columns += ['Bank Address']

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    # Iterate through all movies
    for user in users:

        if user.last_login is not None:
            if user.web_auth_token is None:
                employee_platform = 'Web'
            else:
                employee_platform = 'APP'
        else:
            employee_platform = ''               
         
        row_num += 1
        # Define the data for each cell in the row 

        row = []
        row += [ user.emp_sap_id ]
        row += [ getUserName(user.id) ]
        row += [user.father_name]
        row += [user.mother_name]
        row += [user.date_of_birth]
        row += [user.gender]
        row += [ user.primary_contact_number ]
        row += [ user.official_email ]
        
        row += [ user.cr_address_line_1 ]
        row += [user.cr_address_line_2]
        row += [user.cr_country_name]
        row += [user.cr_state_name]
        row += [user.cr_city_name]
        row += [ user.cr_pincode ]
        
        row += [ user.pr_address_line_1 ]
        row += [user.pr_address_line_2]
        row += [user.pr_country_name]
        row += [user.pr_state_name]
        row += [user.pr_city_name]
        row += [ user.pr_pincode ]
        
        row += [user.date_of_joining]
        row += [ user.role_name ]
        row += [user.organization_name]
        row += [ user.department_name ]
        row += [user.reporting_to_name]
        row += [user.week_of_day]
        row += [user.p_bank_name]
        row += [ user.bank_account_no ]
        row += [user.ifsc_code]
        row += [user.bank_address]
        
       
        # row += [ organization.address + ', ' + organization.pincode ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment

    workbook.save(response)

    return response



@login_required
def exportEmployeeToPdf(request, columns, search, jobs, depts, roles):
    column_list = columns.split (",")
    context = {}
    users = SpUsers.objects.all().filter(user_type=1).exclude(id=1).order_by('-id')

    jobs = jobs.split(",")
    depts = depts.split(",")
    roles = roles.split(",")

    if search != '0':
        users = users.filter(Q(emp_sap_id__icontains = search) | Q(first_name__icontains = search))
    if depts != '0':
        users = users.filter(department_id__in = depts)
    if roles != '0':
        users = users.filter(role_id__in = roles)
    if jobs != '0':
        for user in users:
            if not SpBasicDetails.objects.filter(user_id = user.id, contract_type__in = jobs):
                users = users.exclude(id = user.id)
    users = SpUsers.objects.all().filter(user_type=1).order_by('-id')
    baseurl = settings.BASE_URL
    pdf = render_to_pdf('user-management/employee_pdf_template.html', {'users': users, 'url': baseurl, 'columns' : column_list, 'columns_length' : len(column_list)})
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = 'Employees.pdf'
    content = "attachment; filename=%s" %(filename)
    response['Content-Disposition'] = content
    return response

@login_required
def getUserMap(request):
    user_coordinates = SpUsers.objects.filter(id=request.GET.get('distributor_id')).values('latitude', 'longitude').first()
    
    context = {}
    context['user_coordinates'] = user_coordinates
    context['periphery']        = request.GET.get('periphery')
    context['google_app_key']   = getConfigurationResult('google_app_key')
    template = 'user-management/user-map.html'
    return render(request, template, context) 

@login_required
def importProductVariant(request):
    # workbook object is created 
    # wb_obj = load_workbook('media/operatonal-users.xlsx')

    # sheet_obj = wb_obj.active 
    # m_row = sheet_obj.max_row 
    
    # # Loop will print all values 
    # # of first column  
    
    # for i in range(1, m_row + 1): 
    #     row = [cell.value for cell in sheet_obj[i]] 
    #     print(row)
    #     template = ProductVariantTemplate()
    #     template.store_name = row[0]
    #     template.role = row[1]
    #     template.save()
            
    return HttpResponse('row')     


def updateUserRole(user_id,params):
    role_permissions = SpRolePermissions.objects.filter(role_id=params.POST['role_id'])
    if len(role_permissions):
        SpUserRolePermissions.objects.filter(user_id=user_id).delete()
        for role_permission in role_permissions:
            user_role_permission = SpUserRolePermissions()
            user_role_permission.user_id = user_id
            user_role_permission.role_id = params.POST['role_id']
            user_role_permission.module_id = role_permission.module_id
            user_role_permission.sub_module_id = role_permission.sub_module_id
            user_role_permission.permission_id = role_permission.permission_id
            user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
            # user_role_permission.workflow = role_permission.workflow
            user_role_permission.save()

        
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=params.POST['role_id'])
        if len(role_permission_workflows):
            SpUserRoleWorkflowPermissions.objects.filter(user_id=user_id).delete()
            for role_permission_workflow in role_permission_workflows : 
                user_role_permission_wf = SpUserRoleWorkflowPermissions()
                user_role_permission_wf.user_id = user_id
                user_role_permission_wf.role_id = role_permission_workflow.role_id
                user_role_permission_wf.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_wf.permission_id = role_permission_workflow.permission_id
                user_role_permission_wf.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_wf.level_id = role_permission_workflow.level_id
                user_role_permission_wf.level = role_permission_workflow.level
                user_role_permission_wf.description = role_permission_workflow.description
                user_role_permission_wf.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                user_role_permission_wf.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_wf.status = role_permission_workflow.status
                user_role_permission_wf.save()


def mapProductToUser(user_id):
    product_variants = SpProductVariants.objects.all()
    if len(product_variants):
        
        SpUserProductVariants.objects.filter(user_id=user_id).delete()

        for product_variant in product_variants:
            user_product_variant                            = SpUserProductVariants()
            user_product_variant.user_id                    = user_id
            user_product_variant.product_id                 = product_variant.product_id
            user_product_variant.product_name               = product_variant.product_name
            user_product_variant.product_variant_id         = product_variant.id
            user_product_variant.item_sku_code              = product_variant.item_sku_code
            user_product_variant.variant_quantity           = product_variant.variant_quantity
            user_product_variant.variant_unit_id            = product_variant.variant_unit_id
            user_product_variant.variant_name               = product_variant.variant_name
            user_product_variant.variant_unit_name          = product_variant.variant_unit_name
            user_product_variant.variant_size               = product_variant.variant_size
            user_product_variant.no_of_pouch                = product_variant.no_of_pouch
            user_product_variant.container_size             = product_variant.container_size
            user_product_variant.is_bulk_pack               = product_variant.is_bulk_pack
            user_product_variant.mrp                        = product_variant.mrp
            user_product_variant.sp_distributor             = product_variant.sp_distributor
            user_product_variant.sp_superstockist           = product_variant.sp_superstockist
            user_product_variant.sp_employee                = product_variant.sp_employee
            user_product_variant.container_mrp              = float(product_variant.mrp) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_distributor   = float(product_variant.sp_distributor) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_superstockist = float(product_variant.sp_superstockist) * float(product_variant.no_of_pouch)
            user_product_variant.container_sp_employee      = float(product_variant.sp_employee) * float(product_variant.no_of_pouch)
            user_product_variant.valid_from                 = product_variant.valid_from
            user_product_variant.valid_to                   = product_variant.valid_to
            user_product_variant.status                     = product_variant.status
            user_product_variant.save()


@login_required
def viewUserRolePermission(request,user_id):
    if request.method == "POST":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        context = {}
        role_id = getModelColumnById(SpUsers,user_id,'role_id')
        role = SpRoles.objects.get(id=role_id)
        permissions = SpPermissions.objects.filter(status=1)
        organizations = SpOrganizations.objects.filter(status=1)
        departments = SpDepartments.objects.filter(status=1)

        other_departments = SpDepartments.objects.filter(status=1,organization_id=role.organization_id)
        for department in other_departments : 
            department.other_roles = SpRoles.objects.filter(status=1,department_id=department.id).exclude(id=role.id)

        modules = SpModules.objects.filter(status=1)
        for module in modules : 
            module.sub_modules = SpSubModules.objects.filter(status=1,module_id=module.id)
        
        context['permissions'] = permissions
        context['organizations'] = organizations
        context['modules'] = modules
        context['user_id'] = user_id
        context['role'] = role
        context['other_departments'] = other_departments
        context['first_workflow_level'] =  SpWorkflowLevels.objects.get(priority='first')
        context['last_workflow_level'] =  SpWorkflowLevels.objects.get(priority='last')
        context['middle_workflow_level'] =  SpWorkflowLevels.objects.get(priority='middle')

        
        template = 'user-management/user-role-permission.html'
        return render(request,template,context)

@login_required
def updateUserRolePermission(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
       
        permissions = SpPermissions.objects.filter(status=1)
        sub_modules = SpSubModules.objects.filter(status=1)

        SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id']).delete()
        SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id']).delete()

        for sub_module in sub_modules :
            for permission in permissions :
                var_name = 'permission_'+ str(sub_module.id) + '_' + str(permission.id)
                if var_name in request.POST:
                    role_permission = SpUserRolePermissions()
                    role_permission.user_id = request.POST['user_id']
                    role_permission.role_id = request.POST['role_id']
                    role_permission.module_id = getModelColumnById(SpSubModules,sub_module.id,'module_id')
                    role_permission.sub_module_id = sub_module.id
                    role_permission.permission_id = permission.id
                    role_permission.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                    role_permission.save()
                    total_work_flows_var = 'workflow_'+str(sub_module.id)+'_'+str(permission.id)
                    if request.POST[total_work_flows_var] :

                        role_permission.workflow = request.POST[total_work_flows_var]
                        role_permission.save()

                        total_work_flows = json.loads(request.POST[total_work_flows_var])
                        
                        for total_work_flow in total_work_flows :
                            role_permission_level = SpUserRoleWorkflowPermissions()
                            role_permission_level.user_id = request.POST['user_id']
                            role_permission_level.role_id = request.POST['role_id']
                            role_permission_level.sub_module_id = sub_module.id
                            role_permission_level.permission_id = permission.id
                            role_permission_level.permission_slug = getModelColumnById(SpPermissions,permission.id,'slug')
                            role_permission_level.level_id = total_work_flow['level_id']
                            role_permission_level.level = getModelColumnById(SpWorkflowLevels,total_work_flow['level_id'],'level')
                            role_permission_level.description = total_work_flow['description']
                            if int(total_work_flow['role_id']) > 0 :
                                role_permission_level.workflow_level_dept_id = getModelColumnById(SpRoles,total_work_flow['role_id'],'department_id')
                            else:
                                role_permission_level.workflow_level_dept_id = None
                            role_permission_level.workflow_level_role_id = total_work_flow['role_id']
                            if 'status' in total_work_flow :
                                role_permission_level.status = total_work_flow['status']
                            else:
                                role_permission_level.status = 1

                            role_permission_level.save()

        response['flag']    = True
        response['message'] = "Record has been updated successfully."
        return JsonResponse(response)
        
@login_required
def checkRolePermision(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        if SpRolePermissions.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).exists():
            response['flag'] = True
            response['message'] = "Workflow is already applied to this permission."
        else :
            response['flag'] = False
            response['message'] = "Workflow is not applied to this permission. Please contact administator."
            
        return JsonResponse(response)

@login_required
def saveRolePermisionValidity(request):
    if request.method == "GET":
        response = {}
        response['flag'] = False
        response['message'] = "Method not allowed."
        return JsonResponse(response)
    else:
        response = {}
        role_permission = SpRolePermissions.objects.filter(permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).first()
        role_permission_workflows = SpRoleWorkflowPermissions.objects.filter(role_id=role_permission.role_id, permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id'])

        SpUserRolePermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'],permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()
        
        user_role_permission = SpUserRolePermissions()
        user_role_permission.user_id = request.POST['user_id']
        user_role_permission.role_id = request.POST['role_id']
        user_role_permission.module_id = role_permission.module_id
        user_role_permission.sub_module_id = role_permission.sub_module_id
        user_role_permission.permission_id = role_permission.permission_id
        user_role_permission.permission_slug = getModelColumnById(SpPermissions,role_permission.permission_id,'slug')
        user_role_permission.workflow = role_permission.workflow
        user_role_permission.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        user_role_permission.save()

        if len(role_permission_workflows) :

            SpUserRoleWorkflowPermissions.objects.filter(user_id=request.POST['user_id'],role_id=request.POST['role_id'], permission_id=request.POST['permission_id'], sub_module_id=request.POST['sub_module_id']).delete()

            for role_permission_workflow in role_permission_workflows :
                user_role_permission_level = SpUserRoleWorkflowPermissions()
                user_role_permission_level.user_id = request.POST['user_id']
                user_role_permission_level.role_id = request.POST['role_id']
                user_role_permission_level.sub_module_id = role_permission_workflow.sub_module_id
                user_role_permission_level.permission_id = role_permission_workflow.permission_id
                user_role_permission_level.permission_slug = getModelColumnById(SpPermissions,role_permission_workflow.permission_id,'slug')
                user_role_permission_level.level_id = role_permission_workflow.level_id
                user_role_permission_level.level = role_permission_workflow.level
                user_role_permission_level.description = role_permission_workflow.description
                if int(role_permission_workflow.level_id) > 0 :
                    user_role_permission_level.workflow_level_dept_id = role_permission_workflow.workflow_level_dept_id
                else:
                    user_role_permission_level.workflow_level_dept_id = None

                user_role_permission_level.workflow_level_role_id = role_permission_workflow.workflow_level_role_id
                user_role_permission_level.status = role_permission_workflow.status
                user_role_permission_level.from_date = datetime.strptime(request.POST['from_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.to_date = datetime.strptime(request.POST['to_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
                user_role_permission_level.save()

        response['flag'] = True
        response['message'] = "Record has been updated successfully."
        response['workflow'] = user_role_permission.workflow
            
        return JsonResponse(response)


# add Employee Biometric
@login_required
def addEmployeeBiometricDetails(request):
    if request.method == "POST":
        last_user_id = request.POST['last_user_id']
        user = SpUsers.objects.get(id=request.POST['last_user_id'])
        user.periphery           = request.POST['periphery']
        
        
        if request.POST['geofencing_type']:
            SpBasicDetails.objects.filter(user_id=request.POST['last_user_id']).update(geofencing=request.POST['geofencing_type'])

        for each_shift in request.POST.getlist('working_shift_id[]'):
            working_shift = TblClAllocatedShifts()
            working_shift.user_id = request.POST['last_user_id']
            working_shift.working_shift_id = each_shift
            working_shift.save()
        user.save()
        context={}
        context['college']           = 'college'
        context['document_types']    = []
        context['document_lists']    = []
        context['last_user_id']      = last_user_id 
        context['college_id']        = user.organization_id   
        rooms = TblClRoom.objects.filter(college_id=user.organization_id).values('id', 'room').order_by('id')
        context['rooms']                = rooms 
        context['document_types']       = TblClDocumentTypes.objects.filter().order_by('document_name')
        context['document_lists']       = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).order_by('-id') 
        context['document_list_count']  = document_list_count  = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).count() 
        return render(request,"user-management/add-employee-documents.html",context)
    else:
        context={}

        
        working_shifts         = TblClWorkingShifts.objects.all()
        allocat_shift          = TblClAllocatedShifts.objects.all().values_list("working_shift_id", flat=True)
        
        context['working_shifts']    = working_shifts
        context['allocat_shifts']    = allocat_shift
        context['college'] = 'college'
        return render(request,"user-management/add-employee-biometric-details.html",context)




    
# add student documents
@login_required
def addEmployeeDocuments(request):
    response = {}
    if request.method == "POST":
        try:
            last_user_id = request.POST['last_user_id']
            if TblClEmployeeDocuments.objects.filter(ducument_number=request.POST['document_number']).exists():
                response['error']   = True
                response['message'] = "Document No. already exists."
                return JsonResponse(response)
            elif TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_name=getModelColumnById(TblClDocumentTypes, request.POST['document_type'], 'document_name')).exists():
                response['error']   = True
                response['message'] = "Document already addedd."
                return JsonResponse(response)
            else:
                document                  = TblClEmployeeDocuments()
                document.employee_id       = last_user_id
                document.document_name    = getModelColumnById(TblClDocumentTypes, request.POST['document_type'], 'document_name')
                document.ducument_number  = request.POST['document_number']
                document.is_uploaded      = 0
                document.created_by       = request.user.id
                document.save()

                #-----------------------------notify android block-------------------------------#
                organization_id = getModelColumnById(SpUsers, request.POST['last_user_id'], 'organization_id')
                all_users = SpUsers.objects.filter(organization_id = organization_id, user_type=1).exclude(id=1).values_list("id", flat=True)
                for each_user in all_users:
                    user_id = each_user
                    userFirebaseToken = getModelColumnById(SpUsers,user_id,'firebase_token')
                    employee_name = getUserName(request.user.id)
    
                    message_title = document.document_name+" document scan request."
                    message_body =  document.document_name+" document scan request has been generated by "+employee_name
                    notification_image = ""
    
                    if userFirebaseToken is not None and userFirebaseToken != "" :
                        registration_ids = []
                        registration_ids.append(userFirebaseToken)
                        data_message = {}
                        data_message['id'] = 1
                        data_message['status'] = 'notification'
                        data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                        data_message['image'] = notification_image
                        send_android_notification(message_title,message_body,data_message,registration_ids)

                context={}
                context['college']           = 'college'
                context['college_id']          = getModelColumnById(SpUsers, last_user_id, 'organization_id')
                context['document_lists']    = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).order_by('-id')
                context['document_list_count']  = TblClEmployeeDocuments.objects.filter(employee_id=last_user_id, document_path__isnull=False).count()
                context['last_user_id']      = last_user_id    
                return render(request,"user-management/employee-documents.html",context)
        except Exception as e:
            response['error']            = True
            response['message']          = str(e)
            return HttpResponse(e)
    else:
        #rooms = TblClRoom.objects.filter(college_id=request.GET['college_id']).values('id', 'room').order_by('id')
        context = {}
        context['last_user_id']         = request.GET['employee_id']
        context['college_id']           = getModelColumnById(SpUsers, request.GET['employee_id'], 'organization_id')
        #context['rooms']                = rooms 
        context['document_types']       = TblClDocumentTypes.objects.filter().order_by('document_name')
        context['document_lists']       = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).order_by('-id') 
        context['document_list_count']  = document_list_count  = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).count()       
        return render(request,"user-management/add-employee-documents-details.html",context) 
        

    
@login_required
def employeeDocumentList(request):
    template = 'user-management/add-employee-documents.html'
    response = {}
    if request.method == "POST":
        try:
            if bool(request.FILES.get('qatar_id', False)) == True:
                uploaded_qatar_id = request.FILES['qatar_id']
                qatar = FileSystemStorage()
                qatar_id = qatar.save(uploaded_qatar_id.name, uploaded_qatar_id)
                qatar_id = qatar.url(qatar_id)
            else:
                qatar_id = None
                
            if bool(request.FILES.get('passport_card', False)) == True:        
                uploaded_passport_card = request.FILES['passport_card']
                passport = FileSystemStorage()
                passport_card = passport.save(uploaded_passport_card.name, uploaded_passport_card)
                passport_card = passport.url(passport_card)
            else:
                passport_card = None
            
            if bool(request.FILES.get('resume', False)) == True:        
                uploaded_resume = request.FILES['resume']
                resume_pdf = FileSystemStorage()
                resume = resume_pdf.save(uploaded_resume.name, uploaded_resume)
                resume = resume_pdf.url(resume)
            else:
                resume = None

            if bool(request.FILES.get('educationaldoc', False)) == True:        
                uploaded_educationaldoc = request.FILES['educationaldoc']
                educationaldoc_pdf = FileSystemStorage()
                educationaldoc = educationaldoc_pdf.save(uploaded_educationaldoc.name, uploaded_educationaldoc)
                educationaldoc = educationaldoc_pdf.url(educationaldoc)
            else:
                educationaldoc = None
            if bool(request.FILES.get('offerletter', False)) == True:        
                uploaded_offerletter = request.FILES['offerletter']
                offerletter_pdf = FileSystemStorage()
                offerletter = offerletter_pdf.save(uploaded_offerletter.name, uploaded_offerletter)
                offerletter = offerletter_pdf.url(offerletter)
            else:
                offerletter = None

            if bool(request.FILES.get('visaletter', False)) == True:        
                uploaded_visaletter = request.FILES['visaletter']
                visaletter_pdf = FileSystemStorage()
                visaletter = visaletter_pdf.save(uploaded_visaletter.name, uploaded_visaletter)
                visaletter = visaletter_pdf.url(visaletter)
            else:
                visaletter = None

            documents               = SpUserDocuments()
            documents.user_id       = request.POST['last_user_id']
            documents.qatar_id      = qatar_id
            documents.passport_card = passport_card
            documents.resume           = resume
            documents.educationaldoc         = educationaldoc
            documents.offerletter           = offerletter
            documents.visaletter         = visaletter
            if request.POST['qatar_id_expairy']:
                documents.qatar_id_expairy            =  datetime.strptime(request.POST['qatar_id_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['visa_letter_expairy']:
                documents.visaletter_expairy            =  datetime.strptime(request.POST['visa_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['offer_letter_expairy']:
                documents.offerletter_expairy            =  datetime.strptime(request.POST['offer_letter_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['educational_documents_expairy']:
                documents.educationaldoc_expairy            =  datetime.strptime(request.POST['educational_documents_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['resume_expairy']:
                documents.resume_expairy            =  datetime.strptime(request.POST['resume_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            if request.POST['passport_expairy']:
                documents.passport_card_expairy            =  datetime.strptime(request.POST['passport_expairy'], '%d/%m/%Y').strftime('%Y-%m-%d')
            documents.save()
            response['error'] = False
            response['message'] = "Record has been add successfully."
            return JsonResponse(response)
        except Exception as e:
            response['error'] = True
            response['message'] = e
            return HttpResponse(e)
    return render(request, template)

@login_required
def addEmployeeNewFolder(request):
    if request.method == "POST":
        response = {}
        response['employee_id']   = request.POST['employee_id']
        response['college_id']   = getModelColumnById(SpUsers, request.POST['employee_id'], 'organization_id')
        response['room_id']      = request.POST['room_id']
        response['almira_id']    = request.POST['almira_id']
        response['rack_id']      = request.POST['rack_id']
        response['rack_name']    = getModelColumnById(TblClRack, request.POST['rack_id'], 'rack')
        
        employee_name = getModelColumnById(SpUsers, request.POST['employee_id'], 'first_name')
        employee_name = employee_name.lower()
        
        if request.POST['is_auto'] == '1':
            if request.POST['auto_file_name'] == '1':
                file_name = employee_name
            elif request.POST['auto_file_name'] == '2': 
                file_name = request.POST['employee_id']
            else:
                file_name = str(employee_name)+'_'+str(request.POST['employee_id'])
        else:
            file_name = request.POST['file_name']   

        if TblClEmployeeFileFolder.objects.filter(file_name=file_name, employee_id=request.POST['employee_id'], college_id=response['college_id'], room_id=request.POST['room_id'], almira_id=request.POST['almira_id'], rack_id=request.POST['rack_id']).exists():
            response['error']   = True
            response['message'] = "File Name already exists."
        else:    
            file_folder              = TblClEmployeeFileFolder()
            file_folder.employee_id   = request.POST['employee_id']
            file_folder.college_id   = response['college_id']
            file_folder.room_id      = request.POST['room_id']
            file_folder.almira_id    = request.POST['almira_id']
            file_folder.rack_id      = request.POST['rack_id']
            if request.POST['is_auto'] == '1':   
                file_folder.file_name    = file_name
            else:
                file_folder.file_name    = file_name    
            file_folder.created_by   = request.user.id
            file_folder.save()

            college_name = getModelColumnById(SpOrganizations, file_folder.college_id, 'organization_name')
            room_name    = getModelColumnById(TblClRoom, file_folder.room_id, 'room')
            almira_name  = getModelColumnById(TblClAlmirah, file_folder.almira_id, 'almirah')
            rack_name    = getModelColumnById(TblClRack, file_folder.rack_id, 'rack')
            folder='media/documents/'+str(college_name)+'/'+str(room_name)+'/'+str(almira_name)+'/'+str(rack_name)+'/'+str(file_folder.file_name) 
            try:
                os.mkdir(folder)
            except OSError:
                pass
            else:
                pass
            
        return JsonResponse(response)
    else:    
        context = {}
        context['employee_id']   = request.GET['employee_id']
        context['room_id']      = request.GET['room_id']
        context['almira_id']    = request.GET['almira_id']
        context['rack_id']      = request.GET['rack_id']
        template = 'user-management/add-employee-new-folder.html'
        return render(request,template,context)

@login_required
def addEmployeeNewFile(request):
    if request.method == "POST":
        file_path   = getModelColumnById(TblClEmployeeDocuments, request.POST['document_id'], 'document_path')
        baseurl     = settings.BASE_URL
        file_path   = file_path
        
        destination_folder = str(getModelColumnById(TblClRack, request.POST['rack_id'], 'path'))+'/'+str(request.POST['master_name'])
        print(destination_folder)
        shutil.move(file_path, destination_folder)

        document_file           = TblClEmployeeDocuments.objects.get(id=request.POST['document_id'])
        document_file.document_path = None
        document_file.save()
        
        file                    = TblClEmployeeFolderFiles()
        file.employee_id         = request.POST['employee_id']
        file.college_id         = request.POST['college_id']
        file.room_id            = request.POST['room_id']
        file.room_name          = request.POST['room_name']
        file.almira_id          = request.POST['almira_id']
        file.almira_name        = request.POST['almira_name']
        file.rack_id            = request.POST['rack_id']
        file.rack_name          = request.POST['rack_name']
        file.file_id            = request.POST['id']
        file.file_name          = request.POST['master_name']
        file.docket_no          = request.POST['docket_no']
        file.document_id        = request.POST['document_id']
        file.document_name      = request.POST['document_name']
        file.document_no        = request.POST['document_no']
        file.document_group     = request.POST['document_group']
        file.is_expiry    = request.POST['is_expiry']
        if request.POST['is_expiry'] == '1':   
            file.expiry_date    = datetime.strptime(request.POST['expiry_date'], '%d/%m/%Y').strftime('%Y-%m-%d')
        else:
            file.expiry_date    = None    
        file_path = file_path.split('/')
        file_path = file_path[-1]    
        file.document_path      = str(destination_folder)+'/'+ str(file_path) 
        file.tags               = request.POST['tags'] 
        file.created_by         = request.user.id
        file.save()

        response = {}
        response['id']           = request.POST['id']
        response['type']         = request.POST['type']
        response['college_id']   = request.POST['college_id']
        response['master_name']  = request.POST['master_name']
        response['employee_id']   = request.POST['employee_id']
        response['room_id']      = request.POST['room_id']
        response['room_name']    = request.POST['room_name']
        response['almira_id']    = request.POST['almira_id']
        response['almira_name']  = request.POST['almira_name']
        response['rack_id']      = request.POST['rack_id']
        response['rack_name']    = request.POST['rack_name']
        return JsonResponse(response)
    else:    
        context = {}
        context['document_id']  = request.GET['document_id']
        context['id']           = request.GET['id']
        context['type']         = request.GET['type']
        context['college_id']   = request.GET['college_id']
        context['master_name']  = request.GET['master_name']
        context['employee_id']   = request.GET['employee_id']
        context['room_id']      = request.GET['room_id']
        context['room_name']    = request.GET['room_name']
        context['almira_id']    = request.GET['almira_id']
        context['almira_name']  = request.GET['almira_name']
        context['rack_id']      = request.GET['rack_id']
        context['rack_name']    = request.GET['rack_name']
        context['document_list_count']  = TblClEmployeeDocuments.objects.filter(employee_id=request.GET['employee_id'], document_path__isnull=False).count()
        context['document_details']     = TblClEmployeeDocuments.objects.get(id=request.GET['document_id'], document_path__isnull=False)
        context['document_group_list']  = TblClDocumentGroup.objects.filter(status=1)
        template = 'user-management/create-employee-new-file.html'
        return render(request,template,context)

@login_required
def getEmployeeMasterDetails(request):
    context     = {}
    id          = request.POST['id']
    type        = request.POST['type']
    college_id  = request.POST['college_id']
    last_user_id  = request.POST['last_user_id']
    master_name = request.POST['master_name']
    if type == 'room_list':
        context['master']       = 'room'
        context['rooms']        = TblClRoom.objects.filter(college_id=college_id)
        context['college_id']   = college_id
        context['master_name']  = master_name
        context['last_user_id']  = last_user_id
    elif type == 'room':
        context['master']       = 'almira'
        context['almiras']      = TblClAlmirah.objects.filter(room_id=id)
        context['college_id']   = college_id
        context['master_name']  = master_name
        context['last_user_id']  = last_user_id
    elif type == 'almira':
        context['master']       = 'rack'
        context['racks']      = TblClRack.objects.filter(almira_id=id)
        context['college_id']   = college_id
        context['room_id']      = getModelColumnById(TblClAlmirah, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClAlmirah, id, 'room_name')
        context['almira_id']    = id
        context['master_name']  = getModelColumnById(TblClRoom, context['room_id'], 'room')  
        context['almira_name']  = getModelColumnById(TblClAlmirah, id, 'almirah')
        context['last_user_id']  = last_user_id
    elif type == 'rack':
        context['master']       = 'file'
        
        context['college_id']   = college_id
        context['last_user_id'] = last_user_id
        context['rack_id']      = id
        context['rack_name']    = getModelColumnById(TblClRack, id, 'rack')
        context['room_id']      = getModelColumnById(TblClRack, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClRack, id, 'room_name')
        context['almira_id']    = getModelColumnById(TblClRack, id, 'almira_id') 
        context['almira_name']  = getModelColumnById(TblClRack, id, 'almira_name') 
        context['files']        = TblClEmployeeFileFolder.objects.filter(college_id=college_id, employee_id=last_user_id, room_id=context['room_id'], almira_id=context['almira_id'], rack_id=context['rack_id'])   

        context['master_name']  = '' 
    elif type == 'file':
        context['master']       = 'folder_files'
        
        context['college_id']   = college_id
        context['last_user_id'] = last_user_id
        context['file_id']      = id
        context['file_name']    = getModelColumnById(TblClEmployeeFileFolder, id, 'file_name')
        context['room_id']      = getModelColumnById(TblClEmployeeFileFolder, id, 'room_id')
        context['room_name']    = getModelColumnById(TblClRoom, context['room_id'], 'room')
        context['almira_id']    = getModelColumnById(TblClEmployeeFileFolder, id, 'almira_id') 
        context['almira_name']  = getModelColumnById(TblClAlmirah, context['almira_id'], 'almirah') 
        context['rack_id']      = getModelColumnById(TblClEmployeeFileFolder, id, 'rack_id')
        context['rack_name']    = getModelColumnById(TblClRack, context['rack_id'], 'rack')
        context['folder_files'] = TblClEmployeeFolderFiles.objects.filter(file_id=id, college_id=college_id, employee_id=last_user_id, room_id=context['room_id'], almira_id=context['almira_id'], rack_id=context['rack_id'])    

        context['master_name']  = ''     
    baseurl = settings.BASE_URL
    
    context['baseurl']           = baseurl
    template                     = 'user-management/employee-master-details.html'
    return render(request, template, context)


@login_required
def resetCredential(request,user_id):
    if request.method == "POST":
        response = {}
        try:
            user = SpUsers.objects.get(id=request.POST['user_id'])
            password = make_password(request.POST['new_password'])
            user.password = password
            
            user.save()
            AuthtokenToken.objects.filter(user_id = request.POST['user_id'])
            
            

            if user.id :

                #Save Activity
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'User credentials updated'
                activity    = 'User credentials updated by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('User Management', 'User Management', heading, activity, request.user.id, user_name, 'updateVehiclePass.png', '1', 'web.png')
                
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,user.id,'firebase_token')
                employee_name = getUserName(user.id)

                message_title = "Password reset"
                message_body = "You password has been changed by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(user.id,'SpUsers','User Management','Password reset',message_title,message_body,notification_image,request.user.id,user_name,user.id,employee_name,'password.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['flag'] = True
                response['user_id'] = request.POST['user_id']
                response['message'] = "Record has been updated successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)
    else:
        context = {}
        context['user']     = SpUsers.objects.get(id=user_id)
        template = 'user-management/reset-user-credential.html'
        return render(request, template, context)


@login_required
def resetUserLocation(request):
    if request.method == "POST":
        response = {}
        try:
            user = SpUsers.objects.get(id=request.POST['user_id'])
            # password = make_password(request.POST['new_password'])
            # user.latitude = None
            # user.longitude = None
            user.device_id = None
            user.save()

            if user.id :

                #Save Activity
                user_name   = getUserName(request.user.id)#request.user.first_name+' '+request.user.middle_name+' '+request.user.last_name
                heading     = 'User location reset'
                activity    = 'User location reset by '+user_name+' on '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 
                saveActivity('User Management', 'User Management', heading, activity, request.user.id, user_name, 'updateVehiclePass.png', '1', 'web.png')
                
                #-----------------------------notify android block-------------------------------#
                userFirebaseToken = getModelColumnById(SpUsers,user.id,'firebase_token')
                employee_name = getUserName(user.id)

                message_title = "Location reset"
                message_body = "Your location has been reset by "+user_name
                notification_image = ""

                if userFirebaseToken is not None and userFirebaseToken != "" :
                    registration_ids = []
                    registration_ids.append(userFirebaseToken)
                    data_message = {}
                    data_message['id'] = 1
                    data_message['status'] = 'notification'
                    data_message['click_action'] = 'FLUTTER_NOTIFICATION_CLICK'
                    data_message['image'] = notification_image
                    send_android_notification(message_title,message_body,data_message,registration_ids)
                    #-----------------------------notify android block-------------------------------#

                #-----------------------------save notification block----------------------------#
                saveNotification(user.id,'SpUsers','User Management','Location reset',message_title,message_body,notification_image,request.user.id,user_name,user.id,employee_name,'password.png',2,'app.png',1,1)
                #-----------------------------save notification block----------------------------#

                response['flag'] = True
                response['user_id'] = request.POST['user_id']
                response['message'] = "Location has been reset successfully."
            else:
                response['flag'] = False
                response['message'] = "Failed to save"
        except Exception as e:
            response['error'] = False
            response['message'] = str(e)
        return JsonResponse(response)




def mapUserLeaves(role_id,user_id):
    try:
        leave_policy_id = SpRoleEntityMapping.objects.get(role_id = role_id , entity_type = "leave_policy")
        leave_policy_id = leave_policy_id.entity_id
    except SpRoleEntityMapping.DoesNotExist:
        leave_policy_id = None
    if leave_policy_id:
        leave_policy_dettail = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id)
        balance  = 0
        for policy_detail in leave_policy_dettail:
            leave_polcy_ledger = SpUserLeavePolicyLedger()
            leave_polcy_ledger.user_id = user_id
            leave_polcy_ledger.leave_policy_id = leave_policy_id
            leave_polcy_ledger.leave_type_id = policy_detail.leave_type_id
            current_month = datetime.today().strftime('%m')
            if int(current_month) == 1:
                month_leave_count = policy_detail.year_leave_count / 12
                leave_polcy_ledger.year_leave_count = policy_detail.year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
                
                leave_polcy_ledger.credit =  policy_detail.year_leave_count
                balance += policy_detail.year_leave_count
                leave_polcy_ledger.balance = balance
            else:   
                current_month = 12 - int(current_month)
                sub_leave_count = policy_detail.year_leave_count / 12
                year_leave_count = sub_leave_count*current_month
                month_leave_count = year_leave_count  / current_month
                leave_polcy_ledger.year_leave_count = year_leave_count
                leave_polcy_ledger.month_leave_count = round(month_leave_count,1)
                
                leave_polcy_ledger.credit =  year_leave_count
                balance +=year_leave_count
                leave_polcy_ledger.balance = balance
            
            leave_polcy_ledger.consecutive_leave = policy_detail.consecutive_leave 
            leave_polcy_ledger.save
            
            
            

@login_required
def userTrackingReport(request):
    context = {}
    if request.user.role_id == 0:    
        context['users']        = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
        context['organization'] = SpOrganizations.objects.all()
        context['page_title'] = "User Tracking"
        context['org_latitude'] = getConfigurationResult('org_latitude')
        context['org_longitude'] = getConfigurationResult('org_longitude')
    elif request.user.role_id!=0:
        dipo_id         = str(getModelColumnByColumnId(SpBasicDetails, 'user_id', request.user.id, 'mapped_dipo_id')).split(',')
        user_details    = SpBasicDetails.objects.filter(reduce(operator.or_, (Q(mapped_dipo_id__contains=x) for x in dipo_id))).values_list('user_id',flat=True)
        context['users']        = SpUsers.objects.filter(user_type=1,id__in = user_details).exclude(role_id = 0)
        context['organization'] = SpOrganizations.objects.all()
        context['page_title'] = "User Tracking"
        context['org_latitude'] = getConfigurationResult('org_latitude')
        context['org_longitude'] = getConfigurationResult('org_longitude')
    template = 'user-management/user-tracking-report.html'
    return render(request, template, context)

@login_required
def ajaxUserTracking(request,user_id):
    if 'track_date' in request.GET and request.GET['track_date'] != "" :
            today                   = request.GET['track_date']
            today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    else:
        today                   = date.today()

    distinct_travel_timing     = SpUserTracking.objects.filter(user_id=user_id, sync_date_time__icontains=today, accuracy__lte=50).values('sync_date_time').distinct()

    tracks      = SpUserTracking.objects.filter(user_id=user_id, sync_date_time__in=distinct_travel_timing, accuracy__lte=50).order_by('sync_date_time')
    trackss     = [trackss.id for trackss in tracks]
    len_track   = len(trackss)
    
    if tracks:
        distance_travelled  = SpUserTracking.objects.filter(user_id=user_id, sync_date_time__in=distinct_travel_timing).aggregate(Sum('distance_travelled'))
        distance_travelled  = round(distance_travelled['distance_travelled__sum']*0.001,2)
    else:
        distance_travelled = 0
    
    counter_of_lens = 0
    distance_travelled_list = []
    tracks_list = []
    
    try:
        start_days = SpUserAttendance.objects.filter(user_id = user_id,attendance_date_time__icontains = today).order_by('id').first()
    except SpUserAttendance.DoesNotExist:
        start_days = None
        
    if start_days:
        tracks_dict = {}
        tracks_dict['latitude']     = start_days.latitude
        tracks_dict['longitude']    = start_days.longitude
        tracks_list.append(tracks_dict)
    
    for id, user_last in enumerate(trackss):
        
        counter_of_lens += 1
        track = SpUserTracking.objects.get(id=trackss[id])
        if counter_of_lens < len_track:
            user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
        else:
            user_last_data = SpUserTracking.objects.get(id=trackss[id])
        R = 6373.0
        lat1 = radians(float(track.latitude))
        lon1 = radians(float(track.longitude))
        lat2 = radians(float(user_last_data.latitude))
        lon2 = radians(float(user_last_data.longitude))
        dlon = lon2 - lon1
        dlat = lat2 - lat1
        a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))
        distance = R * c
        meter_distance = float(distance * 1000)
        
        if meter_distance > 10:
            distance_travelled_list.append(meter_distance)
            tracks_dict = {}
            tracks_dict['latitude']     = track.latitude
            tracks_dict['longitude']    = track.longitude
            tracks_list.append(tracks_dict)
    
    context = {}            
    start_time  = ''
    end_time  = ''
    if len(tracks):
        context['tracks']       = tracks_list  #SpUserTracking.objects.filter(user_id=user_id, sync_date_time__in=distinct_travel_timing, accuracy__lte=50).order_by('sync_date_time')
        context['first_track']  = SpUserTracking.objects.filter(user_id=user_id,sync_date_time__contains=today).first()
        context['last_track'] = last_track  = SpUserTracking.objects.filter(user_id=user_id,sync_date_time__contains=today).last()
        
        tracks_dict = {}
        tracks_dict['latitude']     = last_track.latitude
        tracks_dict['longitude']    = last_track.longitude
        tracks_list.append(tracks_dict)
        
        start_day = SpUserAttendance.objects.filter(user_id = user_id,attendance_date_time__icontains = today).order_by('id').first()
        
        if start_day.start_time:
            start_time = str(start_day.start_time)
                
        context['end_day'] = end_day = SpUserAttendance.objects.filter(user_id = user_id,attendance_date_time__icontains = today).order_by('-id').first()
        
        if end_day.end_time:
            end_time = str(end_day.end_time)  
        distance_travelled              = sum(distance_travelled_list)
    context['distance_travelleds']  = round(float(distance_travelled)*0.001,2)
    context['start_time']           = start_time
    context['end_time']             = end_time
    context['user']                 = SpUsers.objects.get(id=user_id)
    template                        = 'user-management/ajax-user-tracking.html'
    return render(request, template, context)    

    


# @login_required
# def userTravelSummary(request):
#     today       = date.today()
    
#     users   = SpUserTracking.objects.filter(sync_date_time__year=today.strftime("%Y"), sync_date_time__month=today.strftime("%m")).values('user_id').distinct()
#     for user in users:
#         user['name'] = getUserName(user['user_id'])

#     user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today.strftime("%Y-%m-%d")).values('user_id').distinct().values('user_id', 'travel_charges')
#     for user_tracking in user_tracking_details:
#         tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today.strftime("%Y-%m-%d"), accuracy__lte=50).order_by('sync_date_time')
#         trackss     = [trackss.id for trackss in tracks]
#         len_track   = len(trackss)
        
#         counter_of_lens = 0
#         distance_travelled_list = []
#         for id, user_last in enumerate(trackss):
            
#             counter_of_lens += 1
#             track = SpUserTracking.objects.get(id=trackss[id])
#             if counter_of_lens < len_track:
#                 user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#             else:
#                 user_last_data = SpUserTracking.objects.get(id=trackss[id])
#             R = 6373.0
#             lat1 = radians(float(track.latitude))
#             lon1 = radians(float(track.longitude))
#             lat2 = radians(float(user_last_data.latitude))
#             lon2 = radians(float(user_last_data.longitude))
#             dlon = lon2 - lon1
#             dlat = lat2 - lat1
#             a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#             c = 2 * atan2(sqrt(a), sqrt(1 - a))
#             distance = R * c
#             meter_distance = float(distance * 1000)
            
#             if meter_distance > 10:
#                 distance_travelled_list.append(meter_distance)
                
#         if len(distance_travelled_list) > 0:
#             distance_travelled                  = sum(distance_travelled_list)
#             user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#             if user_tracking['distance_travelled'] > 0:
#                 user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                 user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#             else:
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0
#             user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#             # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#             user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#         else:
#             user_tracking['distance_travelled']     = 0
#             user_tracking['charges']                = 0
#             user_tracking['total_charges']          = 0
#             user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#             # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#             user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')

#     context = {}
#     context['today_date']               = today.strftime("%d/%m/%Y")
#     context['users']                    = users
#     context['user_tracking_details']    = user_tracking_details
#     context['month_date']               = date.today().strftime("%m/%Y")
#     context['page_title']               = "User Travel Summary"
#     template = 'user-management/user-travel-summary.html'
#     return render(request, template, context)

# @login_required
# def ajaxuserTravelSummary(request):
#     today                   = request.GET['travel_date']
#     today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
#     context = {}
    
#     if request.GET['time_period'] == '1':
#         is_user = 0
#         user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today)
#         if request.GET['user_id']:
#             user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
#         user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
#         for user_tracking in user_tracking_details:
#             tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today, accuracy__lte=50).order_by('sync_date_time')
#             trackss     = [trackss.id for trackss in tracks]
#             len_track   = len(trackss)
            
#             counter_of_lens = 0
#             distance_travelled_list = []
#             for id, user_last in enumerate(trackss):
                
#                 counter_of_lens += 1
#                 track = SpUserTracking.objects.get(id=trackss[id])
#                 if counter_of_lens < len_track:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                 else:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                 R = 6373.0
#                 lat1 = radians(float(track.latitude))
#                 lon1 = radians(float(track.longitude))
#                 lat2 = radians(float(user_last_data.latitude))
#                 lon2 = radians(float(user_last_data.longitude))
#                 dlon = lon2 - lon1
#                 dlat = lat2 - lat1
#                 a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                 c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                 distance = R * c
#                 meter_distance = float(distance * 1000)
                
#                 if meter_distance > 10:
#                     distance_travelled_list.append(meter_distance)
                    
#             if len(distance_travelled_list) > 0:
#                 distance_travelled                  = sum(distance_travelled_list)
#                 user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 if user_tracking['distance_travelled'] > 0:
#                     user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                     user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#                 else:
#                     user_tracking['charges']            = 0
#                     user_tracking['total_charges']      = 0
#                 user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#             else:
#                 user_tracking['distance_travelled'] = 0
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0
#                 user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')    
#         context['user_tracking_details']            = user_tracking_details       
#     else:
#         user_tracking_detail = []
#         if request.GET['user_id']:
#             is_user = 1
#             travel_month = request.GET['travel_month_picker']
#             travel_month = travel_month.split('/')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
            
#             month_list = days_in_months(year,month)
            
#             for months in month_list:
#                 user_trackings={}
#                 user_trackings['month_date'] = months
#                 month_date             = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                
#                 tracks      = SpUserTracking.objects.filter(user_id=request.GET['user_id'], sync_date_time__icontains=month_date, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
                
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=request.GET['user_id'],sync_date_time__month=month,sync_date_time__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(request.GET['user_id'])
#                 user_trackings['user_id']            = request.GET['user_id']
#                 # user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',request.GET['user_id'],'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',request.GET['user_id'],'reporting_to_name')
               
#                 user_tracking_detail.append(user_trackings) 
#         else:
#             is_user = 0    
#             travel_month = request.GET['travel_month_picker']
#             travel_month = travel_month.split('/')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
#             user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__month=month,sync_date_time__year=year).values_list('user_id',flat=True).distinct()
#             if request.GET['user_id']:
#                 user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
#             for user_tracking in user_tracking_details:
#                 user_trackings={}
#                 user_trackings['month_date'] = ''
                
#                 tracks      = SpUserTracking.objects.filter(user_id=user_tracking, sync_date_time__month=month, sync_date_time__year=year, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])   
                    
#                     start_time = datetime.strptime(str(track.sync_date_time), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                     end_time   = datetime.strptime(str(user_last_data.sync_date_time), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                    
#                     if str(start_time) == str(end_time):
#                         R = 6373.0
#                         lat1 = radians(float(track.latitude))
#                         lon1 = radians(float(track.longitude))
#                         lat2 = radians(float(user_last_data.latitude))
#                         lon2 = radians(float(user_last_data.longitude))
#                         dlon = lon2 - lon1
#                         dlat = lat2 - lat1
#                         a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                         c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                         distance = R * c
#                         meter_distance = float(distance * 1000)
                        
#                         if meter_distance > 10:
#                             distance_travelled_list.append(meter_distance)
                        
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_tracking,sync_date_time__month=month,sync_date_time__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_tracking)
#                 user_trackings['user_id']            = user_tracking
#                 # user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking,'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking,'reporting_to_name')
                
#                 user_tracking_detail.append(user_trackings)  
#         context['user_tracking_details']    = user_tracking_detail
#     if request.GET['travel_month_picker']:
#         context['month_date']               = request.GET['travel_month_picker']
#     else:
#         context['month_date']               = date.today().strftime("%m/%Y")
#     context['time_period']              = request.GET['time_period']   
#     context['is_user']      = is_user 
#     template = 'user-management/ajax-user-travel-summary-report.html'
#     return render(request, template, context)


# #get export user summary
# @login_required
# def exportUserTravelSummary(request, travel_date, user_id, travel_month_picker, time_period):
#     today                   = travel_date
          
#     if time_period == '1':
#         is_user = 0
#         user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today)
#         if user_id!='0':
#             user_tracking_details = user_tracking_details.filter(user_id=user_id)
#         user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
#         for user_tracking in user_tracking_details:
#             tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today, accuracy__lte=50).order_by('sync_date_time')
#             trackss     = [trackss.id for trackss in tracks]
#             len_track   = len(trackss)
            
#             counter_of_lens = 0
#             distance_travelled_list = []
#             for id, user_last in enumerate(trackss):
                
#                 counter_of_lens += 1
#                 track = SpUserTracking.objects.get(id=trackss[id])
#                 if counter_of_lens < len_track:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                 else:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                 R = 6373.0
#                 lat1 = radians(float(track.latitude))
#                 lon1 = radians(float(track.longitude))
#                 lat2 = radians(float(user_last_data.latitude))
#                 lon2 = radians(float(user_last_data.longitude))
#                 dlon = lon2 - lon1
#                 dlat = lat2 - lat1
#                 a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                 c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                 distance = R * c
#                 meter_distance = float(distance * 1000)
                
#                 if meter_distance > 10:
#                     distance_travelled_list.append(meter_distance)
                    
#             if len(distance_travelled_list) > 0:
#                 distance_travelled                  = sum(distance_travelled_list)
#                 user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 if user_tracking['distance_travelled'] > 0:
#                     user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                     user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#                 else:
#                     user_tracking['charges']            = 0
#                     user_tracking['total_charges']      = 0   
#                 user_tracking['user_name']          = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']        = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#             else:
#                 user_tracking['distance_travelled'] = 0
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0   
#                 user_tracking['user_name']          = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']        = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         )
#         response['Content-Disposition'] = 'attachment; filename=user_daily_travel_summary.xlsx'.format(
#             date=datetime.now().strftime('%Y-%m-%d'),
#         )
#         workbook = Workbook()

#         # Define some styles and formatting that will be later used for cells
#         header_font = Font(name='Calibri', bold=True)
#         centered_alignment = Alignment(horizontal='left')
#         thin = Side(border_style="thin", color="303030") 
#         black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         wrapped_alignment = Alignment(
#             vertical='top',
#             horizontal='left',
#             wrap_text=True
#         )

#         header_alignment = Alignment(
#             vertical='top',
#             horizontal='center',
#             wrap_text=True
#         )
        
#         # Get active worksheet/tab
#         worksheet = workbook.active
#         worksheet.title = 'User Daily Travel Summary'
#         worksheet.merge_cells('A1:A1') 
        
#         worksheet.page_setup.orientation = 'landscape'
#         worksheet.page_setup.paperSize = 9
#         worksheet.page_setup.fitToPage = True
        
#         worksheet = workbook.worksheets[0]
#         img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
#         img.height = 50
#         img.alignment = 'center'
#         img.anchor = 'A1'
#         worksheet.add_image(img)
        
#         column_length = 6
        
#         worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#         worksheet.cell(row=1, column=2).value = 'User Daily Travel Summary as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
#         worksheet.cell(row=1, column=2).font = header_font
#         worksheet.cell(row=1, column=2).alignment = header_alignment
#         worksheet.cell(row=1, column=column_length).border = black_border
#         worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
#         worksheet.cell(row=1, column=2).fill = PatternFill()

#         # Define the titles for columns
#         # columns = []
#         row_num = 1
#         worksheet.row_dimensions[1].height = 40
        
#         # Define the titles for columns
#         columns = []

#         columns += [ 'Employee Name' ]
#         columns += [ 'Reporting To' ]
#         columns += [ 'Headquarter Name' ]
#         columns += [ 'Distance Travelled in Kilometer' ]
#         columns += [ 'Charges' ]
#         columns += [ 'Total Charges' ]

#         row_num = 2

#         # Assign the titles for each cell of the header
#         for col_num, column_title in enumerate(columns, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = column_title
#             cell.font = header_font
#             cell.alignment = centered_alignment
#             cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#             cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#             column_letter = get_column_letter(col_num)
#             column_dimensions = worksheet.column_dimensions[column_letter]
#             column_dimensions.width = 32

#         for user_tracking in user_tracking_details:
#             row_num += 1
#             # Define the data for each cell in the row 
#             row = []
#             row += [ user_tracking['user_name'] ]
#             row += [ user_tracking['reporting_to_name'] ]
#             if user_tracking['headquarter']:
#                 row += [ user_tracking['headquarter'] ]
#             else:
#                 row += ['']
#             row += [ user_tracking['distance_travelled'] ]
#             row += [ user_tracking['charges'] ]
#             row += [ user_tracking['total_charges'] ]           
        
#             # Assign the data for each cell of the row 
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#                 cell.alignment = wrapped_alignment
#                 cell.border = black_border  

#         wrapped_alignment = Alignment(
#             horizontal='center',
#             wrap_text=True
#         )

#         row_num += 1
#         last_row = row_num
#         worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
#         worksheet.row_dimensions[last_row].height = 20
#         worksheet.cell(row=last_row, column=1).value = 'Generated By sakhiMilk'
#         worksheet.cell(row=last_row, column=1).font = header_font
#         worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#         worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
#         worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

#         workbook.save(response)
#         return response

#     else:
#         user_tracking_detail=[]
#         if user_id!='0':
#             is_user = 1
#             travel_month = travel_month_picker
#             travel_month = travel_month.split('-')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
            
#             month_list = days_in_months(year,month)
            
#             for months in month_list:
#                 user_trackings={}
#                 user_trackings['month_date'] = months
#                 month_date             = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
#                 tracks      = SpUserTracking.objects.filter(user_id=user_id, sync_date_time__icontains=month_date, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
                
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_id,created_at__month=month,created_at__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_id)
#                 user_trackings['user_id']            = user_id
#                 user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',user_id,'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_id,'reporting_to_name')
               
#                 user_tracking_detail.append(user_trackings) 
#         else:
#             is_user = 0
#             travel_month = travel_month_picker
#             travel_month = travel_month.split('-')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
#             user_tracking_detail=[]
#             user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__month=month,sync_date_time__year=year).values_list('user_id',flat=True).distinct()
#             if user_id!='0':
#                 user_tracking_details = user_tracking_details.filter(user_id=user_id)
#             for user_tracking in user_tracking_details:
#                 user_trackings={}
#                 user_trackings['month_date'] = ''
#                 tracks      = SpUserTracking.objects.filter(user_id=user_tracking, sync_date_time__month=month, sync_date_time__year=year, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_tracking,created_at__month=month,created_at__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_tracking)
#                 user_trackings['user_id']            = user_tracking
#                 user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking,'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking,'reporting_to_name')
                
#                 user_tracking_detail.append(user_trackings)
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         )
#         response['Content-Disposition'] = 'attachment; filename=user_monthly_travel_summary.xlsx'.format(
#             date=datetime.now().strftime('%Y-%m-%d'),
#         )
#         workbook = Workbook()

#         # Define some styles and formatting that will be later used for cells
#         header_font = Font(name='Calibri', bold=True)
#         centered_alignment = Alignment(horizontal='left')
#         thin = Side(border_style="thin", color="303030") 
#         black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         wrapped_alignment = Alignment(
#             vertical='top',
#             horizontal='left',
#             wrap_text=True
#         )

#         header_alignment = Alignment(
#             vertical='top',
#             horizontal='center',
#             wrap_text=True
#         )
        
#         # Get active worksheet/tab
#         worksheet = workbook.active
#         worksheet.title = 'User Monthly Travel Summary'
#         worksheet.merge_cells('A1:A1') 
        
#         worksheet.page_setup.orientation = 'landscape'
#         worksheet.page_setup.paperSize = 9
#         worksheet.page_setup.fitToPage = True
        
#         worksheet = workbook.worksheets[0]
#         img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
#         img.height = 50
#         img.alignment = 'center'
#         img.anchor = 'A1'
#         worksheet.add_image(img)
        
#         column_length = 6
#         month_name = datetime(int(travel_month[1]),int(travel_month[0]),1).strftime( '%B' )
#         worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#         if is_user == 1:
#             worksheet.cell(row=1, column=2).value = ''+getUserName(user_id)+' Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
#         else:
#             worksheet.cell(row=1, column=2).value = 'Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
#         worksheet.cell(row=1, column=2).font = header_font
#         worksheet.cell(row=1, column=2).alignment = header_alignment
#         worksheet.cell(row=1, column=column_length).border = black_border
#         worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
#         worksheet.cell(row=1, column=2).fill = PatternFill()

#         # Define the titles for columns
#         # columns = []
#         row_num = 1
#         worksheet.row_dimensions[1].height = 40
        
#         # Define the titles for columns
#         columns = []
#         if is_user == 1:
#             columns += [ 'Date' ]
#         else:    
#             columns += [ 'Employee Name' ]
#         columns += [ 'Reporting To' ]
#         columns += [ 'Headquarter Name' ]
#         columns += [ 'Distance Travelled in Kilometer' ]
#         columns += [ 'Charges' ]
#         columns += [ 'Total Charges' ]

#         row_num = 2

#         # Assign the titles for each cell of the header
#         for col_num, column_title in enumerate(columns, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = column_title
#             cell.font = header_font
#             cell.alignment = centered_alignment
#             cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#             cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#             column_letter = get_column_letter(col_num)
#             column_dimensions = worksheet.column_dimensions[column_letter]
#             column_dimensions.width = 32

#         for user_tracking in user_tracking_detail:
#             row_num += 1
#             # Define the data for each cell in the row 
#             row = []
#             if user_tracking['month_date']:
#                 row += [ user_tracking['month_date'] ]
#             else:
#                 row += [ user_tracking['user_name'] ]    
#             row += [ user_tracking['reporting_to_name'] ]
#             if user_tracking['headquarter']:
#                 row += [ user_tracking['headquarter'] ]
#             else:
#                 row += [''] 
#             row += [ user_tracking['distance_travelled'] ]
#             row += [ user_tracking['charges'] ]
#             row += [ user_tracking['total_charges'] ]           
        
#             # Assign the data for each cell of the row 
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#                 cell.alignment = wrapped_alignment
#                 cell.border = black_border  

#         wrapped_alignment = Alignment(
#             horizontal='center',
#             wrap_text=True
#         )

#         row_num += 1
#         last_row = row_num
#         worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
#         worksheet.row_dimensions[last_row].height = 20
#         worksheet.cell(row=last_row, column=1).value = 'Generated By SakhiMilk'
#         worksheet.cell(row=last_row, column=1).font = header_font
#         worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#         worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
#         worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

#         workbook.save(response)
#         return response     

# @login_required
# def userTravelSummary(request):
#     today       = date.today()
    
#     users   = SpUserTracking.objects.filter(sync_date_time__year=today.strftime("%Y"), sync_date_time__month=today.strftime("%m")).values('user_id').distinct()
#     for user in users:
#         user['name'] = getUserName(user['user_id'])

#     user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today.strftime("%Y-%m-%d")).values('user_id').distinct().values('user_id', 'travel_charges')
#     for user_tracking in user_tracking_details:
#         tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today.strftime("%Y-%m-%d"), accuracy__lte=50).order_by('sync_date_time')
#         trackss     = [trackss.id for trackss in tracks]
#         len_track   = len(trackss)
        
#         counter_of_lens = 0
#         distance_travelled_list = []
#         for id, user_last in enumerate(trackss):
            
#             counter_of_lens += 1
#             track = SpUserTracking.objects.get(id=trackss[id])
#             if counter_of_lens < len_track:
#                 user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#             else:
#                 user_last_data = SpUserTracking.objects.get(id=trackss[id])
#             R = 6373.0
#             lat1 = radians(float(track.latitude))
#             lon1 = radians(float(track.longitude))
#             lat2 = radians(float(user_last_data.latitude))
#             lon2 = radians(float(user_last_data.longitude))
#             dlon = lon2 - lon1
#             dlat = lat2 - lat1
#             a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#             c = 2 * atan2(sqrt(a), sqrt(1 - a))
#             distance = R * c
#             meter_distance = float(distance * 1000)
            
#             if meter_distance > 10:
#                 distance_travelled_list.append(meter_distance)
                
#         if len(distance_travelled_list) > 0:
#             distance_travelled                  = sum(distance_travelled_list)
#             user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#             if user_tracking['distance_travelled'] > 0:
#                 user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                 user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#             else:
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0
#             user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#             # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#             user_tracking['reporting_to_name']     = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#         else:
#             user_tracking['distance_travelled'] = 0
#             user_tracking['charges']            = 0
#             user_tracking['total_charges']      = 0
#             user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#             # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#             user_tracking['reporting_to_name']     = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')

#     context = {}
#     context['today_date']               = today.strftime("%d/%m/%Y")
#     context['users']                    = users
#     context['user_tracking_details']    = user_tracking_details
#     context['month_date']               = date.today().strftime("%m/%Y")
#     context['page_title']               = "User Travel Summary"
#     template = 'user-management/user-travel-summary.html'
#     return render(request, template, context)

@login_required
def userTravelSummary(request):
    today       = date.today()
    if request.user.role_id == 0 or request.user.role_id == 120 or request.user.role_id == 132 or request.user.role_id == 121:
        users   = SpUserTravelHistory.objects.filter().values('user_id').distinct().values('user_id')
        user_tracking_details   = SpUserTravelHistory.objects.filter(treval_date__icontains=today.strftime("%Y-%m-%d")).values('user_id').distinct().values('user_id', 'user_name', 'distance_in_km', 'charge', 'travel_amount')
    else:
        dipo_id         = str(getModelColumnByColumnId(SpBasicDetails, 'user_id', request.user.id, 'mapped_dipo_id')).split(',')
        user_details    = SpBasicDetails.objects.filter(reduce(operator.or_, (Q(mapped_dipo_id__contains=x) for x in dipo_id))).values_list('user_id',flat=True)
        users   = SpUserTravelHistory.objects.filter(user_id__in = user_details).values('user_id').distinct().values('user_id')
        user_tracking_details   = SpUserTravelHistory.objects.filter(treval_date__icontains=today.strftime("%Y-%m-%d"),user_id__in = user_details).values('user_id').distinct().values('user_id', 'user_name', 'distance_in_km', 'charge', 'travel_amount')
    for user in users:
        if user:
            user['name']        = getUserName(user['user_id'])
        else:
            user['name'] = '-'
        #user['emp_sap_id'] = getModelColumnById(SpUsers, user['user_id'], 'emp_sap_id')
   
        
    
    for user_tracking_detail in user_tracking_details:
        user_tracking_detail['emp_sap_id'] = getModelColumnById(SpUsers, user_tracking_detail['user_id'], 'emp_sap_id')
    users   = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context = {}
    context['today_date']               = today.strftime("%d/%m/%Y")
    context['users']                    = users
    print('------------------',users)
    context['user_tracking_details']    = user_tracking_details
    context['month_date']               = date.today().strftime("%m/%Y")
    context['page_title']               = "User Travel Summary"
    template = 'user-management/user-travel-summary.html'
    return render(request, template, context)

# @login_required
# def ajaxuserTravelSummary(request):
#     today                   = request.GET['travel_date']
#     today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
#     context = {}
    
#     if request.GET['time_period'] == '1':
#         is_user = 0
#         user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today)
#         if request.GET['user_id']:
#             user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
#         user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
#         for user_tracking in user_tracking_details:
#             tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today, accuracy__lte=50).order_by('sync_date_time')
#             trackss     = [trackss.id for trackss in tracks]
#             len_track   = len(trackss)
            
#             counter_of_lens = 0
#             distance_travelled_list = []
#             for id, user_last in enumerate(trackss):
                
#                 counter_of_lens += 1
#                 track = SpUserTracking.objects.get(id=trackss[id])
#                 if counter_of_lens < len_track:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                 else:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                 R = 6373.0
#                 lat1 = radians(float(track.latitude))
#                 lon1 = radians(float(track.longitude))
#                 lat2 = radians(float(user_last_data.latitude))
#                 lon2 = radians(float(user_last_data.longitude))
#                 dlon = lon2 - lon1
#                 dlat = lat2 - lat1
#                 a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                 c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                 distance = R * c
#                 meter_distance = float(distance * 1000)
                
#                 if meter_distance > 10:
#                     distance_travelled_list.append(meter_distance)
                    
#             if len(distance_travelled_list) > 0:
#                 distance_travelled                  = sum(distance_travelled_list)
#                 user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 if user_tracking['distance_travelled'] > 0:
#                     user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                     user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#                 else:
#                     user_tracking['charges']            = 0
#                     user_tracking['total_charges']      = 0
#                 user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#             else:
#                 user_tracking['distance_travelled'] = 0
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0
#                 user_tracking['user_name']              = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']            = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']      = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')    
#         context['user_tracking_details']            = user_tracking_details       
#     else:
#         user_tracking_detail = []
#         if request.GET['user_id']:
#             is_user = 1
#             travel_month = request.GET['travel_month_picker']
#             travel_month = travel_month.split('/')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
            
#             month_list = days_in_months(year,month)
            
#             for months in month_list:
#                 user_trackings={}
#                 user_trackings['month_date'] = months
#                 month_date             = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
                
#                 tracks      = SpUserTracking.objects.filter(user_id=request.GET['user_id'], sync_date_time__icontains=month_date, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
                
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=request.GET['user_id'],sync_date_time__month=month,sync_date_time__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(request.GET['user_id'])
#                 user_trackings['user_id']            = request.GET['user_id']
#                 # user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',request.GET['user_id'],'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',request.GET['user_id'],'reporting_to_name')
               
#                 user_tracking_detail.append(user_trackings) 
#         else:
#             is_user = 0    
#             travel_month = request.GET['travel_month_picker']
#             travel_month = travel_month.split('/')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
#             user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__month=month,sync_date_time__year=year).values_list('user_id',flat=True).distinct()
#             if request.GET['user_id']:
#                 user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
#             for user_tracking in user_tracking_details:
#                 user_trackings={}
#                 user_trackings['month_date'] = ''
                
#                 tracks      = SpUserTracking.objects.filter(user_id=user_tracking, sync_date_time__month=month, sync_date_time__year=year, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])   
                    
#                     start_time = datetime.strptime(str(track.sync_date_time), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
#                     end_time   = datetime.strptime(str(user_last_data.sync_date_time), '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d')
                    
#                     if str(start_time) == str(end_time):
#                         R = 6373.0
#                         lat1 = radians(float(track.latitude))
#                         lon1 = radians(float(track.longitude))
#                         lat2 = radians(float(user_last_data.latitude))
#                         lon2 = radians(float(user_last_data.longitude))
#                         dlon = lon2 - lon1
#                         dlat = lat2 - lat1
#                         a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                         c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                         distance = R * c
#                         meter_distance = float(distance * 1000)
                        
#                         if meter_distance > 10:
#                             distance_travelled_list.append(meter_distance)
                        
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_tracking,sync_date_time__month=month,sync_date_time__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_tracking)
#                 user_trackings['user_id']            = user_tracking
#                 # user_trackings['headquarter']         = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking,'headquarter_name')
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking,'reporting_to_name')
                
#                 user_tracking_detail.append(user_trackings)  
#         context['user_tracking_details']    = user_tracking_detail
#     if request.GET['travel_month_picker']:
#         context['month_date']               = request.GET['travel_month_picker']
#     else:
#         context['month_date']               = date.today().strftime("%m/%Y")
#     context['time_period']              = request.GET['time_period']   
#     context['is_user']      = is_user 
#     template = 'user-management/ajax-user-travel-summary-report.html'
#     return render(request, template, context)

@login_required
def ajaxuserTravelSummary(request):
    today                   = request.GET['travel_date']
    today                   = datetime.strptime(str(today), '%d/%m/%Y').strftime('%Y-%m-%d')
    context = {}
    
    if request.GET['time_period'] == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        if request.GET['user_id']:
            user_tracking_details = user_tracking_details.filter(user_id=request.GET['user_id'])
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])
        context['user_tracking_details']    = user_tracking_details       
    else:
        user_tracking_detail = []
        if request.GET['user_id']:
            travel_month = request.GET['travel_month_picker']
            travel_month = travel_month.split('/')
            year  = int(travel_month[1])
            month = int(travel_month[0])
            
            last_day_of_month = calendar.monthrange(year,month)[1]
            if int(month) < 10:
                month = '0'+str(month)
                 
            for x in range(last_day_of_month):
                x = x+1
                if x < 10:
                    x = '0'+str(x)
                travel_date  = str(year)+'-'+str(month)+'-'+str(x)
                travel_dates = str(x)+'/'+str(month)+'/'+str(year)
    
                user_trackings = {}
                
                try:
                    distance_travelled  = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
                except SpUserTracking.DoesNotExist:
                    distance_travelled = None
    
                    
                user_trackings['travel_date']          = travel_dates
                if distance_travelled['distance_travelled__sum']:
                    user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
                else:
                    user_trackings['distance_travelled'] = 0.0
                if user_trackings['distance_travelled'] > 0:
                    travel_charges       = SpUserTracking.objects.filter(user_id=request.GET['user_id'], created_at__icontains=travel_date).values('travel_charges').first()        
                    user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
                    user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
                else:
                    user_trackings['charges']            = 0
                    user_trackings['total_charges']      = 0
                user_trackings['user_name']          = getUserName(request.GET['user_id'])
                user_tracking_detail.append(user_trackings)  
        context['user_tracking_details']    = user_tracking_detail
        
    if request.GET['travel_month_picker']:
        context['month_date']               = request.GET['travel_month_picker']
    else:
        context['month_date']               = date.today().strftime("%m/%Y")
    context['time_period']              = request.GET['time_period']    
    template = 'user-management/ajax-user-travel-summary-report.html'
    return render(request, template, context)

#get export user summary
# @login_required
# def exportUserTravelSummary(request, travel_date, user_id, travel_month_picker, time_period):
#     today                   = travel_date
          
#     if time_period == '1':
#         is_user = 0
#         user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__icontains=today)
#         if user_id!='0':
#             user_tracking_details = user_tracking_details.filter(user_id=user_id)
#         user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
#         for user_tracking in user_tracking_details:
#             tracks      = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], sync_date_time__icontains=today, accuracy__lte=50).order_by('sync_date_time')
#             trackss     = [trackss.id for trackss in tracks]
#             len_track   = len(trackss)
            
#             counter_of_lens = 0
#             distance_travelled_list = []
#             for id, user_last in enumerate(trackss):
                
#                 counter_of_lens += 1
#                 track = SpUserTracking.objects.get(id=trackss[id])
#                 if counter_of_lens < len_track:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                 else:
#                     user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                 R = 6373.0
#                 lat1 = radians(float(track.latitude))
#                 lon1 = radians(float(track.longitude))
#                 lat2 = radians(float(user_last_data.latitude))
#                 lon2 = radians(float(user_last_data.longitude))
#                 dlon = lon2 - lon1
#                 dlat = lat2 - lat1
#                 a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                 c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                 distance = R * c
#                 meter_distance = float(distance * 1000)
                
#                 if meter_distance > 10:
#                     distance_travelled_list.append(meter_distance)
                    
#             if len(distance_travelled_list) > 0:
#                 distance_travelled                  = sum(distance_travelled_list)
#                 user_tracking['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 if user_tracking['distance_travelled'] > 0:
#                     user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
#                     user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
#                 else:
#                     user_tracking['charges']            = 0
#                     user_tracking['total_charges']      = 0   
#                 user_tracking['user_name']          = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']        = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#             else:
#                 user_tracking['distance_travelled'] = 0
#                 user_tracking['charges']            = 0
#                 user_tracking['total_charges']      = 0   
#                 user_tracking['user_name']          = getUserName(user_tracking['user_id'])
#                 # user_tracking['headquarter']        = getModelColumnByColumnId(SpBasicDetails,'user_id',user_tracking['user_id'],'headquarter_name')
#                 user_tracking['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking['user_id'],'reporting_to_name')
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         )
#         response['Content-Disposition'] = 'attachment; filename=user_daily_travel_summary.xlsx'.format(
#             date=datetime.now().strftime('%Y-%m-%d'),
#         )
#         workbook = Workbook()

#         # Define some styles and formatting that will be later used for cells
#         header_font = Font(name='Calibri', bold=True)
#         centered_alignment = Alignment(horizontal='left')
#         thin = Side(border_style="thin", color="303030") 
#         black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         wrapped_alignment = Alignment(
#             vertical='top',
#             horizontal='left',
#             wrap_text=True
#         )

#         header_alignment = Alignment(
#             vertical='top',
#             horizontal='center',
#             wrap_text=True
#         )
        
#         # Get active worksheet/tab
#         worksheet = workbook.active
#         worksheet.title = 'User Daily Travel Summary'
#         worksheet.merge_cells('A1:A1') 
        
#         worksheet.page_setup.orientation = 'landscape'
#         worksheet.page_setup.paperSize = 9
#         worksheet.page_setup.fitToPage = True
        
#         worksheet = workbook.worksheets[0]
#         img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
#         img.height = 50
#         img.alignment = 'center'
#         img.anchor = 'A1'
#         worksheet.add_image(img)
        
#         column_length = 6
        
#         worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#         worksheet.cell(row=1, column=2).value = 'User Daily Travel Summary as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
#         worksheet.cell(row=1, column=2).font = header_font
#         worksheet.cell(row=1, column=2).alignment = header_alignment
#         worksheet.cell(row=1, column=column_length).border = black_border
#         worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
#         worksheet.cell(row=1, column=2).fill = PatternFill()

#         # Define the titles for columns
#         # columns = []
#         row_num = 1
#         worksheet.row_dimensions[1].height = 40
        
#         # Define the titles for columns
#         columns = []

#         columns += [ 'Employee Name' ]
#         columns += [ 'Reporting To' ]
#         columns += [ 'Distance Travelled in Kilometer' ]
#         columns += [ 'Charges' ]
#         columns += [ 'Total Charges' ]

#         row_num = 2

#         # Assign the titles for each cell of the header
#         for col_num, column_title in enumerate(columns, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = column_title
#             cell.font = header_font
#             cell.alignment = centered_alignment
#             cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#             cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#             column_letter = get_column_letter(col_num)
#             column_dimensions = worksheet.column_dimensions[column_letter]
#             column_dimensions.width = 32

#         for user_tracking in user_tracking_details:
#             row_num += 1
#             # Define the data for each cell in the row 
#             row = []
#             row += [ user_tracking['user_name'] ]
#             row += [ user_tracking['reporting_to_name'] ]
            
#             row += [ user_tracking['distance_travelled'] ]
#             row += [ user_tracking['charges'] ]
#             row += [ user_tracking['total_charges'] ]           
        
#             # Assign the data for each cell of the row 
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#                 cell.alignment = wrapped_alignment
#                 cell.border = black_border  

#         wrapped_alignment = Alignment(
#             horizontal='center',
#             wrap_text=True
#         )

#         row_num += 1
#         last_row = row_num
#         worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
#         worksheet.row_dimensions[last_row].height = 20
#         worksheet.cell(row=last_row, column=1).value = ' Generated By Sakhi Milk'
#         worksheet.cell(row=last_row, column=1).font = header_font
#         worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#         worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
#         worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

#         workbook.save(response)
#         return response

#     else:
#         user_tracking_detail=[]
#         if user_id!='0':
#             is_user = 1
#             travel_month = travel_month_picker
#             travel_month = travel_month.split('-')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
            
#             month_list = days_in_months(year,month)
            
#             for months in month_list:
#                 user_trackings={}
#                 user_trackings['month_date'] = months
#                 month_date             = datetime.strptime(str(months), '%d/%m/%Y').strftime('%Y-%m-%d')
#                 tracks      = SpUserTracking.objects.filter(user_id=user_id, sync_date_time__icontains=month_date, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
                
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_id,created_at__month=month,created_at__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_id)
#                 user_trackings['user_id']            = user_id
 
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_id,'reporting_to_name')
               
#                 user_tracking_detail.append(user_trackings) 
#         else:
#             is_user = 0
#             travel_month = travel_month_picker
#             travel_month = travel_month.split('-')
#             year  = int(travel_month[1])
#             month = int(travel_month[0])
#             user_tracking_detail=[]
#             user_tracking_details   = SpUserTracking.objects.filter(sync_date_time__month=month,sync_date_time__year=year).values_list('user_id',flat=True).distinct()
#             if user_id!='0':
#                 user_tracking_details = user_tracking_details.filter(user_id=user_id)
#             for user_tracking in user_tracking_details:
#                 user_trackings={}
#                 user_trackings['month_date'] = ''
#                 tracks      = SpUserTracking.objects.filter(user_id=user_tracking, sync_date_time__month=month, sync_date_time__year=year, accuracy__lte=50).order_by('sync_date_time')
#                 trackss     = [trackss.id for trackss in tracks]
#                 len_track   = len(trackss)
                
#                 counter_of_lens = 0
#                 distance_travelled_list = []
#                 for id, user_last in enumerate(trackss):
                    
#                     counter_of_lens += 1
#                     track = SpUserTracking.objects.get(id=trackss[id])
#                     if counter_of_lens < len_track:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id+1])
#                     else:
#                         user_last_data = SpUserTracking.objects.get(id=trackss[id])
#                     R = 6373.0
#                     lat1 = radians(float(track.latitude))
#                     lon1 = radians(float(track.longitude))
#                     lat2 = radians(float(user_last_data.latitude))
#                     lon2 = radians(float(user_last_data.longitude))
#                     dlon = lon2 - lon1
#                     dlat = lat2 - lat1
#                     a = sin(dlat / 2)**2 + cos(lat1) * cos(lat2) * sin(dlon / 2)**2
#                     c = 2 * atan2(sqrt(a), sqrt(1 - a))
#                     distance = R * c
#                     meter_distance = float(distance * 1000)
                    
#                     if meter_distance > 10:
#                         distance_travelled_list.append(meter_distance)
                        
#                 if len(distance_travelled_list) > 0:
#                     distance_travelled                   = sum(distance_travelled_list)
#                     user_trackings['distance_travelled'] = round(float(distance_travelled)*0.001,2)
#                 else:
#                     user_trackings['distance_travelled'] = 0.0
#                 travel_charges                       = SpUserTracking.objects.filter(user_id=user_tracking,created_at__month=month,created_at__year=year).values('travel_charges').order_by('-id').first()        
#                 user_trackings['charges']            = round(float(travel_charges['travel_charges']),2)
#                 user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
#                 user_trackings['user_name']          = getUserName(user_tracking)
#                 user_trackings['user_id']            = user_tracking
                
#                 user_trackings['reporting_to_name']  = getModelColumnByColumnId(SpUsers,'id',user_tracking,'reporting_to_name')
                
#                 user_tracking_detail.append(user_trackings)
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         )
#         response['Content-Disposition'] = 'attachment; filename=user_monthly_travel_summary.xlsx'.format(
#             date=datetime.now().strftime('%Y-%m-%d'),
#         )
#         workbook = Workbook()

#         # Define some styles and formatting that will be later used for cells
#         header_font = Font(name='Calibri', bold=True)
#         centered_alignment = Alignment(horizontal='left')
#         thin = Side(border_style="thin", color="303030") 
#         black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
#         wrapped_alignment = Alignment(
#             vertical='top',
#             horizontal='left',
#             wrap_text=True
#         )

#         header_alignment = Alignment(
#             vertical='top',
#             horizontal='center',
#             wrap_text=True
#         )
        
#         # Get active worksheet/tab
#         worksheet = workbook.active
#         worksheet.title = 'User Monthly Travel Summary'
#         worksheet.merge_cells('A1:A1') 
        
#         worksheet.page_setup.orientation = 'landscape'
#         worksheet.page_setup.paperSize = 9
#         worksheet.page_setup.fitToPage = True
        
#         worksheet = workbook.worksheets[0]
#         img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
#         img.height = 50
#         img.alignment = 'center'
#         img.anchor = 'A1'
#         worksheet.add_image(img)
        
#         column_length = 6
#         month_name = datetime(int(travel_month[1]),int(travel_month[0]),1).strftime( '%B' )
#         worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
#         if is_user == 1:
#             worksheet.cell(row=1, column=2).value = ''+getUserName(user_id)+' Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
#         else:
#             worksheet.cell(row=1, column=2).value = 'Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
#         worksheet.cell(row=1, column=2).font = header_font
#         worksheet.cell(row=1, column=2).alignment = header_alignment
#         worksheet.cell(row=1, column=column_length).border = black_border
#         worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
#         worksheet.cell(row=1, column=2).fill = PatternFill()

#         # Define the titles for columns
#         # columns = []
#         row_num = 1
#         worksheet.row_dimensions[1].height = 40
        
#         # Define the titles for columns
#         columns = []
#         if is_user == 1:
#             columns += [ 'Date' ]
#         else:    
#             columns += [ 'Employee Name' ]
#         columns += [ 'Reporting To' ]
#         # columns += [ 'Headquarter Name' ]
#         columns += [ 'Distance Travelled in Kilometer' ]
#         columns += [ 'Charges' ]
#         columns += [ 'Total Charges' ]

#         row_num = 2

#         # Assign the titles for each cell of the header
#         for col_num, column_title in enumerate(columns, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = column_title
#             cell.font = header_font
#             cell.alignment = centered_alignment
#             cell.font = Font(size=12, color='FFFFFFFF', bold=True)
#             cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

#             column_letter = get_column_letter(col_num)
#             column_dimensions = worksheet.column_dimensions[column_letter]
#             column_dimensions.width = 32

#         for user_tracking in user_tracking_detail:
#             row_num += 1
#             # Define the data for each cell in the row 
#             row = []
#             if user_tracking['month_date']:
#                 row += [ user_tracking['month_date'] ]
#             else:
#                 row += [ user_tracking['user_name'] ]    
#             row += [ user_tracking['reporting_to_name'] ]
#             # if user_tracking['headquarter']:
#             #     row += [ user_tracking['headquarter'] ]
#             # else:
#             #     row += [''] 
#             row += [ user_tracking['distance_travelled'] ]
#             row += [ user_tracking['charges'] ]
#             row += [ user_tracking['total_charges'] ]           
        
#             # Assign the data for each cell of the row 
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
#                 cell.alignment = wrapped_alignment
#                 cell.border = black_border  

#         wrapped_alignment = Alignment(
#             horizontal='center',
#             wrap_text=True
#         )

#         row_num += 1
#         last_row = row_num
#         worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)
#         worksheet.row_dimensions[last_row].height = 20
#         worksheet.cell(row=last_row, column=1).value = 'Generated By Sakhi Milk'
#         worksheet.cell(row=last_row, column=1).font = header_font
#         worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
#         worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
#         worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

#         workbook.save(response)
#         return response     

@login_required
def exportUserTravelSummary(request, travel_date, user_id, travel_month_picker, time_period):
    today                   = travel_date
          
    if time_period == '1':
        user_tracking_details   = SpUserTracking.objects.filter(created_at__icontains=today)
        if user_id!='0':
            user_tracking_details = user_tracking_details.filter(user_id=user_id)
        user_tracking_details = user_tracking_details.filter().values('user_id').distinct().values('user_id', 'travel_charges')    
        for user_tracking in user_tracking_details:
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_tracking['user_id'], created_at__icontains=today).aggregate(Sum('distance_travelled'))
            user_tracking['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            if user_tracking['distance_travelled'] > 0:
                user_tracking['charges']            = round(float(user_tracking['travel_charges']),2)
                user_tracking['total_charges']      = round(user_tracking['distance_travelled']*float(user_tracking['travel_charges']),2)
            else:
                user_tracking['charges']            = 0
                user_tracking['total_charges']      = 0   
            user_tracking['user_name']          = getUserName(user_tracking['user_id'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_daily_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Employee Daily Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/dashboardLogo2.png')
        img.height = 50
        img.width = 150
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = 'Employee Daily Travel Summary as on Date('+datetime.strptime(str(today), '%Y-%m-%d').strftime('%d/%m/%Y')+')'
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Employee Name' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_details:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['user_name'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By Kanha HRMS'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response

    else:
        travel_month = travel_month_picker
        travel_month = travel_month.split('-')
        year  = int(travel_month[1])
        month = int(travel_month[0])
        
        last_day_of_month = calendar.monthrange(year,month)[1]
        if int(month) < 10:
            month = '0'+str(month)
        user_tracking_detail = []     
        for x in range(last_day_of_month):
            x = x+1
            if x < 10:
                x = '0'+str(x)
            travel_date  = str(year)+'-'+str(month)+'-'+str(x)
            travel_dates = str(x)+'/'+str(month)+'/'+str(year)

            user_trackings = {}
            distance_travelled  = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).aggregate(Sum('distance_travelled'))
            user_trackings['travel_date']          = travel_dates
            if distance_travelled['distance_travelled__sum']:
                user_trackings['distance_travelled'] = round(distance_travelled['distance_travelled__sum']*0.001,2)
            else:
                user_trackings['distance_travelled'] = 0    
            if user_trackings['distance_travelled'] > 0:
                travel_charges       = SpUserTracking.objects.filter(user_id=user_id, created_at__icontains=travel_date).values('travel_charges').first()        
                user_trackings['charges']            = float(travel_charges['travel_charges'])
                user_trackings['total_charges']      = round(user_trackings['distance_travelled']*float(travel_charges['travel_charges']),2)
            else:
                user_trackings['charges']            = 0
                user_trackings['total_charges']      = 0
            user_trackings['user_name']          = getUserName(user_id)
            user_tracking_detail.append(user_trackings)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=user_monthly_travel_summary.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Define some styles and formatting that will be later used for cells
        header_font = Font(name='Calibri', bold=True)
        centered_alignment = Alignment(horizontal='left')
        thin = Side(border_style="thin", color="303030") 
        black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
        wrapped_alignment = Alignment(
            vertical='top',
            horizontal='left',
            wrap_text=True
        )

        header_alignment = Alignment(
            vertical='top',
            horizontal='center',
            wrap_text=True
        )
        
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Employee Monthly Travel Summary'
        worksheet.merge_cells('A1:A1') 
        
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = 9
        worksheet.page_setup.fitToPage = True
        
        worksheet = workbook.worksheets[0]
        img = openpyxl.drawing.image.Image('static/img/dashboardLogo2.png')
        img.height = 50
        img.width = 150
        img.alignment = 'center'
        img.anchor = 'A1'
        worksheet.add_image(img)
        
        column_length = 4
        month_name = datetime(int(travel_month[1]),int(travel_month[0]),1).strftime( '%B' )
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=column_length)
        worksheet.cell(row=1, column=2).value = ''+getUserName(user_id)+' Monthly Travel Summary as on '+str(month_name)+'-'+str(int(travel_month[1]))+''
        worksheet.cell(row=1, column=2).font = header_font
        worksheet.cell(row=1, column=2).alignment = header_alignment
        worksheet.cell(row=1, column=column_length).border = black_border
        worksheet.cell(row=1, column=2).font = Font(size=14, color='303030', bold=True)
        worksheet.cell(row=1, column=2).fill = PatternFill()

        # Define the titles for columns
        # columns = []
        row_num = 1
        worksheet.row_dimensions[1].height = 40
        
        # Define the titles for columns
        columns = []

        columns += [ 'Date' ]
        columns += [ 'Distance Travelled in Kilometer' ]
        columns += [ 'Charges' ]
        columns += [ 'Total Charges' ]

        row_num = 2

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.font = Font(size=12, color='FFFFFFFF', bold=True)
            cell.fill = PatternFill(start_color="4d86bf", end_color="4d86bf", fill_type = "solid")

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32

        for user_tracking in user_tracking_detail:
            row_num += 1
            # Define the data for each cell in the row 
            row = []
            row += [ user_tracking['travel_date'] ]
            row += [ user_tracking['distance_travelled'] ]
            row += [ user_tracking['charges'] ]
            row += [ user_tracking['total_charges'] ]           
        
            # Assign the data for each cell of the row 
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = wrapped_alignment
                cell.border = black_border  

        wrapped_alignment = Alignment(
            horizontal='center',
            wrap_text=True
        )

        row_num += 1
        last_row = row_num
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=4)
        worksheet.row_dimensions[last_row].height = 20
        worksheet.cell(row=last_row, column=1).value = 'Generated By Kanha Dairy'
        worksheet.cell(row=last_row, column=1).font = header_font
        worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
        worksheet.cell(row=last_row, column=1).font = Font(size=12, color='808080', bold=True, underline="single")
        worksheet.cell(row=last_row, column=1).fill = PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type = "solid")

        workbook.save(response)
        return response 

@login_required
def userTrackingReportView(request, user_id, trac_date):
    trac_date = trac_date.replace("-", "/")
    context = {}
    default_user = SpUsers.objects.get(id=user_id)
    user_fname = default_user.first_name
    user_mname = default_user.middle_name
    user_lname = default_user.last_name
    if user_mname:
        default_user_name = user_fname + ' ' + user_mname + ' ' + user_lname
    else:
        default_user_name = user_fname + ' ' + user_lname
    users = SpUsers.objects.filter(user_type=1).exclude(role_id = 0)
    context['default_user_name'] = default_user_name
    context['user_id'] = int(user_id)
    context['trac_date'] = trac_date  
    context['users'] = users
    context['page_title'] = "User Tracking "
    context['org_latitude'] = getConfigurationResult('org_latitude')
    context['org_longitude'] = getConfigurationResult('org_longitude')
    context['organization'] = SpOrganizations.objects.all()
    template = 'user-management/user-tracking-report-view.html'
    return render(request, template, context)
    
# -----------------------------------------------------------------salary sheet patch-----------------------------------------------------------------------------------
@login_required
def employeeSalaySheet(request):
    date    = datetime.now()
    month   = date.month
    year    = date.year
    pay_date = days_in_months(year,month)[-1]
    salary_generate     = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 1).count()
    salary_regenerate   = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 0).count()
    employee            = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month)
    
    for emp in employee:
        emp.users       = SpUsers.objects.get(id = emp.user_id)
        emp.basic       = SpBasicDetails.objects.get(user_id = emp.user_id)
        emp.date_of_birth  =  datetime.strptime(str(emp.basic.date_of_birth), '%Y-%m-%d').strftime('%d-%b-%Y')
        emp.date_of_joining  =  datetime.strptime(str(emp.basic.date_of_joining), '%Y-%m-%d').strftime('%d-%b-%Y')
    first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month).last()
    if first_employee:
        first_employee  =  first_employee.generated_from.strftime('%d/%m/%Y')
        first_employee = 'Salary sheet is updated till' + ' ' + first_employee
        
    template = 'user-management/salary-sheet/salary.html'
    organization = SpOrganizations.objects.all()
    context = {}
    context['employee']             = employee
    context['organization']         = organization
    context['salary_generate']      = salary_generate
    context['salary_regenerate']    = salary_regenerate
    context['first_employee']       = first_employee
    context['pay_date']             =  datetime.strptime(str(pay_date), '%d/%m/%Y').strftime('%d-%b-%Y')
    context['today_date']           = date
    context['page_title']           = "Employee Salary Sheet"
    return render(request, template,context)



def checkHoliday(date,user_id):
    if SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).exists():
        holiday_id = SpHolidays.objects.filter(start_date__month = datetime.strptime(str(date), '%Y-%m-%d').strftime('%m'),start_date__year = datetime.strptime(str(date), '%Y-%m-%d').strftime('%Y'),holiday_status=3,status = 1).values_list('id',flat=True)
        role_id = getModelColumnByColumnId(SpUsers,'id',user_id,'role_id')
        holiday_ids = SpRoleEntityMapping.objects.filter(role_id = role_id,entity_type = 'holiday',entity_id__in = holiday_id).values_list('entity_id',flat=True)
        count = 0
        for holiday_id in holiday_ids:
            last_date = getModelColumnById(SpHolidays,holiday_id,'end_date')
            last_date = last_date.strftime('%Y-%m-%d')
            last_date = datetime.strptime(str(last_date), '%Y-%m-%d')
            from_date = datetime.strptime(str(date), '%Y-%m-%d')#.strftime('%Y-%m-%d')
            delta = last_date - from_date
            start_date = getModelColumnById(SpHolidays,holiday_id,'start_date')
            start_date = start_date.strftime('%Y-%m-%d')
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d')
            current_date = datetime.strptime(str(date), '%Y-%m-%d')
            if current_date>=start_date:
                if delta.days >= 0:
                    count+=1
        if count > 0:
            return True
        else:
            return False 
    else:
        return False 
        
# Employee Payroll Master
@login_required
def saveEmployeePayrollMaster(request):
    if request.method == "POST":
        employee_id             = request.POST['employee_id']
        emp_ctc                 = request.POST['emp_ctc']
        apipercent              = request.POST['apipercent']
        employeehrap            = request.POST['employeehrap']
        employee_pli            = request.POST['employee_pli']
        employee_pli_mt         = request.POST['employee_pli_mt']
        employee_basic          = request.POST['employee_basic']
        employee_basic_mt       = request.POST['employee_basic_mt']
        employee_pf             = request.POST['employee_pf']
        employee_pf_mt          = request.POST['employee_pf_mt']
        employee_gratuity       = request.POST['employee_gratuity']
        employee_gratuity_mt    = request.POST['employee_gratuity_mt']
        employee_tfp            = request.POST['employee_tfp']
        employee_tfp_mt         = request.POST['employee_tfp_mt']
        employee_hra            = request.POST['employee_hra']
        employee_hra_mt         = request.POST['employee_hra_mt']
        employee_spcl_mt        = request.POST['employee_spcl_mt']
        employee_spcl           = request.POST['employee_spcl']
        
        emp_payroll                     = SpPayrollMaster()
        emp_payroll.user_id             = employee_id
        emp_payroll.emp_ctc             = emp_ctc
        emp_payroll.apipercent          = apipercent
        emp_payroll.employeehrap        = employeehrap
        emp_payroll.employee_pli        = employee_pli
        emp_payroll.employee_pli_mt     = employee_pli_mt
        emp_payroll.employee_basic      = employee_basic
        emp_payroll.employee_basic_mt   = employee_basic_mt
        emp_payroll.employee_pf         = employee_pf
        emp_payroll.employee_pf_mt      = employee_pf_mt
        emp_payroll.employee_gratuity   = employee_gratuity
        emp_payroll.employee_gratuity_mt= employee_gratuity_mt
        emp_payroll.employee_tfp        = employee_tfp
        emp_payroll.employee_tfp_mt     = employee_tfp_mt
        emp_payroll.employee_hra        = employee_hra
        emp_payroll.employee_hra_mt     = employee_hra_mt
        emp_payroll.employee_spcl_mt    = employee_spcl_mt
        emp_payroll.employee_spcl       = employee_spcl
        emp_payroll.save()
        
        message  = "Employee Payroll Added Successfully"
        response = {}
        response['error'] = False
        response['message'] = message
        return JsonResponse(response)

@login_required
def updateEmployeePayrollMaster(request):
    if request.method == "POST":
        employee_id             = request.POST['employee_id']
        emp_ctc                 = request.POST['emp_ctc']
        apipercent              = request.POST['apipercent']
        employeehrap            = request.POST['employeehrap']
        employee_pli            = request.POST['employee_pli']
        employee_pli_mt         = request.POST['employee_pli_mt']
        employee_basic          = request.POST['employee_basic']
        employee_basic_mt       = request.POST['employee_basic_mt']
        employee_pf             = request.POST['employee_pf']
        employee_pf_mt          = request.POST['employee_pf_mt']
        employee_gratuity       = request.POST['employee_gratuity']
        employee_gratuity_mt    = request.POST['employee_gratuity_mt']
        employee_tfp            = request.POST['employee_tfp']
        employee_tfp_mt         = request.POST['employee_tfp_mt']
        employee_hra            = request.POST['employee_hra']
        employee_hra_mt         = request.POST['employee_hra_mt']
        employee_spcl_mt        = request.POST['employee_spcl_mt']
        employee_spcl           = request.POST['employee_spcl']
        
        emp_payroll                     = SpPayrollMaster.objects.get(user_id = employee_id)
        emp_payroll.user_id             = employee_id
        emp_payroll.emp_ctc             = emp_ctc
        emp_payroll.apipercent          = apipercent
        emp_payroll.employeehrap        = employeehrap
        emp_payroll.employee_pli        = employee_pli
        emp_payroll.employee_pli_mt     = employee_pli_mt
        emp_payroll.employee_basic      = employee_basic
        emp_payroll.employee_basic_mt   = employee_basic_mt
        emp_payroll.employee_pf         = employee_pf
        emp_payroll.employee_pf_mt      = employee_pf_mt
        emp_payroll.employee_gratuity   = employee_gratuity
        emp_payroll.employee_gratuity_mt= employee_gratuity_mt
        emp_payroll.employee_tfp        = employee_tfp
        emp_payroll.employee_tfp_mt     = employee_tfp_mt
        emp_payroll.employee_hra        = employee_hra
        emp_payroll.employee_hra_mt     = employee_hra_mt
        emp_payroll.employee_spcl_mt    = employee_spcl_mt
        emp_payroll.employee_spcl       = employee_spcl
        emp_payroll.save()
        
        message  = "Employee Payroll Updated Successfully"
        response = {}
        response['error'] = False
        response['message'] = message
        return JsonResponse(response)


@login_required
def employeeSalaySheet(request):
    date    = datetime.now()
    month   = date.month
    year    = date.year
    pay_date = days_in_months(year,month)[-1]
    salary_generate     = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 1).count()
    salary_regenerate   = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 0).count()
    employee            = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month)
    
    for emp in employee:
        emp.users       = SpUsers.objects.get(id = emp.user_id)
        emp.basic       = SpBasicDetails.objects.get(user_id = emp.user_id)
        emp.date_of_birth  =  datetime.strptime(str(emp.basic.date_of_birth), '%Y-%m-%d').strftime('%d-%b-%Y')
        emp.date_of_joining  =  datetime.strptime(str(emp.basic.date_of_joining), '%Y-%m-%d').strftime('%d-%b-%Y')
    first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month).last()
    if first_employee:
        first_employee  =  first_employee.generated_from.strftime('%d/%m/%Y')
        first_employee = 'Salary sheet is updated till' + ' ' + first_employee
        
    template = 'user-management/salary-sheet/salary.html'
    organization = SpOrganizations.objects.all()
    context = {}
    context['employee']             = employee
    context['organization']         = organization
    context['salary_generate']      = salary_generate
    context['salary_regenerate']    = salary_regenerate
    context['first_employee']       = first_employee
    context['pay_date']             =  datetime.strptime(str(pay_date), '%d/%m/%Y').strftime('%d-%b-%Y')
    context['today_date']           = date
    context['page_title']           = "Employee Salary Sheet"
    return render(request, template,context)

def ceckWeekOfDay(date,user_id):
    # print(date)
    user_week_of_day = getModelColumnByColumnId(SpBasicDetails,'user_id',user_id,'week_of_day')
    if user_week_of_day:
        user_week_of_day = user_week_of_day.split(',')
        week_day = datetime.strptime(str(date), '%Y-%m-%d').strftime('%A')
        if week_day in user_week_of_day:
            return True
        else:
            return False

def checkAttendance(date,user_id):
    if SpUserAttendance.objects.filter(user_id=user_id,attendance_date_time__contains=date).exists():
        return True
    else:
        return False

def checkLeavesalary(date, user_id):
    try:
        leaves = SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=user_id).first()
    except SpUserLeaves.DoesNotExist:
        leaves = None
    if leaves:
        return True
    else:
        return False

# @login_required
# def generateSalarySheet(request):
#     context = {}
#     date                = request.GET['salary_date']
#     organization_id     = request.GET['organization_id']
#     currenttoday        = datetime.today()
#     current_month       = currenttoday.month
#     date                = date.split('/')
#     month               = int(date[0])
#     year                = int(date[1]) 
#     todaydays           = datetime.today() 
#     dates                = todaydays.day
#     pay_date = days_in_months(year,month)[-1]
#     salary_generate     = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month)
#     if organization_id:
#         employee_id     = SpUsers.objects.filter(organization_id = organization_id).values_list('id',flat=True)
#         salary_generate = salary_generate.filter(user_id__in=employee_id,attendance_date_time__year = year , attendance_date_time__month = month)
#     salary_generate     = salary_generate.count()
#     if salary_generate == 0:
#         context['message'] = "No attendance recorded"
#         context['flag'] = 1
#         return JsonResponse(context)
#     # regenerate salary
#     if request.GET['status'] == '2':
#         employee        = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month)
#         if organization_id:
#             employee    = employee.filter(organization_id = organization_id)
#         employee.delete()
#     employee_id         = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month).values_list('user_id',flat=True).distinct()
#     day_of_month        = calendar.monthrange(year,month)[1]
#     employee            = SpUsers.objects.filter(user_type=1,status=1,id__in = employee_id).exclude(id=1).order_by('role_name')
#     if organization_id:
#         employee        = employee.filter(organization_id=organization_id)
#     for emp1 in employee:
#         emp             = Spemployeesalarydata()
#         emp.user_id     = emp1.id
#         emp.employee_code = emp1.emp_sap_id
#         emp.employee_name = emp1.first_name+' '+emp1.last_name
#         emp.addhar_name = emp1.first_name+' '+emp1.last_name
#         emp.emp_email   = emp1.official_email 
#         emp.department_name   = emp1.department_name 
#         emp.role_name   = emp1.role_name 
#         # emp.pay_date    = pay_date
#         emp.pay_date    = datetime.strptime(str(pay_date), '%d/%m/%Y').strftime('%Y-%m-%d')
#         #bank details
#         try:
#             bank_details = SpBankDetails.objects.get(user_id = emp1.id)
#         except SpBankDetails.DoesNotExist:
#             bank_details = None
#         if bank_details:
#             emp.account_no = bank_details.bank_account_no
#             emp.ifsc_code = bank_details.ifsc_code
#             emp.bank_address = bank_details.bank_address
#             emp.bank_name = bank_details.bank_name    
#         try:
#             basic_detail  = SpBasicDetails.objects.get(user_id = emp1.id)
#         except SpBasicDetails.DoesNotExist:
#             basic_detail = None
#         if basic_detail:
#             emp.father_name = basic_detail.father_name
#             emp.dob         = basic_detail.date_of_birth
#             emp.gender      = basic_detail.gender
#             emp.date_of_joining = basic_detail.date_of_joining
#             if basic_detail.pf_no:
#                 emp.pf_no   = basic_detail.pf_no
#             if basic_detail.uan:
#                 emp.uan     = basic_detail.uan
#             if basic_detail.pan_no:
#                 emp.pan_no  = basic_detail.pan_no
#             # if basic_detail.Pan_no:
#             #     emp.Pan_no  = basic_detail.pan_no
#             if basic_detail.esi_no:
#                 emp.esi_no  = basic_detail.esi_no
#             if basic_detail.adhaar_no:
#                 emp.aadhar_no  = basic_detail.adhaar_no

#         month_list = days_in_months(year,month)
#         nameMonth = []
#         for month_date in month_list:
#             month_date  = datetime.strptime(str(month_date), '%d/%m/%Y')
#             if current_month == int(month_date.strftime('%m')):
#                 print(int(month_date.strftime('%m')))
#                 if int(dates) >= int(month_date.strftime('%d')):
#                     nameMonth.append(month_date.strftime('%d %b %a' ))
#             else:
#                 nameMonth.append(month_date.strftime('%d %b %a'))
#         SpUserAttendance.objects.filter(user_id = emp1.id, attendance_date_time__year = year, attendance_date_time__month = month).update(is_generated=1)  
#         total_days = 0  
#         if SpUserAttendance.objects.filter(user_id = emp1.id ,attendance_date_time__year = year , attendance_date_time__month = month ).exists():
#             for months in month_list[:len(nameMonth)]:
#                 month_date   = datetime.strptime(str(months),'%d/%m/%Y').strftime('%Y-%m-%d')
#                 if checkAttendance(month_date,emp1.id):
#                     total_days +=1
#                 elif checkHoliday(month_date,emp1.id):
#                     total_days +=1
#                 elif ceckWeekOfDay(month_date,emp1.id):
#                     total_days +=1
#         try:
#             employee_salary_details = SpPayrollMaster.objects.get(user_id = emp1.id) 
#             employee_basic_mtt      =  employee_salary_details.employee_basic_mt
#             employee_hrapp          = employee_salary_details.employeehrap
#             onedaysalary       =  int(round(((employee_salary_details.employee_basic_mt)/day_of_month),0))
#             earn_salary        =  onedaysalary*total_days
            
#             if employee_salary_details.employeehrap>0:
#                 earn_hra       = int(round(((earn_salary*employee_salary_details.employeehrap)/100),0))
#             else:
#                 earn_hra       = 0
#             earn_spcl       = 0
#             grosssalary     = earn_salary+earn_hra+earn_spcl
            
#             basic_spcl      = earn_salary +earn_spcl
#             if earn_salary > 15000:
#                 pfamount    = int(round(earn_salary*12/100))
#             elif basic_spcl >= 15000:
#                 pfamount    = int(round(15000*12/100))
#             else:
#                 pfamount    = int(round(basic_spcl*12/100))
#             earn_esi        = 0
#             itax1           = 0
#             grosssded       = pfamount + earn_esi + itax1
#             Netpay          = grosssalary - grosssded
#             fpf             = int(round(((basic_spcl*8.33)/100),0))
#             empr_pf         = pfamount-fpf
#             employeer_esi   = 0
#             total           = empr_pf+fpf+employeer_esi
            
#         except SpPayrollMaster.DoesNotExist:
#             employee_basic_mtt = 0
#             # day_of_month    = 0
#             employee_hrapp  = 0
#             earn_salary     = 0
#             earn_hra        = 0
#             earn_spcl       = 0
#             grosssalary     = 0
#             pfamount        = 0 
#             earn_esi        = 0
#             itax1           = 0
#             grosssded       = 0
#             Netpay          = 0 
#             empr_pf         = 0
#             fpf             = 0
#             employeer_esi   = 0
#             total           = 0
#         emp.ern_basic       =  earn_salary
#         emp.days_in_month   =  day_of_month
#         emp.employee_basic_mt =  employee_basic_mtt
#         emp.employeehrap    =  employee_hrapp
#         emp.pay_days        =  total_days
#         emp.ern_hra         = earn_hra
#         emp.ern_spl         = earn_spcl
#         emp.grosssalary     = grosssalary 
#         emp.emp_pf          = pfamount
#         emp.emp_esi         = earn_esi
#         emp.itax            = itax1
#         emp.grossded        = grosssded
#         emp.net_pay         = Netpay 
#         emp.fpf             = fpf
#         emp.empr_pf         = empr_pf
#         emp.empr_esi        = employeer_esi
#         emp.total           = total   
#         datess                = request.GET['salary_date']
#         datess                = '01' + '/' + datess 
#         emp.generated_month = datetime.strptime(datess,'%d/%m/%Y').strftime('%Y-%m-%d')
#         emp.generated_to    = datetime.strptime(datess,'%d/%m/%Y').strftime('%Y-%m-%d')
#         if current_month == int(month):
#             emp.generated_from  = datetime.now().strftime('%Y-%m-%d')
#         else:
#             emp.generated_from  = datetime.strptime(str(month_list[-1]), '%d/%m/%Y').strftime('%Y-%m-%d')
#         emp.organization_id = emp1.organization_id
#         emp.save()

#     if request.GET['status'] == '2':
#         msg = 'Salary Sheet Regenerated Succesfully'
#     else:
#         msg = 'Salary Sheet Generated Succesfully'
#     # context = {}
#     context['message']      = msg
#     context['flag']         = 0
#     context['today_date']   = date
#     return JsonResponse(context)

def checkLeavesalary(date, user_id):
    try:
        leaves = SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=user_id).first()
    except SpUserLeaves.DoesNotExist:
        leaves = None
    if leaves:
        return True
    else:
        return False


@login_required
def generateSalarySheet(request):
    context = {}
    date                = request.GET['salary_date']
    organization_id     = request.GET['organization_id']
    currenttoday        = datetime.today()
    current_month       = currenttoday.month
    date                = date.split('/')
    month               = int(date[0])
    year                = int(date[1]) 
    todaydays           = datetime.today() 
    dates                = todaydays.day
    pay_date = days_in_months(year,month)[-1]
    salary_generate     = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month)
    if organization_id:
        employee_id     = SpUsers.objects.filter(organization_id = organization_id).values_list('id',flat=True)
        salary_generate = salary_generate.filter(user_id__in=employee_id,attendance_date_time__year = year , attendance_date_time__month = month)
    salary_generate     = salary_generate.count()
    if salary_generate == 0:
        context['message'] = "No attendance recorded"
        context['flag'] = 1
        return JsonResponse(context)
    # regenerate salary
    if request.GET['status'] == '2':
        employee        = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month)
        if organization_id:
            employee    = employee.filter(organization_id = organization_id)
        employee.delete()
    employee_id         = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month).values_list('user_id',flat=True).distinct()
    day_of_month        = calendar.monthrange(year,month)[1]
    employee            = SpUsers.objects.filter(user_type=1,status=1,id__in = employee_id).exclude(id=1).order_by('role_name')
    if organization_id:
        employee        = employee.filter(organization_id=organization_id)
    for emp1 in employee:
        emp             = Spemployeesalarydata()
        emp.user_id     = emp1.id
        emp.employee_code = emp1.emp_sap_id
        emp.employee_name = emp1.first_name+' '+emp1.last_name
        emp.addhar_name = emp1.first_name+' '+emp1.last_name
        emp.emp_email   = emp1.official_email 
        emp.department_name   = emp1.department_name 
        emp.role_name   = emp1.role_name 
        # emp.pay_date    = pay_date
        emp.pay_date    = datetime.strptime(str(pay_date), '%d/%m/%Y').strftime('%Y-%m-%d')
        #bank details
        try:
            bank_details = SpBankDetails.objects.get(user_id = emp1.id)
        except SpBankDetails.DoesNotExist:
            bank_details = None
        if bank_details:
            emp.account_no = bank_details.bank_account_no
            emp.ifsc_code = bank_details.ifsc_code
            emp.bank_address = bank_details.bank_address
            emp.bank_name = bank_details.bank_name    
        try:
            basic_detail  = SpBasicDetails.objects.get(user_id = emp1.id)
        except SpBasicDetails.DoesNotExist:
            basic_detail = None
        if basic_detail:
            emp.father_name = basic_detail.father_name
            emp.dob         = basic_detail.date_of_birth
            emp.gender      = basic_detail.gender
            emp.date_of_joining = basic_detail.date_of_joining
            if basic_detail.pf_no:
                emp.pf_no   = basic_detail.pf_no
            if basic_detail.uan:
                emp.uan     = basic_detail.uan
            if basic_detail.pan_no:
                emp.pan_no  = basic_detail.pan_no
            # if basic_detail.Pan_no:
            #     emp.Pan_no  = basic_detail.pan_no
            if basic_detail.esi_no:
                emp.esi_no  = basic_detail.esi_no
            if basic_detail.adhaar_no:
                emp.aadhar_no  = basic_detail.adhaar_no

        month_list = days_in_months(year,month)
        nameMonth = []
        for month_date in month_list:
            month_date  = datetime.strptime(str(month_date), '%d/%m/%Y')
            if current_month == int(month_date.strftime('%m')):
               
                if int(dates) >= int(month_date.strftime('%d')):
                    nameMonth.append(month_date.strftime('%d %b %a' ))
            else:
                nameMonth.append(month_date.strftime('%d %b %a'))
        SpUserAttendance.objects.filter(user_id = emp1.id, attendance_date_time__year = year, attendance_date_time__month = month).update(is_generated=1)  
        total_days = 0  
        if SpUserAttendance.objects.filter(user_id = emp1.id ,attendance_date_time__year = year , attendance_date_time__month = month ).exists():
            for months in month_list[:len(nameMonth)]:
                month_date   = datetime.strptime(str(months),'%d/%m/%Y').strftime('%Y-%m-%d')
                if checkAttendance(month_date,emp1.id):
                    total_days +=1
                elif checkHoliday(month_date,emp1.id):
                    total_days +=1
                elif ceckWeekOfDay(month_date,emp1.id):
                    total_days +=1
                elif checkLeavesalary(month_date,emp1.id):
                    total_days +=1
        try:
            employee_salary_details = SpPayrollMaster.objects.get(user_id = emp1.id) 
            employee_basic_mtt      =  employee_salary_details.employee_basic_mt
            employee_hrapp          = employee_salary_details.employeehrap
            onedaysalary       =  round(((employee_salary_details.employee_basic_mt)/day_of_month),2)
            earn_salary        =  onedaysalary*total_days
            print(earn_salary)
            
            if employee_salary_details.employeehrap>0:
                earn_hra       = round(((earn_salary*employee_salary_details.employeehrap)/100),2)
            else:
                earn_hra       = 0
            earn_spcl       = 0
            grosssalary     = round((earn_salary+earn_hra+earn_spcl),2)
            print(grosssalary,emp1.id)
            
            basic_spcl      = earn_salary +earn_spcl
            if earn_salary > 15000:
                pfamount    = round((earn_salary*12/100),2)
            elif basic_spcl >= 15000:
                pfamount    = round((15000*12/100),2)
            else:
                pfamount    = round((basic_spcl*12/100),2)
            
            #calculate emp esic 
            
            if basic_detail.is_esic == 1 :
                earn_esi        = round((grosssalary * 0.75/100),2)
            else:
                earn_esi = 0
                
            itax1           = 0
            grosssded       = pfamount + earn_esi + itax1
            Netpay          = grosssalary - grosssded
            if basic_spcl >= 15000:
                fpf             =round((15000*8.33/100),2)
            else:
                fpf             = round(((basic_spcl*8.33)/100),2)
                
            # fpf             = round(((basic_spcl*8.33)/100),2)
            empr_pf         = pfamount-fpf
            
            #calculate empr esic 
            if basic_detail.is_esic == 1 :
                employeer_esi   = round((grosssalary * 3.25/100),2)
            else :
                employeer_esi = 0
            total           = empr_pf+fpf+employeer_esi
            
        except SpPayrollMaster.DoesNotExist:
            employee_basic_mtt = 0
            # day_of_month    = 0
            employee_hrapp  = 0
            earn_salary     = 0
            earn_hra        = 0
            earn_spcl       = 0
            grosssalary     = 0
            pfamount        = 0 
            earn_esi        = 0
            itax1           = 0
            grosssded       = 0
            Netpay          = 0 
            empr_pf         = 0
            fpf             = 0
            employeer_esi   = 0
            total           = 0
        emp.ern_basic       =  earn_salary
        emp.days_in_month   =  day_of_month
        emp.employee_basic_mt =  employee_basic_mtt
        emp.employeehrap    =  employee_hrapp
        emp.pay_days        =  total_days
        emp.ern_hra         = earn_hra
        emp.ern_spl         = earn_spcl
        emp.grosssalary     = grosssalary 
        emp.emp_pf          = pfamount
        emp.emp_esi         = earn_esi
        emp.itax            = itax1
        emp.grossded        = grosssded
        emp.net_pay         = Netpay 
        emp.fpf             = fpf
        emp.empr_pf         = empr_pf
        emp.empr_esi        = employeer_esi
        emp.total           = total   
        datess                = request.GET['salary_date']
        datess                = '01' + '/' + datess 
        emp.generated_month = datetime.strptime(datess,'%d/%m/%Y').strftime('%Y-%m-%d')
        emp.generated_to    = datetime.strptime(datess,'%d/%m/%Y').strftime('%Y-%m-%d')
        if current_month == int(month):
            emp.generated_from  = datetime.now().strftime('%Y-%m-%d')
        else:
            emp.generated_from  = datetime.strptime(str(month_list[-1]), '%d/%m/%Y').strftime('%Y-%m-%d')
        emp.organization_id = emp1.organization_id
        emp.save()

    if request.GET['status'] == '2':
        msg = 'Salary Sheet Regenerated Succesfully'
    else:
        msg = 'Salary Sheet Generated Succesfully'
    # context = {}
    context['message']      = msg
    context['flag']         = 0
    context['today_date']   = date
    return JsonResponse(context)




@login_required
def ajaxEmployeeSalaySheet(request):
    date                = request.GET['salary_date']
    organization_id     = request.GET['organization_id']
    date                = date.split('/')
    month               = int(date[0])
    year                = int(date[1])
    
    salary_generate = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 1).count()
    salary_regenerate = SpUserAttendance.objects.filter(attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 0).count()
    employee = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month)
    
    if organization_id:
        employee            = employee.filter(organization_id = organization_id)
        employee_id         = employee.filter(organization_id = organization_id).values_list('user_id',flat=True)
        salary_generate     = SpUserAttendance.objects.filter(user_id__in=employee_id,attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 1).count()
        salary_regenerate   = SpUserAttendance.objects.filter(user_id__in=employee_id,attendance_date_time__year = year , attendance_date_time__month = month , is_generated = 0).count()
    
    template = 'user-management/salary-sheet/ajax-salary.html'
    organization = SpOrganizations.objects.all()
    if organization_id:
        first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month,organization_id=organization_id).first()
    else:
        first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = month).last()
    
    if first_employee:
        first_employee  =  first_employee.generated_from.strftime('%d/%m/%Y')
        first_employee = 'Salary sheet is updated till' + ' ' + first_employee
    
        
    context = {}
    context['employee']             = employee
    context['salary_generate']      = salary_generate
    context['salary_regenerate']    = salary_regenerate
    context['organization']         = organization
    context['first_employee']       = first_employee
    context['page_title']           = "Employee Salary Sheet"
    return render(request, template,context)


        
@login_required
def exportSalarySheet(request,organization_id,months,year):

    employee        = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = months)
    
    
    if organization_id != '0':
        employee            = employee.filter(organization_id = organization_id)
        first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = months,organization_id = organization_id).last()
    else:
        first_employee  = Spemployeesalarydata.objects.filter(generated_month__year = year ,generated_month__month = months).last()
    
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=employe-salary-sheet.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='left')
    thin = Side(border_style="thin", color="303030")
    black_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    wrapped_alignment = Alignment(
        vertical='top',
        horizontal='left',
        wrap_text=True
    )

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Employee Salary Sheet'
    worksheet.merge_cells('A1:A1')

    worksheet.page_setup.orientation = 'landscape'
    worksheet.page_setup.paperSize = 9
    worksheet.page_setup.fitToPage = True

    worksheet = workbook.worksheets[0]
    img = openpyxl.drawing.image.Image('static/img/png/dashboardLogo.gif')
   
    img.height = 50
    img.width = 150
    img.alignment = 'center'
    img.anchor = 'A1'
    worksheet.add_image(img)
    if first_employee:
       first_employee  =  first_employee.generated_from.strftime('%d/%m/%Y')
    else:
        first_employee = '-'
    column_length = 21

    worksheet.merge_cells(start_row=1, start_column=2,
                          end_row=1, end_column=column_length)
    worksheet.cell(row=1, column=2).value = 'Salary sheet is updated till ' + first_employee
    worksheet.cell(row=1, column=2).font = header_font
    worksheet.cell(row=1, column=2).alignment = wrapped_alignment
    worksheet.cell(row=1, column=column_length).border = black_border
    worksheet.cell(row=1, column=2).font = Font(
        size=14, color='303030', bold=True)
    worksheet.cell(row=1, column=2).fill = PatternFill()

    # Define the titles for columns
    # columns = []
    row_num = 1
    worksheet.row_dimensions[1].height = 40

    # Define the titles for columns
    columns = []
    columns +=['Employee Code']
    columns +=['Employee Name']

    columns +=['Father Name']
    columns +=['Email']
    columns +=['Date Of Birth']
    columns +=['Gender']
    columns +=['Date Of Joining']
    columns +=['PF No.']
    columns +=['UAN']
    columns +=['PAN No.']
    columns +=['Account No.']
    columns +=['IFSC Code'] 
    columns +=['ESI No.'] 
    columns +=['Name On Aadhar'] 
    columns +=['Aadhar No.'] 
    columns +=['Pay Date'] 
    columns +=['Pay Days'] 
    columns +=['Bank Name'] 
    columns +=['Location Name'] 
    columns +=['State Name'] 
    columns +=['Department'] 

    columns +=['Designation']  
    columns +=['Ern Basic']  
    columns +=['Ern HRA']  
    columns +=['Ern Suppall']  
    columns +=['Gross salary']  
    columns +=['Emp PF']  
    
    columns +=['Emp ESI']  
    columns +=['ITAX 1']  
    columns +=['Grossded']  
    columns +=['Net Pay']  
    columns +=['EMPR PF']
    columns +=['FPF']
    columns +=['EMPR ESI']
    columns +=['Total']
    columns +=['PLI']
    row_num = 2

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.font = Font(size=12, color='FFFFFFFF', bold=True)
        cell.fill = PatternFill(start_color="4d86bf",
                                end_color="4d86bf", fill_type="solid")

        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 32

    for emp in employee:
        row_num += 1
        # Define the data for each cell in the row
        row = []
        row += [emp.employee_code]
        row += [emp.employee_name]
        if emp.father_name:
            row+=[emp.father_name]
        else:
            row+=['-']
        if emp.emp_email:
            row+=[emp.emp_email]
        else:
            row+=['-']
    
        row += [emp.dob]
        row += [emp.gender]
        row += [emp.date_of_joining]
        if emp.pf_no:
            row+=[emp.pf_no]
        else:
            row+=['-']
        if emp.uan:
            row+=[emp.uan]
        else:
            row+=['-']
        if emp.pan_no:
            row+=[emp.pan_no]
        else:
            row+=['-']
        if emp.account_no:
            row+=[emp.account_no]
        else:
            row+=['-']
        if emp.ifsc_code:
            row+=[emp.ifsc_code]
        else:
            row+=['-']
        if emp.esi_no:
            row+=[emp.esi_no]
        else:
            row+=['-']
        row += [emp.employee_name]
        if emp.aadhar_no:
            row+=[emp.aadhar_no]
        else:
            row+=['-']

        row += [emp.pay_date]
        row += [emp.pay_days]
        row += [emp.bank_name]
        if emp.location_name:
            row+=[emp.location_name]
        else:
            row+=['-']
        if emp.state_name:
            row+=[emp.state_name]
        else:
            row+=['-']

        row += [emp.department_name]
        row += [emp.role_name]
        row += [emp.ern_basic]
        row += [emp.ern_hra]
        row += [emp.ern_spl]
        row += [emp.grosssalary]
        row += [emp.emp_pf]
        row += [emp.emp_esi]
        row += [emp.itax]
        row += [emp.grossded]
        row += [emp.net_pay]
        row += [emp.empr_pf]
        row += [emp.fpf]
        row += [emp.empr_esi]
        row += [emp.total]
        row += ['-']
        
      

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.border = black_border

    wrapped_alignment = Alignment(
        horizontal='center',
        wrap_text=True
    )

    row_num += 1
    last_row = row_num
    worksheet.merge_cells(start_row=last_row, start_column=1,
                          end_row=last_row, end_column=21)
    worksheet.row_dimensions[last_row].height = 20
    worksheet.cell(row=last_row, column=1).value = 'Generated By Sakhi Dairy'
    worksheet.cell(row=last_row, column=1).font = header_font
    worksheet.cell(row=last_row, column=1).alignment = wrapped_alignment
    worksheet.cell(row=last_row, column=1).font = Font(
        size=12, color='808080', bold=True, underline="single")
    worksheet.cell(row=last_row, column=1).fill = PatternFill(
        start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

    workbook.save(response)

    return response
    
@login_required
def saveSalarySheet(request):
    
    user_id                 = request.POST.getlist('user_id[]')
    ids                     = request.POST.getlist('id[]')
    paydays                 = request.POST.getlist('paydays[]')
    ernbasic                = request.POST.getlist('ernbasic[]')
    ernhra                  = request.POST.getlist('ernhra[]')
    ernspl                  = request.POST.getlist('ernspl[]')
    grosssalary             = request.POST.getlist('grosssalary[]')
    emppf                   = request.POST.getlist('emppf[]')
    empesi                  = request.POST.getlist('empesi[]')
    itax                    = request.POST.getlist('itax[]')
    grossded                = request.POST.getlist('grossded[]')
    netpay                  = request.POST.getlist('netpay[]')
    emprpf                  = request.POST.getlist('emprpf[]')
    fpf                     = request.POST.getlist('fpf[]')
    empresi                 = request.POST.getlist('empresi[]')
    total                   = request.POST.getlist('total[]')
    for id,user_id in enumerate(user_id):
        salary_sheet = Spemployeesalarydata.objects.get(id=ids[id])
        salary_sheet.pay_days               = paydays[id]
        salary_sheet.ern_basic              = ernbasic[id]
        salary_sheet.ern_hra                = ernhra[id]
        salary_sheet.ern_spl                = ernspl[id]
        salary_sheet.grosssalary            = grosssalary[id]
        salary_sheet.emp_pf                 = emppf[id]
        salary_sheet.emp_esi                = empesi[id]
        salary_sheet.itax                   = itax[id]
        salary_sheet.grossded               = grossded[id]
        salary_sheet.net_pay                = netpay[id]
        salary_sheet.empr_pf                = emprpf[id]
        salary_sheet.fpf                    = fpf[id]
        salary_sheet.empr_esi               = empresi[id]
        salary_sheet.total                  = total[id]
        salary_sheet.generated_from  = datetime.now().strftime('%Y-%m-%d')
        salary_sheet.save()

        
        # salary_sheets                        = SpEmployeeSalaryHistory()
        # salary_sheets.user_id                = salary_sheet.user_id   
        # try:
        #     bank_details = SpBankDetails.objects.get(user_id=salary_sheet.user_id)
        # except SpBankDetails.DoesNotExist:
        #     bank_details = None
        # if bank_details:
        #     salary_sheets.account_no            = bank_details.bank_account_no
        #     salary_sheets.ifsc_code             = bank_details.ifsc_code
        #     salary_sheets.bank_address          = bank_details.bank_address
        #     salary_sheets.bank_name             = bank_details.bank_name
        # salary_sheets.payable_days              = payable_days[id]
        # salary_sheets.total_working_days        = salary_sheet.total_working_days
        # salary_sheets.payable_overtime_days     = payable_overtime_days[id]
        # salary_sheets.total_payable_days        = total_payable_days[id]
        # salary_sheets.gross_salary              = salary_sheet.gross_salary
        # salary_sheets.basic_salary              = basic_salary[id]
        # salary_sheets.emp_ta                    = emp_ta[id]
        # salary_sheets.emp_hra                   = emp_hra[id]
        # salary_sheets.earned_salary             = earned_salary[id]
        # salary_sheets.emp_pf                    = emp_pf[id]
        # salary_sheets.emp_tds                   = emp_tds[id]
        # salary_sheets.total_deduction           = total_deduction[id]
        # salary_sheets.payable_salary            = payable_salary[id]
        # salary_sheets.emp_da                    = emp_da[id]
        # salary_sheets.bank_transfer             = bank_transfer[id]
        # salary_sheets.generated_month           = salary_sheet.generated_month
        # salary_sheets.generated_to              = salary_sheet.generated_to
        # salary_sheets.generated_from            = salary_sheet.generated_from
        # salary_sheets.organization_id           = salary_sheet.organization_id
        # salary_sheets.save()
   
    msg                         = 'Salary Sheet Updated Succesfully'
    context                     = {}
    context['message']          = msg
    return JsonResponse(context)



#Leave ledger
@login_required
def addLeaveLedger(request,employee_id):
    employee = SpUsers.objects.filter(id = employee_id)
    leaves = SpLeaveTypes.objects.filter(status = 1)
    context = {}
    context['employee'] = employee
    context['employee_id'] = employee_id
    context['leaves'] = leaves
    templates = "user-management/add-leave-ledger.html"
    return render(request,templates,context)

@login_required
def saveLeaveLedger(request):
    if request.method == "POST":
        emp_id  = request.POST['emp_id']
        leave_type = request.POST['leave_type']
        no_leaves =  request.POST['no_leaves']
        types =  request.POST['type']
        remark = request.POST['remark']
        
        leavess = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id,leave_type_id = leave_type).last()
        
        leave_types = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave_type, user_id = emp_id).last()
        if leave_types:
            if float(no_leaves) > float(leave_types.balance) and int(request.POST['type']) == 0:
                message  = "Debit is less than no. of leaves"
                response = {}
                response['flag'] = False
                response['message'] = message
                response['emp_id'] = emp_id
                return JsonResponse(response)
        else:
            role_id = getModelColumnById(SpUsers,emp_id,'role_id')
            if int(request.POST['type']) == 0:
                message  = "No balance leave"
                response = {}
                response['flag'] = False
                response['message'] = message
                response['emp_id'] = emp_id
                return JsonResponse(response)
                
            if int(request.POST['type']) == 1:
                message  = "No balance leave"
                response = {}
                response['flag'] = False
                response['message'] = message
                response['emp_id'] = emp_id
                return JsonResponse(response)
            else:
                if SpRoleEntityMapping.objects.filter(role_id=role_id, entity_type = 'leave_policy').exists():
                    led = SpRoleEntityMapping.objects.filter(role_id = role_id, entity_type = 'leave_policy').last()
                    leave_policy_id =led.entity_id
                     
                    leave_ledger = SpLeavePolicyDetails.objects.filter(leave_policy_id = leave_policy_id, leave_type_id= leave_type).last()
                    # leavess = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id).last()
                    # leave_types = SpUserLeavePolicyLedger.objects.filter(leave_type_id = leave_type, user_id = emp_id).last()
                    if leavess:
                        balances = leavess.balance
                    else:
                        balances = 0
                    leave_id = leave_ledger.leave_policy_id
                    leave_typess = leave_ledger.leave_type_id
                    month_leave = leave_ledger.month_leave_count
                    cons_leave = leave_ledger.consecutive_leave
                   
                    
                    
                    l_ledger = SpUserLeavePolicyLedger()
                    l_ledger.user_id = emp_id
                    l_ledger.leave_policy_id = leave_id
                    l_ledger.leave_type_id = leave_typess
                    l_ledger.month_leave_count = month_leave
                    l_ledger.consecutive_leave = cons_leave
                    l_ledger.credit = no_leaves
                    l_ledger.leave_date = datetime.strptime(request.POST['leave_date_id'], '%d/%m/%Y').strftime('%Y-%m-%d')
                    l_ledger.balance = balances + float(no_leaves)
                    l_ledger.remark = remark
                    l_ledger.save()
                    message  = "Leave Ledger Added Successfully"
                    response = {}
                    response['flag'] = True
                    response['message'] = message
                    response['emp_id'] = emp_id
                    return JsonResponse(response)
                else:
                    message  = "There is no leave policy for this Role."
                    response = {}
                    response['flag'] = False
                    response['message'] = message
                    response['emp_id'] = emp_id
                    return JsonResponse(response)   
        leave_policy_id = leavess.leave_policy_id
        leave_debits = leavess.debit
        leave_credit = leavess.credit
        leave_type_id = leave_types.leave_type_id
       
        month_leave_count = leavess.month_leave_count
        consecutive_leave = leavess.consecutive_leave
        balances = leavess.balance
        types =  request.POST['type']
        if types == "1":
            credit = no_leaves
        else:
            debit = no_leaves
        
        ledger = SpUserLeavePolicyLedger()
        ledger.user_id = emp_id
        ledger.leave_policy_id = leave_policy_id
        ledger.leave_type_id = leave_type_id
        
        ledger.month_leave_count = month_leave_count
        ledger.consecutive_leave = consecutive_leave
        if types == '1':
            ledger.credit = no_leaves
            
            ledger.balance = balances + float(credit)
        else:
            ledger.debit = no_leaves
            
            ledger.balance = balances - float(debit)
        ledger.leave_date = datetime.strptime(request.POST['leave_date_id'], '%d/%m/%Y').strftime('%Y-%m-%d')
        ledger.remark = remark
        ledger.save()
        message  = "Leave Ledger Added Successfully"
        response = {}
        response['flag'] = True
        response['message'] = message
        response['emp_id'] = emp_id
        return JsonResponse(response)

 
 
 
@login_required
def ajaxLeaveLedger(request,emp_id):
    if request.method == "POST":
        emp_id  = request.POST['emp_id']
        employee = SpUsers.objects.filter(id = emp_id)
        leave_ledger = SpUserLeavePolicyLedger.objects.filter(user_id = emp_id)
        for leave in leave_ledger:
            leave.leave_policy_name = getModelColumnById(SpLeavePolicies,leave.leave_policy_id,'leave_policy')
            leave.laave_type_name = getModelColumnById(SpLeaveTypes,leave.leave_type_id,'leave_type')
            leave.month_leave_counts = round(leave.month_leave_count,1)
            
    template = 'user-management/ajax-leave-ledger.html'
    context = {}
    context['leave_ledger'] = leave_ledger
    context['employee'] = employee
    context['emp_id'] = emp_id
    return render(request, template,context)

#update user location
@login_required
def clearUserLocation(request):

    response = {}
    if request.method == "GET":
        try:
            id = request.GET.get('id')
            
            data = SpUsers.objects.get(id=id)
            data.latitude = None
            data.longitude = None
            data.save()
                
            user_name   = str(request.user.first_name)+' '+str(request.user.middle_name)+' '+str(request.user.last_name)
            emp_name = ''
            if getModelColumnById(SpUsers, id, 'first_name'):
                emp_name += str(getModelColumnById(SpUsers, id, 'first_name'))
            if getModelColumnById(SpUsers, id, 'middle_name'):
                emp_name += ' '+str(getModelColumnById(SpUsers, id, 'middle_name'))    
            if getModelColumnById(SpUsers, id, 'last_name'):
                emp_name += ' '+str(getModelColumnById(SpUsers, id, 'last_name'))
            heading     = 'User Location Cleared'
            activity    = 'User Location of '+emp_name+' has been cleared by '+user_name+' at '+datetime.now().strftime('%d/%m/%Y | %I:%M %p') 

            saveActivity('Users Management', 'Users', heading, activity, request.user.id, user_name, 'add.png', '1', 'web.png')
            response['error'] = False
            response['message'] = "Record has been updated successfully."
        except ObjectDoesNotExist:
            response['error'] = True
            response['message'] = "Method not allowed"
        except Exception as e:
            response['error'] = True
            response['message'] = e
        return JsonResponse(response)
    return redirect('/users')

def privacyPolicy(request):
    context = {}
    context['page_title'] = "Privacy Policy"

    template = 'privacy-policy.html'
    return render(request, template, context)
    

@login_required
def add_employee_target(request, employee_id):
    if request.method == "POST":
        response = {}
        error_response = {}
        # try:
            # Validate user and request data
        user_exists = SpUsers.objects.filter(id=employee_id).exists()
        financial_year = request.POST.get('financial_year')
        currency_id = request.POST.get('currency_id',0)  

        months = request.POST.getlist('month[]')
        lead_targets = request.POST.getlist('lead_target[]')
        rev_targets = request.POST.getlist('rev_target[]')

        if not user_exists:
            error_response['message'] = "Invalid User"
        if not financial_year:
            error_response['message'] = "Please select a financial year"
        if not months:
            error_response['message'] = "Please enter target month"
        if FinancialYearData.objects.filter(user_id=employee_id,FY_id=int(financial_year)).exists():
            error_response['message'] = "Financial year already exists for this user"


        if error_response:
            response['error'] = True
            response['message'] = error_response['message']
            return JsonResponse(response)

        # Create FinancialYearData instance
       
        target_data = FinancialYearData.objects.create(user_id=employee_id,FY_id=int(financial_year),currency=currency_id)

        # Create FinancialMonthly instances for each month
        for month_year, lead_target, rev_target in zip(months, lead_targets, rev_targets):
            month, year = month_year.split('-')
            month = month.upper()[:3]  # Get first three characters of the month abbreviation
            
            FinancialMonthly.objects.create(
                fy=target_data,
                month=month,
                year=year,
                month_year=month_year,
                lead_target=lead_target,
                revenue_target=rev_target
            )

        response['error'] = False
        response['message'] ='Target Added Succesfully'

        return JsonResponse(response)

        # except Exception as e:
        #     response['error'] = True
        #     response['message'] = str(e)
        #     return JsonResponse(response)

    else: 
        context = {
            'years': SpFinancialYears.objects.all(),
            'employee_id': employee_id,
            'currency': SpCurrencyCode.objects.filter(status=1),
        }
        template = 'user-management/edit-employee/employee-target.html'
        return render(request, template, context)


@login_required
def edit_employee_target(request):
    if request.method == "POST":
        response = {}
        error_response = {}

        employee_id  = request.POST.get('employee_id')
        year_id     = request.POST.get('year_id')
        currency_id = request.POST.get('currency_id',0) 
        

        user_exists = SpUsers.objects.filter(id=employee_id).exists()


        months_ids   = request.POST.getlist('ids[]')
        lead_targets = request.POST.getlist('lead_target[]')
        rev_targets   = request.POST.getlist('rev_target[]')

        if not user_exists:
            error_response['message'] = "Invalid User"

        if error_response:
            response['error'] = True
            response['message'] = error_response['message']
            return JsonResponse(response)

        # Create FinancialYearData instance
        target_data = FinancialYearData.objects.get(user_id=employee_id,id=year_id)
        target_data.currency=currency_id
        target_data.save()

        for month_id in months_ids:
            monthlery = FinancialMonthly.objects.get(fy=target_data,id=month_id)
            monthlery.lead_target = lead_targets[months_ids.index(month_id)]
            monthlery.revenue_target = rev_targets[months_ids.index(month_id)]
            monthlery.save()


        response['error'] = False
        response['message'] = 'Updated Target Successfully'

        return JsonResponse(response)

        # except Exception as e:
        #     response['error'] = True
        #     response['message'] = str(e)
        #     return JsonResponse(response)

    else: 
        # Handle GET request
        employee_id = request.GET.get('employee_id')
        year_id = request.GET.get('year_id')

        # Retrieve the FinancialYearData instance or return 404 if not found
        year_instance = get_object_or_404(FinancialYearData, FY_id=year_id, user_id=employee_id)
        if year_instance.currency != 0:
            currency_obj = SpCurrencyCode.objects.get(id=year_instance.currency)
        else:
            currency_obj = None

        target_data = FinancialMonthly.objects.filter(fy=year_instance).order_by('id')
        context = {
            'year': year_instance,
            'employee_id': employee_id,
            'target_data': target_data,
            'currency_obj':currency_obj,
            'currency': SpCurrencyCode.objects.filter(status=1)
        }
        template = 'user-management/edit-employee/edit-employee-target.html'
        return render(request, template, context)

@login_required
def employee_peformance(request, employee_id):
    context = {}

    single = SpUsers.objects.filter(id=employee_id).first()
    
    fy_years = SpFinancialYears.objects.all().order_by('start_year')
    data = []
    for year in fy_years:
        year.start_year
        year.start_month

        year.end_year
        year.end_month

        if FinancialYearData.objects.filter(user=single,FY=year).exists():
            data.append({
                'id':year.id,
                'start_year': year.start_year,
                'end_year': year.end_year,
                'target_leads': FinancialMonthly.objects.filter(fy__user_id=employee_id, fy__FY_id=year.id).aggregate(Sum('lead_target'))['lead_target__sum'] or 0,
                'revenue_target': FinancialMonthly.objects.filter(fy__user_id=employee_id, fy__FY_id=year.id).aggregate(Sum('revenue_target'))['revenue_target__sum'] or 0,
                'acheive_revenue_target': SpLeadBasic.objects.filter(created_by_id=employee_id, deal_date_time__year=year.start_year).aggregate(Sum('deal_amount'))['deal_amount__sum'] or 0,
                'acheive_lead_target': SpLeadBasic.objects.filter(created_by_id=employee_id, created_at__year=year.start_year).count() or 0

            })

    context = {
        'datas': data,
        'employee_id':employee_id

    }    
    context['employee'] = employee = SpUsers.objects.raw('''SELECT sp_users.*,sp_basic_details.blood_group,sp_basic_details.date_of_birth,sp_basic_details.geofencing,
    sp_basic_details.gender,sp_basic_details.working_shift_name,sp_basic_details.leave_count,sp_basic_details.date_of_joining,sp_basic_details.date_of_joining,sp_basic_details.mother_name,sp_basic_details.father_name,sp_basic_details.aadhaar_nubmer,sp_basic_details.pan_number
     FROM sp_users left join sp_basic_details on sp_basic_details.user_id = sp_users.id
    where sp_users.id = %s''',[employee_id])[0]
    template = 'user-management/employee-performance.html'
    return render(request, template, context)



@login_required
def userSalarySlip(request):
    if request.method == "POST":
        
        user_id = request.POST['employee_id']
        employee_ctc = request.POST['emp_ctc']
        employee_monthly_ctc = request.POST['employee_monthly_ctc']
        ctc_currancy_code = request.POST['ctc_currancy_code']
        
        if request.POST.getlist('fixed_types[]') != []:
            fixed_type =  ','.join(request.POST.getlist('fixed_types[]'))
            per_value =  ','.join(request.POST.getlist('per_values[]'))
            value_on_ctc =  ','.join(request.POST.getlist('values_on_ctc[]'))
            currancy_code =  ','.join(request.POST.getlist('currancy_code[]'))
        else:
            fixed_type =  None
            per_value = None
            value_on_ctc =  None
            currancy_code = None
        if request.POST.getlist('addition_types[]') != []:
            addition_type =  ','.join(request.POST.getlist('addition_types[]'))
        else:
            addition_type = None
        if SpUserSalarySlip.objects.filter(user_id = user_id).exists():
            salary =  SpUserSalarySlip.objects.filter(user_id = user_id).first()
        else:
            salary =  SpUserSalarySlip()
        salary.user_id = user_id
        salary.ctc = employee_ctc
        salary.ctc_currency = ctc_currancy_code
        salary.fixed_pay_currency = currancy_code
        salary.monthly_ctc = employee_monthly_ctc
        salary.fixed_pay_type_ids = fixed_type
        salary.fixed_pay_per_val = per_value
        salary.fixed_pay_converted_val = value_on_ctc
        salary.additional_type = addition_type
        salary.status = 1
        salary.save()
        response = {}
        response['error'] = False
        response['message'] = 'Salary BreakDown Saved Successfully'
        return JsonResponse(response)



def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

import calendar
import itertools
# @login_required
# def downloadSalarySlipPdf(request,month_no,user_id,year):
#     user_id = int(user_id)
#     year = int(year)
#     month = int(month_no)
#     selected_month_name = calendar.month_name[month]
#     # -----------------------------------------------------------
#     organization_name = SpOrganizations.objects.filter().first()
#     role_id         = getModelColumnById(SpUsers,user_id,'role_id')
#     user_basic      = SpUsers.objects.filter(id = user_id).first()
#     address         = SpAddresses.objects.filter(user_id= user_id).first()
#     other_details   = SpBasicDetails.objects.filter(user_id = user_id).first()
#     bank_details    = SpBankDetails.objects.filter(user_id = user_id).first()
#     # ---------------------------------------------------------------
    
#     date = calendar.monthrange(year, month)[1]
#     # total days / paid days ** Start **
#     present_count = 0 
#     absent_count = 0
#     total_count = 0
#     lwp_count = 0

#     month_list = days_in_months(year,month)
#     nameMonth = []
#     for month_date in month_list:
#         month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')

#         if int(date) >= int(month_date.strftime('%d')):
#             nameMonth.append(month_date.strftime('%d %b %a'))

#     for month in month_list[:len(nameMonth)]:

#         total_count += 1
#         month_date = datetime.strptime(str(month), '%d/%m/%Y').strftime('%Y-%m-%d')
        
#         if checkAttendance(month_date, user_id):
#             present_count += 1
#         elif checkLeave(month_date, user_id):
#             absent_count = absent_count +1
#         elif checkHoliday(month_date,role_id):
#             present_count += 1
#         elif ceckWeekOfDay(month_date,user_id):
#             present_count += 1
#         else:
#             lwp_count += 1
    
#     total_days = total_count
#     paid_days = present_count
#     lwp = lwp_count
    
#     # -------------------------------------------------------------------------------
#     sp_user_salary_slip = SpUserSalarySlip.objects.filter(user_id = user_id).first()
#     if sp_user_salary_slip:
       
#         if lwp > 0:
#             get_single_day_ctc = round((sp_user_salary_slip.monthly_ctc/date),2)
#             new_monthly_ctc = sp_user_salary_slip.monthly_ctc - (get_single_day_ctc*lwp)
#         else:
#             new_monthly_ctc = sp_user_salary_slip.monthly_ctc

#         fixed_pay_ids_string = sp_user_salary_slip.fixed_pay_type_ids
#         fixed_pay_ids_list = [int(x) for x in fixed_pay_ids_string.split(",")]

#         fixed_pay_value_string = sp_user_salary_slip.fixed_pay_per_val
#         fixed_pay_val_list = [int(x) for x in fixed_pay_value_string.split(",")]

#         fixed_pay_converted_val_string = sp_user_salary_slip.fixed_pay_converted_val
#         fixed_pay_converted_val_list = [int(x) for x in fixed_pay_converted_val_string.split(",")]

#         fixed_pay_currency_string = sp_user_salary_slip.fixed_pay_currency
#         fixed_pay_currency_list = [int(x) for x in fixed_pay_currency_string.split(",")]

#         additional_pay_ids_string = sp_user_salary_slip.additional_type
#         additional_salary_list = []
#         if sp_user_salary_slip.additional_type:
            
#             additional_pay_ids_list = [int(x) for x in additional_pay_ids_string.split(",")]

#             for i in range(0,len(additional_pay_ids_list)):
#                 additional_salary_list_sub = {}
#                 additional_salary_list_sub['type'] = getModelColumnById(SpSalaryHead,additional_pay_ids_list[i],'salary_head_name')
#                 additional_salary_list.append(additional_salary_list_sub)
       
#         gross_fixed_pay = 0
#         gross_monthly_pay = 0 
#         fixed_salary_list = []

#         for i in range(0,len(fixed_pay_ids_list)):
#             pay_val = round(((new_monthly_ctc*fixed_pay_val_list[i])/100),2)
#             fixed_salary_list_sub = {}
#             fixed_salary_list_sub['type'] = getModelColumnById(SpSalaryHead,fixed_pay_ids_list[i],'salary_head_name')
#             fixed_salary_list_sub['fixed_total_pay'] = fixed_pay_converted_val_list[i]
#             gross_fixed_pay += fixed_pay_converted_val_list[i]
#             fixed_salary_list_sub['total_pay_for_month'] = pay_val
#             gross_monthly_pay += pay_val
#             fixed_salary_list_sub['currancy_type'] =  getModelColumnById(SpCurrencyCode,fixed_pay_currency_list[i],'currency_code')
#             fixed_salary_list.append(fixed_salary_list_sub)
        
#         # --------------------------------------------------------------------------
#         gross_fixed_total_amount = gross_fixed_pay
#         gross_monthly_total_amount = round((gross_monthly_pay),2)
#         gross_currency = getModelColumnById(SpCurrencyCode,sp_user_salary_slip.ctc_currency,'currency_code')
#         # ---------------------------------------------------------------------------------------
#         net_pay_amount = round(( new_monthly_ctc),2)
#         net_pay_amount_currencey =   gross_currency
#         # ---------------------------------------------------------------
#         baseurl = settings.BASE_URL

#         combined_list = list(itertools.zip_longest(fixed_salary_list, additional_salary_list, fillvalue=None))

#         context = {
#             'user_basic': user_basic,
#             'organization_name':organization_name,
#             'address': address,
#             'other_details': other_details,
#             'bank_details': bank_details,
#             'gross_fixed_total_amount': gross_fixed_total_amount,
#             'gross_monthly_total_amount': gross_monthly_total_amount,
#             'gross_currency': gross_currency,
#             'net_pay_amount': net_pay_amount,
#             'net_pay_amount_currency': net_pay_amount_currencey,
#             'combined_list': combined_list,
#             'total_days': total_days,
#             'paid_days': paid_days,
#             'lwp': lwp,
#             'nameMonth': selected_month_name,
#             'url': baseurl,
#             'year': year,
            
#         }

#         template                                = 'user-management/salary-sheet-pdf.html'
#         baseurl     = settings.BASE_URL

#         response = {}
#         filename = 'salary'+'_'+str(month_no)+'_'+str(year)
#         pdf = save_salary_slip_pdf(filename, template, context)

#         if pdf:
#             path = '/media/salary_pdf/'+filename+'.pdf'
#             invoice = SpSalarySlipPdf()
#             invoice.user_id          = int(user_id)
#             invoice.month           = int(month_no)
#             invoice.year            = int(year)
#             invoice.invoice_path     = path
            
#             invoice.status           = 0
#             invoice.save()
#             response['error'] = True
#             response['message'] = 'Salary Slip Genrated Successfully'
#             response['baseurl'] = baseurl
#             response['filename'] = filename
#     else:
#         response = {}
#         response['error'] = False
#         response['message'] = 'Salary slip not found for the given user, month, and year '    
#     return JsonResponse(response)

@login_required
def downloadSalarySlipPdf(request,month_no,user_id,year):
    user_id = int(user_id)
    year = int(year)
    month = int(month_no)
    selected_month_name = calendar.month_name[month]
    # -----------------------------------------------------------
    organization_name = SpOrganizations.objects.filter().first()
    role_id         = getModelColumnById(SpUsers,user_id,'role_id')
    user_basic      = SpUsers.objects.filter(id = user_id).first()
    address         = SpAddresses.objects.filter(user_id= user_id).first()
    other_details   = SpBasicDetails.objects.filter(user_id = user_id).first()
    bank_details    = SpBankDetails.objects.filter(user_id = user_id).first()
    # ---------------------------------------------------------------
    
    date = calendar.monthrange(year, month)[1]
    
    present_count = 0 
    absent_count = 0
    total_count = 0
    lwp_count = 0

    month_list = days_in_months(year,month)
    nameMonth = []
    for month_date in month_list:
        month_date                = datetime.strptime(str(month_date), '%d/%m/%Y')

        if int(date) >= int(month_date.strftime('%d')):
            nameMonth.append(month_date.strftime('%d %b %a'))

    for month in month_list[:len(nameMonth)]:

        total_count += 1
        month_date = datetime.strptime(str(month), '%d/%m/%Y').strftime('%Y-%m-%d')
        
        if checkAttendance(month_date, user_id):
            present_count += 1
        elif checkLeave(month_date, user_id):
            absent_count = absent_count +1
        elif checkHoliday(month_date,role_id):
            present_count += 1
        elif ceckWeekOfDay(month_date,user_id):
            present_count += 1
        else:
            lwp_count += 1
    
    total_days = total_count
    paid_days = present_count
    lwp = lwp_count
    
    # -------------------------------------------------------------------------------
    sp_user_salary_slip = SpUserSalarySlip.objects.filter(user_id = user_id).first()
    if sp_user_salary_slip:
       
        if lwp > 0:
            get_single_day_ctc = round((sp_user_salary_slip.monthly_ctc/date),2)
            new_monthly_ctc = sp_user_salary_slip.monthly_ctc - (get_single_day_ctc*lwp)
        else:
            new_monthly_ctc = sp_user_salary_slip.monthly_ctc

        fixed_pay_ids_string = sp_user_salary_slip.fixed_pay_type_ids
        fixed_pay_ids_list = [int(x) for x in fixed_pay_ids_string.split(",")]

        fixed_pay_value_string = sp_user_salary_slip.fixed_pay_per_val
        fixed_pay_val_list = [int(x) for x in fixed_pay_value_string.split(",")]

        fixed_pay_converted_val_string = sp_user_salary_slip.fixed_pay_converted_val
        fixed_pay_converted_val_list = [int(x) for x in fixed_pay_converted_val_string.split(",")]

        fixed_pay_currency_string = sp_user_salary_slip.fixed_pay_currency
        fixed_pay_currency_list = [int(x) for x in fixed_pay_currency_string.split(",")]

        additional_pay_ids_string = sp_user_salary_slip.additional_type
        additional_salary_list = []
        if sp_user_salary_slip.additional_type:
           
            additional_pay_ids_list = [int(x) for x in additional_pay_ids_string.split(",")]

            for i in range(0,len(additional_pay_ids_list)):
                additional_salary_list_sub = {}
                additional_salary_list_sub['type'] = getModelColumnById(SpSalaryHead,additional_pay_ids_list[i],'salary_head_name')
                additional_salary_list.append(additional_salary_list_sub)
       
        gross_fixed_pay = 0
        gross_monthly_pay = 0 
        fixed_salary_list = []

        for i in range(0,len(fixed_pay_ids_list)):
            pay_val = round(((new_monthly_ctc*fixed_pay_val_list[i])/100),2)
            fixed_salary_list_sub = {}
            fixed_salary_list_sub['type'] = getModelColumnById(SpSalaryHead,fixed_pay_ids_list[i],'salary_head_name')
            fixed_salary_list_sub['fixed_total_pay'] = fixed_pay_converted_val_list[i]
            gross_fixed_pay += fixed_pay_converted_val_list[i]
            fixed_salary_list_sub['total_pay_for_month'] = pay_val
            gross_monthly_pay += pay_val
            fixed_salary_list_sub['currancy_type'] =  getModelColumnById(SpCurrencyCode,fixed_pay_currency_list[i],'currency_code')
            fixed_salary_list.append(fixed_salary_list_sub)
        
        # --------------------------------------------------------------------------
        gross_fixed_total_amount = gross_fixed_pay
        gross_monthly_total_amount = round((gross_monthly_pay),2)
        gross_currency = getModelColumnById(SpCurrencyCode,sp_user_salary_slip.ctc_currency,'currency_code')
        # ---------------------------------------------------------------------------------------
        net_pay_amount = round(( new_monthly_ctc),2)
        net_pay_amount_currencey =   gross_currency
        # ---------------------------------------------------------------
        baseurl = settings.BASE_URL
        # if additional_salary_list:
        combined_list = list(itertools.zip_longest(fixed_salary_list, additional_salary_list, fillvalue=None))
        # else:
        #     combined_list = list(itertools.zip_longest(fixed_salary_list, fillvalue=None))

        context = {
            'user_basic': user_basic,
            'organization_name':organization_name,
            'address': address,
            'other_details': other_details,
            'bank_details': bank_details,
            'gross_fixed_total_amount': gross_fixed_total_amount,
            'gross_monthly_total_amount': gross_monthly_total_amount,
            'gross_currency': gross_currency,
            'net_pay_amount': net_pay_amount,
            'net_pay_amount_currency': net_pay_amount_currencey,
            'combined_list': combined_list,
            'total_days': total_days,
            'paid_days': paid_days,
            'lwp': lwp,
            'nameMonth': selected_month_name,
            'url': baseurl,
            'year': year,
            
        }

        template                                = 'user-management/salary-sheet-pdf.html'
        baseurl     = settings.BASE_URL

        response = {}
        filename = 'salary'+'_'+str(user_id)+'_'+str(month_no)+'_'+str(year)
        pdf = save_salary_slip_pdf(filename, template, context)

        if pdf:
            path = '/media/salary_pdf/'+filename+'.pdf'
            invoice = SpSalarySlipPdf()
            invoice.user_id          = int(user_id)
            invoice.month           = int(month_no)
            invoice.year            = int(year)
            invoice.invoice_path     = path
            
            invoice.status           = 0
            invoice.save()
            response['error'] = True
            response['message'] = 'Salary Slip Genrated Successfully'
            response['baseurl'] = baseurl
            response['filename'] = filename
    else:
        response = {}
        response['error'] = False
        response['message'] = 'Salary slip not found for the given user, month, and year '    
    return JsonResponse(response)

# @login_required
# def save_salary_slip_pdf(file_name, template_src, context_dict={}):
    
#     template = get_template(template_src)
#     html_content = template.render(context_dict)
#     response = BytesIO()
#     pdf = pisa.pisaDocument(BytesIO(html_content.encode('UTF-8')), response)

#     if pdf.err:

#         return '', False
    
#     pdf_file_path = os.path.join(settings.BASE_DIR, 'media', 'salary_pdf', f'{file_name}.pdf')
    
#     os.makedirs(os.path.dirname(pdf_file_path), exist_ok=True)
    
#     try:
#         with open(pdf_file_path, 'wb') as pdf_file:
#             pdf_file.write(response.getvalue())
        
#         return file_name, True
#     except Exception as e:
#         # Print the exception for debugging
#         print(f"Error saving PDF: {e}")
        
#         return '', False

def save_salary_slip_pdf(file_name, template_src, context_dict={}):
    
    template = get_template(template_src)
    html_content = template.render(context_dict)
    response = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_content.encode('UTF-8')), response)

    if pdf.err:

        return '', False
    
    pdf_file_path = os.path.join(settings.BASE_DIR, 'media', 'salary_pdf', f'{file_name}.pdf')
    
    os.makedirs(os.path.dirname(pdf_file_path), exist_ok=True)
    
    try:
        with open(pdf_file_path, 'wb') as pdf_file:
            pdf_file.write(response.getvalue())
        
        return file_name, True
    except Exception as e:
        # Print the exception for debugging
        print(f"Error saving PDF: {e}")
        
        return '', False
@login_required
def printSalarySlip(request, month_no, user_id, year):
    try:

        salary_slip = SpSalarySlipPdf.objects.get(user_id=int(user_id), month=int(month_no), year=int(year))
        
        baseurl = settings.BASE_URL
        pdf_path = salary_slip.invoice_path

        # if not os.path.isfile(pdf_path):
        #     raise FileNotFoundError(f"PDF file not found at path: {pdf_path}")
        
        # with open(pdf_path, 'rb') as file:
        #     pdf_content = file.read()
        
        # response = HttpResponse(pdf_content, content_type='application/pdf')

        # response['Content-Disposition'] = f'attachment; filename="salary-slip-{user_id}-{month_no}-{year}.pdf"'
        if salary_slip:
            pdfpath = baseurl+''+salary_slip.invoice_path
        return pdfpath
    
    except SpSalarySlipPdf.DoesNotExist:
        # Handle the case where the record does not exist
        raise Http404("Salary slip not found for the given user, month, and year")
    except FileNotFoundError as e:
        # Handle the case where the PDF file is not found
        raise Http404(str(e))
    except Exception as e:
        # Handle other potential exceptions
        return HttpResponse(str(e), status=500)
#done
def checkLeave(date, user_id):
    try:
        leaves = SpUserLeaves.objects.filter(leave_from_date__lte=date, leave_to_date__gte=date,leave_status=3,user_id=user_id).first()
    except SpUserLeaves.DoesNotExist:
        leaves = None
    if leaves:
        return getModelColumnById(SpLeaveTypes, leaves.leave_type_id, 'alias')
    else:
        return False
#done
@login_required
def salarySlipList(request):
    user_list = SpUsers.objects.filter(status = 1).exclude(id = 1)
    current_year = datetime.now().year
    current_datetime = datetime.now()

    # Get the current year
    present_year = current_datetime.year
    # Create a list of the last 15 years
    last_15_years = [current_year - i for i in range(15)]
    context = {
        'user_lists':user_list,
        'last_15_years':last_15_years,
        'present_year':present_year
    }
    template = 'user-management/salary-slip/salary-slip.html'
    return render(request, template, context)
@login_required
def getMonthList(request):
    user_id = int(request.GET['user_id'])
    year = int(request.GET['year'])
    salary_slip_month_list = SpSalarySlipPdf.objects.filter(user_id=int(user_id), year=int(year)).values_list('month','invoice_path')
    salary_slip_dict = dict(salary_slip_month_list)
    current_year = datetime.now().year
    current_month = datetime.now().month

    if year == current_year:
        end_month = current_month - 1
    else:
        end_month = 12
    
    months = []
    for month in range(1, end_month + 1):
       
        date = datetime(year, month, 1)
        if month in salary_slip_dict:
            month_generated_status = 1
            invoice_path = salary_slip_dict[month]  # Get the invoice path
        else:
            month_generated_status = 0
            invoice_path = None
    
        
        months.append((month, date.strftime('%B'), month_generated_status,invoice_path))
    context = {
        'months':months
    }
    template = 'user-management/salary-slip/ajax-salary-slip.html'
    return render(request, template, context)
